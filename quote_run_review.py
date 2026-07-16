"""A throwaway Quote-Run review workbook + a note queue Claude works through.

The quote-run twin of `so_review.py`: give a human a place to comment on how
each quote run was captured — every extracted field, every vision complaint,
every uncaptured document line — without touching the live product. The
progress store stays read-only; this builds a SEPARATE, disposable .xlsx you
can scribble in. It exists only as a path to a better parser - it is not part
of the final workbook.

The Quote Runs tab is the canonical row grid. Every run unrolls into real cell
rows, so Excel's normal column filters search the actual data and an edited
Note stays attached to that row:

  RUN      the run file itself (status, template, field count) — linked to Z:
  FIELD    one extracted field per row: name + the value we pulled
  SUSPECT  a vision-QC complaint on a scanned PDF (CHECK VISION / NEEDS HUMAN)
  MISSED   a document line that carries data no pattern captured yet

The loop:
  1. `python quote_run_review.py build` -> writes quote_run_review.xlsx from the
     quote-run scan store: every order's runs, one row per item, with an
     editable "Note" cell containing any open note already recorded.
  2. Filter Quote Runs to the orders/fields you want, type in Note, then
     save and close. The next refresh folds those entries into
     quote_run_review_notes.json for publication (data_push).
  3. `python quote_run_review.py list` shows the OPEN notes. Claude acts on each
     (usually a pattern fix + regression test) and
     `python quote_run_review.py handle <id> "what I did"` marks it handled.
     The next refresh removes it from the active review row so that row is
     ready for another note. Resolved notes remain in the JSON queue and the
     workbook's Resolved tab as history.

Notes may anchor to any displayed row. Field rows anchor by their stable field
name; suspect and missed rows anchor by their text. That text is retained as
context if a later re-parse removes the row (a captured line stops being
MISSED — which is the goal).

Pure logic (store + rows) is import-light and unit-tested; the Excel functions
and the scan-store loader lazy-import their dependencies so the rest of the
module loads without them.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import BACKLOG_DIR
from run_rank import dedupe_runs

REVIEW_STORE_PATH = BACKLOG_DIR / "quote_run_review_notes.json"
DEFAULT_WORKBOOK = BACKLOG_DIR / "quote_run_review.xlsx"

# Return channel for Claude's handled-marks, exactly like so_review's: the note
# queue flows UP to Claude with the published stores (data_push), and Claude
# records each note it resolves in this small TRACKED ledger at the repo root.
# It rides down to the user's machine on the normal Git Update, and "Update QR
# Review" applies it. Append-only {handled: [{id, resolution, handled_at}]}.
HANDLED_MARKS_PATH = Path(__file__).resolve().parent / "quote_run_review_handled.json"

STATUS_OPEN = "open"
STATUS_HANDLED = "handled"
NOTE_SEPARATOR = "\n\n---\n\n"

KIND_RUN = "RUN"
KIND_FIELD = "FIELD"
KIND_SUSPECT = "SUSPECT"
KIND_MISSED = "MISSED"

# Workbook columns (also the read-back contract). ``Note`` renders open notes
# and accepts new human input. ``Add Note`` is recognized only when migrating an
# older workbook.
ADD_NOTE = "Add Note"
HEADERS = ["Order", "Run", "Kind", "Item", "Value", "Note",
           "Status", "Resolution"]
_COL = {h: i for i, h in enumerate(HEADERS)}

REVIEW_SHEET = "Quote Runs"
RESOLVED_SHEET = "Resolved"
RESOLVED_HEADERS = ["Order", "Run", "Item", "Note", "Resolution", "Resolved"]
ROW_KEY_HEADER = "_row_key"


# --------------------------------------------------------------------------- #
# Note queue store                                                             #
# --------------------------------------------------------------------------- #
def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def load_store(path: Optional[Path] = None) -> Dict[str, Any]:
    p = Path(path or REVIEW_STORE_PATH)
    if not p.exists():
        return {"notes": []}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"notes": []}
    if not isinstance(data, dict) or not isinstance(data.get("notes"), list):
        return {"notes": []}
    return data


def save_store(store: Dict[str, Any], path: Optional[Path] = None) -> None:
    p = Path(path or REVIEW_STORE_PATH)
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(store, indent=2, default=str), encoding="utf-8")
    tmp.replace(p)


def _next_id(store: Dict[str, Any]) -> int:
    return max((int(n.get("id", 0)) for n in store["notes"]), default=0) + 1


def record_note(store: Dict[str, Any], order: str, run: str, item: str,
                item_text: str, note: str, when: Optional[str] = None,
                row_key: str = "") -> Optional[Dict[str, Any]]:
    """Append a note for one review row. Idempotent on exact text: if a note
    with the same anchor + text already exists (any status), nothing is added
    and None is returned — so re-syncing the same workbook never duplicates."""
    order, note = str(order).strip(), str(note).strip()
    run, item = str(run).strip(), str(item).strip()
    target_text = str(item_text or "").strip()
    target_key = str(row_key or "").strip()
    if not order or not note:
        return None
    for n in store["notes"]:
        existing_key = str(n.get("row_key", "")).strip()
        if target_key and existing_key:
            same_target = existing_key == target_key
        else:
            same_target = (
                str(n.get("run", "")).strip().casefold() == run.casefold()
                and str(n.get("item", "")).strip().casefold() == item.casefold()
                and str(n.get("item_text", "")).strip().casefold()
                == target_text.casefold()
            )
        if (str(n.get("order")) == order and same_target
                and str(n.get("note")) == note):
            return None
    entry = {"id": _next_id(store), "order": order, "run": run, "item": item,
             "item_text": target_text, "note": note,
             "status": STATUS_OPEN, "created_at": when or _now(),
             "handled_at": None, "resolution": None}
    if target_key:
        entry["row_key"] = target_key
    store["notes"].append(entry)
    return entry


def mark_handled(store: Dict[str, Any], note_id: int, resolution: str,
                 when: Optional[str] = None) -> bool:
    """Mark one note handled with what was done. Claude calls this after acting."""
    for n in store["notes"]:
        if int(n.get("id", -1)) == int(note_id):
            n["status"] = STATUS_HANDLED
            n["resolution"] = str(resolution or "").strip()
            n["handled_at"] = when or _now()
            return True
    return False


def open_notes(store: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [n for n in store["notes"] if n.get("status") != STATUS_HANDLED]


# --------------------------------------------------------------------------- #
# Handled-marks ledger (Claude -> user return channel)                         #
# --------------------------------------------------------------------------- #
def _load_ledger() -> Dict[str, Any]:
    if not HANDLED_MARKS_PATH.exists():
        return {"handled": []}
    try:
        data = json.loads(HANDLED_MARKS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"handled": []}
    return data if isinstance(data.get("handled"), list) else {"handled": []}


def record_handled_mark(note_id: int, resolution: str, when: Optional[str] = None) -> None:
    """Append (or update) a handled-mark in the tracked ledger, so it travels to
    the user's machine on the next Git Update."""
    led = _load_ledger()
    led["handled"] = [m for m in led["handled"] if int(m.get("id", -1)) != int(note_id)]
    led["handled"].append({"id": int(note_id), "resolution": str(resolution or "").strip(),
                           "handled_at": when or _now()})
    HANDLED_MARKS_PATH.write_text(json.dumps(led, indent=2), encoding="utf-8")


def apply_handled_marks(store: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Fold the ledger's handled-marks into a local queue: any note whose id is
    marked (and isn't already handled) becomes handled with the recorded
    resolution. Returns the notes newly closed this call (for reporting)."""
    marks = {int(m["id"]): m for m in _load_ledger()["handled"] if "id" in m}
    closed = []
    for n in store["notes"]:
        m = marks.get(int(n.get("id", -1)))
        if m and n.get("status") != STATUS_HANDLED:
            n["status"] = STATUS_HANDLED
            n["resolution"] = m.get("resolution", "")
            n["handled_at"] = m.get("handled_at") or _now()
            closed.append(n)
    return closed


# --------------------------------------------------------------------------- #
# Rows: every order's runs unrolled + the note recorded for each item          #
# --------------------------------------------------------------------------- #
def _job_sort_key(job: str) -> tuple:
    return (0, int(job), job) if str(job).isdigit() else (1, 0, str(job))


def _base_key(run_file: str, kind: str, item: str, value: str) -> str:
    """The stable anchor for one row. FIELD rows key on the field NAME (the
    value is exactly what a re-parse changes — notes must survive that);
    SUSPECT/MISSED rows key on their text (stable while the problem exists)."""
    run_part = f"run:{str(run_file).strip().casefold()}"
    kind = str(kind).strip().upper()
    if kind == KIND_FIELD:
        return f"{run_part}|field:{str(item).strip().casefold()}"
    if kind == KIND_SUSPECT:
        return f"{run_part}|suspect:{str(value).strip().casefold()}"
    if kind == KIND_MISSED:
        return f"{run_part}|missed:{str(value).strip().casefold()}"
    return run_part


def _row_keys(rows: List[Dict[str, Any]]) -> List[str]:
    """Stable keys for display rows ({order, run, kind, item, value} dicts).
    Duplicate anchors within one order get a count suffix; counts reset per
    order so keys never depend on other orders' rows."""
    counts: Dict[str, int] = {}
    keys: List[str] = []
    current_order = ""
    for row in rows:
        order = str(row.get("order", "")).strip()
        if order and order != current_order:
            current_order = order
            counts = {}
        base = _base_key(row.get("run", ""), row.get("kind", ""),
                         row.get("item", ""), row.get("value", ""))
        counts[base] = counts.get(base, 0) + 1
        keys.append(base if counts[base] == 1 else f"{base}|{counts[base]}")
    return keys


def _order_notes(review_store: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """order# -> open notes for the active review views, in queue order."""
    out: Dict[str, List[Dict[str, Any]]] = {}
    for n in review_store["notes"]:
        if n.get("status") != STATUS_HANDLED:
            out.setdefault(str(n.get("order")), []).append(n)
    return out


def _ordered_fields(fields: Dict[str, Any],
                    field_order: Optional[List[str]]) -> List[str]:
    """Field names in review order: the workbook's CORE_FIELDS order first,
    then whatever else the parse pulled, in parse order."""
    if not field_order:
        return list(fields)
    known = [k for k in field_order if k in fields]
    return known + [k for k in fields if k not in field_order]


def _run_summary(run: Dict[str, Any]) -> str:
    """The RUN header row's Value: health at a glance."""
    fields = run.get("fields") or {}
    parts = [run.get("status", "") or "?", run.get("template", "") or "?",
             f"{len(fields)} fields"]
    tags = run.get("coverage_tags") or []
    if tags:
        parts.append("review: " + ", ".join(tags))
    if run.get("damper"):
        parts.append("DAMPER")
    return " — ".join(p for p in parts if p)


def _suspect_texts(run: Dict[str, Any]) -> List[str]:
    """Vision-QC complaints worth a human ruling. A NEEDS HUMAN run's terminal
    reason already summarizes the suspects, so it replaces them."""
    vision = run.get("vision") or {}
    reason = str(vision.get("human_reason") or "").strip()
    if reason:
        return [reason]
    return [str(s).strip() for s in vision.get("suspect") or [] if str(s).strip()]


def review_rows(records: Dict[str, Dict[str, Any]],
                review_store: Dict[str, Any],
                field_order: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """One display row per run item across every order, newest job first, with
    open notes shown on their lines. Handled notes are intentionally absent so
    the line is ready for another note; their history is retained separately.
    Runs are deduped like quote_runs.xlsx (format dupes / superseded CO-REV
    copies collapse to the active run). A note anchors by its stored row key,
    falling back to exact item text, then the order's first row.
    `group_start` marks the first row of each order. Pure — no Excel."""
    by_order = _order_notes(review_store)
    rows: List[Dict[str, Any]] = []
    orders = sorted((r.get("job", "") for r in records.values() if r.get("runs")),
                    key=_job_sort_key, reverse=True)
    recs = {str(r.get("job", "")): r for r in records.values()}
    for jn in orders:
        rec = recs[str(jn)]
        order_rows: List[Dict[str, Any]] = []
        for run in dedupe_runs(rec.get("runs", [])):
            file = str(run.get("file", ""))
            order_rows.append({"order": str(jn), "run": file, "kind": KIND_RUN,
                               "item": file, "value": _run_summary(run),
                               "path": str(run.get("path", ""))})
            for text in _suspect_texts(run):
                order_rows.append({"order": str(jn), "run": file,
                                   "kind": KIND_SUSPECT,
                                   "item": "(vision suspect)", "value": text})
            fields = run.get("fields") or {}
            for name in _ordered_fields(fields, field_order):
                order_rows.append({"order": str(jn), "run": file,
                                   "kind": KIND_FIELD,
                                   "item": name, "value": fields.get(name, "")})
            for line in run.get("missed_data") or []:
                order_rows.append({"order": str(jn), "run": file,
                                   "kind": KIND_MISSED,
                                   "item": "(uncaptured line)", "value": line})
        if not order_rows:
            continue
        keys = _row_keys(order_rows)
        key_to_idx = {key: i for i, key in enumerate(keys)}
        text_to_idx: Dict[str, int] = {}
        for i, r in enumerate(order_rows):
            for candidate in (str(r["item"]).strip().casefold(),
                              str(r["value"]).strip().casefold()):
                if candidate:
                    text_to_idx.setdefault(candidate, i)
        per_row: Dict[int, List[Dict[str, Any]]] = {}
        for n in by_order.get(str(jn), []):
            stored_key = str(n.get("row_key", "")).strip()
            if stored_key and stored_key in key_to_idx:
                idx = key_to_idx[stored_key]
            else:
                text = str(n.get("item_text", "")).strip().casefold()
                item = str(n.get("item", "")).strip().casefold()
                idx = text_to_idx.get(text, text_to_idx.get(item, 0))
            per_row.setdefault(idx, []).append(n)
        for i, r in enumerate(order_rows):
            here = per_row.get(i, [])
            note_txt = NOTE_SEPARATOR.join(
                f"#{n.get('id')}: {text}" for n in here
                if (text := str(n.get("note", "")).strip()))
            rows.append({**r,
                         "note": note_txt,
                         "status": STATUS_OPEN if here else "",
                         "resolution": "",
                         "row_key": keys[i],
                         "group_start": i == 0})
    return rows


def ingest_edits(review_store: Dict[str, Any],
                 edits: List[Dict[str, Any]], when: Optional[str] = None) -> int:
    """Fold rows read back from the workbook into the queue. Each edit is
    {order, run, item, item_text, note, row_key}. record_note dedups exact
    repeats. Returns how many NEW notes were added."""
    added = 0
    for e in edits:
        if record_note(review_store, e.get("order", ""), e.get("run", ""),
                       e.get("item", ""), e.get("item_text", ""),
                       e.get("note", ""), when=when,
                       row_key=e.get("row_key", "")):
            added += 1
    return added


# --------------------------------------------------------------------------- #
# Excel I/O (lazy openpyxl)                                                     #
# --------------------------------------------------------------------------- #
def _append_text_row(ws, values: List[Any]) -> None:
    """ws.append, but formula-proof. Real document lines can start with ``=``
    (price sums like ``=80240-6773-6056=67,411``), which openpyxl stores as an
    Excel FORMULA — a broken one, so Excel "repairs" the sheet by deleting the
    cell on open. Re-type any such cell as plain text."""
    ws.append(values)
    for cell in ws[ws.max_row]:
        if cell.data_type == "f":
            cell.data_type = "s"


def write_workbook(path: Path, records: Dict[str, Dict[str, Any]],
                   review_store: Dict[str, Any],
                   field_order: Optional[List[str]] = None) -> int:
    """Write the review workbook. Returns the data-row count.

    Colour and grouping are applied as a handful of workbook-level conditional-
    format rules (O(1) each), never styled cell-by-cell, so the large corpus
    does not pay a separate formatting cost for every row. The one exception is
    the RUN rows' hyperlink to the file on Z: (~1 cell per run)."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")      # invites typing
    band_fill = PatternFill("solid", fgColor="F2F2F2")      # alternating orders
    run_fill = PatternFill("solid", fgColor="DDEBF7")       # run header rows
    run_font = Font(bold=True)
    suspect_font = Font(color="C00000", bold=True)          # vision complaints
    missed_fill = PatternFill("solid", fgColor="FFEB9C")    # uncaptured lines
    link_font = Font(color="0563C1", underline="single")

    rows = review_rows(records, review_store, field_order=field_order)
    ncols = len(HEADERS)
    wb = Workbook()

    ws = wb.active
    ws.title = REVIEW_SHEET
    band_col = ncols + 1                                    # hidden parity helper
    row_key_col = ncols + 2
    ws.append(HEADERS + ["_band", ROW_KEY_HEADER])
    parity = 0
    for r in rows:
        if r["group_start"]:
            parity ^= 1                                     # flip per order group
        _append_text_row(ws, [r["order"], r["run"], r["kind"], r["item"],
                              _excel_text(r["value"]), r["note"], r["status"],
                              r["resolution"], parity, r["row_key"]])
        if r["kind"] == KIND_RUN and r.get("path"):
            cell = ws.cell(ws.max_row, _COL["Item"] + 1)
            cell.hyperlink = r["path"]
            cell.font = link_font
    last = ws.max_row
    for c in range(1, ncols + 1):
        ws.cell(1, c).fill = header_fill
        ws.cell(1, c).font = header_font
        ws.cell(1, c).alignment = Alignment(vertical="center")
    ws.cell(1, _COL["Note"] + 1).fill = note_fill
    ws.cell(1, _COL["Note"] + 1).font = Font(color="7F6000", bold=True)
    ws.row_dimensions[1].height = 24
    ws.column_dimensions[get_column_letter(band_col)].hidden = True
    ws.column_dimensions[get_column_letter(row_key_col)].hidden = True

    if last >= 2:
        full = f"A2:{get_column_letter(ncols)}{last}"
        bandL, kindL = get_column_letter(band_col), get_column_letter(_COL["Kind"] + 1)
        noteL = get_column_letter(_COL["Note"] + 1)
        itemL = get_column_letter(_COL["Item"] + 1)
        # Order banding (fill only) + Kind cues coexist: each rule sets only
        # the properties it needs, so they layer without fighting.
        ws.conditional_formatting.add(full, FormulaRule(formula=[f"${bandL}2=1"], fill=band_fill))
        ws.conditional_formatting.add(full, FormulaRule(formula=[f'${kindL}2="{KIND_RUN}"'], fill=run_fill, font=run_font))
        ws.conditional_formatting.add(full, FormulaRule(formula=[f'${kindL}2="{KIND_SUSPECT}"'], font=suspect_font))
        ws.conditional_formatting.add(full, FormulaRule(formula=[f'${kindL}2="{KIND_MISSED}"'], fill=missed_fill))
        # Every displayed row can take a note. Resolved notes leave this grid.
        ws.conditional_formatting.add(
            f"{noteL}2:{noteL}{last}",
            FormulaRule(formula=[f'${itemL}2<>""'], fill=note_fill),
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last}"
    ws.sheet_view.showGridLines = False
    for h, w in {"Order": 10, "Run": 30, "Kind": 10, "Item": 28, "Value": 55,
                 "Note": 50, "Status": 9, "Resolution": 55}.items():
        ws.column_dimensions[get_column_letter(_COL[h] + 1)].width = w

    # --- Resolved tab: compact audit history, separate from active inputs ----
    resolved = wb.create_sheet(RESOLVED_SHEET)
    resolved.append(RESOLVED_HEADERS)
    handled = sorted(
        (n for n in review_store["notes"] if n.get("status") == STATUS_HANDLED),
        key=lambda n: int(n.get("id", 0)),
        reverse=True,
    )
    for n in handled:
        _append_text_row(resolved, [
            str(n.get("order", "")), str(n.get("run", "")),
            str(n.get("item", "") or n.get("item_text", "")), str(n.get("note", "")),
            str(n.get("resolution", "") or ""), str(n.get("handled_at", "") or ""),
        ])
    for c in range(1, len(RESOLVED_HEADERS) + 1):
        resolved.cell(1, c).fill = header_fill
        resolved.cell(1, c).font = header_font
    resolved.freeze_panes = "A2"
    resolved.auto_filter.ref = f"A1:F{max(resolved.max_row, 1)}"
    for c, width in enumerate((10, 30, 28, 50, 60, 20), start=1):
        resolved.column_dimensions[get_column_letter(c)].width = width

    wb.active = 0

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(path))
    except PermissionError:
        raise RuntimeError(
            f"Could not write {path.name} — it looks like it's still open in "
            f"Excel. Close it, then run this again.") from None
    return len(rows)


def _excel_text(value: Any) -> str:
    """Normalize common Excel scalar values without turning 2 into '2.0'."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _note_parts(value: Any) -> List[str]:
    """Split a cell containing several rendered notes back into queue entries."""
    return [part.strip() for part in _excel_text(value).split(NOTE_SEPARATOR) if part.strip()]


def _manual_note_parts(value: Any) -> List[str]:
    """Return human-entered text while ignoring rendered ``#id: note`` history."""
    return [part for part in _note_parts(value)
            if not re.match(r"^#\d+\s*:\s*", part)]


def _review_sheet(wb):
    """The canonical row grid: the Quote Runs tab, or any sheet shaped like it."""
    if REVIEW_SHEET in wb.sheetnames:
        return wb[REVIEW_SHEET]
    for ws in wb.worksheets:
        header = {_excel_text(c.value) for c in ws[1]}
        if {"Order", "Run", "Note"} <= header:
            return ws
    return wb.active


def _layout_needs_upgrade(path: Path) -> bool:
    """Whether this is not yet the direct-Note-input Quote Runs layout."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        wb = load_workbook(str(path), read_only=True, data_only=False)
    try:
        if REVIEW_SHEET not in wb.sheetnames:
            return True
        header = [_excel_text(c.value) for c in wb[REVIEW_SHEET][1]]
        return not all(h in header for h in HEADERS) or ADD_NOTE in header
    finally:
        wb.close()


def read_edits(path: Path) -> List[Dict[str, Any]]:
    """Read edited Note cells (plus legacy Add Note inputs when migrating an
    older workbook), ignoring the rendered ``#id:`` note history."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        wb = load_workbook(str(path), data_only=False)
    try:
        ws = _review_sheet(wb)
        header = [_excel_text(c.value) for c in ws[1]]
        readable_headers = HEADERS + [ADD_NOTE]
        idx = {h: header.index(h) for h in readable_headers if h in header}
        row_key_index = header.index(ROW_KEY_HEADER) if ROW_KEY_HEADER in header else None

        def cell(row: tuple, heading: str) -> str:
            i = idx.get(heading)
            return "" if i is None or i >= len(row) else _excel_text(row[i])

        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
        derived_keys = _row_keys([{
            "order": cell(row, "Order"), "run": cell(row, "Run"),
            "kind": cell(row, "Kind"), "item": cell(row, "Item"),
            "value": cell(row, "Value"),
        } for row in raw_rows])

        edits: List[Dict[str, Any]] = []
        for row, derived_key in zip(raw_rows, derived_keys):
            order = cell(row, "Order")
            if not order:
                continue
            stored_key = ""
            if row_key_index is not None and row_key_index < len(row):
                stored_key = _excel_text(row[row_key_index])
            typed = _manual_note_parts(cell(row, "Note"))
            if ADD_NOTE in header:
                typed.extend(_note_parts(cell(row, ADD_NOTE)))
            for note in typed:
                edits.append({
                    "order": order,
                    "run": cell(row, "Run"),
                    "item": cell(row, "Item"),
                    "item_text": cell(row, "Value"),
                    "note": note,
                    "row_key": stored_key or derived_key,
                })
        return edits
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# CLI                                                                          #
# --------------------------------------------------------------------------- #
def _load_runs() -> Dict[str, Any]:
    import quote_run_scan
    return quote_run_scan.load_progress()


def _core_fields() -> List[str]:
    import quote_run_scan
    return quote_run_scan.CORE_FIELDS


def _cmd_build(args) -> int:
    store = load_store()
    n = write_workbook(Path(args.out), _load_runs(), store, field_order=_core_fields())
    pend = len(open_notes(store))
    print(f"Wrote {args.out} ({n} rows). {pend} open note(s) shown; resolved history "
          f"is on the {RESOLVED_SHEET} tab. Filter the rows on {REVIEW_SHEET} and "
          f"type directly in the yellow Note cells.")
    return 0


def _cmd_open(args) -> int:
    """Open the review workbook in Excel, building it first if it's missing."""
    import os
    path = Path(args.out)
    if not path.exists():
        write_workbook(path, _load_runs(), load_store(), field_order=_core_fields())
        print(f"Built {path}.")
    try:
        os.startfile(str(path))          # Windows: opens in Excel  # type: ignore[attr-defined]
    except AttributeError:               # non-Windows: best-effort
        import subprocess
        subprocess.Popen(["xdg-open", str(path)])
    print(f"Opened {path}.")
    return 0


def _cmd_sync(args) -> int:
    path = Path(args.out)
    if not path.exists():
        print(f"{path} not found — 'Open QR Review' builds it first.", file=sys.stderr)
        return 1
    needs_upgrade = _layout_needs_upgrade(path)
    edits = read_edits(path)
    store = load_store()
    added = ingest_edits(store, edits)
    save_store(store)
    if edits or needs_upgrade:
        try:
            write_workbook(path, _load_runs(), store, field_order=_core_fields())
        except RuntimeError as exc:
            raise RuntimeError(
                f"Recorded {added} new note(s), but could not refresh the Quote Run "
                f"review workbook. {exc} Re-run refresh after closing Excel; "
                f"the notes will not duplicate."
            ) from None
    cleared = f" Captured {len(edits)} edited Note cell(s)." if edits else ""
    upgraded = " Upgraded Quote Runs to the direct-Note layout." if needs_upgrade else ""
    print(f"Recorded {added} new note(s).{cleared}{upgraded} {len(open_notes(store))} open in the "
          f"queue ({REVIEW_STORE_PATH}).")
    return 0


def _cmd_refresh(args) -> int:
    """The 'Update QR Review' action: capture anything typed, fold in Claude's
    handled-marks, and rewrite active rows plus the separate resolved history."""
    path = Path(args.out)
    store = load_store()
    edits: List[Dict[str, Any]] = []
    added = 0
    if path.exists():
        edits = read_edits(path)
        added = ingest_edits(store, edits)              # don't lose un-synced typing
    workbook_note_count = len({
        (str(e.get("order", "")), str(e.get("row_key", "")), str(e.get("note", "")))
        for e in edits
        if str(e.get("order", "")).strip() and str(e.get("note", "")).strip()
    })
    closed = apply_handled_marks(store)                 # Claude's resolutions
    save_store(store)
    try:
        n = write_workbook(path, _load_runs(), store, field_order=_core_fields())
    except RuntimeError as exc:
        if workbook_note_count:
            safe = (
                f"All {workbook_note_count} workbook note(s) are safely recorded in "
                f"{REVIEW_STORE_PATH} ({added} new this run). "
            )
        else:
            safe = f"The note queue was safely saved to {REVIEW_STORE_PATH}. "
        raise RuntimeError(
            f"{safe}The workbook refresh could not finish. {exc}"
        ) from None
    already = ""
    if workbook_note_count and added == 0:
        already = f"; all {workbook_note_count} workbook note(s) were already recorded"
    print(f"Updated {path} ({n} rows). Captured {added} new note(s){already}; "
          f"applied {len(closed)} handled resolution(s); {len(open_notes(store))} still open.")
    for c in closed:
        print(f"  handled #{c['id']} (order {c['order']} {c.get('item', '')}): "
              f"{c.get('resolution', '')}")
    return 0


def _cmd_reparse(args) -> int:
    """The 'Re-parse + Refresh QR Review' action: apply the current parser to
    the WHOLE stored corpus (raw lines + vision transcripts — no Z:, no API,
    seconds), which also refreshes quote_runs.xlsx and the master, then rebuild
    the review sheet so you see the result. Run it right after a Git Update
    that changed the quote-run parser."""
    import quote_run_scan
    rc = quote_run_scan.main(["--reparse-stored"])
    if rc != 0:
        return rc
    return _cmd_refresh(args)


def _cmd_list(args) -> int:
    store = load_store()
    rows = open_notes(store) if not args.all else store["notes"]
    if not rows:
        print("No open notes." if not args.all else "No notes recorded.")
        return 0
    for n in sorted(rows, key=lambda x: int(x.get("id", 0))):
        flag = "x" if n.get("status") == STATUS_HANDLED else " "
        print(f"[{flag}] #{n['id']}  order {n['order']}  run {n.get('run', '')}")
        print(f"      item: {n.get('item', '')}  |  {n.get('item_text', '')}")
        print(f"      note: {n['note']}")
        if n.get("resolution"):
            print(f"      handled: {n['resolution']}")
    return 0


def _cmd_handle(args) -> int:
    # Always record the mark in the tracked ledger (the return channel to the
    # user's machine); also update a local queue if one is present here.
    resolution = " ".join(args.resolution)
    record_handled_mark(args.id, resolution)
    store = load_store()
    if mark_handled(store, args.id, resolution):
        save_store(store)
    print(f"Marked #{args.id} handled (recorded in {HANDLED_MARKS_PATH.name}; it "
          f"applies on the user's next 'Update QR Review').")
    return 0


def _run(func, args) -> int:
    """Run a subcommand, turning an expected error (e.g. the workbook is open in
    Excel) into a clear one-line message rather than a traceback in the log."""
    try:
        return func(args)
    except RuntimeError as e:
        print(str(e), file=sys.stderr)
        return 1


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="write the review workbook from the quote-run store")
    b.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    b.set_defaults(func=_cmd_build)

    o = sub.add_parser("open", help="open the review workbook (building it if missing)")
    o.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    o.set_defaults(func=_cmd_open)

    s = sub.add_parser("sync", help="read your notes back out of the workbook into the queue")
    s.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    s.set_defaults(func=_cmd_sync)

    r = sub.add_parser("refresh", help="capture notes + apply handled-marks + rewrite the sheet")
    r.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    r.set_defaults(func=_cmd_refresh)

    rp = sub.add_parser("reparse", help="re-parse the stored corpus with the current parser, then refresh")
    rp.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    rp.set_defaults(func=_cmd_reparse)

    ls_ = sub.add_parser("list", help="show open notes (--all for handled too)")
    ls_.add_argument("--all", action="store_true")
    ls_.set_defaults(func=_cmd_list)

    h = sub.add_parser("handle", help="mark a note handled with a resolution")
    h.add_argument("id", type=int)
    h.add_argument("resolution", nargs="+")
    h.set_defaults(func=_cmd_handle)

    args = ap.parse_args(argv)
    return _run(args.func, args)


if __name__ == "__main__":
    sys.exit(main())
