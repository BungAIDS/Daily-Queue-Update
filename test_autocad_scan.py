"""Tests for autocad_scan's pure logic + an end-to-end scan on a temp tree.

No pytest needed — run it directly:

    python test_autocad_scan.py

The live cbcinsider/Z: parts can't run off the work machine, but this drawing
filename / suffix-matrix logic is plain Python and is fully checked here.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import autocad_scan as a


def test_parse_drawing():
    assert a.parse_drawing("421314-01.dwg", "421314") == ("01", "dwg")
    assert a.parse_drawing("421314-02.PDF", "421314") == ("02", "pdf")
    assert a.parse_drawing("421314-51.dwg", "421314") == ("51", "dwg")
    # trailing text after the suffix (revisions etc.) still groups by suffix
    assert a.parse_drawing("421314-51 REV B.dwg", "421314") == ("51", "dwg")
    # not a drawing extension / wrong job / no suffix
    assert a.parse_drawing("421314-01.docx", "421314") is None
    assert a.parse_drawing("999999-01.dwg", "421314") is None
    assert a.parse_drawing("421314.dwg", "421314") is None
    # a different job whose number is a prefix must not match
    assert a.parse_drawing("4213140-01.dwg", "421314") is None


def test_revision_letters():
    # Names are [job]-[suffix][revletter]; the rev letter is dropped for grouping.
    assert a.parse_drawing("421314-01A.dwg", "421314") == ("01", "dwg")
    assert a.parse_drawing("421314-51B.pdf", "421314") == ("51", "pdf")
    assert a.parse_drawing("421314-02.dwg", "421314") == ("02", "dwg")  # no rev letter
    # Different revisions of one drawing collapse to a single suffix/column.
    found = a.scan_files(["421314-01A.dwg", "421314-01B.dwg", "421314-51A.pdf"], "421314")
    assert sorted(found) == ["01", "51"]
    assert a.build_record("421314", "AX", "/x", found)["extras"] == {"51": "PDF"}


def test_fmt_exts():
    assert a.fmt_exts({"pdf", "dwg"}) == "PDF+DWG"
    assert a.fmt_exts({"dwg"}) == "DWG"
    assert a.fmt_exts({"pdf"}) == "PDF"
    assert a.fmt_exts(set()) == ""


def test_scan_and_record():
    names = ["421314-01.dwg", "421314-01.pdf", "421314-02.dwg",
             "421314-51.pdf", "421314-35.dwg", "notes.txt", "421314.pdf"]
    found = a.scan_files(names, "421314")
    assert found["01"] == {"dwg", "pdf"}
    assert found["02"] == {"dwg"}
    assert found["51"] == {"pdf"} and found["35"] == {"dwg"}

    rec = a.build_record("421314", "AXIAL", r"Z:\x\y\421314", found)
    assert rec["cw"] == "PDF+DWG" and rec["ccw"] == "DWG"
    assert rec["extras"] == {"35": "DWG", "51": "PDF"}  # sorted numerically
    assert rec["missing_std"] is False


def test_missing_std_flag():
    rec = a.build_record("500", "GL", "/x", a.scan_files(["500-51.dwg"], "500"))
    assert rec["missing_std"] is True
    assert rec["cw"] == "" and rec["ccw"] == ""


def test_suffix_sort_and_union():
    records = {
        "1": {"extras": {"51": "DWG", "9": "PDF"}},
        "2": {"extras": {"100": "PDF", "35": "DWG"}},
    }
    assert a.all_extra_suffixes(records) == ["9", "35", "51", "100"]  # numeric order


def test_job_key():
    assert a.job_key("421314") == "421314"
    assert a.job_key("421314 ACME CORP") == "421314"
    assert a.job_key("LEGACY-AB") == "LEGACY-AB"  # no leading digits -> whole name


def test_is_real_job():
    assert a._is_real_job("403425", 400000, 0)
    assert a._is_real_job("421314", 400000, 0)
    assert not a._is_real_job("133567", 400000, 0)   # below floor (the goofy numbers)
    assert not a._is_real_job("2024", 400000, 0)     # year folder
    assert not a._is_real_job("TEMPLATES", 400000, 0)  # non-numeric
    assert a._is_real_job("405000", 400000, 410000) and not a._is_real_job("415000", 400000, 410000)


def test_iter_job_folders_floor(tmp: Path):
    root = tmp / "JOBS2"
    for typ, inter, leaf in [("AXIAL", "4034", "403425"), ("AXIAL", "4213", "421314"),
                             ("AXIAL", "1335", "133567"), ("MISC", "x", "TEMPLATES"),
                             ("AXIAL", "2024", "2024")]:
        (root / typ / inter / leaf).mkdir(parents=True)
    jobs = sorted(job for job, _t, _p in a.iter_job_folders(root))  # default floor 400000
    assert jobs == ["403425", "421314"], jobs


def test_end_to_end(tmp: Path):
    # Build  <root>/AXIAL/4213/421314/<drawings>  and scan it.
    root = tmp / "JOBS"
    jobdir = root / "AXIAL" / "4213" / "421314"
    jobdir.mkdir(parents=True)
    for fn in ["421314-01.dwg", "421314-01.pdf", "421314-02.pdf", "421314-51.dwg"]:
        (jobdir / fn).write_text("x")
    (jobdir / "random.txt").write_text("x")

    folders = list(a.iter_job_folders(root))
    assert folders and folders[0][0] == "421314" and folders[0][1] == "AXIAL"

    rec = a.scan_one("421314", "AXIAL", jobdir, recursive=False)
    assert rec["cw"] == "PDF+DWG" and rec["ccw"] == "PDF" and rec["extras"] == {"51": "DWG"}

    out = a.write_workbook({"421314": rec}, tmp / "out.xlsx")
    assert out.exists()
    from openpyxl import load_workbook
    ws = load_workbook(out).active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "-51" in headers and "CW (01)" not in headers and "CCW (02)" not in headers
    row = {headers[c - 1]: ws.cell(2, c).value for c in range(1, ws.max_column + 1)}
    assert row["Job #"] == "421314" and row["Extras"] == 1


def test_workbook_color_cells(tmp: Path):
    # Suffix cells carry no text; color is the signal (green=has, red=doesn't).
    recs = {
        "403425": a.build_record("403425", "AX", "/x", a.scan_files(["403425-51.dwg"], "403425")),
        "403500": a.build_record("403500", "AX", "/y", a.scan_files(["403500-01.dwg", "403500-02.dwg"], "403500")),
    }
    out = a.write_workbook(recs, tmp / "colors.xlsx")
    from openpyxl import load_workbook
    ws = load_workbook(out).active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = headers.index("-51") + 1
    has, hasnt = ws.cell(2, col), ws.cell(3, col)  # rows sorted by job: 403425, 403500
    assert has.value in (None, "") and hasnt.value in (None, "")
    assert (has.fill.fgColor.rgb or "").endswith("C6EFCE"), has.fill.fgColor.rgb   # green
    assert (hasnt.fill.fgColor.rgb or "").endswith("FFC7CE"), hasnt.fill.fgColor.rgb  # red


def main() -> int:
    passed = 0
    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)
        for name, fn in sorted(globals().items()):
            if not name.startswith("test_") or not callable(fn):
                continue
            (fn(tmp) if "tmp" in fn.__code__.co_varnames else fn())
            print(f"  ok  {name}")
            passed += 1
    print(f"\n{passed} tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
