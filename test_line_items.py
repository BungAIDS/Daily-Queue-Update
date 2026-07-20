"""Tests for the sales-order line-item capture / normalization / search logic.

No pytest needed — run it directly:

    python test_line_items.py

Real SO PDFs only exist on the work machine, so extraction is tested against
synthetic reconstructed-text lines shaped like the dumps (priced item rows,
an Additional Features section, totals/footer noise, CO-history lines). The
normalization, tagging, store, and search logic is pure Python and fully
checked here.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import line_items as li


SO_LINES = [
    "Sales Order 421314",
    "Sold To: ACME CORP   Ship To: ACME CORP PLANT 2",
    "Customer PO: 4500123  Terms: Net 30",
    "Design 53 Industrial Exhauster",
    "Qty Design Size Arrangement MotorPos Class",
    "1 53 365 DB W 2",
    "1 BASE FAN SIZE 365 ARR. 1 CW THD 1,847.00 1,847.00",
    "1 SPARK RESISTANT CONST. TYPE B 312.00",
    "SS SHAFT SLEEVE 184.00",
    "SHAFT SEAL - CERAMIC FELT N/C",
    "EPOXY COATED INTERIOR & EXTERIOR 96.50",
    "Fan suitable for 3600 rpm Motor",          # unpriced, outside a section
    "ADDITIONAL FEATURES",
    "EXTENDED LUBE LINES W/ ZERK FITTINGS",
    "VIBRATION ISOLATORS - RUBBER IN SHEAR",
    "2 ACCESS DOOR, QUICK OPEN",
    "NOTES:",
    "CO#1 5/21/26 DG - ADDED SHAFT SEAL PER CUSTOMER",
    "TOTAL BILLING 2,439.50",
    "Freight Allowed",
    "Page 1 of 2",
]


# Verbatim line sequences from real CBC Sales-Order dumps (jobs 421473 and
# 421314, 2026-06-11 discovery) — the regression base the rules were fitted to.
REAL_LINES_473 = [
    "Chicago Blower Corporation Sales Order",
    "Date:",
    "Design 16A SW",
    "Order # Rep Ref. # Customer P.O. # Fan Serial Number:",
    "421473 7074-49840-00-AI26 7074-49840-00-AI26",
    "Sold To: Ship To:",
    "INNO-VENT INDUSTRIAL INC. INNO-VENT INDUSTRIAL INC.",
    "Total Commission: - BUY SELL w/ comm to dest rep (Rev 0%)",
    "Qty Design Size Arrangement Motor Pos Class Rotation Discharge % Width Wheel Type",
    "1 D16A 37 A/9H Z 3 CW TH 100 LS",
    "CFM OV SP RPM BHP Design Temp Max Temp Elev. Density BHP @ 41° F",
    "34000 4552 21 963 184.4 70 95 1352 0.0714 194.99",
    "Type Price Freight Markup Net Comm.",
    "Base Fan (Base fan Suitable for 447T motor frame, L 69,648.00",
    "Inquiry Num: 909-26-1604)",
    "Motor (Model ECP84407TR-5 complete with : C 11,435.83 457.00 1,903.00 13,796.00 571.00",
    "Vendor: Baldor/Reliance",
    "200 HP, 1800 RPM, Enclosure: TEFC Severe",
    "Duty",
    "447T, Cast Iron, Foot Mounted, 3/60/575, F1,",
    "1.15 SF",
    "Quote Num: 1101260457",
    "Mounting Charge L 1,336.00",
    "Drive N 4,366.00 437.00",
    "Constant Speed, SF: 1.3",
    "CBC Mount",
    "Mounting Charge L INC",
    "Access Door, Quick Clamp L 617.00",
    "Door Location: @9:00",
    "Belt Guard, Painted Safety Yellow L 2,338.00",
    "Extended Grease Fittings L 280.00",
    "Housing Drain with Plug L 129.00",
    "Housing, Heavy Duty L 4,754.00",
    "Include 3D STEP Drawings L 991.00",
    "Inlet, Flanged, Punched L 312.00",
    "v1.8.1.5 -1-",
    "Chicago Blower Corporation Sales Order (cont.)",
    "Order # Rep Ref. # Customer P.O. # Page 2 of 2",
    "421473 7074-49840-00-AI26 7074-49840-00-AI26",
    "Mechanical Run Test, Not Available L STD",
    "Outlet, Flanged, Punched L STD",
    "Shaft and Bearing Guard, Painted Safety Yellow L 805.00",
    "Weights on drawing, Inquiry Num: 909-26-1812 L",
    "Inlet flanged C 776.00 213.00 989.00 64.00",
    "Vendor: FlexibleCompensators",
    "Quote Num: 34592",
    "Ship Loose",
    "Product: Expansion Joint",
    "Freight to CBC N 125.00",
    "Ship Loose Charge N 49.00 5.00",
    "List Total Each 81,710.00 0.38 31,050.00 2,484.00",
    "Lead Time: 76 working days Product 51,866.00",
    "See Additional Features / Notes below Freight",
    "Total 51,866.00",
    "Sales Tax (NT)",
    "Buy/Sell Deduction -3,631.00",
    "Total Billing 48,235.00",
    "Additional Features / Notes:",
    "E-Mail Prints to: hmallette@inno-vent.ca",
    "Run Test - N/A (Send to Sales for Run Test Availability Check if CBC is mounting the motor)",
    "Fan Drawings:",
    "Emailed Mailed",
    "Fan Drawings Both",
    "O & M X",
    "Motor Prints X",
    "Motor Data Sheets X",
    "Buyout Prints (e.g. silencer, filter, etc.) X",
    "Other",
    "v1.8.1.5 -2-",
]

REAL_LINES_314 = [
    "Chicago Blower Corporation CO #2 Sales Order",
    "Design 34 Vaneaxial Belt Drive",
    "Qty Design Size Arrangement Motor Pos Class Rotation Discharge % Width Wheel Type",
    "1 D34 15 N/A N/A 2 100 M",
    "CFM OV SP RPM BHP Design Temp Max Temp Elev. Density BHP @ 70° F",
    "3200 2607 2 3344 2.52 70 70 0 0.075 2.52",
    "Type Price Freight Markup Net Comm.",
    "Base Fan (Base fan, Suitable for 3600rpm Motor, L 5,761.00",
    "Inquiry Num: 317-26-1510)",
    "Motor C 254.83 5.00 62.00 322.00 19.00",
    "Vendor: Toshiba or equivalent",
    "3 HP, 3600 RPM, Enclosure: TEFC Premium",
    "182T, Cast Iron, Foot Mounted, 3/60/230/460,",
    "F1, 1.15 SF",
    "Model #0032SDSR41A-P, PN 06-6-0020-01",
    "VFD Suitable",
    "CBC Mount",
    "Mounting Charge L 479.00",
    "Drive (Drive Set, Constant Speed (CBC mounted), N 360.00 36.00",
    "Belt Guard, Painted Safety Yellow L 945.00",
    "Mechanical Run Test, Standard L STD",
    "Wheel, Steel L STD",
    "IVD C 2,750.79 440.00 3,191.00 132.00",
    "Vendor: Ruskin",
    "Quote Num: 042426GCM2",
    "Ship Loose",
    "Product: Damper",
    "Freight to CBC N 223.00",
    "List Total Each 8,181.00 0.5000 4,090.00 818.00",
    "Lead Time: 60 working days Product 8,346.00",
    "Total Billing 9,564.00",
    "Additional Features / Notes:",
    "C/O #2 5/15/26 ECR: CORRECTED TOTAL BILLING.",
    "CO#1 050826 AMF - CORRECTED CLASS",
    "CASH IN ADVANCE",
    "NO TAXES",
    "Total Billing $ $9,023.00",
    "Run Test - Required",
]


def test_real_so_std_inc_and_bare_type_letter():
    items = {it["norm"]: it for it in li.extract_items(REAL_LINES_473)}
    # STD / INC in the price column (after the L/C/N type letter) are items...
    assert items["MECHANICAL RUN TEST NOT AVAILABLE"]["price"] == "STD"
    assert "OUTLET FLANGED PUNCHED" in items
    inc = {it["norm"]: it for it in li.extract_items(["Temporary Fan Feature L INC"])}
    assert inc["TEMPORARY FAN FEATURE"]["price"] == "INC"
    assert "MOUNTING CHARGE" not in items
    # ...and so is a row whose price column is simply empty (trailing bare L).
    assert any(n.startswith("WEIGHTS ON DRAWING") for n in items), items.keys()
    # But "INC." at the end of a company name never makes an address an item.
    assert not any("INNO" in n for n in items), items.keys()


def test_real_so_type_letter_stripped_and_price_column():
    items = {it["norm"]: it for it in li.extract_items(REAL_LINES_473)}
    bg = items["BELT GUARD PAINTED SAFETY YELLOW"]   # no trailing " L" in norm
    assert bg["ptype"] == "L" and bg["price"] == "2,338.00", bg
    # Multi-column money tail keeps the LEFTMOST (the Price column), not Comm.
    motor = items["MOTOR MODEL ECP84407TR 5 COMPLETE WITH"]
    assert motor["price"] == "11,435.83", motor


def test_real_so_noise_excluded():
    for lines in (REAL_LINES_473, REAL_LINES_314):
        joined = " | ".join(it["norm"] for it in li.extract_items(lines))
        for bad in ("LIST TOTAL", "LEAD TIME", "CUSTOMS", "FAN DRAWINGS",
                    "EMAILED", "BUYOUT PRINTS", "CHICAGO BLOWER", "DEDUCTION",
                    "COMMISSION", "CORRECTED CLASS"):
            assert bad not in joined, (bad, joined)
        # The CFM/RPM performance values row (numbers only) is not an item,
        # and neither is the spec-table row under the Qty/Design/Size header.
        assert "34000" not in joined and "D16A" not in joined, joined
        assert "3200 2607" not in joined and "D34" not in joined, joined


def test_real_so_details_attached():
    items = {it["norm"]: it for it in li.extract_items(REAL_LINES_314)}
    motor = items["MOTOR"]
    det = " | ".join(motor["details"])
    assert "Toshiba" in det and "3 HP" in det and "VFD Suitable" in det, det
    assert "VFD" in motor["tags"], motor["tags"]  # detail lines drive tags too
    # Page furniture between items must never attach as a detail.
    items473 = {it["norm"]: it for it in li.extract_items(REAL_LINES_473)}
    det = " | ".join(items473["INLET FLANGED PUNCHED"]["details"])
    assert "Chicago" not in det and "7074" not in det and "v1.8" not in det, det


def test_real_so_ivd_and_buyout_tagging():
    items = {it["norm"]: it for it in li.extract_items(REAL_LINES_314)}
    ivd = items["INLET VANE DAMPER"]            # IVD abbreviation expanded
    assert "DAMPER" in ivd["tags"] and "INLET VANES" in ivd["tags"], ivd
    assert any("Ruskin" in d for d in ivd["details"]), ivd["details"]
    assert ivd["attributes"]["component"] == "INLET VANE DAMPER"
    assert "used_on" not in ivd["attributes"]
    assert ivd["attributes"]["damper_subcategory"] == "INLET VANE DAMPER"
    assert ivd["attributes"]["quote_number"] == "042426GCM2"
    assert not ivd.get("review_flags")
    # The flanged expansion-joint buyouts tag FLEX CONNECTOR via "Product:".
    items473 = {it["norm"]: it for it in li.extract_items(REAL_LINES_473)}
    assert "FLEX CONNECTOR" in items473["INLET FLANGED"]["tags"]
    assert {"BEARINGS", "MOTOR"} <= set(items473["EXTENDED GREASE FITTINGS"]["tags"])
    assert "EXTENDED LUBE" not in items473["EXTENDED GREASE FITTINGS"]["tags"]
    assert "HEAVY DUTY" in items473["HOUSING HEAVY DUTY"]["tags"]
    assert "3D STEP DRAWINGS" in items473["INCLUDE 3D STEP DRAWINGS"]["tags"]


def test_real_so_search_reaches_details():
    store = li.load_store(Path("/nonexistent/store.json"))
    li.record_job(store, "421314", li.extract_items(REAL_LINES_314))
    li.record_job(store, "421473", li.extract_items(REAL_LINES_473))
    assert [h["job"] for h in li.search(store, ["toshiba"])] == ["421314"]
    assert [h["job"] for h in li.search(store, ["baldor"])] == ["421473"]
    assert [h["job"] for h in li.search(store, ["200 HP"])] == ["421473"]
    assert {h["job"] for h in li.search(store, [], tag="DAMPER")} == {"421314"}


def test_priced_lines_captured():
    items = li.extract_items(SO_LINES)
    raws = [it["raw"] for it in items]
    assert any("BASE FAN" in r for r in raws), raws
    assert any("SPARK RESISTANT" in r for r in raws), raws
    assert any("SHAFT SLEEVE" in r for r in raws), raws
    assert any("CERAMIC FELT" in r for r in raws), raws  # N/C counts as priced


def test_section_lines_captured_unpriced():
    items = {it["norm"]: it for it in li.extract_items(SO_LINES)}
    assert any("ZERK FITTINGS" in n for n in items), items.keys()
    assert any("VIBRATION ISOLATORS" in n for n in items), items.keys()
    it = next(v for n, v in items.items() if "ACCESS DOOR" in n)
    assert it["section"] == "ADDITIONAL FEATURES"
    assert it["qty"] == "2"


def test_noise_skipped():
    norms = " | ".join(it["norm"] for it in li.extract_items(SO_LINES))
    for bad in ("TOTAL BILLING", "FREIGHT", "PAGE 1", "SOLD TO", "CUSTOMER PO",
                "CO#1", "ADDED SHAFT SEAL PER CUSTOMER", "NOTES"):
        assert bad not in norms, (bad, norms)
    # The spec-table row ("1 53 365 DB W 2") has no price and no section.
    assert "365 DB" not in norms, norms


def test_section_closes_at_notes():
    # The CO-history line sits after NOTES:, so even though section capture was
    # on, nothing after the section_end marker is captured.
    items = li.extract_items(["ACCESSORIES", "BELT GUARD", "NOTES:", "SOME NOTE TEXT"])
    assert [it["norm"] for it in items] == ["BELT GUARD"], items


def test_price_and_qty_parsed():
    items = {it["norm"]: it for it in li.extract_items(SO_LINES)}
    spark = next(v for n, v in items.items() if "SPARK" in n)
    assert spark["price"] == "312.00" and spark["qty"] == "1", spark
    base = next(v for n, v in items.items() if "BASE FAN" in n)
    assert base["price"] in ("1,847.00", "1847.00"), base  # double money tail stripped
    assert "1,847" not in base["norm"], base
    seal = next(v for n, v in items.items() if "CERAMIC" in n)
    assert seal["price"].upper().replace("/", "") == "NC", seal


def test_bare_integers_are_not_prices():
    # "3600" in "suitable for 3600 rpm Motor" must not read as a price column.
    desc, price = li.split_price_tail("FAN SUITABLE FOR 3600 RPM MOTOR")
    assert price == "" and desc.endswith("MOTOR")
    desc, price = li.split_price_tail("DAMPER 1,250.00")
    assert price == "1,250.00" and desc == "DAMPER"
    desc, price = li.split_price_tail("GUARD $96.50 EACH")
    assert price == "$96.50" and desc == "GUARD"


def test_normalization_converges_variants():
    # Different order-entry spellings of the same option meet at one norm.
    a = li.normalize_text("SS SHAFT SLEEVE 184.00")
    b = li.normalize_text("1  Stainless Steel Shaft Sleeve")
    c = li.normalize_text("S/S SHAFT SLEEVE - N/C")
    assert a == b == c == "STAINLESS STEEL SHAFT SLEEVE", (a, b, c)
    assert li.normalize_text("316SS sleeve") == "316 STAINLESS STEEL SLEEVE"
    assert li.normalize_text("304SST Airstream") == "304 STAINLESS STEEL AIRSTREAM"
    assert li.normalize_text("316LSS housing") == "316L STAINLESS STEEL HOUSING"
    assert li.normalize_text("SST exterior") == "STAINLESS STEEL EXTERIOR"
    assert li.normalize_text("Wheel Aluminium AMCA B") == "WHEEL ALUMINUM AMCA B"
    # "EXT" stays unexpanded (ambiguous with EXTERIOR).
    assert li.normalize_text("EXT LUBE LINES W/ ZERKS") == "EXT LUBE LINES WITH ZERKS"
    assert li.normalize_text("W/DRAIN") == "WITH DRAIN"


def test_tagging():
    assert "SHAFT SEAL" in li.tag_item("SHAFT SEAL CERAMIC FELT")
    assert "SHAFT SEAL" in li.tag_item(li.normalize_text("Teflon shaft seal"))
    assert "SPARK RESISTANT" in li.tag_item("SPARK RESISTANT CONSTRUCTION TYPE B")
    assert "STAINLESS STEEL" in li.tag_item(li.normalize_text("316SS wheel"))
    assert "MATERIALS" in li.tag_item(li.normalize_text("304SST Airstream"))
    assert "COATING" in li.tag_item("EPOXY COATED INTERIOR AND EXTERIOR")
    assert "VIBRATION ISOLATION" in li.tag_item("VIBRATION ISOLATORS RUBBER IN SHEAR")
    assert "DRIVE COMPONENTS" in li.tag_item(li.normalize_text("Drive"))
    assert "VIBRATION ISOLATION" in li.tag_item(li.normalize_text("Vibration Base"))
    assert "INLET VANES" in li.tag_item(li.normalize_text("Inlet Volume Control Automatic"))
    assert "COATING" not in li.tag_item(li.normalize_text("Passivation of Welds"))
    assert "STAINLESS STEEL" in li.tag_item(li.normalize_text("Passivation of Welds"))
    assert "LINING" in li.tag_item(li.normalize_text("Firmex liners on blades and housing scroll"))
    assert "COUPLING" in li.tag_item(li.normalize_text("Flexible Coupling Falk Type T Steelflex"))
    assert "EXTREME TEMP" in li.tag_item(li.normalize_text("High Temp Fan"))
    assert "LOW LEAKAGE" in li.tag_item(li.normalize_text("Low Leak IVC"))
    assert "ALUMINUM" in li.tag_item(li.normalize_text("Wheel Aluminium AMCA B"))
    assert "MATERIALS" in li.tag_item(li.normalize_text("Wheel Aluminium AMCA B"))
    assert "WHEEL" in li.tag_item(li.normalize_text("Percent Width 78.7%"))
    assert "LIFTING LUGS" in li.tag_item(li.normalize_text("Lifting Lugs"))
    assert "NAMEPLATE" in li.tag_item(li.normalize_text("Fan Nameplate without Chicago Blower Name"))
    assert "PACKAGING" in li.tag_item(li.normalize_text("ISPM Wood Inspection Stamp"))
    assert "SHIPPING" in li.tag_item(li.normalize_text("Ship Loose Freight Included"))
    assert "ACTUATOR" in li.tag_item(li.normalize_text("Actuator for IVC Bettis #RPED100"))
    assert "3D STEP DRAWINGS" in li.tag_item(li.normalize_text("3D STEP File Drawings"))
    assert "DRIVE COMPONENTS" in li.tag_item(li.normalize_text("Motor Sheave/Bushing 3B5V74/B"))
    assert "SPLIT HOUSING" in li.tag_item(li.normalize_text("Drawing and split housing released into production"))
    assert "SPLIT HOUSING" in li.tag_item(li.normalize_text("Shipping Split"))
    assert "DRIVE COMPONENTS" in li.tag_item(li.normalize_text("Drive (Max/Min RPM: 1531/1531, 3 belts: B112"))
    assert "WHEEL" in li.tag_item(li.normalize_text("110% Effective Diameter"))
    assert "WHEEL" in li.tag_item(li.normalize_text("Cast Hub with Straight Bore"))
    assert "SPECIAL CONSTRUCTION" in li.tag_item(li.normalize_text("Tie Rod Support"))
    assert "SPECIAL CONSTRUCTION" in li.tag_item(li.normalize_text("Loc Tite on the set screw threads"))
    assert "SPECIAL CONSTRUCTION" in li.tag_item(li.normalize_text("Continuous Weld Airstream"))
    assert "SPECIAL CONSTRUCTION" in li.tag_item(li.normalize_text("Earthing Boss"))
    assert "INSPECTION" in li.tag_item(li.normalize_text("Customer Final Inspection"))
    certifications = li.tag_item(li.normalize_text("General Mill Certifications"))
    assert "CERTIFICATION" in certifications
    assert "INSPECTION" not in certifications
    assert "LABEL" in li.tag_item(li.normalize_text("FEI Label Inquiry Num"))
    assert "BALANCE" in li.tag_item(li.normalize_text("G2.5 Balance"))
    assert "BALANCE" in li.tag_item(li.normalize_text("Welded Balance Weights"))
    assert "BALANCE" not in li.tag_item(li.normalize_text("Balance Report"))
    assert "BALANCE" not in li.tag_item(li.normalize_text(
        "All Chicago Blower wheels are precision balanced"))
    assert "BEARINGS" in li.tag_item(li.normalize_text("Bearings, Standard"))
    assert "BEARINGS" in li.tag_item(li.normalize_text("Bearings, Split Pillow Block"))
    assert "BEARINGS" in li.tag_item(li.normalize_text("Repair Bearings (Pair), 2-3/16 Bore"))
    guard_tags = li.tag_item(li.normalize_text("Shaft, Bearing, and Coupling Guard, Painted Safety Yellow"))
    assert "SHAFT/BEARING/COUPLING GUARD" in guard_tags
    assert "BEARINGS" not in guard_tags
    assert "BEARINGS" not in li.tag_item(li.normalize_text("Motor with insulated bearings"))
    assert "BEARINGS" not in li.tag_item(li.normalize_text(
        "Paint Interior Wheel Exterior Motor Base Channel Base and Bearing Base"))
    assert li.tag_item("SOMETHING NOBODY EVER ORDERED") == []


def test_tags_label():
    items = li.extract_items(SO_LINES)
    label = li.tags_label(items)
    assert "SHAFT SEAL" in label and "SPARK RESISTANT" in label, label
    assert li.tags_label([]) == ""
    assert li.tags_label([{"raw": "X", "norm": "MYSTERY OPTION", "tags": []}]) == "(1 items)"


def test_rules_extension_file(tmp: Path):
    ext = tmp / "rules_ext.json"
    ext.write_text('{"tags": {"SHAFT GROUNDING": ["grounding\\\\s*ring"]},'
                   ' "abbreviations": {"GRND": "GROUNDING"}}', encoding="utf-8")
    rules = li.load_rules(ext)
    norm = li.normalize_text("SHAFT GRND RING", rules)
    assert norm == "SHAFT GROUNDING RING"
    assert "SHAFT GROUNDING" in li.tag_item(norm, rules)
    # Defaults are still present (extended, not replaced).
    assert "SHAFT SEAL" in li.tag_item("SHAFT SEAL", rules)


def test_store_roundtrip_and_merge(tmp: Path):
    path = tmp / "store.json"
    store = li.load_store(path)
    items = li.extract_items(SO_LINES)
    li.record_job(store, "421314", items, customer="ACME CORP", co_number=1,
                  so_pdf="Z:/SO/421314.pdf")
    li.save_store(store, path)
    store2 = li.load_store(path)
    assert store2["jobs"]["421314"]["customer"] == "ACME CORP"
    assert len(store2["jobs"]["421314"]["items"]) == len(items)
    # A sparse re-record (archive scan: no customer known) keeps the metadata.
    li.record_job(store2, "421314", items)
    assert store2["jobs"]["421314"]["customer"] == "ACME CORP"
    assert store2["jobs"]["421314"]["co_number"] == 1
    assert store2["jobs"]["421314"]["so_pdf"] == "Z:/SO/421314.pdf"


def _seeded_store() -> dict:
    store = li.load_store(Path("/nonexistent/line_items.json"))
    li.record_job(store, "421314", li.extract_items(SO_LINES), customer="ACME CORP")
    li.record_job(store, "421999", li.extract_items(
        ["1 BASE FAN 900.00", "TEFLON SHAFT SEAL 55.00", "BELT GUARD 80.00"]),
        customer="ZEECO")
    return store


def test_search_and_terms():
    store = _seeded_store()
    # AND (default): both jobs have a shaft seal...
    hits = li.search(store, ["shaft seal"])
    assert [h["job"] for h in hits] == ["421999", "421314"]  # newest job first
    # ...but only one also has spark-resistant construction.
    hits = li.search(store, ["shaft seal", "spark"])
    assert [h["job"] for h in hits] == ["421314"]
    # OR mode.
    hits = li.search(store, ["spark", "belt guard"], any_mode=True)
    assert {h["job"] for h in hits} == {"421314", "421999"}
    # Tag filter (canonical lookup): only the tagged items come back.
    hits = li.search(store, [], tag="SHAFT SEAL")
    assert {h["job"] for h in hits} == {"421314", "421999"}
    assert all("SHAFT SEAL" in it["tags"] for h in hits for it in h["matches"])
    # Tag + term AND at the JOB level — they may sit on different items.
    hits = li.search(store, ["spark"], tag="SHAFT SEAL")
    assert [h["job"] for h in hits] == ["421314"]
    raws = " | ".join(it["raw"] for it in hits[0]["matches"])
    assert "SPARK" in raws and "CERAMIC FELT" in raws, raws
    # Search hits normalized text, so "stainless" finds the raw "SS" item;
    # terms are ANDed, each may land on a different word of the same item.
    hits = li.search(store, ["stainless", "sleeve"])
    assert [h["job"] for h in hits] == ["421314"]


def test_search_fuzzy():
    store = _seeded_store()
    assert li.search(store, ["cermic"]) == []          # typo: no substring hit
    hits = li.search(store, ["cermic"], fuzzy=0.8)     # fuzzy catches it
    assert [h["job"] for h in hits] == ["421314"]


def test_ai_cache_applied():
    store = li.load_store(Path("/nonexistent/line_items.json"))
    items = li.extract_items(["Mystery Option L 12.00"])
    li.record_job(store, "421314", items)
    rec = store["jobs"]["421314"]
    it = rec["items"][0]
    assert it["norm"] in li.unknown_norms(store)
    assert it["review_flags"] == ["UNTAGGED"]
    store["ai_tags"][it["norm"]] = ["BASE FAN"]
    li.apply_ai_cache(rec["items"], store)
    assert "BASE FAN" in it["tags"]
    assert "review_flags" not in it
    # Once cached it's never sent to the API again.
    assert it["norm"] not in li.unknown_norms(store)


def test_ai_cache_does_not_override_current_rules():
    store = li.load_store(Path("/nonexistent/line_items.json"))
    items = li.extract_items(["Base Fan L 900.00"])
    assert items[0]["tags"] == ["BASE FAN"]
    store["ai_tags"][items[0]["norm"]] = ["UNITARY BASE"]
    li.apply_ai_cache(items, store)
    assert items[0]["tags"] == ["BASE FAN"]


def test_renormalize_uses_current_rules():
    store = _seeded_store()
    store["jobs"]["421314"]["items"].append(
        {"raw": "Prints:", "norm": "PRINTS", "qty": "", "price": "",
         "ptype": "", "section": "", "details": [], "tags": ["DRAWINGS"], "attributes": {}}
    )
    store["jobs"]["421314"]["items"].extend(li.extract_items([
        "Motor C 254.83",
        "3 HP, 3600 RPM, Enclosure: TEFC Premium",
    ]))
    # Sabotage the stored derived fields; renormalize must rebuild from raw.
    for it in store["jobs"]["421314"]["items"]:
        it["norm"], it["tags"] = "WRONG", ["BOGUS"]
        it["attributes"], it["review_flags"] = {"bad": "value"}, ["STALE"]
    n = li.renormalize_store(store)
    assert n >= len(store["jobs"]["421314"]["items"])
    norms = [it["norm"] for it in store["jobs"]["421314"]["items"]]
    assert "STAINLESS STEEL SHAFT SLEEVE" in norms, norms
    assert "PRINTS" not in norms
    motor = next(it for it in store["jobs"]["421314"]["items"] if it["norm"] == "MOTOR")
    assert motor["attributes"]["motor_hp"] == "3"
    assert "STALE" not in (motor.get("review_flags") or [])


def test_audit_untagged_uses_current_rules():
    store = li.load_store(Path("/nonexistent/line_items.json"))
    items = li.extract_items(["Drive L 100.00", "Mystery Option L 12.00"])
    items.append({"raw": "Product:", "norm": "PRODUCT", "qty": "", "price": "",
                  "ptype": "", "section": "", "details": [], "tags": []})
    li.record_job(store, "421000", items)
    rows = li.audit_untagged(store)
    assert [r["norm"] for r in rows] == ["MYSTERY OPTION"], rows


def test_audit_review_groups_marked_templates():
    store = li.load_store(Path("/nonexistent/line_items.json"))
    li.record_job(store, "421000", li.extract_items([
        "Mystery Option L 12.00",
        "Actuator Bettis #RPED200 C 5,192.00",
        "Extended Grease Fittings L 280.00",
    ]))
    rows = {r["norm"]: r for r in li.audit_review(store)}
    assert rows["MYSTERY OPTION"]["review_flags"] == ["UNTAGGED"]
    assert any("USED ON REVIEW" in flag for flag in rows["ACTUATOR BETTIS #RPED200"]["review_flags"])
    assert any("UNCLEAR GREASE TARGET" in flag for flag in rows["EXTENDED GREASE FITTINGS"]["review_flags"])


def test_actuator_attributes():
    items = {it["norm"]: it for it in li.extract_items([
        "Actuator for IVC, Bettis #RPED100 Double Acting C 5,054.00 809.00",
        "Pneumatic Actuator. Complete with a VAC",
        "Vendor: Novaspect",
        "Product: Actuator",
        "Operation: Automatic",
        "Actuator Manufacturer: Bettis RPD",
        "Actuator Supplied By: CBC (Mounted)",
        "Fail Position Upon Loss of Supply/Air/Power: In",
        "Place",
        "Fail Position Upon Loss of Signal: Closed",
        "Actuator Mounting: Bracket and actuator",
    ])}
    actuator = items["ACTUATOR FOR IVC BETTIS #RPED100 DOUBLE ACTING"]
    attrs = actuator["attributes"]
    assert {"ACTUATOR", "DAMPER", "INLET VANES"} <= set(actuator["tags"])
    assert "MOUNTING" not in actuator["tags"]
    assert attrs["component"] == "ACTUATOR"
    assert attrs["used_on"] == "IVC"
    assert "ivc_subcategory" not in attrs and "ivc_component" not in attrs
    assert attrs["vendor"] == "Novaspect"
    assert "product" not in attrs
    assert attrs["manufacturer"] == "BETTIS"
    assert attrs["model"] == "RPED100"
    assert attrs["size"] == "100"
    assert attrs["operation"] == "Automatic"
    assert "supplied_by" not in attrs
    assert attrs["mounting"] == "Bracket and actuator"
    assert attrs["fail_position_upon_loss_of_power"] == "In Place"
    assert attrs["fail_position_upon_loss_of_signal"] == "Closed"
    assert not actuator.get("review_flags")


def test_shared_damper_actuator_attributes_keep_explicit_owners():
    items = li.extract_items([
        "Outlet Damper Actuator C 6,537.14",
        "Bettis #RPED150 Double Acting Pneumatic Actuator with 4-20mA Transmitter",
        "Vendor: Novaspect",
        "Quote Num: N-583671-2",
        "Product: Actuator",
        "Prespin and Fresh Air Damper Actuators C 13,254.28",
        "QTY (2) - Bettis #RPED200 Double Acting Pneumatic Actuator with 4-20mA Transmitter",
        "Vendor: Novaspect",
        "Quote Num: N-583671-2",
        "Product: Actuator",
    ])
    outlet, shared = items
    assert outlet["attributes"]["used_on"] == "OUTLET DAMPER"
    assert outlet["attributes"]["model"] == "RPED150"
    assert outlet["attributes"]["size"] == "150"
    assert shared["attributes"]["used_on"] == "PRESPIN DAMPER, FRESH AIR DAMPER"
    assert shared["attributes"]["model"] == "RPED200"
    assert shared["attributes"]["size"] == "200"
    assert shared["attributes"]["quantity"] == "2"
    for item in items:
        assert item["attributes"]["transmitter"] == "4-20MA"
        assert item["attributes"]["quote_number"] == "N-583671-2"
        assert not item.get("review_flags")


def test_shared_damper_actuator_accepts_reverse_order_and_ampersand():
    item = li.extract_items([
        "Fresh Air & Prespin Damper Actuators C 13,254.28",
        "QTY (2) - Bettis #RPED200 Double Acting Pneumatic Actuator",
    ])[0]
    assert item["attributes"]["used_on"] == "FRESH AIR DAMPER, PRESPIN DAMPER"


def test_manual_damper_operation_does_not_create_an_actuator():
    item = li.extract_items([
        "Fresh Air Damper C 2,100.00",
        "Operation: Manual",
    ])[0]
    attrs = item["attributes"]
    assert attrs["component"] == "FRESH AIR DAMPER"
    assert attrs["operation"] == "Manual"
    assert not any(key.startswith("actuator_") for key in attrs)


def test_drive_attributes():
    items = {it["norm"]: it for it in li.extract_items([
        "Drive (Max/Min RPM: 1531/1531, 3 belts: B112, C 409.00 125.00",
        "Motor Sheave/Bushing: 3B5V74/B (1 7/8\"), Fan",
        "Sheave/Bushing: 3B5V86/B (2 3/16\"), Actual SF:",
        "1.31, Actual CD: 44.34)",
        "Constant Speed, SF: 1.3",
        "CBC Mount",
    ])}
    drive = items["DRIVE MAX MIN RPM 1531 1531 3 BELTS B112"]
    attrs = drive["attributes"]
    assert "DRIVE COMPONENTS" in drive["tags"]
    assert "V-BELT DRIVE" not in drive["tags"]
    assert "MOTOR" not in drive["tags"]
    assert "MOUNTING" not in drive["tags"]
    assert attrs["component"] == "DRIVE COMPONENTS"
    assert attrs["drive_subcategory"] == "SHEAVE/BUSHING"
    assert attrs["belt_qty"] == "3"
    assert attrs["belt"] == "B112"
    assert attrs["max_rpm"] == "1531"
    assert attrs["min_rpm"] == "1531"
    assert attrs["drive_sheave_bushing"] == '3B5V74/B (1 7/8")'
    assert attrs["driven_sheave_bushing"] == '3B5V86/B (2 3/16")'
    assert attrs["drive_sheave"] == "3B5V74"
    assert attrs["drive_bushing"] == 'B (1 7/8")'
    assert attrs["driven_sheave"] == "3B5V86"
    assert attrs["driven_bushing"] == 'B (2 3/16")'
    assert attrs["actual_sf"] == "1.31"
    assert attrs["actual_cd"] == "44.34"
    assert attrs["service_factor"] == "1.3"
    assert attrs["mounting"] == "CBC MOUNT"


def test_vfd_attributes_are_motor_details():
    motor = li.extract_items([
        "Motor C 254.83",
        "Vendor: Toshiba or equivalent",
        "3 HP, 3600 RPM, Enclosure: TEFC Premium",
        "VFD Suitable",
    ])[0]
    assert {"MOTOR", "VFD"} <= set(motor["tags"])
    assert motor["attributes"]["component"] == "MOTOR"
    assert motor["attributes"]["motor_hp"] == "3"
    assert motor["attributes"]["motor_rpm"] == "3600"
    assert motor["attributes"]["motor_enclosure"] == "TEFC PREMIUM"
    assert motor["attributes"]["vfd_context"] == "MOTOR"
    assert motor["attributes"]["motor_vfd_suitability"] == "VFD SUITABLE"

    inverter = li.derive_item_fields({
        "raw": "Duty, Inverter Duty 15:1 CT, 20:1 VT",
        "details": [],
    })
    assert "VFD" in inverter["tags"]
    assert inverter["attributes"]["component"] == "MOTOR"
    assert inverter["attributes"]["motor_vfd_suitability"] == "INVERTER DUTY"
    assert inverter["attributes"]["motor_vfd_speed_range"] == "15:1 CT, 20:1 VT"

    by_others = li.derive_item_fields({
        "raw": "VFD by others, motor complete with AEGIS bearing protection ring",
        "details": [],
    })
    assert "VFD" in by_others["tags"]
    assert by_others["attributes"]["component"] == "MOTOR"
    assert by_others["attributes"]["vfd_supplied_by"] == "OTHERS"

    not_suitable = li.extract_items([
        "Motor C 254.83",
        "Not VFD Suitable",
        "VFD Controlled",
    ])[0]
    assert not_suitable["attributes"]["motor_vfd_suitability"] == "NOT VFD SUITABLE"
    assert not_suitable["attributes"]["motor_vfd_operation"] == "VFD CONTROLLED"


def test_motor_core_attributes_and_unclassified_detail_review():
    motor = li.extract_items([
        "Motor (Model ECP84407TR-5 complete with : C 11,435.83",
        "Vendor: Baldor/Reliance",
        "200 HP, 1800 RPM, Enclosure: TEFC Severe",
        "Duty",
        "447T, Cast Iron, Foot Mounted, 3/60/575, F1,",
        "1.15 SF",
        "PN 06-6-0020-01",
        "Mounted by Others",
        "Special winding instruction we have not categorized",
    ])[0]
    attrs = motor["attributes"]
    assert attrs["component"] == "MOTOR"
    assert attrs["vendor"] == "Baldor/Reliance"
    assert attrs["motor_hp"] == "200"
    assert attrs["motor_rpm"] == "1800"
    assert attrs["motor_enclosure"] == "TEFC SEVERE DUTY"
    assert attrs["motor_duty"] == "SEVERE DUTY"
    assert attrs["motor_frame"] == "447T"
    assert attrs["motor_frame_material"] == "CAST IRON"
    assert attrs["motor_mounting"] == "FOOT MOUNTED"
    assert attrs["motor_phase"] == "3"
    assert attrs["motor_frequency_hz"] == "60"
    assert attrs["motor_voltage"] == "575"
    assert attrs["motor_conduit_box_location"] == "F1"
    assert attrs["motor_service_factor"] == "1.15"
    assert attrs["motor_model"] == "ECP84407TR-5"
    assert attrs["motor_part_number"] == "06-6-0020-01"
    assert attrs["motor_mounted_by"] == "OTHERS"

    assert motor.get("review_flags") == [
        "UNCLASSIFIED DETAIL: Special winding instruction we have not categorized"
    ]


def test_motor_special_attributes_join_wrapped_nameplate_commitments():
    motor = li.extract_items([
        "Motor (Model 0604XDSB41A-P) C 11,435.83",
        "Vendor: Toshiba",
        "60 HP, 1800 RPM, Enclosure: TEFC Premium",
        "364T, Cast Iron, Foot Mounted, 3/60/460, F1,",
        "1.15 SF",
        "Motors meet NEMA GM 7E-TA spec and are",
        "silicone free.",
        'Motors will be re-nameplated as "IEEE841',
        'features only" due to the AEGIS shaft',
        "grounding ring, which voids Div 2 capabilities.",
        'We will ADD "Meets GM 7E-TA Spec"',
        "nameplate.",
    ])[0]
    attrs = motor["attributes"]
    assert attrs["motor_shaft_grounding"] == "SHAFT GROUNDING RING"
    assert attrs["motor_duty"] == "IEEE 841 FEATURES ONLY"
    assert attrs["special_attribute"] == [
        "MOTORS MEET NEMA GM 7E-TA SPEC AND ARE SILICONE FREE.",
        'MOTORS WILL BE RE-NAMEPLATED AS "IEEE841 FEATURES ONLY" DUE TO THE '
        "AEGIS SHAFT GROUNDING RING, WHICH VOIDS DIV 2 CAPABILITIES.",
        'WE WILL ADD "MEETS GM 7E-TA SPEC" NAMEPLATE.',
    ]
    assert not motor.get("review_flags")

    exception = li.extract_items([
        "Motor C 10,000.00",
        "Toshiba will provide nameplate to read:",
        "\ufffdCOMPLIANCE WITH GM7E-TA WITH",
        "EXCEPTION PER TIC-2025-02-SGM7A-TE-R02\ufffd.",
        "Toshiba standard is laser etched not raised",
        "letters.",
    ])[0]
    assert exception["attributes"]["special_attribute"] == [
        'TOSHIBA WILL PROVIDE NAMEPLATE TO READ: "COMPLIANCE WITH GM7E-TA WITH '
        'EXCEPTION PER TIC-2025-02-SGM7A-TE-R02".',
        "TOSHIBA STANDARD IS LASER ETCHED NOT RAISED LETTERS.",
    ]
    assert not exception.get("review_flags")


def test_selected_drive_table_attributes():
    items = li.extract_items([
        "1515/1515 B116 3 3TB80 Q1 3B5V94 B 1.49 45.24 436.00",
        "Notes:",
        "* Selected Drive",
        "Specified minimum belt service factor: 1.3",
        "Center Distance with allowance for install and take-up: 43.5 - 46.5",
    ])
    drive = items[0]
    attrs = drive["attributes"]
    assert "DRIVE COMPONENTS" in drive["tags"]
    assert "V-BELT DRIVE" not in drive["tags"]
    assert attrs["drive_subcategory"] == "SELECTED DRIVE TABLE"
    assert attrs["belt"] == "B116"
    assert attrs["belt_qty"] == "3"
    assert attrs["drive_sheave_bushing"] == "3TB80 Q1"
    assert attrs["driven_sheave_bushing"] == "3B5V94 B"
    assert attrs["drive_sheave"] == "3TB80"
    assert attrs["drive_bushing"] == "Q1"
    assert attrs["driven_sheave"] == "3B5V94"
    assert attrs["driven_bushing"] == "B"
    assert attrs["actual_sf"] == "1.49"
    assert attrs["actual_cd"] == "45.24"
    assert attrs["selected_drive"] == "YES"
    assert attrs["service_factor"] == "1.3"
    assert attrs["center_distance_range"] == "43.5 - 46.5"
    bare = li.extract_items([
        "1515/1515 B116 3 3B5V80 B 3B5V94 B 1.49 45.24 428.00",
    ])[0]
    assert "DRIVE COMPONENTS" in bare["tags"]
    assert bare["attributes"]["belt"] == "B116"
    assert bare["attributes"]["drive_sheave_bushing"] == "3B5V80 B"
    assert bare["attributes"]["driven_sheave_bushing"] == "3B5V94 B"

    compact = li.extract_items([
        "*2567/2567 A52 1 AK46 AK30H H 1.25 21.14 99.00",
        "Notes:",
        "* Selected Drive",
        "Motor conduit box in F1 position.",
        "Specified minimum belt service factor: 1.2",
        "Center Distance range assumes adjustable base per drawing 7-2-94.",
    ])[0]
    compact_attrs = compact["attributes"]
    assert "DRIVE COMPONENTS" in compact["tags"]
    assert "DRAWINGS" not in compact["tags"]
    assert "MOTOR" not in compact["tags"]
    assert "SPECIAL CONSTRUCTION" not in compact["tags"]
    assert compact_attrs["selected_drive"] == "YES"
    assert compact_attrs["drive_sheave"] == "AK46"
    assert "drive_bushing" not in compact_attrs
    assert compact_attrs["driven_sheave"] == "AK30H"
    assert compact_attrs["driven_bushing"] == "H"
    assert compact_attrs["actual_sf"] == "1.25"
    assert compact_attrs["actual_cd"] == "21.14"

    center = li.extract_items([
        "Center Distance with allowance for install and take-up: 32.24 - 35.24",
    ])[0]
    assert "DRIVE COMPONENTS" in center["tags"]
    assert "V-BELT DRIVE" not in center["tags"]
    assert center["attributes"]["drive_subcategory"] == "CENTER DISTANCE"
    assert center["attributes"]["center_distance_range"] == "32.24 - 35.24"

    wheel_bushing = li.extract_items([
        "Wheel, Taper Lock Bushing (Bore: 1-3/8) L 981.00",
    ])[0]
    assert "WHEEL" in wheel_bushing["tags"]
    assert "V-BELT DRIVE" not in wheel_bushing["tags"]
    assert "DRIVE COMPONENTS" not in wheel_bushing["tags"]
    assert wheel_bushing["attributes"]["component"] == "WHEEL"
    assert wheel_bushing["attributes"]["wheel_feature"] == "TAPER LOCK BUSHING"
    assert wheel_bushing["attributes"]["wheel_bore"] == '1-3/8"'

    warranty = li.derive_item_fields({
        "raw": "year warranty from date of shipment **Drive sets will have standard warranty offered by vendor",
        "details": [],
    })
    assert warranty["tags"] == ["WARRANTY"]

    anti_short = li.extract_items([
        "Vertical mounting plate with 4 studs at 90 degrees. L 1,057.00",
        "Tack and weld conduit box to housing. Anti-short",
        "bushing. Extend motor leads, includes flexible",
        "conduit and labor, Inquiry Num: 15-9-2056",
    ])[0]
    assert "SPECIAL CONSTRUCTION" in anti_short["tags"]
    assert "V-BELT DRIVE" not in anti_short["tags"]
    assert "DRIVE COMPONENTS" not in anti_short["tags"]


def test_vibration_isolation_attributes_and_motor_bearing_cleanup():
    spring = li.extract_items([
        "Isolator, Spring 1\" Deflection (364 - 405 Frame) C 511.00",
        "Furnished By: CBC",
        "Ship Direct (Freight Included)",
    ])[0]
    assert {"SHIPPING", "VIBRATION ISOLATION"} <= set(spring["tags"])
    assert spring["attributes"]["component"] == "VIBRATION ISOLATION"
    assert spring["attributes"]["vibration_isolation_type"] == "SPRING ISOLATOR"
    assert spring["attributes"]["isolation_deflection"] == '1"'
    assert spring["attributes"]["isolation_frame"] == "364 - 405"
    assert spring["attributes"]["furnished_by"] == "CBC"

    rubber = li.extract_items([
        "Isolator, Rubber 1/2\" Deflection (0 - 326 Frame) C 314.00",
        "Furnished By: CBC",
    ])[0]
    assert rubber["attributes"]["vibration_isolation_type"] == "RUBBER ISOLATOR"
    assert rubber["attributes"]["isolation_deflection"] == '1/2"'

    base = li.extract_items(["Spring type vibration base with Isolators (Ship Direct) L 1,346.00"])[0]
    assert "VIBRATION ISOLATION" in base["tags"]
    assert base["attributes"]["vibration_isolation_type"] == "VIBRATION BASE, SPRING ISOLATOR"

    motor_bearing = li.extract_items([
        "Motor (EM4400T-G C 5,386.17",
        "M2F Add Isolated Bearings for NEMA 404 to 405",
    ])[0]
    assert "MOTOR" in motor_bearing["tags"]
    assert "VIBRATION ISOLATION" not in motor_bearing["tags"]

    warranty = li.derive_item_fields({
        "raw": "carry only the warranty passed onto us by the component manufacturer. Motor has warranty. Isolators have 1 year warranty",
        "details": [],
    })
    assert "WARRANTY" in warranty["tags"]
    assert "VIBRATION ISOLATION" not in warranty["tags"]


def test_warranty_attributes_do_not_leak_component_tags():
    warranty = li.derive_item_fields({
        "raw": "carry only the warranty passed onto us by the component manufacturer. Motor has a standard 3-year warranty. Isolators have 1 year warranty",
        "details": [],
    })
    assert warranty["tags"] == ["WARRANTY"]
    attrs = warranty["attributes"]
    assert attrs["note_type"] == "WARRANTY"
    assert attrs["warranty_duration"] == "3 YEARS, 1 YEAR"
    assert attrs["warranty_scope"] == "MOTOR, VIBRATION ISOLATION"
    assert attrs["warranty_source"] == "COMPONENT MANUFACTURER"

    motor = li.derive_item_fields({
        "raw": "Motor (ABB C 26,730.00",
        "details": [
            "Warranty 3 years, S.S. Nameplate, IP55 Protection",
            "300HP, 1200RPM, 449TC Frame, Inverter Duty 20:1 VT, 4:1 CT",
        ],
    })
    assert "MOTOR" in motor["tags"]
    assert "WARRANTY" not in motor["tags"]
    assert motor["attributes"]["motor_warranty"] == "3 YEARS"


def test_unitary_base_attributes():
    channel = li.extract_items(["8\" Channel Base, Inquiry Num: 253-24-1651 L 2,215.00"])[0]
    assert "UNITARY BASE" in channel["tags"]
    assert channel["attributes"]["component"] == "UNITARY BASE"
    assert channel["attributes"]["unitary_base_type"] == "CHANNEL BASE"
    assert channel["attributes"]["unitary_base_size"] == '8"'

    common = li.extract_items([
        "Special Product Design (Base Fan with CW and CCW L 24,032.00",
        "Wheel and Housing, Dual Shaft Motor and Common",
        "Unitary Base, Inquiry Num: 333-26-1551)",
    ])[0]
    assert "UNITARY BASE" in common["tags"]
    assert common["attributes"]["unitary_base_type"] == "UNITARY BASE"
    assert common["attributes"]["unitary_base_detail"] == "COMMON UNITARY BASE"

    outlet_note = li.extract_items([
        "slip outlet-to extend past channel base by 2\", L 1,641.00",
    ])[0]
    assert {"OUTLET", "UNITARY BASE"} <= set(outlet_note["tags"])
    assert outlet_note["attributes"]["unitary_base_detail"] == "OUTLET EXTENDS PAST CHANNEL BASE"
    assert outlet_note["attributes"]["unitary_base_clearance"] == '2"'


def test_inquiry_number_attributes():
    items = {it["norm"]: it for it in li.extract_items([
        "3D Drawings, InquiryNum:333-25-1622 L",
        "Access Door, Quick Clamp @ 10:30 position, Inquiry L 645.00",
        "Num: 352-23-2696",
        "Base Fan (Base fan, Suitable for 3600rpm Motor, L 5,761.00",
        "Inquiry Num: 317-26-1510)",
    ])}
    drawing_attrs = items["3D DRAWINGS INQUIRYNUM 333 25 1622"]["attributes"]
    assert "3D STEP DRAWINGS" in items["3D DRAWINGS INQUIRYNUM 333 25 1622"]["tags"]
    assert drawing_attrs["inquiry_num"] == "333-25-1622"
    assert drawing_attrs["component"] == "3D STEP DRAWINGS"
    assert "drawing_type" not in drawing_attrs and "drawing_scope" not in drawing_attrs
    assert items["ACCESS DOOR QUICK CLAMP 10 30 POSITION INQUIRY"]["attributes"]["inquiry_num"] == "352-23-2696"
    assert items["BASE FAN BASE FAN SUITABLE FOR 3600RPM MOTOR"]["attributes"]["inquiry_num"] == "317-26-1510"


def test_numeric_continuations_capture_clock_and_split_inquiry():
    actuator = li.extract_items([
        "Inlet Volume Control, Low Leak, Automatic L 3,531.00",
        "Handle Location (viewed from inlet side):",
        "@3:00",
        "Actuator Manufacturer: By Others",
        "Fail Position Upon Loss of Supply/Air/Power:",
        "Will Advise",
    ])[0]
    assert "@3:00" in actuator["details"]
    assert actuator["attributes"]["handle_location"] == "3:00"
    assert actuator["attributes"]["fail_position_upon_loss_of_power"] == "Will Advise"
    flags = actuator.get("review_flags") or []
    assert not any("Handle Location" in flag or "Fail Position" in flag for flag in flags)

    outlet = li.extract_items([
        "Outlet, Flanged, Unpunched, Inquiry Num: 361-26- L 250.00",
        "2440",
    ])[0]
    assert outlet["details"] == ["2440"]
    assert outlet["attributes"]["inquiry_num"] == "361-26-2440"


def test_421967_review_notes_are_general_parser_rules():
    items = li.extract_items([
        "Access Door, Quick Clamp L 525.00",
        "Door Location: @9:00",
        "Include 3D STEP Drawings L NC",
        "Inlet Volume Control Handle Location, Non-standard L 1,014.00",
        "IVC handle location for Discharge",
        "Inlet Volume Control, Low Leak, Automatic L 3,531.00",
        "Handle Location (viewed from inlet side):",
        "@3:00",
        "Actuator Manufacturer: By Others",
        "Fail Position Upon Loss of Supply/Air/Power:",
        "Will Advise",
        "Fail Position Upon Loss of Signal: Will Advise",
        "Inlet, Flanged, Punched (with IVC) L 1,559.00",
        "Lifting Lugs L STD",
        "Mechanical Run Test, Standard L STD",
        "Outlet, Flanged, Punched L STD",
        "Stainless steel nameplate L 88.00",
        "Outlet, Flanged, Unpunched, Inquiry Num: 361-26- L 250.00",
        "2440",
        "Additional Features / Notes:",
        "SHIP WITH FANS REFERENCED BY CUSTOMER PO(s) 23971 (nOTE THERE ARE TWO SLIGHTLY DIFFERENT CBC FAN ON THIS",
        "SINGLE PURCHASE ORDER(",
        "Run Test - Required",
        "Fan Drawings:",
    ])
    by_component = {item.get("attributes", {}).get("component"): item for item in items
                    if item.get("attributes", {}).get("component")}

    door = by_component["ACCESS DOOR"]
    assert door["attributes"]["door_type"] == "QUICK CLAMP"
    assert door["attributes"]["door_location"] == "9:00"
    assert by_component["3D STEP DRAWINGS"]["attributes"] == {
        "component": "3D STEP DRAWINGS"
    }

    ivc_lines = [item for item in items if item.get("attributes", {}).get("used_on") == "IVC"]
    assert len(ivc_lines) == 2
    handle = next(item for item in ivc_lines if "HANDLE LOCATION" in item["norm"])
    actuator = next(item for item in ivc_lines if "LOW LEAK" in item["norm"])
    assert handle["attributes"]["handle_location"] == "NON-STD"
    assert "handle_location_reference" not in handle["attributes"]
    assert actuator["attributes"]["fail_position_upon_loss_of_power"] == "Will Advise"
    assert actuator["attributes"]["fail_position_upon_loss_of_signal"] == "Will Advise"
    assert "ivc_subcategory" not in actuator["attributes"]

    inlet = by_component["INLET"]
    assert inlet["tags"] == ["FLANGE", "INLET"]
    assert inlet["attributes"]["flanged"] == "YES"
    assert inlet["attributes"]["punched"] == "YES"
    assert "flange_type" not in inlet["attributes"]
    assert "used_on" not in inlet["attributes"] and "ivc_relation" not in inlet["attributes"]
    assert "flange_scope" not in inlet["attributes"]

    lugs = by_component["HOUSING"]
    assert lugs["attributes"] == {"component": "HOUSING", "lifting_lugs": "YES"}
    run_tests = [item for item in items
                 if item.get("attributes", {}).get("component") == "MECHANICAL RUN TEST"]
    assert len(run_tests) == 2
    assert all(item["attributes"] == {"component": "MECHANICAL RUN TEST"}
               for item in run_tests)
    outlet_lines = [item for item in items
                    if item.get("attributes", {}).get("component") == "OUTLET"]
    assert len(outlet_lines) == 2 and all(
        "flange_scope" not in item["attributes"] for item in outlet_lines)
    assert {item["attributes"]["punched"] for item in outlet_lines} == {"YES", "NO"}
    assert by_component["NAMEPLATE"]["attributes"] == {
        "component": "NAMEPLATE", "material": "STAINLESS STEEL"
    }
    ship_with = by_component["SHIP WITH"]
    assert ship_with["details"] == ["SINGLE PURCHASE ORDER("]
    assert ship_with["attributes"] == {
        "component": "SHIP WITH",
        "instruction": (
            "SHIP WITH FANS REFERENCED BY CUSTOMER PO(s) 23971 "
            "(nOTE THERE ARE TWO SLIGHTLY DIFFERENT CBC FAN ON THIS "
            "SINGLE PURCHASE ORDER("
        ),
    }
    assert not any(item["norm"] == "SINGLE PURCHASE ORDER" for item in items)
    assert not any(item.get("review_flags") for item in items)


def test_421959_repair_order_wrapping_material_and_ispm():
    items = li.extract_items([
        "Chicago Blower Corporation Sales Order For Repair Parts",
        "Replacement Shaft for fan manufactured by Chicago N 9,123.00 1,825.00",
        "Blower Oceania (Qty: 1), Inquiry Num: 990-26-",
        "1764RP",
        "Replacement Wheel for fan manufactured by N 15,318.00 3,064.00",
        "Chicago Blower Oceania (Qty: 1), Inquiry Num:",
        "990-26-1764RP",
        "ISPM Wood Inspection Stamp L INC",
        "Additional Features / Notes:",
        "Construction SS316",
        "SERIAL NUMBER UNKNOWN",
        "Customer assumes responsibility for fit-up and suitability of part(s) due to unknown serial number",
        "Special Invoicing:",
    ])
    assert len(items) == 5
    by_component = {item.get("attributes", {}).get("component"): item for item in items
                    if item.get("attributes", {}).get("component")}

    for component in ("REPLACEMENT SHAFT", "REPLACEMENT WHEEL"):
        part = by_component[component]
        assert part["attributes"]["manufacturer"] == "Chicago Blower Oceania"
        assert part["attributes"]["inquiry_num"] == "990-26-1764RP"
        assert "spare_part_type" not in part["attributes"]
        assert "spare_part_component" not in part["attributes"]
        assert not part.get("review_flags")
    assert by_component["REPLACEMENT SHAFT"]["details"][-1] == "1764RP"
    assert by_component["REPLACEMENT WHEEL"]["details"][-1] == "990-26-1764RP"

    inspection = by_component["INSPECTION"]
    assert inspection["attributes"]["inspection_subcategory"] == "ISPM WOOD STAMP"
    parts_order = by_component["PARTS-ONLY ORDER"]
    assert "PARTS-ONLY ORDER" in parts_order["tags"]
    assert parts_order["attributes"]["parts_only"] == "YES"
    assert parts_order["attributes"]["material"] == "STAINLESS STEEL"
    assert parts_order["attributes"]["material_grade"] == "316 SS"
    assert parts_order["attributes"]["material_scope"] == "ORDER PARTS"
    assert "BASE FAN" not in by_component
    assert not any(item["norm"] == "SERIAL NUMBER UNKNOWN" for item in items)
    responsibility = next(item for item in items if "ASSUMES RESPONSIBILITY" in item["norm"])
    assert responsibility["tags"] == ["MISC NOTE"]


def test_422004_customer_motor_and_requested_component_schemas():
    items = li.extract_items([
        "Base Fan L 8,000.00",
        "Motor (Customer Provided) N NA",
        "Vendor: Baldor",
        "15 HP, 1800 RPM, Enclosure: TEFC Premium",
        "254T, Cast Iron, Foot Mounted",
        "3/60/230/460, F1, 1.15 SF",
        "Model: EM2333T",
        "Mounted by Others",
        "Mounted by Others L",
        "Extended Grease Fittings L 100.00",
        "Housing Drain with Plug L STD",
        "Housing, Heavy Duty L 250.00",
        "Inlet, Slip L STD",
        "Outlet, Flanged, Punched L STD",
        "Belt Guard, Painted Safety Yellow L 500.00",
        "Additional Features / Notes:",
        "All Chicago Blower wheels are precision balanced; however when a precision balanced wheel is installed in a fan, the wheel may have",
        "to be phase balanced to compensate for any imbalance or eccentricity in other components in the drive train. If the motor and drives",
        "are not mounted by Chicago Blower, we cannot check the assembled rotating system for vibration, and phase balance by others may",
        "Special Invoicing:",
    ], order_context={"arrangement": "A/9S"})
    by_component = {}
    for item in items:
        component = item.get("attributes", {}).get("component")
        if component:
            by_component.setdefault(component, []).append(item)

    motor = by_component["MOTOR"][0]
    assert motor["attributes"]["motor_supplied_by"] == "CUSTOMER"
    assert motor["attributes"]["motor_hp"] == "15"
    assert motor["attributes"]["motor_rpm"] == "1800"
    assert motor["attributes"]["motor_model"] == "EM2333T"
    assert motor["attributes"]["vendor"] == "Baldor"
    assert not motor.get("review_flags")
    mounting = next(item for item in by_component["MOTOR"] if item["norm"] == "MOUNTED BY OTHERS")
    assert mounting["attributes"]["motor_mounted_by"] == "OTHERS"
    assert not mounting.get("review_flags")

    grease = by_component["BEARINGS"][0]
    assert grease["attributes"]["extended_grease_fittings"] == "YES"
    assert not grease.get("review_flags")
    assert by_component["HOUSING DRAIN"][0]["attributes"]["size"] == "STD"
    assert by_component["HOUSING"][0]["attributes"]["heavy_duty"] == "YES"
    assert by_component["INLET"][0]["attributes"] == {
        "component": "INLET", "flanged": "NO", "punched": "NO",
        "inlet_subcategory": "SLIP",
    }
    assert by_component["OUTLET"][0]["attributes"] == {
        "component": "OUTLET", "flanged": "YES", "punched": "YES",
    }
    guard_attrs = by_component["BELT GUARD"][0]["attributes"]
    assert guard_attrs["coating"] == "PAINTED SAFETY YELLOW"
    assert not any(key.startswith("coating_") for key in guard_attrs)
    disclaimer = next(item for item in items if item["norm"].startswith(
        "ALL CHICAGO BLOWER WHEELS ARE PRECISION BALANCED"
    ))
    assert disclaimer["tags"] == ["MISC NOTE"]
    assert len(disclaimer["details"]) == 2
    assert not disclaimer.get("review_flags")


def test_422003_accessory_boundaries_motor_mods_and_shared_metadata():
    items = li.extract_items([
        "Motor C 4,000.00",
        "Vendor: Baldor",
        "Product: 11355340 W/ MOD N10 (NP CHANGE); MOD N20 (CUSTOMER TAG (BM- 101) AND MOD B220 (INSULATED NDE BEARING))",
        "Paint: CBC Standard Black L STD",
        "G2.5 Balance L 341.00",
        "Housing Drain with Plug L STD",
        "Inlet, Flanged, Punched without IVC L STD",
        "Outlet, Flanged, Punched L STD",
        "Threaded Plug for Conduit Box Opening, Inquiry Num: 333-26-1000 L 50.00",
        "Inlet Silencer C 2,500.00",
        "Painted to match fan",
        "VAW to provide PDF, AutoCAD and 3D drawings",
        "Ship direct, best way PP&A, mark with THM job",
        "number MJ26-821.",
        "TAG Silencer SIL-101",
        "Vendor: VAW Systems",
        "Product: Inlet Silencer",
        "Outlet Expantion Joint C 3,000.00",
        "Fiberglass Sound Pillow, 9 inch F/F",
        "Painted BUB",
        "Punched to match outlet flange",
        "Ship Direct, PP&A",
        "FlexCom drawing for the record",
        "Mark box and drawing with THM job number MJ26-821",
        "Vendor: FlexibleCompensators",
        "Product: Expansion Joint",
        "List Total 10,000.00",
        "Lead Time for Inlet Silencer 8 weeks",
        "RH painted to match fan",
        "VAW drawing after approval",
        "Additional Features / Notes:",
        "Job Name: THM MJ26-821",
        "Location: Howard Energy/Spitzer H-101",
        "CBC to show fan plus accessory weight, motor weight and total weight on fan drawing",
        "Special Invoicing:",
    ])
    by_component = {item.get("attributes", {}).get("component"): item for item in items
                    if item.get("attributes", {}).get("component")}

    motor = by_component["MOTOR"]
    assert motor["attributes"]["product"] == "11355340"
    assert motor["attributes"]["motor_mod_n10"] == "NP CHANGE"
    assert motor["attributes"]["motor_mod_n20"] == "CUSTOMER TAG BM-101"
    assert motor["attributes"]["motor_mod_b220"] == "INSULATED NDE BEARING"

    assert by_component["PAINT"]["attributes"]["component"] == "PAINT"
    balance = by_component["BALANCE"]["attributes"]
    assert balance["balance_grade"] == "G2.5" and "balance_type" not in balance
    assert by_component["HOUSING DRAIN"]["attributes"]["size"] == "STD"
    threaded = by_component["THREADED PLUG FOR CONDUIT BOX OPENING"]
    assert threaded["attributes"] == {
        "inquiry_num": "333-26-1000",
        "component": "THREADED PLUG FOR CONDUIT BOX OPENING",
    }

    silencer = by_component["INLET SILENCER"]
    assert silencer["attributes"]["coating"] == "PAINTED TO MATCH FAN"
    assert silencer["attributes"]["drawing_requirement"] == "PDF, AUTOCAD, 3D"
    assert silencer["attributes"]["shipping_method"] == "SHIP DIRECT"
    assert silencer["attributes"]["job_number"] == "MJ26-821"
    assert silencer["attributes"]["tag"] == "SIL-101"
    assert silencer["attributes"]["vendor"] == "VAW Systems"
    assert "product" not in silencer["attributes"]
    assert not silencer.get("review_flags")

    expansion = by_component["OUTLET EXPANSION JOINT"]
    assert expansion["attributes"]["insulation"] == "FIBERGLASS SOUND PILLOW"
    assert expansion["attributes"]["face_to_face"] == '9"'
    assert expansion["attributes"]["coating"] == "PAINTED BUB"
    assert expansion["attributes"]["punched_to_match"] == "OUTLET FLANGE"
    assert expansion["attributes"]["drawing_requirement"] == "DRAWING FOR RECORD"
    assert expansion["attributes"]["shipping_method"] == "SHIP DIRECT"
    assert expansion["attributes"]["job_number"] == "MJ26-821"
    assert expansion["attributes"]["vendor"] == "FlexibleCompensators"
    assert "product" not in expansion["attributes"]
    assert all("RH painted" not in detail and "VAW drawing after" not in detail
               for detail in expansion["details"])
    assert not expansion.get("review_flags")

    weights = by_component["FAN DRAWING WEIGHTS"]
    assert weights["attributes"] == {"component": "FAN DRAWING WEIGHTS"}
    assert not any(item["norm"].startswith("JOB NAME") or item["norm"].startswith("LOCATION")
                   for item in items)


def test_order_context_survives_store_renormalization():
    items = li.extract_items([
        "Inlet Silencer C 1,000.00",
        "Product: Inlet Silencer",
        "Outlet Expansion Joint C 1,000.00",
        "Product: Expansion Joint",
        "Extended Grease Fittings L 100.00",
        "Additional Features / Notes:",
        "Job Name: THM MJ26-821",
        "Special Invoicing:",
    ], order_context={"arrangement": "A/9S"})
    store = {"jobs": {}, "ai_tags": {}}
    li.record_job(store, "422003", items, arrangement="A/9S", job_number="MJ26-821")
    li.renormalize_store(store)
    record = store["jobs"]["422003"]
    assert record["arrangement"] == "A/9S"
    assert record["job_number"] == "MJ26-821"
    by_component = {item.get("attributes", {}).get("component"): item
                    for item in record["items"]}
    assert by_component["BEARINGS"]["attributes"]["extended_grease_fittings"] == "YES"
    assert by_component["INLET SILENCER"]["attributes"]["job_number"] == "MJ26-821"
    assert by_component["OUTLET EXPANSION JOINT"]["attributes"]["job_number"] == "MJ26-821"

    legacy_items = li.extract_items([
        "Inlet Silencer C 1,000.00",
        "Product: Inlet Silencer",
    ])
    legacy_items.append({
        "raw": "Job Name: THM MJ26-822",
        "norm": "JOB NAME THM MJ26 822",
        "tags": ["MISC NOTE"],
        "attributes": {},
        "details": [],
    })
    legacy_store = {"jobs": {"422005": {"items": legacy_items}}, "ai_tags": {}}
    li.renormalize_store(legacy_store)
    legacy_record = legacy_store["jobs"]["422005"]
    assert legacy_record["job_number"] == "MJ26-822"
    assert len(legacy_record["items"]) == 1
    assert legacy_record["items"][0]["attributes"]["job_number"] == "MJ26-822"

    repair_items = li.extract_items([
        "Chicago Blower Corporation Sales Order For Repair Parts",
        "Replacement Shaft N 500.00",
        "Additional Features / Notes:",
        "Construction SS316",
        "Special Invoicing:",
    ])
    legacy_store = {"jobs": {"421959": {"items": repair_items}}, "ai_tags": {}}
    li.renormalize_store(legacy_store)
    legacy = legacy_store["jobs"]["421959"]
    assert legacy["parts_only"] is True
    assert any(item.get("attributes", {}).get("component") == "PARTS-ONLY ORDER"
               for item in legacy["items"])


def test_line_item_rescan_uses_atomic_order_context_records():
    import line_items_scan as scanner
    import sales_orders

    captured = []
    original_load = scanner.li.load_store
    original_record = scanner.li.record_jobs_atomic
    original_parse = sales_orders.parse_sales_order_pdf
    scanner.li.load_store = lambda: {"jobs": {}, "ai_tags": {}}
    scanner.li.record_jobs_atomic = lambda rows: (
        captured.extend(dict(row) for row in rows) or len(rows)
    )
    sales_orders.parse_sales_order_pdf = lambda _path: {
        "line_items": [{"raw": "Extended Grease Fittings", "tags": ["BEARINGS"]}],
        "arrangement": "A/9S",
        "parts_only": False,
        "job_number": "MJ26-821",
    }
    try:
        assert scanner.scan([("422004", Path("422004.pdf"), 2)], True, 0) == 0
    finally:
        scanner.li.load_store = original_load
        scanner.li.record_jobs_atomic = original_record
        sales_orders.parse_sales_order_pdf = original_parse

    assert len(captured) == 1
    assert captured[0]["job"] == "422004"
    assert captured[0]["co_number"] == 2
    assert captured[0]["arrangement"] == "A/9S"
    assert captured[0]["parts_only"] is False
    assert captured[0]["job_number"] == "MJ26-821"


def test_document_mark_ship_to_and_continuation_metadata():
    recon = [
        "Order # Rep Ref. # Customer P.O. # Fan Serial Number:",
        "421967 NATIONWIDE BOILER - For DARLENE PO 23971",
        "INGREDIENTS - FAN For BOILER NO 1",
        "Sold To: Ship To:",
        "NATIONWIDE BOILER, INC DARLINGING INREDIENTS COMPANY",
        "42400 CHRISTY ST. 11946 CARPENTER ROAD",
        "FREMONT, CA 94538 CROWS LANDING, CA 95313",
        "UNITED STATES OF AMERICA UNITED STATES OF AMERICA",
        "TRAFFIC SEE BELOW",
        "Mark (shipping documents):",
        "FD FAN for BOILER No 1 - NATIONWIDE BOILERP.O. PO 23971",
    ]
    tables = [[[
        "Order#\n421967",
        "RepRef.#\nNATIONWIDEBOILER-ForDARLENE\nINGREDIENTS-FANForBOILERNO1",
        "CustomerP.O.#\nPO23971",
        "FanSerialNumber:",
    ], [
        "SoldTo:\nNATIONWIDEBOILER,INC\n42400CHRISTYST.\nFREMONT,CA94538\nUNITEDSTATESOFAMERICA",
        "ShipTo:\nDARLINGINGINREDIENTSCOMPANY\n11946CARPENTERROAD\nCROWSLANDING,CA95313\nUNITEDSTATESOFAMERICA\nTRAFFICSEEBELOW",
    ], [
        "Mark(shippingdocuments):\nFDFANforBOILERNo1-NATIONWIDEBOILERP.O.PO23971",
    ]]]
    facts = {item["document_fact"]: item
             for item in li.document_fact_items_from_tables(tables, recon)}
    assert set(facts) == {"MARK", "SHIP TO"}
    mark = facts["MARK"]
    assert mark["attributes"]["component"] == "MARK"
    assert mark["attributes"]["mark_text"] == (
        "FD FAN for BOILER No 1 - NATIONWIDE BOILERP.O. PO 23971"
    )
    assert "rep_reference" not in mark["attributes"]
    ship = facts["SHIP TO"]
    assert ship["attributes"]["ship_to_company"] == "DARLINGING INREDIENTS COMPANY"
    assert ship["attributes"]["ship_to_address"] == (
        "11946 CARPENTER ROAD | CROWS LANDING, CA 95313"
    )
    assert ship["attributes"]["ship_to_country"] == "UNITED STATES OF AMERICA"
    assert ship["attributes"]["ship_to_instruction"] == "TRAFFIC SEE BELOW"
    assert li.derive_item_fields(ship)["attributes"] == ship["attributes"]

    legacy_mark = li.document_fact_item(
        "MARK", {"mark_text": "KEEP ME", "rep_reference": "REMOVE ME"},
        ["KEEP ME", "Rep Ref.: REMOVE ME"],
    )
    legacy_ship = li.document_fact_item(
        "SHIP TO", {
            "ship_to_company": "ACME",
            "ship_to_address_1": "1 MAIN ST",
            "ship_to_address_2": "CHICAGO, IL 60601",
            "ship_to_instruction_1": "CALL AHEAD",
        },
        ["ACME", "1 MAIN ST", "CHICAGO, IL 60601", "CALL AHEAD"],
    )
    store = {"jobs": {"1": {"items": [legacy_mark, legacy_ship]}}, "ai_tags": {}}
    li.renormalize_store(store)
    clean_mark, clean_ship = store["jobs"]["1"]["items"]
    assert clean_mark["document_attributes"] == {"mark_text": "KEEP ME"}
    assert clean_mark["details"] == ["KEEP ME"]
    assert clean_ship["document_attributes"]["ship_to_address"] == (
        "1 MAIN ST | CHICAGO, IL 60601"
    )
    assert clean_ship["document_attributes"]["ship_to_instruction"] == "CALL AHEAD"
    assert not any(key.endswith("_1") or key.endswith("_2")
                   for key in clean_ship["document_attributes"])

    continuation_lines = [
        "Inlet, Flanged, Punched (with IVC) L 1,559.00",
        "Chicago Blower Corporation Sales Order (cont.)",
        "Order # Rep Ref. # Customer P.O. # Page # of #",
        "421967 NATIONWIDE BOILER - For PO 23971",
        "DARLENE INGREDIENTS - FAN For",
        "BOILER NO 1",
        "Lifting Lugs L STD",
    ]
    continuation_table = [[[
        "Order #\n421967",
        "Rep Ref. #\nNATIONWIDE BOILER - For\nDARLENE INGREDIENTS - FAN For\nBOILER NO 1",
        "Customer P.O. #\nPO 23971",
        "Page # of #",
    ]]]
    clean = li.strip_continuation_metadata(continuation_lines, continuation_table)
    assert "421967 NATIONWIDE BOILER - For PO 23971" not in clean
    inlet = li.extract_items(clean)[0]
    assert inlet["details"] == []


def test_drawing_note_attributes():
    items = li.extract_items([
        "Add COG and Weights to the fan drawing, Inquiry L",
        "Num: 333-25-1622",
        "Add BOM to fan drawing, Inquiry Num: 333-25-1622 L",
        "Plan View on Customer Drawing, Inquiry Num: 333-25-1622 L",
        "ADDITIONAL FEATURES",
        "Static and Dymanic Loads on drawing, Inquiry Num: L 500.00",
        "ADD CG TO FAN DRAWING",
        "SHOW GROUNDING LUGS ON FAN DRAWING",
        "Please update the fan / drawing to show the inlet box rotated 270 degrees from the motor side to the 9 o'clock position.",
        "Add tag to be included on the expansion joint drawing: 1A-EXJ-AXS812",
        "Should add DO NOT STACK and ISPM Stamping to drawing.",
        "Change Subject line on Drawing Transmittal to include tag information as it appears on the order face, Inquiry Num: 364-20-200",
        "Only change on drawings are for CAT purposes added a 1E2999 Symbols",
        "DRAWING AND SPLIT HOUSING AND RELEASED INTO PRODUCTION",
    ])
    by_raw = {it["raw"]: it for it in items}
    weights = by_raw["Add COG and Weights to the fan drawing, Inquiry L"]["attributes"]
    assert weights["inquiry_num"] == "333-25-1622"
    assert weights["drawing_type"] == "WEIGHTS/COG/LOADS"
    assert weights["drawing_scope"] == "FAN"
    assert by_raw["Add BOM to fan drawing, Inquiry Num: 333-25-1622 L"]["attributes"]["drawing_type"] == "BOM"
    assert by_raw["Plan View on Customer Drawing, Inquiry Num: 333-25-1622 L"]["attributes"]["drawing_type"] == "PLAN VIEW/CUSTOMER DRAWING"
    assert by_raw["Static and Dymanic Loads on drawing, Inquiry Num: L 500.00"]["attributes"]["drawing_type"] == "WEIGHTS/COG/LOADS"
    assert by_raw["ADD CG TO FAN DRAWING"]["attributes"]["drawing_type"] == "WEIGHTS/COG/LOADS"
    assert by_raw["SHOW GROUNDING LUGS ON FAN DRAWING"]["attributes"]["drawing_type"] == "GROUNDING/LUGS"
    assert by_raw["Please update the fan / drawing to show the inlet box rotated 270 degrees from the motor side to the 9 o'clock position."]["attributes"]["drawing_type"] == "ORIENTATION/ARRANGEMENT"
    expansion_drawing = by_raw[
        "Add tag to be included on the expansion joint drawing: 1A-EXJ-AXS812"
    ]["attributes"]
    assert expansion_drawing["component"] == "EXPANSION JOINT"
    assert expansion_drawing["drawing_requirement"] == "TAG/MARKING"
    assert by_raw["Should add DO NOT STACK and ISPM Stamping to drawing."]["attributes"]["drawing_type"] == "PACKAGING/MARKING"
    assert by_raw["Change Subject line on Drawing Transmittal to include tag information as it appears on the order face, Inquiry Num: 364-20-200"]["attributes"]["drawing_type"] == "TAG/MARKING, DRAWING TRANSMITTAL"
    assert by_raw["Only change on drawings are for CAT purposes added a 1E2999 Symbols"]["attributes"]["drawing_type"] == "CUSTOMER SYMBOLS/NOTES"
    split = by_raw["DRAWING AND SPLIT HOUSING AND RELEASED INTO PRODUCTION"]
    assert "SPLIT HOUSING" in split["tags"]
    assert "HOUSING" not in split["tags"]
    assert split["attributes"]["drawing_type"] == "SPLIT HOUSING"


def test_split_housing_attributes():
    items = {it["raw"]: it for it in li.extract_items([
        "Horizontal Split Housing L 3,567.00",
        "Pie Wedge Split Housing L 2,303.00",
        "Shipping Split L 1,000.00",
        "Housing split for shipment L 1,000.00",
    ])}
    horizontal = items["Horizontal Split Housing L 3,567.00"]
    assert horizontal["tags"] == ["SPLIT HOUSING"]
    assert horizontal["attributes"]["component"] == "SPLIT HOUSING"
    assert horizontal["attributes"]["split_type"] == "HORIZONTAL"
    assert items["Pie Wedge Split Housing L 2,303.00"]["attributes"]["split_type"] == "PIE WEDGE"
    assert items["Shipping Split L 1,000.00"]["attributes"]["split_type"] == "SHIPPING"
    assert items["Housing split for shipment L 1,000.00"]["attributes"]["split_type"] == "SHIPPING"


def test_explosion_proof_split_detail_line():
    motor = li.extract_items([
        "Motor C 633.61",
        "Vendor: WEG",
        "Enclosure: Premium Explosion",
        "Proof",
    ])[0]
    assert "MOTOR" in motor["tags"]
    assert "EXPLOSION PROOF" not in motor["tags"]
    assert motor["attributes"]["component"] == "MOTOR"
    assert motor["attributes"]["motor_enclosure"] == "EXPLOSION PROOF"

    warranty = li.extract_items([
        "Exclusive 3 Year Warranty L INC",
        "Enclosure: Premium Explosion",
        "Proof",
    ])[0]
    assert warranty["tags"] == ["WARRANTY"]
    assert "motor_enclosure" not in warranty["attributes"]


def test_flange_scope_motor_and_non_wheel_end():
    motor = li.extract_items(["C-Flange Motor Replate L 100.00"])[0]
    assert "MOTOR" in motor["tags"]
    assert "MOUNTING" not in motor["tags"]
    assert "FLANGE" not in motor["tags"]
    assert motor["attributes"]["component"] == "MOTOR"
    assert motor["attributes"]["motor_mounting"] == "C-FLANGE"
    assert motor["attributes"]["motor_nameplate"] == "YES"
    assert motor["attributes"]["motor_nameplate_action"] == "REPLATE"
    assert motor["attributes"]["flange_scope"] == "MOTOR"

    housing = li.extract_items(["Non-Wheel End Flange Reinforcing Gussets L 200.00"])[0]
    assert "FLANGE" in housing["tags"]
    assert "HOUSING" in housing["tags"]
    assert "WHEEL" not in housing["tags"]
    assert housing["attributes"]["flange_scope"] == "HOUSING"
    assert housing["attributes"]["flange_location"] == "NON-WHEEL END"


def test_unpunched_flange_is_not_punched():
    # 421966 item 15: "Unpunched" contains "Punched", and the substring check
    # used to stamp punched=YES on it.
    outlet = li.extract_items(["Outlet, Flanged, Unpunched, Inquiry Num: 361-26- L 250.00"])[0]
    assert outlet["attributes"]["flanged"] == "YES"
    assert outlet["attributes"]["punched"] == "NO"
    inlet = li.extract_items(["Inlet, Flanged, Unpunched L 100.00"])[0]
    assert inlet["attributes"]["flanged"] == "YES"
    assert inlet["attributes"]["punched"] == "NO"
    assert "flange_type" not in inlet["attributes"]
    assert "inlet_subcategory" not in inlet["attributes"]

    removed = li.extract_items([
        "Remove outlet flange, Inquiry Num: 317-23-1131 L",
    ])[0]
    assert removed["attributes"]["component"] == "OUTLET"
    assert removed["attributes"]["flanged"] == "NO"
    assert removed["attributes"]["punched"] == "NO"
    assert removed["attributes"]["flange_instruction"] == "REMOVE"
    assert "flange_type" not in removed["attributes"]

    removed_flanged = li.extract_items([
        "Remove flanged Outlet, Inquiry Num: 317-25-1715 L",
    ])[0]
    assert removed_flanged["attributes"]["component"] == "OUTLET"
    assert removed_flanged["attributes"]["flanged"] == "NO"
    assert removed_flanged["attributes"]["punched"] == "NO"
    assert "flange_type" not in removed_flanged["attributes"]
    # The punched wording keeps its original attributes.
    punched = li.extract_items(["Outlet, Flanged, Punched L STD"])[0]
    assert punched["attributes"]["punched"] == "YES"


def test_canonical_component_names_from_prose():
    # Lines the SO names only in prose now carry a canonical `component`, so two
    # lines for the same thing merge (so_hierarchy) instead of standing as two.
    door = li.extract_items(["Access Door, Quick Clamp L 525.00",
                             "Door Location: @9:00"])[0]
    assert door["attributes"]["component"] == "ACCESS DOOR"
    assert door["attributes"]["door_location"] == "9:00"
    assert door["attributes"]["door_type"] == "QUICK CLAMP"

    pct = li.extract_items(["Percent Width (85%) L 1,252.00"])[0]
    assert pct["attributes"]["component"] == "PERCENT WIDTH"
    assert pct["attributes"]["pct_width_customer"] == "85%"
    assert pct["attributes"]["pct_width_rounded"] == "85%"
    # CBC builds to the nearest 5%: a customer value rounds to that step (a lower
    # value that rounds the same is the same construction).
    assert li.extract_items(["Percent Width (83%) L 1.00"])[0]["attributes"]["pct_width_rounded"] == "85%"
    assert li.extract_items(["Percent Width (82%) L 1.00"])[0]["attributes"]["pct_width_rounded"] == "80%"

    shrink = li.extract_items(['Shrink wrap, Wheel Dia. 21" to 36-1/2" L 278.00'])[0]
    assert shrink["attributes"]["component"] == "SHRINK WRAP"
    assert shrink["attributes"]["shrink_wrap_range"] == '21" to 36-1/2"'

    ship = li.extract_items(["SHIP WITH FANS REFERENCED BY CUSTOMER PO(s) 23971 L STD"])[0]
    assert ship["attributes"] == {
        "component": "SHIP WITH",
        "instruction": "SHIP WITH FANS REFERENCED BY CUSTOMER PO(s) 23971",
    }

    outlet = li.extract_items(["Outlet, Flanged, Punched L STD"])[0]
    assert outlet["attributes"]["component"] == "OUTLET"

    # Run test and mechanical run test are the ONE mechanical run test: same
    # component, and the bare "Run Test" synonym is canonicalized so two such
    # lines don't false-conflict on testing_type when they merge.
    mrt = li.extract_items(["Mechanical Run Test, Standard L STD"])[0]
    assert mrt["attributes"]["component"] == "MECHANICAL RUN TEST"
    # An unpriced "Run Test - Required" parses only inside a feature section.
    section = li.extract_items(["Additional Features / Notes:", "Run Test - Required"])
    run = next(i for i in section if "RUN TEST" in i["norm"])
    assert run["attributes"]["component"] == "MECHANICAL RUN TEST"
    assert not any(k in run["attributes"] for k in
                   ("testing_type", "testing_standard", "testing_required"))

    # A line already tied to a component (the IVC, via used_on) is NOT renamed —
    # but its handle location is still lifted to an attribute.
    handle = li.extract_items(["Inlet Volume Control Handle Location, Non-standard L 1,014.00",
                               "IVC handle location for Discharge"])[0]
    assert handle["attributes"].get("used_on") == "IVC"
    assert handle["attributes"]["handle_location"] == "NON-STD"
    assert "handle_location_reference" not in handle["attributes"]
    assert not handle.get("review_flags")


def test_housing_modified_flange_and_length_attributes():
    thickness = li.extract_items(['Housing Flange Thickness, 2-7/8", Inquiry Num: L 250.00'])[0]
    assert {"FLANGE", "HOUSING"} <= set(thickness["tags"])
    assert thickness["attributes"]["flange_scope"] == "HOUSING"
    assert thickness["attributes"]["housing_subcategory"] == "MODIFIED FLANGE"
    assert thickness["attributes"]["housing_feature"] == "FLANGE THICKNESS"

    bolt_pattern = li.extract_items([
        'Drum Housing Bolt Pattern, QTY (12) 9/16" L',
        'Diameter Holes, 34" B.C. Diameter, Straddle',
    ])[0]
    assert "HOUSING" in bolt_pattern["tags"]
    assert bolt_pattern["attributes"]["housing_subcategory"] == "MODIFIED FLANGE"
    assert bolt_pattern["attributes"]["housing_feature"] == "BOLT PATTERN"

    length = li.extract_items(['Housing Length 30-3/4" Flange to Flange, Inquiry L 650.00'])[0]
    assert {"FLANGE", "HOUSING"} <= set(length["tags"])
    assert length["attributes"]["housing_subcategory"] == "DIMENSION NOTE"
    assert length["attributes"]["housing_dimension"] == "FLANGE-TO-FLANGE LENGTH"

    mixing_box = li.extract_items(['Mixing Box Length 30" Flange to Flange L 650.00'])[0]
    assert "HOUSING" not in mixing_box["tags"]
    assert "housing_subcategory" not in mixing_box["attributes"]


def test_housing_construction_and_nameplate_mount_attributes():
    square = li.extract_items(["Housing, Square L 284.00"])[0]
    assert square["attributes"]["housing_subcategory"] == "SQUARE"

    universal = li.extract_items(["Housing, Universal Housing L 75.00"])[0]
    assert universal["attributes"]["housing_subcategory"] == "UNIVERSAL"

    stiffener = li.extract_items(["Housing Stiffner Plates, Inquiry Num: 311-19-2159 L 114.00"])[0]
    assert stiffener["attributes"]["housing_subcategory"] == "STIFFENER PLATES"

    casing = li.extract_items(["Housing, Casing Extension for Outlet L 3,213.00"])[0]
    assert casing["attributes"]["housing_subcategory"] == "CASING EXTENSION"
    assert casing["attributes"]["used_on"] == "OUTLET"

    nameplate = li.extract_items([
        "316 SS Nameplate, Riveted to Housing with 316 SS L 1,179.00",
        "Hardware, Inquiry Num: 253-26-1710",
    ])[0]
    assert {"HOUSING", "MATERIALS", "NAMEPLATE", "STAINLESS STEEL"} <= set(nameplate["tags"])
    assert nameplate["attributes"]["housing_subcategory"] == "NAMEPLATE LOCATION"
    assert nameplate["attributes"]["mount_location"] == "HOUSING"
    assert nameplate["attributes"]["nameplate_mount_location"] == "HOUSING"
    assert nameplate["attributes"]["nameplate_mounting"] == "RIVETED"
    assert nameplate["attributes"]["material_scope"] == "NAMEPLATE, HARDWARE"


def test_conduit_box_location_is_motor_not_housing():
    conduit = li.extract_items([
        "mount conduit box as close to the housing as L",
        "possible., Inquiry Num: 311-24-1312",
    ])[0]
    assert "MOTOR" in conduit["tags"]
    assert "HOUSING" not in conduit["tags"]
    assert conduit["attributes"]["component"] == "MOTOR"
    assert conduit["attributes"]["motor_feature"] == "CONDUIT BOX LOCATION"
    assert conduit["attributes"]["motor_conduit_box_location"] == "CLOSE TO HOUSING"

    mixed = li.extract_items([
        "HOUSING FLANGE THICKNESS, DRUM HOUSING BOLT PATTERN AND MOTOR CONDUIT BOX HUGGING HOUSING L 100.00",
    ])[0]
    assert {"HOUSING", "MOTOR", "SPECIAL CONSTRUCTION"} <= set(mixed["tags"])
    assert mixed["attributes"]["component"] == "MOTOR"
    assert mixed["attributes"]["motor_conduit_box_location"] == "HUGGING HOUSING"
    assert mixed["attributes"]["housing_subcategory"] == "MODIFIED FLANGE"
    assert "motor_mounting" not in mixed["attributes"]


def test_packaging_housing_reference_does_not_tag_housing():
    packaging = li.extract_items([
        "For all CAT D/38 Fans: bolted fan to the skid with N 122.00",
        "scrap wood for blocking under the fan housing,",
    ])[0]
    assert "PACKAGING" in packaging["tags"]
    assert "HOUSING" not in packaging["tags"]


def test_flex_connector_flange_scope_and_attrs():
    flex = li.extract_items([
        "Inlet Flanged Expansion Joint L 100.00",
        "Vendor: FlexCom",
        "Product: Expansion Joint",
    ])[0]
    assert "FLEX CONNECTOR" in flex["tags"]
    assert "FLANGE" not in flex["tags"]
    assert flex["attributes"]["component"] == "INLET EXPANSION JOINT"
    assert flex["attributes"]["vendor"] == "FlexCom"
    assert "product" not in flex["attributes"]
    assert "flex_connector_type" not in flex["attributes"]
    assert "used_on" not in flex["attributes"]
    assert "flange_scope" not in flex["attributes"]


def test_flexible_coupling_is_coupling_subcategory():
    coupling = li.extract_items([
        "Flexible Coupling Falk Type T Steelflex, Size 1070T, Clearance, Horizontal Split Cover T10, Set Screws, CBC Mount L 1,000.00",
    ])[0]
    assert "COUPLING" in coupling["tags"]
    assert "FLEXIBLE COUPLING" not in coupling["tags"]
    attrs = coupling["attributes"]
    assert attrs["component"] == "COUPLING"
    assert attrs["coupling_subcategory"] == "FLEXIBLE COUPLING"
    assert attrs["coupling_type"] == "FALK TYPE T STEELFLEX"
    assert attrs["size"] == "1070T"
    assert attrs["fit"] == "CLEARANCE"
    assert attrs["cover_type"] == "HORIZONTAL SPLIT COVER T10"
    assert attrs["set_screws"] == "YES"
    assert attrs["mounting"] == "CBC MOUNT"


def test_search_reaches_inquiry_attributes():
    store = li.load_store(Path("/nonexistent/line_items.json"))
    li.record_job(store, "421463", li.extract_items([
        "3D Drawings, InquiryNum:333-25-1622 L",
    ]))
    assert [h["job"] for h in li.search(store, ["333-25-1622"])] == ["421463"]


def test_inquiry_counts():
    store = li.load_store(Path("/nonexistent/line_items.json"))
    li.record_job(store, "421463", li.extract_items([
        "3D Drawings, InquiryNum:333-25-1622 L",
        "Add COG and Weights to the fan drawing, Inquiry L",
        "Num: 333-25-1622",
    ]))
    li.record_job(store, "421464", li.extract_items([
        "Plan View on Customer Drawing, Inquiry Num: 333-25-1622 L",
    ]))
    counts = {num: (jobs, items, job_list) for num, jobs, items, job_list in li.inquiry_counts(store)}
    assert counts["333-25-1622"] == (2, 3, ["421464", "421463"])


def test_material_attributes():
    wheel = li.extract_items([
        "Wheel and Hub, 316 SS Construction, Inquiry Num: L 15,589.00",
    ])[0]
    attrs = wheel["attributes"]
    assert {"MATERIALS", "STAINLESS STEEL", "WHEEL"} <= set(wheel["tags"])
    assert attrs["material"] == "STAINLESS STEEL"
    assert attrs["material_grade"] == "316 SS"
    assert attrs["material_scope"] == "WHEEL AND HUB"

    base = li.extract_items([
        "Base Fan (Base Fan, Design 53 Size F1, 304SST L 9,212.00",
        "Airstream, Arrangement 4 (Less Motor), Inquiry",
        "Num: 333-26-1680)",
    ])[0]
    attrs = base["attributes"]
    assert {"BASE FAN", "MATERIALS", "STAINLESS STEEL"} <= set(base["tags"])
    assert "MOTOR" not in base["tags"]
    assert attrs["material"] == "STAINLESS STEEL"
    assert attrs["material_grade"] == "304 SS"
    assert attrs["material_scope"] == "BASE FAN, AIRSTREAM"

    base_with_motor, customer_motor = li.extract_items([
        "Base Fan L 3,684.00",
        "Motor (Customer Provided) N NA",
        "Vendor: CBC Option",
        "5 HP, 1800 RPM, Enclosure: TEFC",
        "184T, Cast Iron, 3/60/230/460, F1, 1.15 SF",
        "Mounted by Others",
    ])
    assert base_with_motor["tags"] == ["BASE FAN"]
    assert "vendor" not in base_with_motor["attributes"]
    assert customer_motor["price"] == "NA"
    assert customer_motor["tags"] == ["MOTOR"]
    assert customer_motor["attributes"]["component"] == "MOTOR"
    assert customer_motor["attributes"]["motor_supplied_by"] == "CUSTOMER"
    assert customer_motor["attributes"]["motor_hp"] == "5"
    assert customer_motor["attributes"]["motor_rpm"] == "1800"
    assert customer_motor["attributes"]["vendor"] == "CBC Option"
    assert not customer_motor.get("review_flags")

    alum = li.extract_items(["Wheel, Aluminum (AMCA B) L INC"])[0]
    attrs = alum["attributes"]
    assert {"ALUMINUM", "MATERIALS", "WHEEL"} <= set(alum["tags"])
    assert attrs["material"] == "ALUMINUM"
    assert attrs["material_scope"] == "WHEEL"

    grades = li.extract_items([
        "Housing and Base, 304L SS / 316L SS Construction L 1,200.00",
    ])[0]
    attrs = grades["attributes"]
    assert attrs["material_grade"] == "304L SS, 316L SS"
    assert attrs["material_scope"] == "HOUSING AND BASE"


def test_material_scope_requires_material():
    report = li.extract_items(["CBC Runtest Results and Wheel Balance Report, L"])[0]
    assert "BALANCE" not in report["tags"]
    assert "material_scope" not in report["attributes"]


def test_balance_attributes():
    g25 = li.extract_items(["G2.5 Balance L 341.00"])[0]
    assert "BALANCE" in g25["tags"]
    assert g25["attributes"]["component"] == "BALANCE"
    assert "balance_type" not in g25["attributes"]
    assert g25["attributes"]["balance_grade"] == "G2.5"

    g10 = li.extract_items([
        "Special Product Design (EMD Wheel PN 40214482 N 4,393.00",
        "(CBC PN 08-5-4281) Includes G1.0 Balance on",
        "clearance fit arbor, Inquiry Num: 410-13-237)",
    ])[0]
    assert {"BALANCE", "WHEEL"} <= set(g10["tags"])
    assert "balance_type" not in g10["attributes"]
    assert g10["attributes"]["balance_grade"] == "G1.0"

    welded = li.extract_items([
        "Welded Balance Weights, Inquiry Num: 339-26-2098 L 126.00",
    ])[0]
    assert "BALANCE" in welded["tags"]
    assert welded["attributes"]["balance_type"] == "WELDED BALANCE WEIGHTS"
    assert "balance_grade" not in welded["attributes"]


def test_low_leakage_and_extreme_temperature_attributes():
    low_leak = li.extract_items([
        "Inlet Volume Control, Low Leak, Automatic L 3,639.00",
        "Size 4014 Low Leak IVC with rotating ring arm only",
        "@ 9 o'clock when viewed from the inlet side.",
    ])[0]
    assert {"INLET VANES", "LOW LEAKAGE"} <= set(low_leak["tags"])
    assert low_leak["attributes"]["leakage_class"] == "LOW LEAKAGE"
    assert low_leak["attributes"]["used_on"] == "IVC"
    assert "ivc_subcategory" not in low_leak["attributes"]
    assert low_leak["attributes"]["operation"] == "Automatic"
    assert low_leak["attributes"]["ivc_size"] == "4014"
    assert low_leak["attributes"]["ivc_feature"] == "ROTATING RING ARM"
    assert low_leak["attributes"]["ivc_arm_position"] == "9 O'CLOCK"

    cold_fan = li.extract_items([
        "Base Fan Suitable for -45 C Temperature L 21,101.00",
    ])[0]
    assert {"BASE FAN", "EXTREME TEMP"} <= set(cold_fan["tags"])
    assert cold_fan["attributes"]["temperature_service"] == "LOW TEMPERATURE"
    assert cold_fan["attributes"]["temperature_rating"] == "-45C"

    hot_ivc = li.extract_items([
        "Fan and Low Leakage IVC suitable for 175F, Inquiry L",
        "Num: 9-14-164",
    ])[0]
    assert {"EXTREME TEMP", "LOW LEAKAGE", "INLET VANES"} <= set(hot_ivc["tags"])
    assert hot_ivc["attributes"]["temperature_service"] == "HIGH TEMPERATURE"
    assert hot_ivc["attributes"]["temperature_rating"] == "175F"
    assert hot_ivc["attributes"]["leakage_class"] == "LOW LEAKAGE"

    motor_grease = li.extract_items([
        "Motor with High Temperature Grease C 1,000.00",
    ])[0]
    assert "MOTOR" in motor_grease["tags"]
    assert "EXTREME TEMP" not in motor_grease["tags"]
    assert motor_grease["attributes"]["component"] == "MOTOR"
    assert motor_grease["attributes"]["grease_type"] == "HIGH TEMPERATURE GREASE"

    shaft_seal = li.extract_items([
        "John Crane Double Carbon Shaft Seal, High Temp N 6,040.00",
    ])[0]
    assert "SHAFT SEAL" in shaft_seal["tags"]
    assert "EXTREME TEMP" not in shaft_seal["tags"]
    assert shaft_seal["attributes"]["component"] == "SHAFT SEAL"
    assert shaft_seal["attributes"]["temperature_service"] == "HIGH TEMPERATURE"


def test_inlet_vane_manual_locking_quadrant_attributes():
    manual = li.extract_items([
        "Inlet Volume Control, 350 F, Manual, with Locking L 2,175.00",
        "Quadrant",
        "Handle Location (viewed from inlet side):",
    ])[0]
    assert {"DAMPER", "INLET VANES"} <= set(manual["tags"])
    assert "ivc_subcategory" not in manual["attributes"]
    assert manual["attributes"]["operation"] == "Manual"
    assert manual["attributes"]["ivc_feature"] == "LOCKING QUADRANT"
    assert manual["attributes"]["used_on"] == "IVC"


def test_heavy_duty_component_attributes():
    wheel = li.extract_items(["Wheel, Heavy Duty L 302.00"])[0]
    assert {"HEAVY DUTY", "WHEEL"} <= set(wheel["tags"])
    assert wheel["attributes"]["component"] == "WHEEL"
    assert wheel["attributes"]["heavy_duty"] == "YES"

    housing = li.extract_items(["Housing, Heavy Duty L 4,754.00"])[0]
    assert {"HEAVY DUTY", "HOUSING"} <= set(housing["tags"])
    assert housing["attributes"]["component"] == "HOUSING"
    assert housing["attributes"]["heavy_duty"] == "YES"

    severe_motor = li.extract_items(["Motor TEFC Severe Duty C 1,200.00"])[0]
    assert "HEAVY DUTY" not in severe_motor["tags"]
    assert "heavy_duty" not in severe_motor["attributes"]


def test_bearing_attributes():
    repair = li.extract_items([
        'Repair Bearings (Pair), 2-3/16" Bore (Qty: 1), N 552.00 110.00',
        "Inquiry Num: 340-25-943RP",
    ])[0]
    assert "BEARINGS" in repair["tags"]
    assert repair["attributes"]["bearing_type"] == "REPAIR BEARINGS"
    assert repair["attributes"]["bearing_bore"] == '2-3/16"'
    assert repair["attributes"]["inquiry_num"] == "340-25-943RP"

    split = li.extract_items(["Bearings, Split Pillow Block L 2,387.00"])[0]
    assert "BEARINGS" in split["tags"]
    assert split["attributes"]["component"] == "BEARINGS"
    assert split["attributes"]["bearing_type"] == "SPLIT PILLOW BLOCK"

    spare = li.extract_items(["Spare Bearings, Inquiry Num: 340-26-1112 L 3,583.00"])[0]
    assert "BEARINGS" in spare["tags"]
    assert spare["attributes"]["bearing_type"] == "SPARE BEARINGS"
    assert spare["attributes"]["inquiry_num"] == "340-26-1112"

    adder = li.extract_items(["Bearing ADDER for 200,00 hours, Inquiry Num: L"])[0]
    assert "BEARINGS" in adder["tags"]
    assert adder["attributes"]["component"] == "BEARINGS"
    assert adder["attributes"]["bearing_life_adder"] == "YES"
    assert "bearing_type" not in adder["attributes"]
    assert adder["attributes"]["bearing_life_hours"] == "200,000"

    complete = li.extract_items(["Bearing adder for 200,000 hours L 400.00"])[0]
    assert complete["attributes"]["bearing_life_hours"] == "200,000"


def test_shaft_bearing_guard_tag_uses_primary_line():
    guard = li.extract_items(["Shaft and Bearing Guard, Painted Safety Yellow L 949.00"])[0]
    assert "SHAFT/BEARING/COUPLING GUARD" in guard["tags"]

    grease = li.extract_items([
        "Extended Grease Fittings - Shipped Loose, Inquiry L 349.00",
        "Num: 6-5-1551",
        "Size 16-1/2 Shaft and Bearing Guard, Inquiry Num: L -940.00",
    ])[0]
    assert "BEARINGS" in grease["tags"]
    assert "MOTOR" not in grease["tags"]
    assert "EXTENDED LUBE" not in grease["tags"]
    assert "SHAFT/BEARING/COUPLING GUARD" not in grease["tags"]


def test_lube_accessories_map_to_bearing_or_motor():
    assert "EXTENDED LUBE" not in li.load_rules(refresh=True)["tags"]

    unknown = li.extract_items(["Extended Lube Lines with Zerk Fittings L 100.00"])[0]
    assert {"BEARINGS", "MOTOR"} <= set(unknown["tags"])
    assert unknown["attributes"]["extended_grease_fittings"] == "YES"
    assert unknown["attributes"]["component_review"] == "UNCLEAR GREASE TARGET - VERIFY MOTOR/BEARINGS/ARRANGEMENT"
    assert any("UNCLEAR GREASE TARGET" in flag for flag in unknown["review_flags"])

    motor = li.extract_items(["Motor Grease Lines L 100.00"])[0]
    assert "MOTOR" in motor["tags"]
    assert "BEARINGS" not in motor["tags"]
    assert "review_flags" not in motor
    assert motor["attributes"]["extended_grease_fittings"] == "YES"

    bearing = li.extract_items(["Extended Grease Leads to Fan Bearings L 100.00"])[0]
    assert "BEARINGS" in bearing["tags"]
    assert "MOTOR" not in bearing["tags"]
    assert "review_flags" not in bearing
    assert bearing["attributes"]["component"] == "BEARINGS"

    motor_bearing = li.extract_items(["Motor Bearing Grease Fittings L 100.00"])[0]
    assert motor_bearing["tags"] == ["MOTOR"]
    assert motor_bearing["attributes"]["component"] == "MOTOR"
    assert "review_flags" not in motor_bearing

    arrangement_9 = li.extract_items(
        ["Extended Grease Fittings L 100.00"],
        order_context={"arrangement": "A/9S"},
    )[0]
    assert arrangement_9["tags"] == ["BEARINGS"]
    assert arrangement_9["attributes"]["component"] == "BEARINGS"
    assert not arrangement_9.get("review_flags")

    arrangement_4 = li.extract_items(
        ["Extended Grease Fittings L 100.00"],
        order_context={"arrangement": "A/4"},
    )[0]
    assert arrangement_4["tags"] == ["MOTOR"]
    assert arrangement_4["attributes"]["component"] == "MOTOR"
    assert not arrangement_4.get("review_flags")

    items = {it["norm"]: it for it in li.extract_items([
        "Motor C 1,000.00",
        "Extended Grease Leads",
        "Extended Grease Leads L 100.00",
    ])}
    assert "MOTOR" in items["MOTOR"]["tags"]
    assert "BEARINGS" not in items["MOTOR"]["tags"]
    assert {"BEARINGS", "MOTOR"} <= set(items["EXTENDED GREASE LEADS"]["tags"])
    assert any("UNCLEAR GREASE TARGET" in flag for flag in items["EXTENDED GREASE LEADS"]["review_flags"])


def test_assembly_note_is_misc_note_not_component():
    assembly = li.extract_items([
        "Assembly, Assemble Panel, Wheel, Shaft, and L INC",
        "Bearings",
    ])[0]
    assert "MISC NOTE" in assembly["tags"]
    assert "WHEEL" not in assembly["tags"]
    assert "BEARINGS" not in assembly["tags"]
    assert assembly["attributes"]["note_type"] == "ASSEMBLY"


def test_paint_line_is_canonical_fan_paint_component():
    paint = li.extract_items([
        "Paint: Interior, Wheel, Exterior, Motor Base, L 983.00",
        "Channel Base and Bearing Base",
    ])[0]
    assert paint["tags"] == ["COATING"]
    assert paint["attributes"]["component"] == "PAINT"
    assert "coating_context" not in paint["attributes"]
    assert paint["attributes"]["coating_category"] == "PAINT"
    assert "WHEEL" not in paint["tags"]

    paint_with_warranty_detail = li.extract_items([
        "Paint: Exterior, Motor Base L 714.00",
        "Extended Warranty: Chicago Blower standard warranty extended to 24 months.",
    ])[0]
    assert paint_with_warranty_detail["tags"] == ["COATING"]

    pre_coating = li.extract_items(["Pre-Coating Assembly/Disassembly L 1,185.00"])[0]
    assert pre_coating["tags"] == ["COATING"]
    assert pre_coating["attributes"]["coating_category"] == "PRE-COATING PROCESS"
    assert pre_coating["attributes"]["coating_process"] == "PRE-COATING ASSEMBLY/DISASSEMBLY"

    special = li.extract_items([
        "Special Paint, exterior and airstream: SSPC-SP10, N 5,310.00 220.00",
        "Zinc Rich Epoxy primer, 3 mils, mid coat epoxy 7 mil, top coat poly, 3-mil",
        "(RAL 6019 if not available use RAL 7035), Inquiry Num: 253-24-1651",
    ])[0]
    assert special["attributes"]["coating_category"] == "SPECIAL COATING"
    assert special["attributes"]["coating_type"] == "EPOXY"
    assert special["attributes"]["coating_color"] == "RAL 6019"
    assert special["attributes"]["alternate_coating_color"] == "RAL 7035"

    veg_oil = li.extract_items([
        "Airstream to be unpainted and Coated with L 714.00",
        "vegetable oil, Inquiry Num: 317-26-1059",
    ])[0]
    assert veg_oil["attributes"]["coating_category"] == "SPECIAL COATING"
    assert veg_oil["attributes"]["coating_type"] == "VEGETABLE OIL"
    assert veg_oil["attributes"]["coating_state"] == "UNPAINTED"

    unpainted = li.extract_items(["Wheel, Cast Aluminum CCW (Unpainted) (Bore: L 739.00"])[0]
    assert unpainted["attributes"]["coating_category"] == "UNPAINTED"

    note = li.extract_items([
        "Specification and updated coating Note # B to show wheel is un-coated L 100.00",
    ])[0]
    assert note["attributes"]["coating_category"] == "COATING NOTE"


def test_accessory_coating_is_attribute_not_fan_coating():
    belt = li.extract_items([
        "Belt Guard, Painted Safety Yellow L 1,581.00",
        "CBC Mount",
        "Tach Hole in Guard: with Plug",
        "Fan End",
        "Motor End",
    ])[0]
    assert belt["tags"] == ["BELT GUARD"]
    attrs = belt["attributes"]
    assert attrs["component"] == "BELT GUARD"
    assert attrs["coating"] == "PAINTED SAFETY YELLOW"
    assert not any(key.startswith("coating_") for key in attrs)
    assert attrs["mounting"] == "CBC MOUNT"
    assert attrs["tach_hole"] == "WITH PLUG"
    assert attrs["tach_hole_location"] == "FAN END, MOTOR END"

    shaft_guard = li.extract_items(["Shaft and Bearing Guard, Painted Safety Yellow L 949.00"])[0]
    assert "SHAFT/BEARING/COUPLING GUARD" in shaft_guard["tags"]
    assert "COATING" not in shaft_guard["tags"]
    assert shaft_guard["attributes"]["coating"] == "PAINTED SAFETY YELLOW"


def test_passivation_is_stainless_treatment_not_coating():
    passivation = li.extract_items([
        "Passivation of Welds (Passivation of Welds L 1,412.00",
        "(Airstream and Exterior), Inquiry Num: 333-26-1234",
    ])[0]
    assert "COATING" not in passivation["tags"]
    assert {"MATERIALS", "STAINLESS STEEL"} <= set(passivation["tags"])
    attrs = passivation["attributes"]
    assert attrs["material"] == "STAINLESS STEEL"
    assert attrs["material_treatment"] == "PASSIVATION OF WELDS"
    assert attrs["material_scope"] == "AIRSTREAM, EXTERIOR, WELDS"


def test_accessory_coating_does_not_make_silencer_fan_coating():
    silencer = li.extract_items([
        "Inlet Silencer for 85 dBA @ 3' Outdoors, Model C 2,267.00 363.00",
        "includes Leg, 2 coats of paint",
        "Vendor: VAW Systems",
        "Product: Inlet Silencer",
    ])[0]
    assert "SILENCER" in silencer["tags"]
    assert "COATING" not in silencer["tags"]
    assert silencer["attributes"]["coating"] == "2 COATS OF PAINT"
    assert not any(key.startswith("coating_") for key in silencer["attributes"])


def test_admin_notes_are_misc_notes():
    note = li.extract_items([
        "Slow Pay Addition N 339.00",
        "3-coat paint system to match fan.",
    ])[0]
    assert note["tags"] == ["MISC NOTE"]
    assert note["attributes"] == {"note_type": "ADMIN"}


def test_component_materials_do_not_count_as_fan_materials():
    actuator = li.extract_items([
        "Bettis #RPED150 Double Acting Pneumatic Actuator C 6,537.14 1,046.00",
        "Fisher #67CFR Filter/ Regulator, SS Tubing.",
        "Vendor: Novaspect",
        "Product: Actuator",
    ])[0]
    attrs = actuator["attributes"]
    assert "ACTUATOR" in actuator["tags"]
    assert not {"MATERIALS", "STAINLESS STEEL"} & set(actuator["tags"])
    assert attrs["component"] == "ACTUATOR"
    assert attrs["ss_tubing"] == "YES"
    assert "component_material" not in attrs
    assert "component_material_scope" not in attrs
    assert "material" not in attrs

    motor = li.extract_items([
        "Motor (Multimounting IE3 5.5 HP 2P 112M 3Ph C 620.49",
        "112M, Cast Aluminum, Foot Mounted",
        "Vendor: Weg",
    ])[0]
    attrs = motor["attributes"]
    assert "MOTOR" in motor["tags"]
    assert not {"MATERIALS", "ALUMINUM"} & set(motor["tags"])
    assert attrs["component"] == "MOTOR"
    assert attrs["component_material"] == "ALUMINUM"
    assert attrs["component_material_scope"] == "MOTOR"
    assert "material" not in attrs


def test_data_branch_noise_skipped():
    lines = [
        "ADDITIONAL FEATURES",
        "Product:",
        "Product 7,623.00",
        "Prints",
        "Prints:",
        "Warranty Exclusive 3 Year",
        "Paymode-X invoice processing system N 52.00",
        "Includes Paymode-X invoice processing system",
        "be necessary. For more information on this process, contact Chicago Blower.",
        "Additional Shipping Notes must use BOL provided by CH Robinson",
        "Do Not Stack sticker on all 4 sides of skid",
        "Drive L 100.00",
    ]
    items = {it["norm"]: it for it in li.extract_items(lines)}
    assert "DRIVE" in items
    paymode = next(v for n, v in items.items() if "PAYMODE" in n)
    assert paymode["tags"] == ["MISC NOTE"]
    assert paymode["attributes"]["note_type"] == "ADMIN"
    for bad in ("PRODUCT", "PRINTS", "WARRANTY", "SHIPPING NOTES", "DO NOT STACK"):
        assert not any(bad in n for n in items), items


def test_performance_header_noise_skipped():
    items = li.extract_items([
        "MAX TEMP, ELEVATION, DENSITY, COLD START BHP, MOTOR DESCRIPTION, FRAME, HP, INLET BOX, MIXING BOX, FRESH AIR",
    ])
    assert items == []


def test_used_on_requires_damper_context():
    flange = li.extract_items(["Outlet, Flanged, Punched L STD"])[0]
    assert "used_on" not in flange["attributes"]
    damper = li.extract_items(["Outlet Damper, Opposed L 100.00"])[0]
    assert "OUTLET" not in damper["tags"]
    assert damper["attributes"]["component"] == "OUTLET DAMPER"
    assert "used_on" not in damper["attributes"]
    volume = li.extract_items(["Outlet Volume Control, Manual L 691.00"])[0]
    assert "DAMPER" in volume["tags"]
    assert "OUTLET" not in volume["tags"]
    assert volume["attributes"]["used_on"] == "OUTLET DAMPER"
    discharge = li.extract_items(["Actuator for Discharge Damper, Bettis #RPED200 C 5,192.00"])[0]
    assert discharge["attributes"]["used_on"] == "OUTLET DAMPER"
    fresh_air = li.extract_items(["Change Actuator Location for the FA damper to be on motor side of fan L 100.00"])[0]
    assert fresh_air["attributes"]["used_on"] == "FRESH AIR DAMPER"
    assert "used_on_review" not in fresh_air["attributes"]
    prespin = li.extract_items(["Actuator for Pre-spin Damper, Bettis #RPED200 C 5,192.00"])[0]
    assert prespin["attributes"]["used_on"] == "PRESPIN DAMPER"

    mounted_damper = li.extract_items([
        "Fresh Air Damper, Opposed Blade, Without Stuffing L 13,946.00",
        "Boxes (Shipped Loose), Mounted on Oversized Inlet Box, Automatic",
        "Product: Actuator",
        "Vendor: Bettis RPD",
    ])[0]
    assert "DAMPER" in mounted_damper["tags"]
    assert "ACTUATOR" in mounted_damper["tags"]
    assert "INLET" not in mounted_damper["tags"]
    assert mounted_damper["attributes"]["component"] == "FRESH AIR DAMPER"
    assert "used_on" not in mounted_damper["attributes"]
    assert mounted_damper["attributes"]["actuator_manufacturer"] == "BETTIS"
    assert mounted_damper["attributes"]["actuator_operation"] == "Automatic"
    assert "product" not in mounted_damper["attributes"]
    assert "vendor" not in mounted_damper["attributes"]
    assert mounted_damper["attributes"]["mount_location"] == "INLET BOX"


def test_without_ivc_does_not_tag_inlet_vanes_or_used_on():
    item = li.extract_items(["Inlet, Flanged, Punched (without IVC) L 468.00"])[0]
    assert {"INLET", "FLANGE"} <= set(item["tags"])
    assert "INLET VANES" not in item["tags"]
    assert "used_on" not in item["attributes"]


def test_inlet_cone_width_does_not_tag_wheel_without_wheel_wording():
    item = li.extract_items(["Aluminum Inlet Cone, Inlet Cone 100% Width Cartoned (Unpainted) L 752.00"])[0]
    assert {"INLET", "MATERIALS", "ALUMINUM"} <= set(item["tags"])
    assert "WHEEL" not in item["tags"]
    assert item["attributes"]["material_scope"] == "INLET CONE"
    assert item["attributes"]["coating_scope"] == "INLET"


def test_inlet_subcategory_attributes():
    open_inlet = li.extract_items(["Inlet, Open L STD"])[0]
    assert open_inlet["attributes"]["inlet_subcategory"] == "OPEN"

    slip = li.extract_items(["Inlet, Slip N STD"])[0]
    assert slip["attributes"]["inlet_subcategory"] == "SLIP"

    bell = li.extract_items(["Inlet, Bell L 656.00"])[0]
    assert bell["attributes"]["inlet_subcategory"] == "BELL"

    tube = li.extract_items(["Inlet, Tube L STD"])[0]
    assert tube["attributes"]["inlet_subcategory"] == "TUBE"

    cone = li.extract_items([
        "INLET CONE 72%, D62 SW, SIZE 200 (Customer N 102.00 10.00",
        "Part #: 90515, CBC Part #: 62-0-0062)",
    ])[0]
    assert cone["attributes"]["inlet_subcategory"] == "INLET CONE"
    assert cone["attributes"]["inlet_cone_width_percent"] == "72"
    assert cone["attributes"]["inlet_size"] == "200"


def test_inlet_flange_direction_and_box_attributes():
    punched = li.extract_items(["Inlet, Flanged, Punched (without IVC) L 468.00"])[0]
    assert punched["attributes"]["component"] == "INLET"
    assert punched["attributes"]["flanged"] == "YES"
    assert punched["attributes"]["punched"] == "YES"
    assert "inlet_subcategory" not in punched["attributes"]
    assert punched["attributes"]["ivc_relation"] == "WITHOUT IVC"

    with_ivc = li.extract_items(["Inlet, Flanged, Punched (with IVC) L 1,453.00"])[0]
    assert with_ivc["tags"] == ["FLANGE", "INLET"]
    assert with_ivc["attributes"]["component"] == "INLET"
    assert with_ivc["attributes"]["flanged"] == "YES"
    assert with_ivc["attributes"]["punched"] == "YES"
    assert "ivc_relation" not in with_ivc["attributes"]
    assert "used_on" not in with_ivc["attributes"]

    bolted = li.extract_items(['Inlet, Flanged, Standard Bolted (Inlet Dia 10") N STD'])[0]
    assert bolted["attributes"]["flanged"] == "YES"
    assert bolted["attributes"]["punched"] == "NO"
    assert bolted["attributes"]["inlet_feature"] == "STANDARD BOLTED"

    direction = li.extract_items(["Inlet Direction: Vertical Inlet Down L STD"])[0]
    assert direction["attributes"]["inlet_subcategory"] == "DIRECTION"
    assert direction["attributes"]["inlet_direction"] == "VERTICAL INLET DOWN"

    box = li.extract_items([
        "Inlet Box, Bolt-on (Shipped Loose), Oversized Inlet L 7,161.00",
        "Box Size 300 (Inlet Box, Bolt-on (Shipped Loose),",
        "Bolt-On Inlet Box Position: @ 0",
    ])[0]
    assert box["attributes"]["component"] == "INLET BOX"
    assert box["attributes"]["bolt_on"] == "YES"
    assert box["attributes"]["inlet_box_size"] == "OVERSIZED 300"
    assert "inlet_subcategory" not in box["attributes"]
    assert "inlet_box_type" not in box["attributes"]
    assert box["attributes"]["inlet_box_position"] == "0"
    assert box["attributes"]["shipping_state"] == "SHIPPED LOOSE"


def test_mixing_box_is_not_inlet_category():
    mixing = li.extract_items([
        "Mixing Box with Flanged FGR Port (Shipped Loose) L 8,297.00",
        "(Mixing Box With Flanged FGR Port (Shipped Loose), Suitable for Size 300 Inlet Box)",
    ])[0]
    assert "MIXING BOX" in mixing["tags"]
    assert "INLET" not in mixing["tags"]
    assert mixing["attributes"]["component"] == "MIXING BOX"
    assert mixing["attributes"]["flange_scope"] == "MIXING BOX"
    assert mixing["attributes"]["flange_type"] == "FLANGED"
    assert mixing["attributes"]["fgr_port"] == "YES"
    assert "mixing_box_feature" not in mixing["attributes"]
    assert mixing["attributes"]["used_on"] == "INLET BOX"
    assert mixing["attributes"]["used_on_size"] == "300"


def test_ship_via_component_list_is_shipping_note():
    item = li.extract_items([
        "Damper, Inlet Volume Control, Outlet Damper, Inlet Silencer, Expansion Joint and Updated Ship Via, Additional L 100.00",
    ])[0]
    assert item["tags"] == ["SHIPPING"]
    assert "used_on" not in item["attributes"]
    assert "component" not in item["attributes"]


def test_inspection_subcategories_keep_ispm_and_overseas_crate():
    ispm = li.derive_item_fields({"raw": "ISPM Wood Inspection Stamp L INC", "details": []})
    assert {"INSPECTION", "PACKAGING"} <= set(ispm["tags"])
    assert ispm["attributes"]["inspection_subcategory"] == "ISPM WOOD STAMP"
    assert ispm["attributes"]["inspection_scope"] == "PACKAGING"

    requirement = li.derive_item_fields({
        "raw": "All international shipments are to be packaged using suitable materials meeting IPPC (or) ISPM code requirements or equivalent (pest",
        "details": [],
    })
    assert {"INSPECTION", "PACKAGING", "SHIPPING"} <= set(requirement["tags"])
    assert requirement["attributes"]["component"] == "INSPECTION"
    assert requirement["attributes"]["inspection_subcategory"] == "ISPM REQUIREMENTS"
    assert requirement["attributes"]["inspection_scope"] == "PACKAGING"

    overseas = li.derive_item_fields({
        "raw": "Order is shipping overseas; verify order is complete. For parts orders verify counts and sign off on inspection/crate report.",
        "details": [],
    })
    assert {"INSPECTION", "PACKAGING", "SHIPPING"} <= set(overseas["tags"])
    assert overseas["attributes"]["inspection_subcategory"] == "OVERSEAS CRATE REPORT"

    booth_job_name = li.extract_items([
        "AMCA B Spark Resistant Construction, Inquiry Num: L 1,338.00",
        "421906 Fuel Nozzle Inspection Booth 4222426",
    ])[0]
    assert "SPARK RESISTANT" in booth_job_name["tags"]
    assert "INSPECTION" not in booth_job_name["tags"]


def test_insulation_splits_motor_details_from_fan_insulation():
    motor = li.extract_items([
        "Motor (WEG SPECIAL BUILD C 4,364.67 131.00",
        "F insulation",
        "NDE insulated bearing",
        "DE AEGIS RING",
        "Vendor: Weg",
    ])[0]
    assert "MOTOR" in motor["tags"]
    assert "INSULATION" not in motor["tags"]
    assert motor["attributes"]["motor_insulation_class"] == "F"
    assert motor["attributes"]["motor_insulated_bearing"] == "NDE"
    assert motor["attributes"]["motor_shaft_grounding"] == "AEGIS RING"

    variants = li.extract_items([
        "Motor C 1,000.00",
        "M2F Add isolated bearings for NEMA 400 to 500",
        "M39B 10 Internal AEGIS bearing protection ring",
    ])[0]
    assert variants["attributes"]["motor_insulated_bearing"] == "YES"
    assert variants["attributes"]["motor_shaft_grounding"] == "AEGIS RING"
    assert "review_flags" not in variants

    plug_panel = li.extract_items([
        "Plug Panel, Insulation 8\" L 531.00",
        "Insulated By: CBC",
    ])[0]
    assert {"INSULATION", "SPECIAL CONSTRUCTION"} <= set(plug_panel["tags"])
    assert plug_panel["attributes"]["insulation_scope"] == "PLUG PANEL"
    assert plug_panel["attributes"]["insulation_thickness"] == "8\""
    assert plug_panel["attributes"]["insulated_by"] == "CBC"

    housing_lagging = li.extract_items(["Housing Lagging 2\" L 531.00"])[0]
    assert {"HOUSING", "INSULATION"} <= set(housing_lagging["tags"])
    assert housing_lagging["attributes"]["insulation_scope"] == "HOUSING"
    assert housing_lagging["attributes"]["insulation_type"] == "LAGGING"
    assert housing_lagging["attributes"]["insulation_thickness"] == "2\""

    housing_fiberglass = li.extract_items(["Fan Housing with 2\" Fiberglass Insulation L 531.00"])[0]
    assert {"HOUSING", "INSULATION"} <= set(housing_fiberglass["tags"])
    assert housing_fiberglass["attributes"]["insulation_scope"] == "HOUSING"
    assert housing_fiberglass["attributes"]["insulation_material"] == "FIBERGLASS"


def test_label_attributes_and_detail_tag_cleanup():
    warning = li.extract_items([
        "Only apply CBC's warning label to the fan. (Except L 100.00",
        "D/47 will have the vendor's motor nameplate",
        "applied as well.) All other stickers/labels and",
        "nameplate will be placed in a bag and shipped with",
        "fan., Inquiry Num: 15-14-2975",
    ])[0]
    assert warning["tags"] == ["LABEL"]
    assert warning["attributes"]["label_type"] == "WARNING LABEL"
    assert warning["attributes"]["label_scope"] == "FAN"
    assert warning["attributes"]["label_handling"] == "OTHER LABELS/NAMEPLATE BAGGED"
    assert warning["attributes"]["related_nameplate_handling"] == "VENDOR MOTOR NAMEPLATE APPLIED"

    barcode = li.derive_item_fields({"raw": "Shipping barcode label, Inquiry Num: 15-14-2975 L", "details": []})
    assert {"LABEL", "SHIPPING"} <= set(barcode["tags"])
    assert barcode["attributes"]["label_type"] == "SHIPPING BARCODE LABEL"
    assert barcode["attributes"]["label_scope"] == "SHIPPING"

    motor = li.extract_items([
        "Motor (Multimounting IE3 10 HP 2P 132S 3Ph C 864.10",
        "Mods: Add tropicalization, RETIE label",
    ])[0]
    assert {"LABEL", "MOTOR"} <= set(motor["tags"])
    assert "MOUNTING" not in motor["tags"]
    assert motor["attributes"]["label_type"] == "RETIE LABEL"
    assert motor["attributes"]["label_scope"] == "MOTOR"
    assert motor["attributes"]["motor_mounting"] == "MULTIMOUNTING"

    marked = li.derive_item_fields({"raw": "Mark all items with LOT/SER Number, Project, & Heat Lot number", "details": []})
    assert "LABEL" in marked["tags"]
    assert marked["attributes"]["label_type"] == "ITEM MARKING"

    spec_nameplate = li.extract_items([
        "125HP,1785RPM,3PH,60HZ,444T,TEFC,F1 L 6,476.60",
        "M15B Replace Nameplate - RR #47",
    ])[0]
    assert {"MOTOR", "NAMEPLATE"} <= set(spec_nameplate["tags"])
    assert "MOUNTING" not in spec_nameplate["tags"]
    assert spec_nameplate["attributes"]["component"] == "MOTOR"
    assert spec_nameplate["attributes"]["motor_nameplate_action"] == "REPLACE NAMEPLATE"


def test_lifting_lugs_and_lining_attributes():
    lugs = li.extract_items(["Lifting Lugs @12:00, Inquiry Num: 374-26-320 L 463.00"])[0]
    assert "LIFTING LUGS" in lugs["tags"]
    assert lugs["attributes"]["component"] == "HOUSING"
    assert "housing_feature" not in lugs["attributes"]
    assert lugs["attributes"]["lifting_lugs"] == "YES"
    assert "lug_position" not in lugs["attributes"]

    silencer = li.extract_items([
        "Aeroacoustic Silentflow inlet silencer C 100.00",
        "Lifting lugs",
    ])[0]
    assert "SILENCER" in silencer["tags"]
    assert "LIFTING LUGS" not in silencer["tags"]

    housing_scroll = li.extract_items([
        "Firmex Liners on Blades and Housing Scroll, Inquiry L 10,602.00",
        "Num: 300-25-3241",
    ])[0]
    assert {"HOUSING", "LINING", "WHEEL"} <= set(housing_scroll["tags"])
    assert housing_scroll["attributes"]["lining_type"] == "FIRMEX"
    assert housing_scroll["attributes"]["lining_scope"] == "HOUSING SCROLL, WHEEL BLADES"

    scroll_side_sheet = li.extract_items([
        "Firmex Liners, On Scroll, side sheet, Wheel Blades, L 36,295.00",
    ])[0]
    assert {"HOUSING", "LINING", "WHEEL"} <= set(scroll_side_sheet["tags"])
    assert scroll_side_sheet["attributes"]["lining_scope"] == "HOUSING SCROLL, SIDE SHEET, WHEEL BLADES"

    flex = li.extract_items([
        "Outlet Expansion Joint C 1,466.00",
        "10Ga. A36 Flow Liner",
        "Product: Expansion Joint",
    ])[0]
    assert "FLEX CONNECTOR" in flex["tags"]
    assert "LINING" not in flex["tags"]
    assert flex["attributes"]["component"] == "OUTLET EXPANSION JOINT"
    assert flex["attributes"]["feature"] == "FLOW LINER"


def test_screen_and_shaft_cooler_attributes():
    inlet = li.extract_items(["Inlet Screen, Standard, 304 SS Construction L 975.00"])[0]
    attrs = inlet["attributes"]
    assert "SCREEN" in inlet["tags"]
    assert attrs["component"] == "INLET SCREEN"
    assert "screen_subcategory" not in attrs
    assert attrs["screen_feature"] == "STANDARD"
    assert "used_on" not in attrs
    assert attrs["material_scope"] == "INLET SCREEN"

    outlet = li.extract_items(["Outlet Screen, Outlet Screen (Shipped Loose) N 55.00"])[0]
    assert outlet["attributes"]["component"] == "OUTLET SCREEN"
    assert "used_on" not in outlet["attributes"]
    assert outlet["attributes"]["shipping_state"] == "SHIPPED LOOSE"

    wrapped = li.extract_items([
        "Inlet Screen, Standard (Inlet Screen, Oversized L 781.00",
        "(D10 size 33, D51 size 365), Inquiry Num: 317-24-",
        "388)",
    ])[0]
    assert wrapped["attributes"]["component"] == "INLET SCREEN"
    assert wrapped["attributes"]["screen_feature"] == "OVERSIZED, STANDARD"
    assert wrapped["attributes"]["inquiry_num"] == "317-24-388"
    assert "screen_subcategory" not in wrapped["attributes"]
    assert "used_on" not in wrapped["attributes"]
    assert not wrapped.get("review_flags")

    silencer = li.extract_items([
        "VAW Inlet Silencer With Piezometer tube C 5,716.00",
        'thermowell port, trash screen, with Support Legs for 38" centerline Height',
        "Product: Inlet Silencer",
    ])[0]
    assert {"SCREEN", "SILENCER"} <= set(silencer["tags"])
    assert silencer["attributes"]["component"] == "INLET SILENCER"
    assert silencer["attributes"]["screen_subcategory"] == "TRASH SCREEN"
    assert "used_on" not in silencer["attributes"]

    shaft = li.extract_items(["Shaft Cooler, Cast Aluminum Construction L 271.00"])[0]
    attrs = shaft["attributes"]
    assert {"ALUMINUM", "MATERIALS", "SHAFT COOLER"} <= set(shaft["tags"])
    assert attrs["component"] == "SHAFT COOLER"
    assert "shaft_cooler" not in attrs
    assert "shaft_cooler_type" not in attrs
    assert attrs["shaft_cooler_construction"] == "CAST"
    assert attrs["material_scope"] == "SHAFT COOLER"


def test_weather_cover_attributes_and_reference_cleanup():
    motor = li.extract_items([
        "Motor (5hp 3470 rpm 3ph 184tc with drip cover C 342.91",
        "CEM3613T\\M8A Install Dripcover on TEFC or ODP)",
        "Vendor: Baldor/Reliance",
    ])[0]
    assert {"MOTOR", "WEATHER COVER"} <= set(motor["tags"])
    assert motor["attributes"]["component"] == "MOTOR"
    assert "motor_enclosure" not in motor["attributes"]
    assert motor["attributes"]["weather_cover_type"] == "DRIP COVER"
    assert motor["attributes"]["weather_cover_scope"] == "MOTOR"

    rainhood = li.extract_items([
        "Bolt-On 3 Sided Inlet Rainhood with Inlet Screen L 1,383.00",
        "CBC Mount",
    ])[0]
    assert {"SCREEN", "WEATHER COVER"} <= set(rainhood["tags"])
    assert rainhood["attributes"]["component"] == "WEATHER COVER"
    assert rainhood["attributes"]["weather_cover_type"] == "RAINHOOD"
    assert rainhood["attributes"]["weather_cover_used_on"] == "INLET"
    assert rainhood["attributes"]["weather_cover_feature"] == "INLET SCREEN"
    assert rainhood["attributes"]["mounting"] == "CBC MOUNT"

    inlet_hood = li.extract_items([
        "Inlet Hood Model VWH5-36 includes 1 set H&G each C 2,640.00",
        "Hood std.",
        "Vendor: VAW Systems",
        "Ship Direct",
        "Product: Inlet Silencer",
    ])[0]
    assert {"SILENCER", "SHIPPING", "WEATHER COVER"} <= set(inlet_hood["tags"])
    assert inlet_hood["attributes"]["component"] == "INLET SILENCER"
    assert inlet_hood["attributes"]["model"] == "VWH5-36"
    assert "weather_cover_type" not in inlet_hood["attributes"]
    assert "weather_cover_scope" not in inlet_hood["attributes"]

    drawing_note = li.derive_item_fields({
        "raw": "Engineering to show fan plus accessory weight, including rainhood, motor weight and total assembled weight on the drawing.",
        "details": [],
    })
    assert "DRAWINGS" in drawing_note["tags"]
    assert "WEATHER COVER" not in drawing_note["tags"]


def test_silencer_spare_parts_and_spark_resistant_attributes():
    silencer = li.extract_items([
        "VAW Inlet Silencer model 12VRSB-S81 and with C 5,629.00",
        "VWH1-16x18 inlet rainhood with galvanized screen.",
        "Designed for 85 dBA @ 3 feet, Pressure Drop: .19",
        "Vendor: VAW Systems",
        "Ship Direct",
        "Product: Inlet Silencer",
    ])[0]
    attrs = silencer["attributes"]
    assert {"SILENCER", "SHIPPING", "WEATHER COVER"} <= set(silencer["tags"])
    assert attrs["component"] == "INLET SILENCER"
    assert attrs["model"] == "12VRSB-S81"
    assert attrs["noise_target"] == "85 DBA"
    assert attrs["pressure_drop"] == ".19"
    assert attrs["rain_hood"] == "YES"
    assert attrs["screen_subcategory"] == "RAINHOOD SCREEN"
    assert "used_on" not in attrs
    assert not any(key.startswith("weather_cover_") for key in attrs)
    assert attrs["shipping_method"] == "SHIP DIRECT"
    assert attrs["coating"] == "GALVANIZED"
    assert "product" not in attrs

    discharge = li.extract_items([
        "CIB - Circular Discharge silencer C 1,616.00",
        "Model: 06VCIB-V99-SN4400",
        "Unit Wt: 63 lbs",
        "Vendor: VAW Systems",
        "Product: Outlet Silencer",
    ])[0]
    attrs = discharge["attributes"]
    assert "SILENCER" in discharge["tags"]
    assert attrs["component"] == "OUTLET SILENCER"
    assert attrs["silencer_type"] == "CIRCULAR DISCHARGE SILENCER"
    assert attrs["model"] == "06VCIB-V99-SN4400"

    ispm = li.derive_item_fields({
        "raw": "ISPM Wood Inspection Stamp L INC",
        "details": ["One (1) Aeroacoustic Silentflow Model 6-TA-2B discharge silencer as per submittal."],
    })
    assert ispm["tags"] == ["INSPECTION", "PACKAGING"]
    assert ispm["attributes"] == {
        "component": "INSPECTION",
        "inspection_subcategory": "ISPM WOOD STAMP",
        "inspection_scope": "PACKAGING",
    }

    repair = li.extract_items([
        "Repair Shaft (Qty: 1), Inquiry Num: 340-25-943RP N 768.00",
    ])[0]
    assert "SPARE PARTS" in repair["tags"]
    assert repair["attributes"]["component"] == "REPAIR SHAFT"
    assert "spare_part_type" not in repair["attributes"]
    assert "spare_part_component" not in repair["attributes"]

    replacement = li.extract_items([
        "Replacement Drive Set (Ship Direct) (Qty: 1), N 692.00",
    ])[0]
    assert {"DRIVE COMPONENTS", "SPARE PARTS", "SHIPPING"} <= set(replacement["tags"])
    assert "V-BELT DRIVE" not in replacement["tags"]
    assert replacement["attributes"]["component"] == "REPLACEMENT DRIVE SET"

    replacement_bearings = li.extract_items([
        'Replacement Bearings, 2 7/16" Bore (Pair) (Qty: 1), N 1,166.00',
    ])[0]
    assert replacement_bearings["attributes"]["component"] == "REPLACEMENT BEARINGS"

    spare_motor = li.extract_items([
        "Spare Motor C 1,700.00",
        "304 SS Shaft",
        "Product: Motor",
    ])[0]
    assert spare_motor["attributes"]["component"] == "SPARE MOTOR"

    vendor_named_spare_motor = li.extract_items([
        "Spare WEG or equivalent C 619.07",
        "3 HP, 3600 RPM, Enclosure: Premium Explosion Proof",
        "182T, Cast Iron, Foot Mounted, 3/60/230/460, F1, 1.15 SF",
    ])[0]
    assert vendor_named_spare_motor["attributes"]["component"] == "SPARE MOTOR"
    assert "spare_part_review" not in vendor_named_spare_motor["attributes"]

    replacement_isolators = li.extract_items([
        "Replacement Isolators (SHIP DIRECT) (Qty: 1), N 736.00",
    ])[0]
    assert replacement_isolators["attributes"]["component"] == "REPLACEMENT ISOLATORS"

    unknown_part = li.extract_items([
        "Replacement Thermocouple, Type J, (Pair) (Qty: 1), N 2,659.00",
    ])[0]
    assert "component" not in unknown_part["attributes"]
    assert any("UNCATEGORIZED REPAIR/SPARE PART" in flag
               for flag in unknown_part.get("review_flags") or [])

    repair_ivc = li.extract_items([
        "Repair IVC Linkage (Qty: 1), Inquiry Num: 373-24-1000RP N 1,622.00",
    ])[0]
    assert repair_ivc["attributes"]["component"] == "REPAIR IVC"
    assert repair_ivc["attributes"]["applies_to"] == "IVC"
    assert "used_on" not in repair_ivc["attributes"]

    support_lugs = li.extract_items([
        "Support Lugs, Fan Inlet and Outlet, Inquiry Num: L 898.00",
        "421837 SPARE FAN FOR SN 413585 1220",
    ])[0]
    assert "SPARE PARTS" not in support_lugs["tags"]

    spark_wheel = li.extract_items(["Wheel, Aluminum (AMCA B) L INC"])[0]
    assert {"ALUMINUM", "MATERIALS", "SPARK RESISTANT", "WHEEL"} <= set(spark_wheel["tags"])
    assert spark_wheel["attributes"]["spark_resistant"] == "YES"
    assert spark_wheel["attributes"]["spark_resistant_type"] == "AMCA B"

    spark = li.extract_items(["Spark Resistant Construction, AMCA Type C, 650 Deg F L 1,686.00", "Max"])[0]
    assert "SPARK RESISTANT" in spark["tags"]
    assert spark["attributes"]["spark_resistant_type"] == "AMCA C"
    assert spark["attributes"]["temperature_rating"] == "650F"


def test_silencer_leading_numeric_detail_is_structured_not_reviewed():
    silencer = li.extract_items([
        'Inlet Silencer C 14,054.00',
        'One (1) Aeroacoustic "Silentflow" Model IB10-5.0-4 inlet silencer',
        '85 dBA at Position #3, 3 ft from fan, 5 ft above',
        'drop of at 19,447 CFM.',
        '- Model IB10-5.0-4 inlet silencer',
        'Vendor: Aeroacoustic Corp',
        'Quote Num: 260259',
        'Product: Inlet Silencer',
    ])[0]
    attrs = silencer["attributes"]
    assert attrs["component"] == "INLET SILENCER"
    assert attrs["brand"] == "AEROACOUSTIC"
    assert attrs["model"] == "IB10-5.0-4"
    assert attrs["noise_target"] == "85 DBA"
    assert attrs["quote_number"] == "260259"
    assert not silencer.get("review_flags")


def test_special_construction_stainless_and_testing_attributes():
    weld = li.extract_items(["Continuous Weld Airstream, Inquiry Num: 253-24-1651 L 4,744.00"])[0]
    assert "SPECIAL CONSTRUCTION" in weld["tags"]
    assert weld["attributes"]["special_construction_type"] == "CONTINUOUS WELD"
    assert weld["attributes"]["special_construction_scope"] == "AIRSTREAM"

    for wording in ("Silicone-Free Caulk", "Silicone-free Caulk", "Silicone Free Caulking"):
        caulk = li.extract_items([f"{wording}, Inquiry Num: 317-24-1723 L"])[0]
        assert "SPECIAL CONSTRUCTION" in caulk["tags"]
        assert caulk["attributes"]["special_construction_type"] == "CAULKING"
        assert caulk["attributes"]["special_construction_detail"] == "SILICONE-FREE"
        assert caulk["attributes"]["inquiry_num"] == "317-24-1723"
        assert not caulk.get("review_flags")

    code_weld = li.extract_items(["AWS D14.6 Code Welding on Rotating Components L 3,357.00"])[0]
    assert code_weld["attributes"]["special_construction_type"] == "CODE WELDING"
    assert code_weld["attributes"]["special_construction_scope"] == "ROTATING COMPONENTS"
    assert code_weld["attributes"]["welding_code"] == "AWS D14.6"

    effective = li.extract_items(["110% Effective Diameter, Inquiry Num: 909-26-465 L 1,000.00"])[0]
    assert "WHEEL" in effective["tags"]
    assert effective["attributes"]["component"] == "WHEEL"
    assert effective["attributes"]["special_construction_type"] == "EFFECTIVE DIAMETER"
    assert effective["attributes"]["effective_diameter_percent"] == "110"
    assert effective["attributes"]["wheel_feature"] == "EFFECTIVE DIAMETER"
    assert effective["attributes"]["wheel_effective_diameter_percent"] == "110"

    cast_hub = li.extract_items([
        "Cast Hub with Straight Bore (Original CAT Design, L",
        "PN 512-1835), Inquiry Num: 311-19-1989",
    ])[0]
    assert {"SPECIAL CONSTRUCTION", "WHEEL"} <= set(cast_hub["tags"])
    assert cast_hub["attributes"]["component"] == "WHEEL"
    assert cast_hub["attributes"]["special_construction_type"] == "CAST HUB"
    assert cast_hub["attributes"]["wheel_feature"] == "CAST HUB"
    assert cast_hub["attributes"]["wheel_hub_construction"] == "CAST HUB"
    assert cast_hub["attributes"]["wheel_hub_bore"] == "STRAIGHT BORE"

    pressure = li.extract_items(["Outlet Pressure Tap L 432.00"])[0]
    assert {"OUTLET", "SPECIAL CONSTRUCTION"} <= set(pressure["tags"])
    assert pressure["attributes"]["special_construction_type"] == "PRESSURE TAP"
    assert pressure["attributes"]["special_construction_scope"] == "OUTLET"

    conduit = li.extract_items([
        "Motor Conduit Box Location L STD",
        "Viewed from Outlet: @8:00",
    ])[0]
    assert conduit["tags"] == ["MOTOR"]
    assert conduit["attributes"]["component"] == "MOTOR"
    assert conduit["attributes"]["motor_conduit_box_position"] == "8:00 VIEWED FROM OUTLET"

    knockout = li.extract_items([
        "Rotate motor conduit box so knockout faces L",
        "downward, Inquiry Num: 374-26-1537",
    ])[0]
    assert "MOTOR" in knockout["tags"]
    assert "SPECIAL CONSTRUCTION" not in knockout["tags"]
    assert knockout["attributes"]["motor_conduit_box_orientation"] == "KNOCKOUT FACES DOWNWARD"

    run_test = li.extract_items([
        "Mechanical Run Test, 2 Hour - Vibration readings L 2,288.00",
        "taken every 15 minutes",
        "Customer witness",
    ])[0]
    assert "TESTING" in run_test["tags"]
    assert run_test["attributes"]["component"] == "MECHANICAL RUN TEST"
    assert "testing_type" not in run_test["attributes"]
    assert run_test["attributes"]["testing_duration"] == "2 HOUR"
    assert run_test["attributes"]["testing_measurements"] == "VIBRATION READINGS"
    assert run_test["attributes"]["witnessed"] == "CUSTOMER"

    voltage_test = li.derive_item_fields({
        "raw": "Mechanical Run Test 380V / 400V / 415V /460V via Delta connection.",
        "details": [],
    })
    assert voltage_test["attributes"]["testing_voltage"] == "380V, 400V, 415V, 460V"

    hardware = li.derive_item_fields({
        "raw": "3/8-16 hardware to mount housing to Drive cover Plate.",
        "details": ["D37/D38 Run Test, Grey, Sticker provided in separate bag."],
    })
    assert "TESTING" not in hardware["tags"]

    stainless = li.extract_items(["Airstream, 304 SS Construction L 1,000.00"])[0]
    assert {"MATERIALS", "STAINLESS STEEL"} <= set(stainless["tags"])
    assert stainless["attributes"]["material_grade"] == "304 SS"
    assert stainless["attributes"]["material_scope"] == "AIRSTREAM"


def test_shaft_seal_sleeve_and_shipping_attributes():
    seal = li.extract_items(["Shaft Seal (Not Gas Tight), 304 SS Construction L 461.00"])[0]
    attrs = seal["attributes"]
    assert {"MATERIALS", "SHAFT SEAL", "STAINLESS STEEL"} <= set(seal["tags"])
    assert attrs["component"] == "SHAFT SEAL"
    assert attrs["shaft_seal_type"] == "NOT GAS TIGHT"
    assert attrs["material_scope"] == "SHAFT SEAL"

    sleeve = li.extract_items(["SS Shaft Sleeve L 184.00"])[0]
    attrs = sleeve["attributes"]
    assert {"MATERIALS", "SHAFT SLEEVE", "STAINLESS STEEL"} <= set(sleeve["tags"])
    assert attrs["component"] == "SHAFT SLEEVE"
    assert attrs["shaft_sleeve"] == "YES"
    assert attrs["used_on"] == "SHAFT"
    assert attrs["material_scope"] == "SHAFT SLEEVE"

    damper = li.extract_items([
        "Outlet Damper, Opposed Blade, With Stuffing Boxes (Shipped Loose) L 10,545.00",
    ])[0]
    assert "DAMPER" in damper["tags"]
    assert "SHIPPING" in damper["tags"]
    assert "SHAFT SEAL" not in damper["tags"]
    assert damper["attributes"]["shipping_state"] == "SHIPPED LOOSE"

    motor_lip = li.derive_item_fields({
        "raw": "Duty, Double Lip Seals Both Ends, CE Mark, IP55 Protection, Inverter Duty 15:1 Constant Torque",
        "details": [],
    })
    assert "SHAFT SEAL" not in motor_lip["tags"]

    direct = li.extract_items(["Replacement Drive Set (Ship Direct) (Qty: 1), N 692.00"])[0]
    assert {"DRIVE COMPONENTS", "SHIPPING"} <= set(direct["tags"])
    assert "V-BELT DRIVE" not in direct["tags"]
    assert direct["attributes"]["shipping_method"] == "SHIP DIRECT"

    split_direct = li.derive_item_fields({
        "raw": 'Isolators, Floor- Spring 1" Deflection (OCT1) (Ship N 683.00',
        "details": ["Direct), Inquiry Num: 362-26-1261"],
    })
    assert {"SHIPPING", "VIBRATION ISOLATION"} <= set(split_direct["tags"])
    assert split_direct["attributes"]["shipping_method"] == "SHIP DIRECT"

    shipping_info = li.derive_item_fields({
        "raw": "SHIPPING INFORMATION AND INVOICE INSTRUCTIONS",
        "details": [],
    })
    assert shipping_info["tags"] == ["SHIPPING"]
    assert shipping_info["attributes"]["shipping_instruction"] == "SHIPPING/INVOICE INSTRUCTIONS"

    aux = li.derive_item_fields({
        "raw": "All auxiliary items except motor are to be ship loose.",
        "details": [],
    })
    assert aux["tags"] == ["SHIPPING"]
    assert aux["attributes"]["shipping_state"] == "SHIPPED LOOSE"
    assert aux["attributes"]["shipping_instruction"] == "SHIP AUXILIARY ITEMS LOOSE"
    assert aux["attributes"]["shipping_scope"] == "AUXILIARY ITEMS EXCEPT MOTOR"

    warranty = li.derive_item_fields({
        "raw": "1 Year warranty from date of shipment. Drive sets will have standard warranty offered by vendor.",
        "details": [],
    })
    assert "SHIPPING" not in warranty["tags"]

    ship_date_warranty = li.derive_item_fields({
        "raw": "Extended Warranty: Chicago Blower's standard",
        "details": ["warranty extended to 24 months from ship date."],
    })
    assert "SHIPPING" not in ship_date_warranty["tags"]


def test_drain_attributes():
    housing = li.extract_items(["Housing Drain with Plug, 304 SS Construction L 328.00"])[0]
    attrs = housing["attributes"]
    assert "HOUSING" not in housing["tags"]
    assert attrs["component"] == "HOUSING DRAIN"
    assert attrs["size"] == "STD"
    assert "drain_type" not in attrs and "drain_closure" not in attrs
    assert attrs["material"] == "STAINLESS STEEL"
    assert attrs["material_grade"] == "304 SS"
    assert attrs["material_scope"] == "HOUSING DRAIN"

    housing_plural = li.extract_items(["Housing Drains with Plugs, Inquiry Num: 333-26-1234 L 252.00"])[0]
    assert housing_plural["tags"] == ["DRAIN"]
    assert housing_plural["attributes"]["component"] == "HOUSING DRAIN"
    assert housing_plural["attributes"]["size"] == "STD"

    inlet = li.extract_items(['Inlet Box, Drain Plug 3/4" Diameter L 51.00'])[0]
    assert "INLET" not in inlet["tags"]
    assert inlet["attributes"]["component"] == "INLET BOX DRAIN"
    assert inlet["attributes"]["size"] == '3/4"'

    motor = li.extract_items([
        "Motor (CEM3711T-10hp motor C 618.96 25.00",
        "M7A Add Condensation Drain Holes - Vertical Shaft",
        "Down)",
    ])[0]
    assert motor["attributes"]["component"] == "MOTOR CONDUIT BOX DRAIN"
    assert motor["attributes"]["size"] == "STD"

    motor_ss = li.extract_items(["Motor (Conduit Box Drain, 304 SS Construction C 100.00"])[0]
    attrs = motor_ss["attributes"]
    assert attrs["component"] == "MOTOR CONDUIT BOX DRAIN"
    assert attrs["material"] == "STAINLESS STEEL"
    assert attrs["material_grade"] == "304 SS"
    assert attrs["material_scope"] == "MOTOR CONDUIT BOX DRAIN"


def test_standalone_actuator_used_on_review():
    actuator = li.extract_items([
        "Bettis #RPED150 Double Acting Pneumatic Actuator C 6,537.14 1,046.00",
        "Vendor: Novaspect",
        "Product: Actuator",
    ])[0]
    attrs = actuator["attributes"]
    assert attrs["component"] == "ACTUATOR"
    assert "used_on" not in attrs
    assert attrs["used_on_review"] == "INCONCLUSIVE - INLET/OUTLET/PRESPIN/IVC"


def test_tag_counts():
    store = _seeded_store()
    counts = {t: (j, i) for t, j, i in li.tag_counts(store)}
    assert counts["SHAFT SEAL"][0] == 2   # both jobs
    assert counts["SPARK RESISTANT"][0] == 1


def test_bounded_corpus_parser_improvements():
    silencer = li.extract_items([
        "INLET SILENCER C 1,000.00",
        "CIB - Circular Inlet silencer",
        "Product: Inlet Silencer",
    ])[0]
    assert silencer["attributes"]["component"] == "INLET SILENCER"

    multi = li.derive_item_fields({
        "raw": "Filter Box, Inquiry Num: 123-26-1234, 456-26-5678 L 100.00",
        "details": [],
    })
    assert multi["attributes"]["inquiry_num"] == ["123-26-1234", "456-26-5678"]
    single = li.derive_item_fields({
        "raw": "Filter, Inquiry Num: 123-26-1234 L 100.00",
        "details": [],
    })
    assert single["attributes"]["inquiry_num"] == "123-26-1234"
    assert "NAN" not in li.normalize_text("Filter NaN Box L 100.00")

    cases = {
        "T Rails": ("UNITARY BASE", "unitary_base_type", "T RAILS"),
        "Offset Base": ("UNITARY BASE", "unitary_base_type", "OFFSET BASE"),
        "Panel Mounting Holes": ("PANEL MOUNTING HOLES", "mounting_holes", "YES"),
        "Mounting Feet": ("MOUNTING", "mounting_type", "FEET"),
        "Mounting Lugs": ("MOUNTING", "mounting_type", "LUGS"),
        "Inlet Venturi": ("INLET", "inlet_subcategory", "VENTURI"),
        "Blade Side Guard": ("BLADE SIDE GUARD", "guard_type", "BLADE SIDE"),
        "Filter Box": ("FILTER BOX", "filter_type", "FILTER BOX"),
        "Piezometer Ring": ("PIEZOMETER RING", "piezometer_ring", "YES"),
    }
    for wording, (component, key, value) in cases.items():
        attrs = li.derive_item_fields({"raw": wording, "details": []})["attributes"]
        assert attrs["component"] == component, wording
        assert attrs[key] == value, wording


def test_structured_pdf_rows_preserve_item_and_detail_sources():
    rows = [
        {"text": "Inlet Damper Actuator L 1,000.00",
         "source": {"page": 2, "row": 11, "top": 110}},
        {"text": "Vendor: Siemens", "source": {"page": 2, "row": 12, "top": 120}},
    ]
    item = li.extract_items(rows)[0]
    assert item["source"]["page"] == 2
    assert item["source"]["row"] == 11
    assert item["source"]["source_text"] == "Inlet Damper Actuator L 1,000.00"
    assert item["details"] == ["Vendor: Siemens"]
    assert item["detail_sources"][0]["row"] == 12


def test_nan_cells_do_not_survive_in_raw_or_details():
    items = li.extract_items([
        "Filter NaN Box L 100.00",
        "Vendor: NaN Greenheck",
    ])
    assert "NAN" not in items[0]["raw"].upper()
    assert all("NAN" not in detail.upper() for detail in items[0]["details"])


def test_inventory_xlsx(tmp: Path):
    import find_orders
    from openpyxl import load_workbook
    store = _seeded_store()
    hits = li.search(store, [])  # no filter -> the full inventory
    out = find_orders.write_xlsx(hits, tmp / "inv.xlsx")
    ws = load_workbook(out).active
    assert ws.cell(1, 1).value == "Job #"
    assert ws.max_row == 1 + sum(len(h["matches"]) for h in hits)


def test_feature_matrix_sheet(tmp: Path):
    import find_orders
    from openpyxl import load_workbook
    store = _seeded_store()
    out = find_orders.write_xlsx(li.search(store, []), tmp / "inv.xlsx", store)
    ws = load_workbook(out)["Feature Matrix"]
    headers = [c.value for c in ws[1]]
    rows = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    assert {"421314", "421999"} <= set(rows)
    # Green ✓ where the order has the feature, red (blank) where it doesn't.
    seal = headers.index("SHAFT SEAL") + 1
    spark = headers.index("SPARK RESISTANT") + 1
    assert ws.cell(rows["421314"], seal).value == "✓"
    assert ws.cell(rows["421999"], seal).value == "✓"
    assert ws.cell(rows["421314"], spark).value == "✓"
    assert ws.cell(rows["421999"], spark).value is None
    assert str(ws.cell(rows["421999"], spark).fill.start_color.rgb).endswith("FFC7CE")
    assert str(ws.cell(rows["421999"], seal).fill.start_color.rgb).endswith("C6EFCE")


def test_feature_matrix_full_profile_when_filtered(tmp: Path):
    # A search narrows the hits, but each matched job's matrix row still shows
    # its WHOLE feature profile (from the store), not just the matched items.
    import find_orders
    from openpyxl import load_workbook
    store = _seeded_store()
    hits = li.search(store, ["spark"])           # only 421314, one matched item
    out = find_orders.write_xlsx(hits, tmp / "f.xlsx", store)
    ws = load_workbook(out)["Feature Matrix"]
    headers = [c.value for c in ws[1]]
    assert ws.max_row == 2 and ws.cell(2, 1).value == "421314"
    assert ws.cell(2, headers.index("SHAFT SEAL") + 1).value == "✓"  # not the spark item


def _mini_pdf(lines: list, path: Path) -> None:
    """A minimal one-page text PDF (Helvetica, one string per line) — just
    enough for pdfplumber to extract real words back out, so the whole
    parse_sales_order_pdf -> extract_items path runs against an actual PDF."""
    content = ["BT", "/F1 10 Tf"]
    y = 760
    for ln in lines:
        esc = str(ln).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        content.append(f"1 0 0 1 40 {y} Tm ({esc}) Tj")
        y -= 14
    content.append("ET")
    stream = "\n".join(content).encode("latin-1", "replace")
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R "
        b"/Resources << /Font << /F1 5 0 R >> >> >>",
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, body in enumerate(objs, start=1):
        offsets.append(len(out))
        out += f"{i} 0 obj\n".encode() + body + b"\nendobj\n"
    xref_pos = len(out)
    out += f"xref\n0 {len(objs) + 1}\n".encode() + b"0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_pos}\n%%EOF").encode()
    path.write_bytes(bytes(out))


def test_order_verification_report_is_not_parsed_as_sales_order(tmp: Path):
    from sales_orders import parse_sales_order_pdf

    pdf = tmp / "421507 - Sales Order CO2.pdf"
    _mini_pdf([
        "Order Verification Report",
        "Design Info",
        "D95 Backward Curved SW, SIZE 270, A/4, CW, TH, 44.5%, WHEEL TYPE Backward Curved",
        "Performance",
        "CFM 12689, RPM 3570, DESIGN TEMP 95, MAX TEMP 95, ELEV 965",
    ], pdf)
    parsed = parse_sales_order_pdf(pdf)
    assert not parsed["design_desc"] and not parsed["size"]
    assert not parsed["line_items"] and not parsed["co_history"]


def test_legacy_summary_handles_class_and_wrapped_wheel_type():
    from sales_orders import _legacy_spec_from_text

    parsed = _legacy_spec_from_text([
        "Design Info",
        "D38 CAPB, SIZE 1400, A/4 With Base for foot mounted motor, WHEEL SIZE 13 x 3-1/4,",
        "INLET DIAMETER 8, CCW, TH, 100%, WHEEL TYPE",
        "Radial Blade",
        "Base Fan, 13 x 3-1/4 RB Wheel",
        "Performance",
        "RPM 3600, DESIGN TEMP 70, MAX TEMP 70",
    ])
    assert parsed["design_desc"] == "CAPB"
    assert parsed["size"] == "1400" and parsed["arrangement"] == "A/4"
    assert parsed["rotation"] == "CCW" and parsed["discharge"] == "TH"
    assert parsed["pct_width"] == "100" and parsed["wheel_type"] == "RB"

    classed = _legacy_spec_from_text([
        "Design Info",
        "D1903 PFD, SIZE 3612 (1800 RPM or less), C/HD, A/4, CCW, DB, 95%, WHEEL TYPE Airfoil",
        "Performance",
        "RPM 1780, DESIGN TEMP 120, MAX TEMP 120",
        "CSIV10C Chicago Blower Corporation Page 1 of 4",
    ])
    assert classed["size"] == "3612 (1800 RPM or less)"
    assert classed["fan_class"] == "C/HD" and classed["arrangement"] == "A/4"
    assert classed["wheel_type"] == "AF"
    assert classed["design_temp"] == "120" and classed["max_temp"] == "120"

    direct = _legacy_spec_from_text([
        "Design Info",
        "Design 36A SQAD Dual Direct Drive, SIZE 33, Arrangement 4, CW, UB, 78.7%",
        "Base Fan with CW and CCW Wheel and Housing, Dual Shaft Motor and Common Unitary Base",
        "Performance",
        "CFM 48000, DESIGN TEMP 105, MAX TEMP 105",
    ])
    assert direct == {
        "design_desc": "SQAD Dual Direct Drive", "size": "33",
        "arrangement": "Arrangement 4", "motor_pos": "N/A", "fan_class": "N/A",
        "rotation": "CW", "discharge": "UB", "pct_width": "78.7",
        "wheel_type": "", "design_temp": "105", "max_temp": "105",
    }


def test_unpriced_legacy_base_fan_is_captured():
    items = li.extract_items([
        "Base Fan with CW and CCW Wheel and Housing, Dual Shaft Motor and Common Unitary Base",
    ])
    assert len(items) == 1
    assert {"BASE FAN", "HOUSING", "UNITARY BASE", "WHEEL"} <= set(items[0]["tags"])


def test_repair_missing_sales_order_summaries_is_fill_only(tmp: Path):
    from sales_orders import repair_missing_sales_order_summaries

    pdf = tmp / "421618 - Sales Order CO1.pdf"
    _mini_pdf([
        "Chicago Blower Corporation Sales Order",
        "Order# RepRef#",
        "421618 987",
        "Design Info",
        "D95 Backward Curved SW, SIZE 200, A/4, CCW, UB, 80.9%, WHEEL TYPE Backward Curved",
        "Performance",
        "CFM 3178, DESIGN TEMP 300, MAX TEMP 300",
    ], pdf)
    job = {
        "job": "421618",
        "so_pdf": str(pdf),
        "so_size": "KEEP",
        "so_special_temp": "0",
    }
    assert repair_missing_sales_order_summaries([job]) == 1
    assert job["so_size"] == "KEEP"                 # known values are never replaced
    assert job["so_rotation"] == "CCW" and job["so_pct_width"] == "80.9"
    assert job["so_special_temp"] == "300"          # derived fields are reconciled
    assert repair_missing_sales_order_summaries([job]) == 0   # now complete / idempotent

    complete = {
        "job": "421973",
        "so_design_desc": "SQB SW Belt Drive",
        "so_size": "27",
        "so_arrangement": "A/9SL",
        "so_design_temp": "300",
        "so_max_temp": "300",
        "so_special_temp": "0",
    }
    assert repair_missing_sales_order_summaries([complete]) == 1
    assert complete["so_special_temp"] == "300"
    assert repair_missing_sales_order_summaries([complete]) == 0

    standard_pdf = tmp / "421619 - Sales Order CO1.pdf"
    _mini_pdf([
        "Chicago Blower Corporation Sales Order",
        "Order# RepRef#",
        "421619 987",
        "Design Info",
        "D95 Backward Curved SW, SIZE 245, A/4, CCW, UB, 101.6%, WHEEL TYPE Backward Curved",
        "Performance",
        "CFM 3178, DESIGN TEMP 77, MAX TEMP 77",
    ], standard_pdf)
    known_low = {
        "job": "421619",
        "so_pdf": str(standard_pdf),
        "so_special_temp": "-45",
    }
    assert repair_missing_sales_order_summaries([known_low]) == 1
    assert known_low["so_special_temp"] == "-45"     # a parser default cannot erase it


def test_pdf_roundtrip(tmp: Path):
    # End to end: a real (tiny) PDF through parse_sales_order_pdf.
    try:
        import pdfplumber  # noqa: F401
    except ImportError:
        print("      (pdfplumber not installed - skipping the pdf round-trip)")
        return
    from sales_orders import parse_sales_order_pdf
    pdf = tmp / "421314 - Sales Order CO#1.pdf"
    _mini_pdf(SO_LINES, pdf)
    res = parse_sales_order_pdf(pdf)
    norms = [it["norm"] for it in res["line_items"]]
    assert any("SPARK RESISTANT" in n for n in norms), norms
    assert any("ZERK FITTINGS" in n for n in norms), norms     # section capture
    assert not any("TOTAL" in n for n in norms), norms
    assert res["co_history"] and "CO#1" in res["co_history"][0].replace(" ", "")


def test_co_history_joins_wrapped_description_lines():
    from sales_orders import _co_history_from_lines

    lines = [
        "***UNAPPROVED ORDER - DRAWINGS ONLY***",
        "CO#2 071326 DG - ADDED OUTLET DAMPER AND",
        "IVC ACTUATORS PER CUSTOMER REQUEST",
        "CO#1 070826 AMF - CORRECTED MOTOR VENDOR",
        "__________________________",
        "Design Info",
    ]
    assert _co_history_from_lines(lines) == [
        "CO#2 071326 DG - ADDED OUTLET DAMPER AND IVC ACTUATORS PER CUSTOMER REQUEST",
        "CO#1 070826 AMF - CORRECTED MOTOR VENDOR",
    ]


def main() -> int:
    passed = 0
    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)
        for name, fn in sorted(globals().items()):
            if not name.startswith("test_") or not callable(fn):
                continue
            (fn(tmp) if "tmp" in fn.__code__.co_varnames else fn())
            print(f"  ok  {name}")
            passed += 1
    print(f"\n{passed} tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
