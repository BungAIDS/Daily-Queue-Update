"""Email an existing Excel report — no scrape, no diff, no AI, no stages.

    python send_excel.py                 # newest report that HAS an AI overview
    python send_excel.py --dry-run       # show what would be sent, send nothing
    python send_excel.py "C:\\path\\queue_2026-06-08.xlsx"   # send a specific file

Use this when you already have a report you're happy with and just want it in
your inbox. It regenerates nothing.

With no path, it scans OUTPUT_DIR and picks the most recent report that actually
contains an AI overview — so a later "no-AI" report (e.g. from `main.py --no-ai`)
won't be sent by mistake. It prints the folder it searched and the full path it
chose, so you can always see exactly which file went out.

The email is rebuilt to match a normal daily run exactly: the briefing text,
anomalies, action items, and counts line are read back out of the Changes tab
and sent through the same emailer the daily run uses, with the workbook attached.
"""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from config import EMAIL_TO, OUTPUT_DIR
from emailer import _send, send_daily_briefing

# Changes-tab section headers -> the diff key whose count the email reports.
# Matched by prefix; the trailing "(N)" carries the count.
_SECTIONS = [
    ("New orders", "new"),
    ("Returning orders", "returning"),
    ("Completed / Removed", "removed"),
    ("Changed orders", "changed"),
    ("Persistent orders", "persistent"),
]

# Markers that mean a report's "AI Briefing" is a placeholder, not real output
# (no API key, AI stage not run yet, API failed, etc.).
_PLACEHOLDER_MARKERS = (
    "skip", "not generated", "unavailable", "no briefing",
    "no anthropic", "failed", "in the attached report",
)


def _candidates() -> List[Path]:
    """All reports in OUTPUT_DIR, newest (by save time) first."""
    return sorted(OUTPUT_DIR.glob("queue_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)


def _report_date(path: Path) -> str:
    """The date in the filename (queue_YYYY-MM-DD.xlsx), else today."""
    m = re.search(r"\d{4}-\d{2}-\d{2}", path.stem)
    return m.group(0) if m else date.today().isoformat()


def _is_placeholder(briefing_text: str) -> bool:
    t = (briefing_text or "").strip().lower()
    if not t:
        return True
    # Real briefings are prose; placeholders are short parenthesized notes.
    return t.startswith("(") and any(m in t for m in _PLACEHOLDER_MARKERS)


def _parse_changes_tab(path: Path) -> Tuple[Dict[str, Any], Dict[str, int]]:
    """Reconstruct the briefing dict and the section counts from the Changes
    tab, so the email can be rebuilt to match a normal run exactly."""
    ws = load_workbook(path, read_only=True)["Changes"]
    briefing_text = ""
    anomalies: list[str] = []
    action_items: list[dict] = []
    counts: Dict[str, int] = {}
    mode: str | None = None

    for row in ws.iter_rows(values_only=True):
        a = row[0]

        if isinstance(a, str):
            section = next((key for prefix, key in _SECTIONS if a.startswith(prefix)), None)
            if section is not None:
                m = re.search(r"\((\d+)\)", a)
                counts[section] = int(m.group(1)) if m else 0
                mode = None
                continue
            if a.startswith("AI Briefing"):
                mode = "briefing"
                continue
            if a.startswith("Anomalies"):
                mode = "anomalies"
                continue
            if a.startswith("Top Action Items"):
                mode = "action_header"  # the row after this is the column header
                continue

        if mode == "briefing" and isinstance(a, str) and a.strip():
            briefing_text = a.strip()
            mode = None
        elif mode == "anomalies" and isinstance(a, str) and a.lstrip().startswith("-"):
            anomalies.append(a.lstrip("- ").strip())
        elif mode == "action_header":
            mode = "action_data"  # skip the Rank/Job #/Reason header row
        elif mode == "action_data":
            rank, job, reason = (list(row) + [None, None, None])[:3]
            if not job and not rank:
                mode = None
            else:
                action_items.append({"rank": rank, "job": str(job or ""), "reason": reason or ""})

    briefing = {
        "briefing": briefing_text or "(briefing in the attached report)",
        "anomalies": anomalies,
        "action_items": action_items,
    }
    return briefing, counts


def _select() -> Tuple[Path, Dict[str, Any], Dict[str, int], bool]:
    """Pick the report to send: the most recent one that has a real AI overview,
    falling back to the most recent overall if none do. Returns
    (path, briefing, counts, has_ai)."""
    cands = _candidates()
    if not cands:
        raise SystemExit(f"No queue_*.xlsx found in {OUTPUT_DIR}")

    newest_fallback = None
    for p in cands:  # newest first
        try:
            briefing, counts = _parse_changes_tab(p)
        except Exception:
            continue
        has_ai = not _is_placeholder(briefing["briefing"])
        if newest_fallback is None:
            newest_fallback = (p, briefing, counts, has_ai)
        if has_ai:
            return p, briefing, counts, True

    # Nothing parsed at all -> use the newest file with empty briefing.
    if newest_fallback is None:
        p = cands[0]
        return p, {"briefing": "(briefing in the attached report)", "anomalies": [], "action_items": []}, {}, False
    return newest_fallback


def _send_report(path: Path, briefing: Dict[str, Any], counts: Dict[str, int]) -> None:
    # send_daily_briefing only needs each section's length, so size throwaway
    # lists to the parsed counts — the email's "Counts:" line then matches.
    diff = {k: [None] * counts.get(k, 0)
            for k in ("new", "returning", "removed", "changed", "persistent")}
    try:
        send_daily_briefing(briefing, diff, path, _report_date(path))
    except Exception as e:  # never fail to send just because parsing drifted
        print(f"(Could not rebuild the full briefing body: {e}; sending file with a short note.)")
        _send(EMAIL_TO, f"Daily Queue Briefing — {_report_date(path)}",
              f"Daily queue report attached: {path.name}", attachment=path)


def main() -> int:
    if not EMAIL_TO:
        print("EMAIL_TO is not set in .env — nothing to send to.")
        return 1

    dry_run = "--dry-run" in sys.argv
    paths = [a for a in sys.argv[1:] if not a.startswith("--")]

    if paths:
        path = Path(paths[0])
        if not path.exists():
            print(f"File not found: {path}")
            return 1
        try:
            briefing, counts = _parse_changes_tab(path)
            has_ai = not _is_placeholder(briefing["briefing"])
        except Exception:
            briefing, counts, has_ai = {"briefing": "", "anomalies": [], "action_items": []}, {}, False
    else:
        path, briefing, counts, has_ai = _select()

    mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    print(f"Searched folder: {OUTPUT_DIR}")
    print(f"Selected report: {path}")
    print(f"  last saved:    {mtime}")
    print(f"  AI overview:   {'present' if has_ai else 'MISSING (placeholder only)'}")
    if not has_ai:
        print("  WARNING: this report has no AI overview. If you expected one, the report")
        print("           you want may be elsewhere (pass its full path), or run the AI stage.")

    if dry_run:
        print("\n--dry-run: nothing was sent. Re-run without --dry-run to send.")
        return 0

    _send_report(path, briefing, counts)
    print(f"\nSent {path.name} to {EMAIL_TO}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
