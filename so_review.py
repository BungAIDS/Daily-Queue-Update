"""A throwaway Sales-Order review workbook + a note queue Claude works through.

Purpose: give a human a place to comment on how each Sales-Order line item was
captured and what should be done with it, without touching the live product.
The co-authored master stays read-only; this builds a SEPARATE, disposable
.xlsx you can scribble in. It exists only as a path to a better parser - it is
not part of the final workbook.

The Sales Order tab is the canonical row grid. Every hierarchy row is a real
cell row, so Excel's normal column filters search the actual data and an edited
Note stays attached to that row without a picker or formula projection.

The loop:
  1. `python so_review.py build`  -> writes sales_order_review.xlsx from the
     line-items store: every order's component hierarchy (so_hierarchy), one
     row per row, with an editable "Note" cell containing any open note already
     recorded.
  2. Filter Sales Order to the orders/rows you want, type in Note, then save
     and close. The workbook remains the durable draft; the next refresh or
     review folds those entries into so_review_notes.json for publication.
  3. `python so_review.py list` shows the OPEN notes. Claude acts on each and
     `python so_review.py handle <id> "what I did"` marks it handled with a
     resolution. The next refresh removes it from the active review row so that
     row is ready for another note. Resolved notes remain in the JSON queue and
     the workbook's Resolved tab as history.

Notes may anchor to any displayed hierarchy row. Source rows use their stable
order + item number; component, attribute, and review rows use their displayed
hierarchy text. That text is retained as context if a later re-parse removes or
renames the row.

Pure logic (store + rows) is import-light and unit-tested; the two Excel
functions lazy-import openpyxl so the rest of the module loads without it.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterator, List, Optional

import so_hierarchy
from config import BACKLOG_DIR

REVIEW_STORE_PATH = BACKLOG_DIR / "so_review_notes.json"
DEFAULT_WORKBOOK = BACKLOG_DIR / "sales_order_review.xlsx"
PARSER_METRICS_PATH = BACKLOG_DIR / "so_review_parser_metrics.json"

# Return channel for Claude's handled-marks. The note queue flows UP to Claude
# with the other published stores (data_push), but that push is one-way, so
# Claude records each note it resolves in this small TRACKED ledger at the repo
# root. It rides down to the user's machine on the normal Git Update, and the
# "Update SO Review" action applies it. Handled notes leave the active review
# rows but remain in the queue and the workbook's Resolved history tab.
# Append-only {handled: [{id, resolution, handled_at}]} so it cannot conflict.
HANDLED_MARKS_PATH = Path(__file__).resolve().parent / "so_review_handled.json"

STATUS_OPEN = "open"
STATUS_HANDLED = "handled"
NOTE_SEPARATOR = "\n\n---\n\n"

# Workbook columns (also the read-back contract). ``Note`` renders open notes
# and accepts new human input. ``Add Note`` is recognized only when migrating an
# older workbook.
NOTES_ADD_NOTE = "Add Note"
HEADERS = ["Order", "Item", "Kind", "Hierarchy", "Price", "Note",
           "Status", "Resolution"]
_COL = {h: i for i, h in enumerate(HEADERS)}

# New workbooks keep the canonical flat grid directly on Sales Order. The other
# names/constants remain only so notes can still be recovered from old picker
# workbooks during the transition.
BROWSE_SHEET = "Sales Order"
NOTES_SHEET = "Notes"
RESOLVED_SHEET = "Resolved"
ORDERS_SHEET = "Orders"
BROWSE_HEADERS = ["Item", "Kind", "Hierarchy", "Price", "Existing Note", "Add Note"]
BROWSE_ADD_NOTE = "Add Note"
RESOLVED_HEADERS = ["Order", "Item", "Line", "Note", "Resolution", "Resolved"]
ROW_KEY_HEADER = "_row_key"
BROWSE_HEADER_ROW = 3
BROWSE_FIRST_ROW = 4


# --------------------------------------------------------------------------- #
# Note queue store                                                             #
# --------------------------------------------------------------------------- #
def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def load_store(path: Optional[Path] = None) -> Dict[str, Any]:
    p = Path(path or REVIEW_STORE_PATH)
    if not p.exists():
        return {"notes": []}
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"notes": []}
    if not isinstance(data, dict) or not isinstance(data.get("notes"), list):
        return {"notes": []}
    return data


def save_store(store: Dict[str, Any], path: Optional[Path] = None) -> None:
    p = Path(path or REVIEW_STORE_PATH)
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(json.dumps(store, indent=2, default=str), encoding="utf-8")
    tmp.replace(p)


def _next_id(store: Dict[str, Any]) -> int:
    return max((int(n.get("id", 0)) for n in store["notes"]), default=0) + 1


def record_note(store: Dict[str, Any], order: str, item_no: Any, item_text: str,
                note: str, when: Optional[str] = None,
                row_key: str = "") -> Optional[Dict[str, Any]]:
    """Append a note for (order, item). Idempotent on exact text: if a note with
    the same order + item + text already exists (any status), nothing is added
    and None is returned — so re-syncing the same workbook never duplicates."""
    order, note = str(order).strip(), str(note).strip()
    item = str(item_no).strip()
    target_text = str(item_text or "").strip()
    target_key = str(row_key or "").strip()
    if not order or not note:
        return None
    for n in store["notes"]:
        existing_item = str(n.get("item_no", "")).strip()
        existing_key = str(n.get("row_key", "")).strip()
        if target_key and existing_key:
            same_target = existing_key == target_key
        elif item or existing_item:
            same_target = existing_item == item
        else:
            same_target = (
                str(n.get("item_text", "")).strip().casefold()
                == target_text.casefold()
            )
        if (str(n.get("order")) == order and same_target
                and str(n.get("note")) == note):
            return None
    entry = {"id": _next_id(store), "order": order, "item_no": item,
             "item_text": target_text, "note": note,
             "status": STATUS_OPEN, "created_at": when or _now(),
             "handled_at": None, "resolution": None}
    if target_key:
        entry["row_key"] = target_key
    store["notes"].append(entry)
    return entry


def mark_handled(store: Dict[str, Any], note_id: int, resolution: str,
                 when: Optional[str] = None) -> bool:
    """Mark one note handled with what was done. Claude calls this after acting."""
    for n in store["notes"]:
        if int(n.get("id", -1)) == int(note_id):
            n["status"] = STATUS_HANDLED
            n["resolution"] = str(resolution or "").strip()
            n["handled_at"] = when or _now()
            return True
    return False


def open_notes(store: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [n for n in store["notes"] if n.get("status") != STATUS_HANDLED]


def parser_review_metrics(line_items_store: Dict[str, Any]) -> Dict[str, int]:
    """Counts behind the workbook's red MARKED FOR REVIEW rows."""
    review_rows = 0
    jobs_with_review = 0
    flagged_items = 0
    parser_flags = 0
    for record in (line_items_store.get("jobs") or {}).values():
        items = (record or {}).get("items") or []
        job_rows = sum(
            row.get("kind") == so_hierarchy.KIND_REVIEW
            for row in so_hierarchy.tree_rows(items)
        )
        review_rows += job_rows
        jobs_with_review += int(job_rows > 0)
        flagged_items += sum(bool(item.get("review_flags")) for item in items)
        parser_flags += sum(len(item.get("review_flags") or []) for item in items)
    return {
        "review_rows": review_rows,
        "jobs_with_review": jobs_with_review,
        "flagged_items": flagged_items,
        "parser_flags": parser_flags,
    }


def record_parser_metrics(before: Dict[str, int], after: Dict[str, int],
                          item_count: int, job_count: int,
                          path: Optional[Path] = None) -> Dict[str, Any]:
    """Append one reparse comparison so parser progress remains measurable."""
    destination = Path(path or PARSER_METRICS_PATH)
    try:
        history = json.loads(destination.read_text(encoding="utf-8")) \
            if destination.exists() else {"runs": []}
    except (json.JSONDecodeError, OSError):
        history = {"runs": []}
    if not isinstance(history, dict) or not isinstance(history.get("runs"), list):
        history = {"runs": []}
    entry = {
        "recorded_at": _now(),
        "items": int(item_count),
        "jobs": int(job_count),
        "before": dict(before),
        "after": dict(after),
        "review_rows_reduced": int(before.get("review_rows", 0)
                                    - after.get("review_rows", 0)),
    }
    history["runs"].append(entry)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temp = destination.with_suffix(destination.suffix + ".tmp")
    temp.write_text(json.dumps(history, indent=2), encoding="utf-8")
    temp.replace(destination)
    return entry


# --------------------------------------------------------------------------- #
# Handled-marks ledger (Claude -> user return channel)                         #
# --------------------------------------------------------------------------- #
def _load_ledger() -> Dict[str, Any]:
    if not HANDLED_MARKS_PATH.exists():
        return {"handled": []}
    try:
        data = json.loads(HANDLED_MARKS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {"handled": []}
    return data if isinstance(data.get("handled"), list) else {"handled": []}


def record_handled_mark(note_id: int, resolution: str, when: Optional[str] = None) -> None:
    """Append (or update) a handled-mark in the tracked ledger, so it travels to
    the user's machine on the next Git Update."""
    led = _load_ledger()
    led["handled"] = [m for m in led["handled"] if int(m.get("id", -1)) != int(note_id)]
    led["handled"].append({"id": int(note_id), "resolution": str(resolution or "").strip(),
                           "handled_at": when or _now()})
    HANDLED_MARKS_PATH.write_text(json.dumps(led, indent=2), encoding="utf-8")


def apply_handled_marks(store: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Fold the ledger's handled-marks into a local queue: any note whose id is
    marked (and isn't already handled) becomes handled with the recorded
    resolution. Returns the notes newly closed this call (for reporting)."""
    marks = {int(m["id"]): m for m in _load_ledger()["handled"] if "id" in m}
    closed = []
    for n in store["notes"]:
        m = marks.get(int(n.get("id", -1)))
        if m and n.get("status") != STATUS_HANDLED:
            n["status"] = STATUS_HANDLED
            n["resolution"] = m.get("resolution", "")
            n["handled_at"] = m.get("handled_at") or _now()
            closed.append(n)
    return closed


# --------------------------------------------------------------------------- #
# Rows: every order's hierarchy + the note recorded for each line item         #
# --------------------------------------------------------------------------- #
def _job_sort_key(job: str) -> tuple:
    return (0, int(job), job) if str(job).isdigit() else (1, 0, str(job))


def _hierarchy_row_keys(rows: List[Dict[str, Any]]) -> List[str]:
    """Stable keys for source and derived rows, including component context."""
    parent = ""
    counts: Dict[str, int] = {}
    keys: List[str] = []
    current_order = ""
    for row in rows:
        order = str(row.get("order", "")).strip()
        if order and order != current_order:
            current_order = order
            parent = ""
            counts = {}
        kind = str(row.get("kind", "")).strip()
        hierarchy = str(row.get("hierarchy", "")).strip()
        item = str(row.get("item_no", "")).strip()
        if kind == so_hierarchy.KIND_COMPONENT:
            parent = hierarchy
        if item:
            base = f"item:{item}"
        else:
            base = f"{kind.casefold()}|{parent.casefold()}|{hierarchy.casefold()}"
        counts[base] = counts.get(base, 0) + 1
        keys.append(base if counts[base] == 1 else f"{base}|{counts[base]}")
    return keys


def _order_notes(review_store: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """order# -> open notes for the active review views, in queue order."""
    out: Dict[str, List[Dict[str, Any]]] = {}
    for n in review_store["notes"]:
        if n.get("status") != STATUS_HANDLED:
            out.setdefault(str(n.get("order")), []).append(n)
    return out


def iter_review_rows(line_items_store: Dict[str, Any],
                     review_store: Dict[str, Any]) -> Iterator[Dict[str, Any]]:
    """One display row per hierarchy row across every order, newest job first,
    with open notes shown on their lines. Handled notes are intentionally absent
    so the line is ready for another note; their history is retained separately.
    A source note anchors by item number; a component/attribute/review note
    anchors by its displayed hierarchy text. If that target no longer exists,
    it re-anchors to the order's first row. `group_start` marks the first row of
    each order. Rows are yielded one order at a time so large workbooks do not
    require a second full-backlog list in memory. Pure — no Excel."""
    by_order = _order_notes(review_store)
    jobs = (line_items_store.get("jobs") or {})
    for jn in sorted(jobs, key=_job_sort_key, reverse=True):
        items = (jobs[jn] or {}).get("items") or []
        tree = so_hierarchy.tree_rows(items)
        if not tree:
            continue
        display_rows = [{
            "kind": tr["kind"],
            "hierarchy": so_hierarchy.indent_text(tr),
            "item_no": tr.get("item_no", ""),
        } for tr in tree]
        row_keys = _hierarchy_row_keys(display_rows)
        # item # -> the tree row index that carries it (a SOURCE row under a
        # component, or a single-line COMPONENT row).
        item_to_idx = {str(tr["item_no"]): i for i, tr in enumerate(tree)
                       if tr.get("item_no") != ""}
        text_to_idx: Dict[str, int] = {}
        key_to_idx = {key: i for i, key in enumerate(row_keys)}
        for i, tr in enumerate(tree):
            key = so_hierarchy.indent_text(tr).strip().casefold()
            if key:
                text_to_idx.setdefault(key, i)
        # Attach by stable item number when available, otherwise by the exact
        # displayed hierarchy text. The order's first row is the safe fallback.
        per_row: Dict[int, List[Dict[str, Any]]] = {}
        for n in by_order.get(str(jn), []):
            stored_key = str(n.get("row_key", "")).strip()
            item = str(n.get("item_no", "")).strip()
            if stored_key and stored_key in key_to_idx:
                idx = key_to_idx[stored_key]
            elif item:
                idx = item_to_idx.get(item, 0)
            else:
                key = str(n.get("item_text", "")).strip().casefold()
                idx = text_to_idx.get(key, 0)
            per_row.setdefault(idx, []).append(n)
        for i, tr in enumerate(tree):
            here = per_row.get(i, [])
            note_txt = NOTE_SEPARATOR.join(
                f"#{n.get('id')}: {text}" for n in here
                if (text := str(n.get("note", "")).strip()))
            item_no = tr.get("item_no")
            yield {
                "order": str(jn),
                "item_no": item_no if item_no != "" else "",
                "kind": tr["kind"],
                "hierarchy": so_hierarchy.indent_text(tr),
                "price": tr.get("price", ""),
                "note": note_txt,
                "status": STATUS_OPEN if here else "",
                "resolution": "",
                "annotatable": True,
                "row_key": row_keys[i],
                "group_start": i == 0,
            }


def review_rows(line_items_store: Dict[str, Any],
                review_store: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Materialized review rows for callers that need random access."""
    return list(iter_review_rows(line_items_store, review_store))


def ingest_edits(review_store: Dict[str, Any],
                 edits: List[Dict[str, Any]], when: Optional[str] = None) -> int:
    """Fold rows read back from the workbook into the queue. Each edit is
    {order, item_no, item_text, note}. Source rows carry an item number; derived
    rows intentionally leave it blank and anchor by item_text. record_note dedups
    exact repeats. Returns how many NEW notes were added."""
    added = 0
    for e in edits:
        if record_note(review_store, e.get("order", ""), e.get("item_no", ""),
                        e.get("item_text", ""), e.get("note", ""), when=when,
                        row_key=e.get("row_key", "")):
            added += 1
    return added


# --------------------------------------------------------------------------- #
# Excel I/O (lazy openpyxl)                                                     #
# --------------------------------------------------------------------------- #
def _write_only_text_cell(ws, value: Any, *, fill=None, font=None,
                          alignment=None):
    from openpyxl.cell import WriteOnlyCell

    cell = WriteOnlyCell(ws, value=value)
    if cell.data_type == "f":
        cell.data_type = "s"
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    return cell


def _append_text_row(ws, values: List[Any]) -> None:
    """ws.append, but formula-proof. Hierarchy or note text starting with ``=``
    (e.g. a quoted price sum) would be stored by openpyxl as an Excel FORMULA —
    a broken one, so Excel "repairs" the sheet by deleting the cell on open.
    Re-type any such cell as plain text."""
    if getattr(ws.parent, "write_only", False):
        ws.append([_write_only_text_cell(ws, value) for value in values])
        return
    ws.append(values)
    for cell in ws[ws.max_row]:
        if cell.data_type == "f":
            cell.data_type = "s"


def write_workbook(path: Path, line_items_store: Dict[str, Any],
                   review_store: Dict[str, Any], *,
                   progress: Optional[Callable[[str], None]] = None) -> int:
    """Write the review workbook. Returns the data-row count.

    Rows are generated and written as streams. Colour and grouping are applied
    as a handful of workbook-level conditional-format rules (O(1) each), never
    styled cell-by-cell, so the large backlog stays bounded in memory.

    Sales Order is one canonical, filterable grid of real rows. Resolved keeps
    handled-note history out of the active inputs."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")      # invites typing
    band_fill = PatternFill("solid", fgColor="F2F2F2")      # alternating orders
    comp_fill = PatternFill("solid", fgColor="DDEBF7")      # component header rows
    comp_font = Font(bold=True)
    review_font = Font(color="C00000", bold=True)           # parser review rows

    def report(message: str) -> None:
        if progress is not None:
            progress(message)

    path = Path(path)
    lock_path = path.with_name(f"~${path.name}")
    if lock_path.exists():
        raise RuntimeError(
            f"Could not write {path.name} — it is open in Excel. Close it, then "
            "run this again."
        )

    ncols = len(HEADERS)
    wb = Workbook(write_only=True)

    # --- Sales Order: canonical data + input grid -----------------------------
    notes = wb.create_sheet(BROWSE_SHEET)
    band_col = ncols + 1                                    # hidden parity helper
    row_key_col = ncols + 2
    notes.freeze_panes = "A2"
    notes.sheet_view.showGridLines = False
    notes.row_dimensions[1].height = 24
    notes.column_dimensions[get_column_letter(band_col)].hidden = True
    notes.column_dimensions[get_column_letter(row_key_col)].hidden = True
    for h, w in {"Order": 10, "Item": 6, "Kind": 11, "Hierarchy": 60, "Price": 12,
                 "Note": 50, "Status": 10, "Resolution": 55}.items():
        notes.column_dimensions[get_column_letter(_COL[h] + 1)].width = w

    header_cells = []
    for column, value in enumerate(HEADERS + ["_band", ROW_KEY_HEADER], start=1):
        if column <= ncols:
            fill = note_fill if column == _COL["Note"] + 1 else header_fill
            font = (Font(color="7F6000", bold=True)
                    if column == _COL["Note"] + 1 else header_font)
            alignment = Alignment(vertical="center")
        else:
            fill = font = alignment = None
        header_cells.append(_write_only_text_cell(
            notes, value, fill=fill, font=font, alignment=alignment
        ))
    notes.append(header_cells)

    report("Building Sales Order review rows...")
    parity = 0
    row_count = 0
    for row_count, r in enumerate(
            iter_review_rows(line_items_store, review_store), start=1):
        if r["group_start"]:
            parity ^= 1                                     # flip per order group
        _append_text_row(notes, [r["order"], r["item_no"], r["kind"], r["hierarchy"],
                                 r["price"], r["note"], r["status"], r["resolution"],
                                 parity, r["row_key"]])
        if row_count % 25_000 == 0:
            report(f"  {row_count:,} rows written...")
    last = row_count + 1

    if last >= 2:
        full = f"A2:{get_column_letter(ncols)}{last}"
        bandL, kindL = get_column_letter(band_col), get_column_letter(_COL["Kind"] + 1)
        noteL = get_column_letter(_COL["Note"] + 1)
        hierarchyL = get_column_letter(_COL["Hierarchy"] + 1)
        # Order banding (fill only) + Kind cues (font only) coexist: each rule
        # sets only the properties it needs, so they layer without fighting.
        notes.conditional_formatting.add(full, FormulaRule(formula=[f"${bandL}2=1"], fill=band_fill))
        notes.conditional_formatting.add(full, FormulaRule(formula=[f'${kindL}2="COMPONENT"'], fill=comp_fill, font=comp_font))
        notes.conditional_formatting.add(full, FormulaRule(formula=[f'${kindL}2="REVIEW"'], font=review_font))
        # Every displayed hierarchy row can take a note. Resolved notes leave
        # this active grid.
        notes.conditional_formatting.add(
            f"{noteL}2:{noteL}{last}",
            FormulaRule(formula=[f'${hierarchyL}2<>""'], fill=note_fill),
        )
    notes.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last}"

    # --- Resolved tab: compact audit history, separate from active inputs ----
    resolved = wb.create_sheet(RESOLVED_SHEET)
    resolved.freeze_panes = "A2"
    for c, width in enumerate((10, 7, 65, 50, 60, 20), start=1):
        resolved.column_dimensions[get_column_letter(c)].width = width
    resolved.append([
        _write_only_text_cell(resolved, value, fill=header_fill, font=header_font)
        for value in RESOLVED_HEADERS
    ])
    handled = sorted(
        (n for n in review_store["notes"] if n.get("status") == STATUS_HANDLED),
        key=lambda n: int(n.get("id", 0)),
        reverse=True,
    )
    resolved_count = 0
    for resolved_count, n in enumerate(handled, start=1):
        _append_text_row(resolved, [
            str(n.get("order", "")), str(n.get("item_no", "")),
            str(n.get("item_text", "")), str(n.get("note", "")),
            str(n.get("resolution", "") or ""), str(n.get("handled_at", "") or ""),
        ])
    resolved.auto_filter.ref = f"A1:F{resolved_count + 1}"
    wb.active = notes

    path.parent.mkdir(parents=True, exist_ok=True)
    building = path.with_name(f"{path.stem}.building{path.suffix}")
    report(f"Saving {row_count:,} rows to {path.name}...")
    try:
        wb.save(str(building))
    except PermissionError:
        raise RuntimeError(
            f"Could not write {building.name}. Close Excel, then run this again."
        ) from None
    try:
        building.replace(path)
    except PermissionError:
        raise RuntimeError(
            f"Could not replace {path.name} because it is open in Excel. The "
            f"completed replacement is preserved as {building.name}."
        ) from None
    report(f"Saved {path.name}.")
    return row_count


def _notes_sheet(wb):
    """The canonical row grid, with legacy workbook fallbacks."""
    for name in (BROWSE_SHEET, NOTES_SHEET, "Line Items"):
        if name in wb.sheetnames:
            ws = wb[name]
            header = {_excel_text(c.value) for c in ws[1]}
            if {"Order", "Note"} <= header:
                return ws
    for ws in wb.worksheets:
        header = {str(c.value or "") for c in ws[1]}
        if {"Order", "Note"} <= header:
            return ws
    return wb.active


def _browse_layout_needs_upgrade(path: Path) -> bool:
    """Whether this is not yet the canonical Sales Order row-grid layout."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        wb = load_workbook(str(path), read_only=True, data_only=False)
    try:
        if BROWSE_SHEET not in wb.sheetnames:
            return True
        header = [_excel_text(c.value) for c in wb[BROWSE_SHEET][1]]
        return not all(h in header for h in HEADERS) or NOTES_ADD_NOTE in header
    finally:
        wb.close()


def _excel_text(value: Any) -> str:
    """Normalize common Excel scalar values without turning 2 into '2.0'."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _note_parts(value: Any) -> List[str]:
    """Split a cell containing several rendered notes back into queue entries."""
    return [part.strip() for part in _excel_text(value).split(NOTE_SEPARATOR) if part.strip()]


def _manual_note_parts(value: Any) -> List[str]:
    """Return human-entered text while ignoring rendered ``#id: note`` history."""
    return [part for part in _note_parts(value)
            if not re.match(r"^#\d+\s*:\s*", part)]


def read_edits(path: Path) -> List[Dict[str, Any]]:
    """Read dedicated inputs and recover manual overrides in rendered Note cells."""
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
        )
        wb = load_workbook(str(path), data_only=False)
    try:
        notes = _notes_sheet(wb)
        header = [_excel_text(c.value) for c in notes[1]]
        readable_headers = HEADERS + [NOTES_ADD_NOTE]
        idx = {h: header.index(h) for h in readable_headers if h in header}
        row_key_index = header.index(ROW_KEY_HEADER) if ROW_KEY_HEADER in header else None

        def notes_value(row: tuple, heading: str) -> str:
            i = idx.get(heading)
            return "" if i is None or i >= len(row) else _excel_text(row[i])

        edits: List[Dict[str, Any]] = []
        workbook_rows: List[Dict[str, Any]] = []
        raw_rows = list(notes.iter_rows(min_row=2, values_only=True))
        for row in raw_rows:
            workbook_rows.append({
                "order": notes_value(row, "Order"),
                "kind": notes_value(row, "Kind"),
                "hierarchy": notes_value(row, "Hierarchy"),
                "item_no": notes_value(row, "Item"),
            })
        derived_row_keys = _hierarchy_row_keys(workbook_rows)
        notes_row_keys: Dict[int, str] = {}
        for sheet_row, (row, derived_key) in enumerate(zip(raw_rows, derived_row_keys), start=2):
            order = notes_value(row, "Order")
            item_no = notes_value(row, "Item")
            stored_key = ""
            if row_key_index is not None and row_key_index < len(row):
                stored_key = _excel_text(row[row_key_index])
            row_key = stored_key or derived_key
            notes_row_keys[sheet_row] = row_key
            if order:
                source = "sales_order" if notes.title == BROWSE_SHEET else "notes"
                typed = _manual_note_parts(notes_value(row, "Note"))
                if NOTES_ADD_NOTE in header:
                    typed.extend(_note_parts(notes_value(row, NOTES_ADD_NOTE)))
                for note in typed:
                    edits.append({
                        "order": order,
                        "item_no": item_no,
                        "item_text": notes_value(row, "Hierarchy"),
                        "note": note,
                        "row_key": row_key,
                        "source": source,
                    })

        # New workbooks already use Sales Order as the canonical grid. The
        # remainder only recovers Add Note entries from the old formula picker.
        if notes.title == BROWSE_SHEET or BROWSE_SHEET not in wb.sheetnames:
            return edits
        browse = wb[BROWSE_SHEET]
        browse_header = [_excel_text(c.value) for c in browse[BROWSE_HEADER_ROW]]
        legacy_formula_input = BROWSE_ADD_NOTE not in browse_header and "Note" in browse_header
        if BROWSE_ADD_NOTE in browse_header:
            add_note_col = browse_header.index(BROWSE_ADD_NOTE) + 1
        elif legacy_formula_input:
            # Before Add Note existed, typing in Sales Order replaced the Note
            # formula. A literal in that old column is recoverable; formulas are
            # merely the normal browse display and must not be imported.
            add_note_col = browse_header.index("Note") + 1
        else:
            return edits
        pending = [
            (row_no, _excel_text(browse.cell(row_no, add_note_col).value))
            for row_no in range(BROWSE_FIRST_ROW, browse.max_row + 1)
            if _excel_text(browse.cell(row_no, add_note_col).value)
            and not (
                legacy_formula_input
                and str(browse.cell(row_no, add_note_col).value).startswith("=")
            )
        ]
        if not pending:
            return edits

        selected_order = _excel_text(browse["B1"].value)
        if not selected_order:
            raise RuntimeError(
                "Sales Order has unsynced Add Note entries but no selected order. "
                "Select the order again, save, close Excel, and retry."
            )
        if ORDERS_SHEET not in wb.sheetnames:
            raise RuntimeError("The hidden Sales Order row map is missing; Add Note entries were not imported.")

        window = None
        for order, start_row, row_count in wb[ORDERS_SHEET].iter_rows(
                min_row=1, max_col=3, values_only=True):
            if _excel_text(order) == selected_order:
                window = (int(start_row), int(row_count))
                break
        if window is None:
            raise RuntimeError(
                f"Selected order {selected_order} is not in the workbook row map; "
                "Add Note entries were not imported."
            )

        start_row, row_count = window
        for output_row, note in pending:
            offset = output_row - BROWSE_FIRST_ROW
            if offset < 0 or offset >= row_count:
                raise RuntimeError(
                    f"Sales Order row {output_row} is outside order {selected_order}; "
                    "its Add Note entry was not imported."
                )
            source_row = start_row + offset
            source_order = _excel_text(notes.cell(source_row, idx["Order"] + 1).value)
            item_no = _excel_text(notes.cell(source_row, idx["Item"] + 1).value)
            hierarchy = _excel_text(notes.cell(source_row, idx["Hierarchy"] + 1).value)
            if source_order != selected_order or not hierarchy:
                raise RuntimeError(
                    f"Sales Order row {output_row} is not a review row for "
                    f"order {selected_order}; its Add Note entry was not imported."
                )
            edits.append({
                "order": source_order,
                "item_no": item_no,
                "item_text": hierarchy,
                "note": note,
                "row_key": notes_row_keys.get(source_row, ""),
                "source": "sales_order",
            })
        return edits
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# CLI                                                                          #
# --------------------------------------------------------------------------- #
def _load_line_items() -> Dict[str, Any]:
    import line_items
    return line_items.load_store()


def _console_progress(message: str) -> None:
    print(message, flush=True)


def _cmd_build(args) -> int:
    store = load_store()
    n = write_workbook(
        Path(args.out), _load_line_items(), store, progress=_console_progress
    )
    pend = len(open_notes(store))
    print(f"Wrote {args.out} ({n} rows). {pend} open note(s) shown; resolved history "
          f"is on the {RESOLVED_SHEET} tab. Filter the real rows on Sales Order and "
          f"type directly in the yellow Note cells.")
    return 0


def _cmd_open(args) -> int:
    """Open the review workbook in Excel, building it first if it's missing."""
    import os
    path = Path(args.out)
    if not path.exists():
        write_workbook(
            path, _load_line_items(), load_store(), progress=_console_progress
        )
        print(f"Built {path}.")
    try:
        os.startfile(str(path))          # Windows: opens in Excel  # type: ignore[attr-defined]
    except AttributeError:               # non-Windows: best-effort
        import subprocess
        subprocess.Popen(["xdg-open", str(path)])
    print(f"Opened {path}.")
    return 0


def _cmd_sync(args) -> int:
    path = Path(args.out)
    if not path.exists():
        print(f"{path} not found — 'Open SO Review' builds it first.", file=sys.stderr)
        return 1
    needs_upgrade = _browse_layout_needs_upgrade(path)
    edits = read_edits(path)
    store = load_store()
    added = ingest_edits(store, edits)
    save_store(store)
    if edits or needs_upgrade:
        try:
            write_workbook(
                path, _load_line_items(), store, progress=_console_progress
            )
        except RuntimeError as exc:
            raise RuntimeError(
                f"Recorded {added} new note(s), but could not refresh the Sales Order "
                f"review workbook. {exc} Re-run refresh after closing Excel; "
                f"the notes will not duplicate."
            ) from None
    cleared = f" Captured {len(edits)} edited Note cell(s)." if edits else ""
    upgraded = " Upgraded Sales Order to the filterable row-grid layout." if needs_upgrade else ""
    print(f"Recorded {added} new note(s).{cleared}{upgraded} {len(open_notes(store))} open in the "
          f"queue ({REVIEW_STORE_PATH}).")
    return 0


def _cmd_refresh(args) -> int:
    """The 'Update SO Review' action: capture anything typed, fold in Claude's
    handled-marks, and rewrite active rows plus the separate resolved history."""
    path = Path(args.out)
    store = load_store()
    edits: List[Dict[str, Any]] = []
    added = 0
    if path.exists():
        edits = read_edits(path)
        added = ingest_edits(store, edits)              # don't lose un-synced typing
    workbook_note_count = len({
        (str(e.get("order", "")),
         str(e.get("row_key") or e.get("item_no") or e.get("item_text") or ""),
         str(e.get("note", "")))
        for e in edits
        if str(e.get("order", "")).strip()
        and (str(e.get("item_no", "")).strip()
             or str(e.get("item_text", "")).strip())
        and str(e.get("note", "")).strip()
    })
    closed = apply_handled_marks(store)                 # Claude's resolutions
    save_store(store)
    try:
        n = write_workbook(
            path, _load_line_items(), store, progress=_console_progress
        )
    except RuntimeError as exc:
        if workbook_note_count:
            safe = (
                f"All {workbook_note_count} workbook note(s) are safely recorded in "
                f"{REVIEW_STORE_PATH} ({added} new this run). "
            )
        else:
            safe = f"The note queue was safely saved to {REVIEW_STORE_PATH}. "
        raise RuntimeError(
            f"{safe}The workbook refresh could not finish. {exc}"
        ) from None
    already = ""
    if workbook_note_count and added == 0:
        already = f"; all {workbook_note_count} workbook note(s) were already recorded"
    print(f"Updated {path} ({n} rows). Captured {added} new note(s){already}; "
          f"applied {len(closed)} handled resolution(s); {len(open_notes(store))} still open.")
    for c in closed:
        print(f"  handled #{c['id']} (order {c['order']} item {c['item_no']}): "
              f"{c.get('resolution', '')}")
    return 0


def _cmd_reparse(args) -> int:
    """The 'Re-parse + Refresh SO Review' action: re-derive EVERY stored order's
    components/attributes with the current parser (this is what applies the
    latest SO-parser changes to the whole backlog), then rebuild the sheet so you
    see the result. Run it right after a Git Update that changed the parser."""
    import line_items
    from process_lock import data_file_lock
    with data_file_lock(line_items.store_path(), label="line-items renormalization"):
        store_li = line_items.load_store()
        before = parser_review_metrics(store_li)
        n = line_items.renormalize_store(store_li)
        after = parser_review_metrics(store_li)
        line_items.save_store(store_li)
        metric = record_parser_metrics(
            before, after, n, len(store_li.get("jobs") or {})
        )
    print(f"Re-parsed {n} item(s) across {len(store_li.get('jobs') or {})} order(s) "
          f"with the current parser; rebuilding the review sheet...")
    reduction = metric["review_rows_reduced"]
    direction = f"{reduction:,} fewer" if reduction >= 0 else f"{-reduction:,} more"
    print(f"Marked-for-review rows: {before['review_rows']:,} -> "
          f"{after['review_rows']:,} ({direction}); recorded in {PARSER_METRICS_PATH}.")
    return _cmd_refresh(args)


def _cmd_list(args) -> int:
    store = load_store()
    rows = open_notes(store) if not args.all else store["notes"]
    if not rows:
        print("No open notes." if not args.all else "No notes recorded.")
        return 0
    for n in sorted(rows, key=lambda x: int(x.get("id", 0))):
        flag = "x" if n.get("status") == STATUS_HANDLED else " "
        print(f"[{flag}] #{n['id']}  order {n['order']}  item {n['item_no']}")
        print(f"      line: {n.get('item_text', '')}")
        print(f"      note: {n['note']}")
        if n.get("resolution"):
            print(f"      handled: {n['resolution']}")
    return 0


def _cmd_handle(args) -> int:
    # Always record the mark in the tracked ledger (the return channel to the
    # user's machine); also update a local queue if one is present here.
    resolution = " ".join(args.resolution)
    record_handled_mark(args.id, resolution)
    store = load_store()
    if mark_handled(store, args.id, resolution):
        save_store(store)
    print(f"Marked #{args.id} handled (recorded in {HANDLED_MARKS_PATH.name}; it "
          f"applies on the user's next 'Update SO Review').")
    return 0


def _run(func, args) -> int:
    """Run a subcommand, turning an expected error (e.g. the workbook is open in
    Excel) into a clear one-line message rather than a traceback in the log."""
    try:
        return func(args)
    except RuntimeError as e:
        print(str(e), file=sys.stderr)
        return 1


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="write the review workbook from the line-items store")
    b.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    b.set_defaults(func=_cmd_build)

    o = sub.add_parser("open", help="open the review workbook (building it if missing)")
    o.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    o.set_defaults(func=_cmd_open)

    s = sub.add_parser("sync", help="read your notes back out of the workbook into the queue")
    s.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    s.set_defaults(func=_cmd_sync)

    r = sub.add_parser("refresh", help="capture notes + apply handled-marks + rewrite the sheet")
    r.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    r.set_defaults(func=_cmd_refresh)

    rp = sub.add_parser("reparse", help="re-derive all stored orders with the current parser, then refresh")
    rp.add_argument("--out", default=str(DEFAULT_WORKBOOK))
    rp.set_defaults(func=_cmd_reparse)

    ls_ = sub.add_parser("list", help="show open notes (--all for handled too)")
    ls_.add_argument("--all", action="store_true")
    ls_.set_defaults(func=_cmd_list)

    h = sub.add_parser("handle", help="mark a note handled with a resolution")
    h.add_argument("id", type=int)
    h.add_argument("resolution", nargs="+")
    h.set_defaults(func=_cmd_handle)

    args = ap.parse_args(argv)
    return _run(args.func, args)


if __name__ == "__main__":
    sys.exit(main())
