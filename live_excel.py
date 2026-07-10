"""Render the live master workbook by driving the desktop Excel app through COM —
the same no-password, use-the-signed-in-app trick emailer.py uses for Outlook.

Why COM and not openpyxl: the daily report (excel_writer.py) is written with
openpyxl, which *replaces the whole file* — that can't touch a Microsoft 365
co-authored workbook without kicking everyone out / conflicting. Driving the real
Excel application means edits flow through Excel itself, which syncs them to
OneDrive/SharePoint so coworkers see them live (cursors and all).

What it writes: one worksheet per `live_sheets.Sheet` model (Live Queue, Changes,
History, Line Items, Similar Orders + its hidden Similar Data sheet). The
*content* lives in live_sheets.py (pure, tested); this
module is the generic renderer — bulk-write the values, then map each cell's
named fill/font to real Excel colors, add hyperlinks, freeze panes, and
AutoFilter. The named styles mirror excel_writer so the live master and the daily
report look the same.

Everything is lazy-imported and best-effort: a failed Excel update logs and the
poll cycle carries on (state + notifications still happen).
"""
from __future__ import annotations

import contextlib
import logging
import os
import threading
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl.utils import get_column_letter

from live_sheets import Sheet

log = logging.getLogger(__name__)

_XL_UNDERLINE_SINGLE = 2  # xlUnderlineStyleSingle
_XL_VALIDATE_LIST = 3     # xlValidateList (the picker cell's dropdown)
_XL_EXPRESSION = 2        # xlExpression (conditional formatting)
_XL_EQUAL = 3             # xlEqual — a real Operator value (ignored for xlExpression)
_XL_LIST_SEPARATOR = 5    # Application.International index for the list separator
_XL_CONTINUOUS = 1        # xlContinuous border line style
_XL_MEDIUM = -4138        # xlMedium border weight
_XL_RIGHT = -4152         # xlRight horizontal alignment
_XL_CENTER = -4108        # xlCenter horizontal alignment
_BORDER_EDGES = (7, 8, 9, 10)  # xlEdgeLeft, xlEdgeTop, xlEdgeBottom, xlEdgeRight


def _bgr(rgb_hex: str) -> int:
    """Excel COM colors are BGR longs; convert an 'RRGGBB' hex string."""
    r, g, b = int(rgb_hex[0:2], 16), int(rgb_hex[2:4], 16), int(rgb_hex[4:6], 16)
    return (b << 16) | (g << 8) | r


# Named fills -> Excel BGR (same RGB values as excel_writer's PatternFills).
_FILL_RGB = {
    "header": "305496",
    "overdue": "FFC7CE", "duetoday": "F8CBAD", "soon": "FFEB9C", "new": "D9D9D9",
    "overdue_new": "F4A5A8", "duetoday_new": "F4B183", "soon_new": "F5D750",
    "chg1": "D9D9D9", "chg2": "BFBFBF", "chg3": "A6A6A6", "chg4": "8C8C8C",  # darker per later instance
    "dwg_yes": "C6EFCE", "dwg_no": "FFC7CE",
    "sep": "808080",   # the vertical divider column between the two matrices
}
_FILL = {k: _bgr(v) for k, v in _FILL_RGB.items()}

# Live Queue search bar colors (mirror excel_writer's Full Queue search): a bright
# yellow row highlight + red box on the match, and a pale input cell.
_SEARCH_HIT_FILL = _bgr("FFFF00")
_SEARCH_BOX_FILL = _bgr("FFF2CC")
_SEARCH_RED = _bgr("C00000")
_NOTE_GRAY = _bgr("808080")

# Named fonts -> (rgb color or None, bold, underline, size or None).
_FONT = {
    "header": ("FFFFFF", True, False, None),
    "section": (None, True, False, 12),
    "link": ("0563C1", False, True, None),
    "drive_run": ("C55A11", True, False, None),
    "drive_run_link": ("C55A11", True, True, None),
    "red": ("C00000", True, False, None),
    "note": ("808080", False, False, None),   # muted gray (e.g. the 'last updated' stamp)
}

# "History" is a RESERVED worksheet name in Excel (shared-workbook change
# tracking), so the archived-orders tab is "Order History".
SHEET_ORDER = ["Live Queue", "Changes", "Order History", "Line Items"]

# Last-rendered fingerprint per sheet, so a tab is only repainted when its
# content actually changed — otherwise a coworker's active filter/scroll on that
# tab would be reset every poll. Reset on process start (re-renders once).
_RENDER_CACHE: Dict[str, int] = {}

# Sheets whose freeze pane is already set this process. Freeze panes persist in
# the workbook and setting one requires Activating the sheet — which would yank a
# coworker's view to that tab — so we only ever do it once per sheet.
_FROZEN: set = set()

# Sheets whose search-highlight conditional format has been applied successfully
# this process. Tracked so we keep retrying until it lands once (a busy Excel can
# reject the first attempt), then only re-apply when rows/the below-block change.
_SEARCH_CF_DONE: set = set()

# Per-sheet column count at last AutoFit. AutoFit scans the whole sheet, so on a
# full-repaint tab (Changes) we only re-fit on first render or when the column
# count changes — not every poll.
_AUTOFIT_NCOLS: Dict[str, int] = {}

# Upsert tabs (Live Queue, Order History) whose header row + initial autofit are
# already done this process, so we don't rewrite the header every cycle.
_HEADER_DONE: set = set()

# Last-rendered 'below' block per sheet, so it's only re-drawn (and repositioned)
# when the live data row count shifts or the block's content changes.
_BELOW_LAST: Dict[str, Any] = {}

_XL_UP = -4162           # xlUp
_XL_NONE = -4142         # xlColorIndexNone
_XL_AUTO = -4105         # xlColorIndexAutomatic

_XL_CALC_MANUAL = -4135     # xlCalculationManual
_XL_CALC_AUTOMATIC = -4105  # xlCalculationAutomatic


@contextlib.contextmanager
def _tuned(app):
    """Suspend screen repaints, event handlers, save/co-authoring prompts and
    AUTOMATIC recalculation for one render pass, restoring each afterward.

    Driving Excel cell-by-cell with these ON is what makes a render expensive and
    leaky over a long session: every single .Value/.Interior write repaints the
    window AND triggers Excel to recompute the WHOLE workbook — including the ~12K
    =HYPERLINK() formulas and the conditional-format rules on the Order History
    matrices. That's the bulk of the CPU, and the recalc/redraw caches it churns
    are a steady source of the memory growth. With recalc held to manual we do all
    the writes, then Calculate() once at the end.

    Each setting is application-level (so it also covers a coworker's other open
    files on this box) and is restored in the finally, so it's only suspended for
    the few seconds a render takes. Every get/set is best-effort: a busy or
    co-authoring Excel can reject one, and that must never abort the update."""
    saved: Dict[str, Any] = {}
    for attr, off in (("ScreenUpdating", False), ("EnableEvents", False),
                      ("DisplayAlerts", False), ("Calculation", _XL_CALC_MANUAL)):
        try:
            saved[attr] = getattr(app, attr)
            setattr(app, attr, off)
        except Exception:  # noqa: BLE001 - never let a rejected tuning abort the render
            pass
    try:
        yield
    finally:
        # Recompute once now that every write is in (manual calc skipped them), then
        # restore each setting to what it was before — including Calculation, so we
        # don't leave the user's Excel stuck on manual.
        try:
            app.Calculate()
        except Exception:  # noqa: BLE001
            pass
        # Restore to known-GOOD values, not the values seen on entry: if a prior
        # render was force-killed mid-tune, ScreenUpdating/Calculation were left
        # OFF (the classic "the whole Excel window went blank, but the data is
        # there") — restoring the stuck value would keep it broken. Resetting to
        # the normal live state self-heals it on the next successful render.
        for attr, good in (("ScreenUpdating", True), ("EnableEvents", True),
                           ("DisplayAlerts", True), ("Calculation", _XL_CALC_AUTOMATIC)):
            try:
                setattr(app, attr, good)
            except Exception:  # noqa: BLE001
                pass


def _refresh_volatile(wb, sheet: Sheet) -> None:
    """Write just the sheet's volatile cells (the 'Last updated' stamp) to their
    positions — no Cells.Clear/repaint — so the stamp stays current without
    disturbing a coworker's filter/scroll. Only runs on a render-cache hit, where
    the layout is unchanged, so each volatile cell is still where it was drawn."""
    vol = [(r, c, cell)
           for r, row in enumerate(sheet.grid, start=1)
           for c, cell in enumerate(row, start=1) if getattr(cell, "volatile", False)]
    if not vol:
        return
    try:
        ws = _get_or_make_sheet(wb, sheet.name)
        for r, c, cell in vol:
            ws.Cells(r, c).Value = cell.value
    except Exception as e:  # noqa: BLE001
        log.debug("volatile refresh failed (%s)", e)


def _fingerprint(sheet: Sheet) -> int:
    parts = [sheet.name, sheet.freeze or "", sheet.autofilter_a1 or "",
             str(sheet.hidden), str(sheet.picker)]
    for row in sheet.grid:
        for cell in row:
            # Volatile cells (e.g. the 'Last updated' stamp) change every cycle;
            # hashing their VALUE would defeat the render cache and force a full
            # repaint each poll. Keep their structure/style in the hash but ignore
            # the value, and refresh them in place instead (see _refresh_volatile).
            value = "\x00VOLATILE" if getattr(cell, "volatile", False) else cell.value
            parts.append(f"{value}|{cell.fill}|{cell.font}|{cell.link}|"
                         f"{cell.number_format}|{cell.center}")
        parts.append(";")
    return hash("\n".join(parts))


# --------------------------------------------------------------------------- #
# COM plumbing                                                                 #
# --------------------------------------------------------------------------- #
def _get_excel():
    """Attach to a running Excel, or start one. Lazy COM import (Windows-only)."""
    import win32com.client  # type: ignore
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
    except Exception:  # noqa: BLE001 - not running; launch it
        app = win32com.client.Dispatch("Excel.Application")
    # Best-effort: when Excel is busy (co-authoring sync, a dialog, you typing in
    # a cell) setting Visible is rejected. It's only a nicety, so never let it
    # abort the update — Excel is already visible if it was running.
    try:
        app.Visible = True
    except Exception:  # noqa: BLE001
        pass
    return app


def _find_workbook(app, path: Path):
    """The already-open workbook matching `path`, or open it. Co-authored files
    live in OneDrive/SharePoint, so FullName may be a URL — match on file name
    first, full path second.

    The open is strictly NON-INTERACTIVE. A locked file (a stale co-authoring
    lock, OneDrive mid-sync, an exclusive holder) would otherwise pop Excel's
    modal 'File in Use' dialog, which blocks this open AND every later COM call
    until a human dismisses it — with nobody at the desk that dead-ends the rest
    of the day's writes (see 2026-07-01, 16:42 onward). With DisplayAlerts off,
    Notify:=False makes a locked open FAIL instead; the caller's retry/skip path
    logs it and simply reopens on a later poll once the lock clears."""
    name = path.name
    target = os.path.normcase(str(path))
    for w in app.Workbooks:
        try:
            if w.Name == name or os.path.normcase(w.FullName) == target:
                return w
        except Exception:  # noqa: BLE001
            continue
    try:
        app.DisplayAlerts = False       # a locked file must fail, not show a modal
    except Exception:  # noqa: BLE001
        pass
    try:
        # UpdateLinks=0 and IgnoreReadOnlyRecommended silence the other
        # open-time prompts; Notify=False turns 'in use' into an error.
        wb = app.Workbooks.Open(str(path), UpdateLinks=0,
                                IgnoreReadOnlyRecommended=True, Notify=False)
    finally:
        try:
            app.DisplayAlerts = True
        except Exception:  # noqa: BLE001
            pass
    if getattr(wb, "ReadOnly", False):
        # A read-only handle would swallow our writes silently — and the caller
        # would then commit row signatures for rows that never reached the sheet.
        # Fail the cycle instead; a later poll retries for a read/write open.
        try:
            wb.Close(SaveChanges=False)
        except Exception:  # noqa: BLE001
            pass
        raise RuntimeError(f"{name} opened READ-ONLY (locked by another user/sync?); "
                           "skipping this cycle rather than writing to a dead copy")
    return wb


def _get_or_make_sheet(wb, name: str):
    for s in wb.Worksheets:
        if s.Name == name:
            return s
    ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
    try:
        ws.Name = name
    except Exception:  # noqa: BLE001 - invalid/reserved name; don't leave a stray
        try:
            ws.Application.DisplayAlerts = False
            ws.Delete()
            ws.Application.DisplayAlerts = True
        except Exception:  # noqa: BLE001
            pass
        raise
    return ws


# --------------------------------------------------------------------------- #
# Styling                                                                      #
# --------------------------------------------------------------------------- #
def _apply_font(rng, font_name: Optional[str]) -> None:
    spec = _FONT.get(font_name or "")
    if not spec:
        return
    color, bold, underline, size = spec
    f = rng.Font
    f.Bold = bool(bold)
    if color:
        f.Color = _bgr(color)
    f.Underline = _XL_UNDERLINE_SINGLE if underline else False
    if size:
        f.Size = size


def _apply_run(ws, r: int, c1: int, c2: int, fill: Optional[str],
               font: Optional[str], center: bool) -> None:
    rng = ws.Range(ws.Cells(r, c1), ws.Cells(r, c2))
    if fill and fill in _FILL:
        rng.Interior.Color = _FILL[fill]
    if font:
        _apply_font(rng, font)
    if center:
        rng.HorizontalAlignment = -4108  # xlCenter


def _style_row(ws, r: int, cells: List) -> None:
    """Style one model row: collapse adjacent cells that share (fill, font,
    center) into a single Range (few COM calls), then add hyperlinks and number
    formats per cell."""
    n = len(cells)
    c = 0
    while c < n:
        key = (cells[c].fill, cells[c].font, cells[c].center)
        if key == (None, None, False):
            c += 1
            continue
        c2 = c
        while c2 + 1 < n and (cells[c2 + 1].fill, cells[c2 + 1].font,
                              cells[c2 + 1].center) == key:
            c2 += 1
        _apply_run(ws, r, c + 1, c2 + 1, *key)
        c = c2 + 1
    for i, cell in enumerate(cells, start=1):
        if cell.link:
            try:
                link = str(cell.link)
                if link.startswith("#"):
                    # Internal link — jump to a sheet/cell in this workbook (e.g.
                    # Live Queue 'Similar' -> that order's group on Similar Data).
                    ws.Hyperlinks.Add(Anchor=ws.Cells(r, i), Address="",
                                      SubAddress=link[1:])
                else:
                    ws.Hyperlinks.Add(Anchor=ws.Cells(r, i), Address=link)
                _apply_font(ws.Cells(r, i), cell.font)  # keep our link style
            except Exception:  # noqa: BLE001 - a bad path shouldn't stop the row
                pass
        if cell.number_format:
            ws.Cells(r, i).NumberFormat = cell.number_format
        if cell.comment:
            try:
                cc = ws.Cells(r, i)
                if cc.Comment is not None:
                    cc.Comment.Delete()
                cc.AddComment(str(cell.comment))
            except Exception:  # noqa: BLE001 - a hover note is never worth failing on
                pass


def _pad(row: List, ncols: int) -> List[Any]:
    # Trailing cells are padded with None (a truly EMPTY cell), not "" — so a long
    # title/header in an earlier column can spill over them instead of being
    # clipped (Excel won't overflow text into a cell that holds an empty string).
    vals = [(cell.value if cell.value is not None else "") for cell in row]
    return vals + [None] * (ncols - len(vals))


def _overflow_col_widths(sheet: Sheet) -> Dict[int, float]:
    """For each column that holds an 'overflow' cell (free text meant to overrun,
    e.g. the Changes tab's 'What changed'), the width it should take from its
    NON-overflow content only — so AutoFit doesn't widen the column to fit the long
    text and it overruns the empty cells to its right instead. {col(1-based): width}.
    Measured from the model, so no extra COM round-trips."""
    overflow_cols = {i for row in sheet.grid
                     for i, c in enumerate(row, start=1) if c.overflow}
    widths: Dict[int, int] = {col: 0 for col in overflow_cols}
    for row in sheet.grid:
        for i, c in enumerate(row, start=1):
            if i in widths and not c.overflow and c.value not in (None, ""):
                widths[i] = max(widths[i], len(str(c.value)))
    return {col: min(max(w + 2, 10), 60) for col, w in widths.items()}


def render_sheet(app, wb, sheet: Sheet) -> None:
    """Write one Sheet model into its worksheet: clear, bulk-write values, style,
    freeze, AutoFilter, autofit."""
    ws = _get_or_make_sheet(wb, sheet.name)
    nrows, ncols = sheet.nrows, sheet.ncols
    kept_pick = None
    if sheet.picker:  # what the user picked/typed must survive the repaint
        with contextlib.suppress(Exception):
            kept_pick = ws.Range(sheet.picker["cell"]).Value
    try:
        if ws.AutoFilterMode:
            ws.AutoFilterMode = False
    except Exception:  # noqa: BLE001
        pass
    ws.Cells.Clear()  # bot-owned sheet — full repaint keeps it correct
    try:
        ws.Cells.UnMerge()   # drop any merges from a previous (taller) repaint
    except Exception:  # noqa: BLE001
        pass
    if nrows == 0 or ncols == 0:
        return

    ws.Range(ws.Cells(1, 1), ws.Cells(nrows, ncols)).Value = [
        _pad(row, ncols) for row in sheet.grid
    ]
    for r, row in enumerate(sheet.grid, start=1):
        if row:
            _style_row(ws, r, row)

    # Formula cells ("=...") are re-assigned via .Formula: the bulk .Value write
    # usually parses them, but .Formula is deterministic and always takes EN-US
    # argument separators regardless of the machine's locale.
    for r, row in enumerate(sheet.grid, start=1):
        for c, cell in enumerate(row, start=1):
            if isinstance(cell.value, str) and cell.value.startswith("="):
                with contextlib.suppress(Exception):
                    ws.Cells(r, c).Formula = cell.value

    if sheet.picker:
        _apply_picker(ws, sheet.picker, kept_pick)

    # Merge any cell that spans columns (e.g. the Changes 'Job #' header over its
    # blank spacer). The covered cells stay in the grid as positional spacers, so
    # this only removes the wall — it never shifts the columns to its right.
    for r, row in enumerate(sheet.grid, start=1):
        for c, cell in enumerate(row, start=1):
            if getattr(cell, "colspan", 1) > 1:
                try:
                    ws.Range(ws.Cells(r, c), ws.Cells(r, c + cell.colspan - 1)).Merge()
                except Exception:  # noqa: BLE001
                    pass

    if sheet.autofilter_a1:
        try:
            ws.Range(sheet.autofilter_a1).AutoFilter()
        except Exception:  # noqa: BLE001
            pass

    # AutoFit scans the whole sheet — the one O(sheet) op in a repaint. Column
    # widths only need recomputing when the column set changes, so skip it when
    # the count is unchanged from the last fit (column widths persist in Excel).
    if _AUTOFIT_NCOLS.get(sheet.name) != ncols:
        try:
            ws.UsedRange.WrapText = False        # let long titles spill, not wrap
            ws.UsedRange.Columns.AutoFit()
            for col in range(1, ncols + 1):
                if ws.Columns(col).ColumnWidth > 60:
                    ws.Columns(col).ColumnWidth = 60
            # Column 1 holds the section titles and the 'last updated' stamp; cap it
            # so those don't balloon the column — they overflow into the empty cells
            # to the right instead. Its real data (Job #, times) is short.
            if ws.Columns(1).ColumnWidth > 12:
                ws.Columns(1).ColumnWidth = 12
            # Re-narrow any column whose width was driven by an 'overflow' cell (free
            # text like 'What changed'): size it to its other content so the text
            # overruns the empty cells to its right rather than widening the column.
            for col, width in _overflow_col_widths(sheet).items():
                ws.Columns(col).ColumnWidth = width
            _AUTOFIT_NCOLS[sheet.name] = ncols
        except Exception:  # noqa: BLE001
            pass

    # Freeze panes persist in the workbook and setting one steals focus to this
    # tab, so do it only once per sheet (not every repaint).
    if sheet.freeze and sheet.name not in _FROZEN:
        try:
            ws.Activate()
            app.ActiveWindow.FreezePanes = False
            ws.Range(sheet.freeze).Select()
            app.ActiveWindow.FreezePanes = True
            _FROZEN.add(sheet.name)
        except Exception:  # noqa: BLE001
            pass

    # Honor the model's visibility both ways, so a sheet that was hidden in an
    # earlier build resurfaces when its model stops being hidden.
    with contextlib.suppress(Exception):
        ws.Visible = 0 if sheet.hidden else -1  # xlSheetHidden / xlSheetVisible


def _apply_picker(ws, picker: Dict[str, str], kept_pick: Any = None) -> None:
    """The Similar Orders-style input cell: restore what the user had picked
    before the repaint, style it like the search box, and (re)attach its
    dropdown list. Typing values not on the list stays allowed (any order #)."""
    try:
        box = ws.Range(picker["cell"])
        if kept_pick not in (None, ""):
            box.Value = kept_pick
        box.Interior.Color = _SEARCH_BOX_FILL
        box.Font.Bold = True
        box.HorizontalAlignment = _XL_CENTER
        _box_border(box, _SEARCH_RED)
        if picker.get("comment") and box.Comment is None:
            box.AddComment(picker["comment"])
        v = box.Validation
        with contextlib.suppress(Exception):
            v.Delete()
        v.Add(Type=_XL_VALIDATE_LIST, AlertStyle=1, Operator=1,
              Formula1=picker["source"])
        v.ShowError = False       # free typing allowed — not just the list
        v.InCellDropdown = True
    except Exception as e:  # noqa: BLE001 - the dropdown is a nicety
        log.debug("picker setup failed on %s (%s)", picker.get("cell"), e)


# --------------------------------------------------------------------------- #
# Incremental upsert renderer (Live Queue + Order History)                     #
#                                                                             #
# Rows are keyed on the order number: append new ones, update changed ones in  #
# place, delete departed ones (Live Queue). No Cells.Clear(), so a coworker's  #
# filter/sort/scroll survives — only an add/remove shifts an active filter.    #
# --------------------------------------------------------------------------- #
def _norm_key(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _write_header(ws, headers: List[str], row: int = 1) -> None:
    ncols = len(headers)
    rng = ws.Range(ws.Cells(row, 1), ws.Cells(row, ncols))
    rng.Value = [list(headers)]
    try:
        # Keep titles on one horizontal line (no wrap/rotation) so a header never
        # drives the column width or hides behind a tall wrapped cell.
        rng.WrapText = False
        rng.Orientation = 0
    except Exception:  # noqa: BLE001
        pass
    _apply_run(ws, row, 1, ncols, "header", "header", False)


def _box_border(target, color: int, weight: int = _XL_MEDIUM) -> None:
    """Draw a solid colored box around `target` (a Range or a FormatCondition):
    its four outer edges, in `color`."""
    for edge in _BORDER_EDGES:
        b = target.Borders(edge)
        b.LineStyle = _XL_CONTINUOUS
        b.Weight = weight
        b.Color = color


def _write_search_bar(ws, key_col: int) -> None:
    """Row 1 (above the header): a 'Search:' label and an input cell sitting right
    above the Job # column. Typing an order # (or just its last few digits) there
    lights up the matching row via the conditional format in _apply_search_cf — no
    macros, mirroring the daily report's Full Queue search."""
    lbl = ws.Cells(1, max(key_col - 1, 1))
    lbl.Value = "Search:"
    try:
        lbl.Font.Bold = True
        lbl.HorizontalAlignment = _XL_RIGHT
    except Exception:  # noqa: BLE001
        pass
    box = ws.Cells(1, key_col)             # the cell the conditional format watches;
    try:                                   # never overwrite its value (what you typed)
        box.Interior.Color = _SEARCH_BOX_FILL
        box.Font.Bold = True
        box.HorizontalAlignment = _XL_CENTER
        _box_border(box, _SEARCH_RED)
        if box.Comment is None:
            box.AddComment("Type an order # (or just its last few digits, e.g. the "
                           "last 3) here and press Enter. The matching row below "
                           "lights up yellow with a red box. Clear this cell to "
                           "remove the highlight.")
    except Exception:  # noqa: BLE001
        pass
    hint = ws.Cells(1, key_col + 1)
    hint.Value = "← type an order # or its last 3 digits to highlight its row (clear to remove)"
    try:
        hint.Font.Italic = True
        hint.Font.Color = _NOTE_GRAY
    except Exception:  # noqa: BLE001
        pass


def _cf_separators(ws) -> List[str]:
    """Argument separators to try in a conditional-format formula, likeliest
    first: Excel's reported LOCAL list separator (some locales need ';'), then a
    comma. A CF formula handed to Excel via .Add can require the local separator,
    and we don't know the locale up front — so callers try each until one sticks."""
    seps: List[str] = []
    try:
        local = ws.Application.International(_XL_LIST_SEPARATOR)
        if local:
            seps.append(str(local))
    except Exception:  # noqa: BLE001
        pass
    if "," not in seps:
        seps.append(",")
    return seps


def _apply_search_cf(ws, key_col: int, ncols: int, first_data_row: int,
                     last_row: int) -> bool:
    """Highlight (yellow fill + red box) the whole data row whose Job # ENDS WITH the
    search cell (row 1, the key column) — so typing the full order # OR just its last
    few digits (e.g. the last 3) lights up the row. Bounded to the data + a buffer for
    later appends — a whole-column CF is rejected at scale. Returns True if the rule was
    applied. INDEX(col, ROW()) reads each row's Job # with only absolute parts, so the
    rule can't be mis-translated against the active cell; RIGHT(job, LEN(typed)) takes
    the matching tail and &"" coerces both sides to text so a job stored as text still
    matches digits typed in the box.

    The formula's argument separator is locale-sensitive: a CF formula handed to
    Excel may need the LOCAL list separator (e.g. ';' in some regions) rather than
    ','. We don't know which up front, so try Excel's reported separator first and
    fall back to a comma — whichever Excel accepts wins."""
    key = get_column_letter(key_col)
    bottom = max(last_row + 3000, first_data_row)   # buffer for appends
    rng = ws.Range(ws.Cells(first_data_row, 1), ws.Cells(bottom, ncols))

    last_e = None
    for sep in _cf_separators(ws):
        # Match if the Job # ENDS WITH whatever's typed: RIGHT(job, LEN(typed))
        # == typed. Typing the whole order # still matches (a string ends with
        # itself); typing just the last few digits (e.g. the last 3) matches its
        # row too. Both sides coerced to text (&"") so a job stored as text still
        # matches digits typed in the box.
        formula = (f'=AND(${key}$1<>""{sep}'
                   f'RIGHT(INDEX(${key}:${key}{sep}ROW())&""{sep}'
                   f'LEN(${key}$1&""))=${key}$1&"")')
        try:
            rng.FormatConditions.Delete()
            # Pass Type, Operator, Formula1 BY POSITION with a REAL Operator value.
            # Operator is ignored for xlExpression, but it must be a concrete value:
            # omitting it (or pythoncom.Missing) makes late-bound Excel drop Formula1,
            # so Excel sees no formula and raises "parameter not optional". A real
            # Operator keeps Formula1 as positional arg #3 so it actually arrives.
            hit = rng.FormatConditions.Add(_XL_EXPRESSION, _XL_EQUAL, formula)
            hit.Interior.Color = _SEARCH_HIT_FILL
            hit.Font.Bold = True
            try:
                _box_border(hit, _SEARCH_RED)
            except Exception:  # noqa: BLE001 - keep the fill even if the box fails
                pass
            return True
        except Exception as e:  # noqa: BLE001 - try the next separator
            last_e = e

    # WARNING (not debug) + diagnostics: legacy sharing / sheet protection BLOCK
    # conditional formatting outright, and the probe (the simplest possible CF)
    # tells us whether ANY rule can be added — if even it fails, CF is blocked in
    # this Excel/workbook rather than our formula being wrong.
    shared = protected = probe = "?"
    try:
        shared = ws.Parent.MultiUserEditing
    except Exception:  # noqa: BLE001
        pass
    try:
        protected = ws.ProtectContents
    except Exception:  # noqa: BLE001
        pass
    try:
        c = ws.Cells(first_data_row, 1)
        c.FormatConditions.Delete()
        c.FormatConditions.Add(_XL_EXPRESSION, _XL_EQUAL, "=TRUE")
        c.FormatConditions.Delete()
        probe = "ok"
    except Exception as pe:  # noqa: BLE001
        probe = f"failed({pe})"
    log.warning("Search highlight conditional formatting failed (%s) "
                "[shared=%s protected=%s trivial-CF=%s]", last_e, shared, protected, probe)
    return False


def _read_keymap(ws, key_col: int, start_row: int = 2):
    """{order# -> row} and the last LIVE data row. The live data is a contiguous
    block from `start_row`; we bulk-read the key column and STOP at the first blank
    cell — the one-row gap above the 'Removed' block — so that block's real Job #s
    (drawn below the gap) are never read as live rows. Robust to a coworker
    re-sorting the live data (rows found by key, not a remembered index)."""
    bottom = ws.Cells(ws.Rows.Count, key_col).End(_XL_UP).Row   # last non-empty (incl. block)
    keymap: Dict[str, int] = {}
    last = start_row - 1
    if bottom >= start_row:
        data = ws.Range(ws.Cells(start_row, key_col), ws.Cells(bottom, key_col)).Value
        if not isinstance(data, tuple):       # single cell -> scalar
            data = ((data,),)
        for i, row in enumerate(data, start=start_row):
            v = row[0] if isinstance(row, (list, tuple)) else row
            k = _norm_key(v)
            if not k:
                break          # gap -> end of live data; ignore the block below it
            keymap[k] = i
            last = i
    return keymap, last


def _write_row(ws, r: int, cells: List, ncols: int) -> None:
    """Overwrite one row's values + styling in place. Clears the row's prior
    fill/font/links first so a cleared style doesn't linger."""
    vals = [(c.value if c.value is not None else "") for c in cells]
    if len(vals) < ncols:
        vals += [""] * (ncols - len(vals))
    rng = ws.Range(ws.Cells(r, 1), ws.Cells(r, ncols))
    rng.Value = [vals]
    try:
        rng.Interior.ColorIndex = _XL_NONE
        rng.Font.Bold = False
        rng.Font.Underline = False
        rng.Font.ColorIndex = _XL_AUTO
        rng.Hyperlinks.Delete()
        rng.ClearComments()
    except Exception:  # noqa: BLE001
        pass
    _style_row(ws, r, cells)


def _render_below(ws, live_last: int, ncols: int, below: Dict[str, Any]) -> None:
    """Draw the 'Removed since this morning' block right below the live data
    (whose last row is `live_last`, passed in from apply_upserts). Rows are full
    styled Cells — same columns and look as the Live Queue — written under a
    one-row blank gap, so the keymap scan stops at the gap and never reads the
    block's Job #s as live rows."""
    try:
        used = ws.UsedRange
        used_bottom = used.Row + used.Rows.Count - 1
    except Exception:  # noqa: BLE001
        used_bottom = live_last
    rows = below.get("rows") or []                # list of [Cell, ...]
    header_cells = below.get("header_cells") or []
    title_row = live_last + 2                      # row live_last+1 stays a blank gap
    header_row = title_row + 1
    bottom = max(used_bottom, header_row + len(rows))
    # Clear from the gap down, FULL WIDTH — wipes any stale/older block (and rows
    # an append may have overwritten the old block with) and lets the block's text
    # spill into the empty cells to its right.
    if bottom >= live_last + 1:
        try:
            ws.Range(ws.Cells(live_last + 1, 1), ws.Cells(bottom, ncols)).Clear()
        except Exception:  # noqa: BLE001
            pass
    title = ws.Cells(title_row, 1)
    title.Value = f"{below.get('title', 'Removed')} ({len(rows)})"
    try:
        title.Font.Bold = True
        title.Font.Size = 12
    except Exception:  # noqa: BLE001
        pass
    if header_cells:                               # header row, styled like the board's
        hvals = [c.value for c in header_cells]
        ws.Range(ws.Cells(header_row, 1), ws.Cells(header_row, len(hvals))).Value = [hvals]
        _style_row(ws, header_row, header_cells)
    if not rows:
        return
    # Pre-format the AM/PM text columns (Added = 1, Removed = last) as Text BEFORE
    # writing, so a time label isn't coerced into a serial (e.g. 0.40416667).
    top, bot = header_row + 1, header_row + len(rows)
    for tcol in (1, ncols):
        try:
            ws.Range(ws.Cells(top, tcol), ws.Cells(bot, tcol)).NumberFormat = "@"
        except Exception:  # noqa: BLE001
            pass
    # Data rows: bulk values, then full styling (fills, fonts, links, comments).
    for i, cells in enumerate(rows, start=top):
        vals = [(c.value if c.value is not None else "") for c in cells]
        ws.Range(ws.Cells(i, 1), ws.Cells(i, len(vals))).Value = [vals]
        _style_row(ws, i, cells)


def apply_upserts(app, wb, name: str, headers: List[str], ops: List,
                  key_col: int, allow_delete: bool, freeze: str | None = None,
                  sort_col: int | None = None, text_cols: List[int] | None = None,
                  below: Dict[str, Any] | None = None, header_row: int = 1,
                  search: bool = False) -> int:
    """Apply append/update/delete ops to a keyed sheet. Returns rows touched.
    When `sort_col` is set, the data is re-sorted ascending by that column after
    any change (Live Queue: the '#' board-position column, to match cbcinsider).
    `text_cols` are pre-formatted as Text before any data is written. `below`
    renders a small static section under the data (orders removed today).
    `header_row` is the row the column headers sit on (data starts just below it);
    `search` draws a search bar on row 1 above the header and a conditional format
    that highlights the row whose Job # (the key column) matches what's typed."""
    ws = _get_or_make_sheet(wb, name)
    ncols = len(headers)
    first_data_row = header_row + 1
    first_time = name not in _HEADER_DONE
    if first_time:
        # Once per process: wipe the sheet so a previous run's (possibly
        # different-schema) content can't collide with the keyed upsert. The
        # watcher resets the stored sigs at startup to match, so this cycle
        # rebuilds the tab; every later cycle is incremental.
        ws.Cells.Clear()
        _write_header(ws, headers, row=header_row)
        _HEADER_DONE.add(name)

    # Re-assert Text on the AM/PM columns EVERY render, before any value is
    # written — so a time label like "9:42 AM" can never be coerced into a serial
    # (which shows as e.g. 0.40416667). Doing it once at startup wasn't enough: a
    # single failed/lost format set leaves later writes to corrupt the cell.
    for col in (text_cols or []):
        try:
            ws.Columns(col).NumberFormat = "@"   # Text — must precede any write
        except Exception:  # noqa: BLE001
            pass

    keymap, last_row = _read_keymap(ws, key_col, start_row=first_data_row)
    deletes = [k for kind, k, _ in ops if kind == "delete"]
    updates = [(k, c) for kind, k, c in ops if kind == "update"]
    appends = [(k, c) for kind, k, c in ops if kind == "append"]

    for r in sorted((keymap[k] for k in deletes if k in keymap), reverse=True):
        try:
            ws.Rows(r).Delete()
        except Exception:  # noqa: BLE001
            pass
    if deletes:
        keymap, last_row = _read_keymap(ws, key_col, start_row=first_data_row)

    for k, cells in updates:
        r = keymap.get(k)
        if r:
            _write_row(ws, r, cells, ncols)
        else:
            appends.append((k, cells))   # vanished from the sheet — re-add it
    for k, cells in appends:
        r = keymap.get(k)
        if r:
            # Already on the sheet — a prior write of this key landed but its
            # signature wasn't committed (the render failed afterwards, so it was
            # re-planned as an append). Update in place instead of adding a
            # duplicate row; this keeps re-tried writes idempotent.
            _write_row(ws, r, cells, ncols)
            continue
        last_row += 1
        keymap[k] = last_row              # so a duplicate key within this batch can't re-add
        _write_row(ws, last_row, cells, ncols)

    # Keep the tab sorted (Live Queue by the '#' board-position column) and
    # re-extend AutoFilter. Sort the DATA ROWS ONLY (below the header) so the
    # header never moves; blanks fall to the bottom under ascending order.
    touched = bool(deletes or updates or appends)
    # A STRUCTURAL change (rows added/removed) is the only thing that requires the
    # search conditional format to be re-laid; an in-place value update doesn't
    # shift any row. Re-applying CF every poll (on plain updates) is what lets
    # Excel fragment one rule into thousands of sub-range rules over a day — a real
    # source of the unbounded memory growth — so gate the re-apply on structure.
    structural = bool(deletes or appends)
    if touched and last_row >= first_data_row:
        try:
            if ws.AutoFilterMode:
                ws.AutoFilterMode = False
            if sort_col and last_row >= first_data_row + 1:
                ws.Range(ws.Cells(first_data_row, 1), ws.Cells(last_row, ncols)).Sort(
                    Key1=ws.Cells(first_data_row, sort_col), Order1=1, Header=2)  # xlAscending, xlNo
            ws.Range(ws.Cells(header_row, 1), ws.Cells(last_row, ncols)).AutoFilter()
        except Exception:  # noqa: BLE001
            pass

    if first_time:
        try:
            ws.UsedRange.Columns.AutoFit()
            for col in range(1, ncols + 1):
                w = ws.Columns(col).ColumnWidth
                # Leave room for the AutoFilter dropdown arrow (~3 units) so it
                # never covers the header title, then cap very wide columns.
                hdr = str(headers[col - 1]) if col - 1 < len(headers) else ""
                ws.Columns(col).ColumnWidth = min(60, max(w, len(hdr) + 3))
        except Exception:  # noqa: BLE001
            pass
    # Draw the search bar EVERY cycle: it's three cheap cells and never overwrites
    # the input cell's value (so what you typed survives), which makes it self-heal
    # — if a render is ever interrupted mid-rebuild (e.g. Excel busy), a later cycle
    # still paints it instead of leaving row 1 blank until the next restart. The
    # autofit above (first render only) runs first, so the hint text spilling right
    # never widens a data column. The highlight CF is applied LAST (see below).
    if search:
        try:
            _write_search_bar(ws, key_col)
        except Exception as e:  # noqa: BLE001
            log.debug("search bar render failed (%s)", e)

    if freeze and name not in _FROZEN:
        try:
            ws.Activate()
            app.ActiveWindow.FreezePanes = False
            ws.Range(freeze).Select()
            app.ActiveWindow.FreezePanes = True
            _FROZEN.add(name)
        except Exception:  # noqa: BLE001
            pass

    # The 'Removed since this morning' block sits right after the live data, so it
    # must be re-drawn (repositioned) whenever the row count shifted — else new
    # appends overwrite it — as well as when its own content changed. Best-effort.
    below_ran = False
    if below is not None and (deletes or appends or below != _BELOW_LAST.get(name)):
        try:
            _render_below(ws, last_row, ncols, below)
            _BELOW_LAST[name] = below
            below_ran = True
        except Exception as e:  # noqa: BLE001
            log.debug("below-block render failed (%s)", e)

    # Apply the search highlight LAST — _render_below's Clear() on the rows beneath
    # the live data wipes any CF there, so applying it earlier (before the block)
    # left nothing highlighting. Re-apply only on a STRUCTURAL change (rows added/
    # removed), when the block redrew, or until it first lands (a busy Excel can
    # reject the first try); plain value updates leave every row where it was, so
    # they don't need the rule re-laid — re-applying it then is pure churn that
    # fragments the CF and grows memory.
    if search and (first_time or structural or below_ran or name not in _SEARCH_CF_DONE):
        if _apply_search_cf(ws, key_col, ncols, first_data_row, last_row):
            _SEARCH_CF_DONE.add(name)

    return len(updates) + len(appends) + len(deletes)


def reset_sheet(name: str) -> None:
    """Forget that a sheet's header/freeze are done, so the next render rebuilds
    it from scratch — used when its column schema changed (e.g. a new DWG suffix
    or feature tag appeared in the Order History matrices)."""
    _HEADER_DONE.discard(name)
    _FROZEN.discard(name)
    _SEARCH_CF_DONE.discard(name)


def _cell_to_value(cell) -> Any:
    """A bulk-writable value for an Order History cell. A hyperlink becomes a
    =HYPERLINK() formula so links survive the bulk write (Hyperlinks.Add per cell
    would crawl). The first build pays this once; after that we only append."""
    if cell.link:
        url = str(cell.link).replace('"', '""')
        disp = str(cell.value if cell.value is not None else "").replace('"', '""')
        return f'=HYPERLINK("{url}","{disp}")'
    return cell.value if cell.value is not None else ""


def _write_oh_row(ws, r: int, cells: List, ncols: int) -> None:
    vals = [_cell_to_value(c) for c in cells]
    if len(vals) < ncols:
        vals += [""] * (ncols - len(vals))
    ws.Range(ws.Cells(r, 1), ws.Cells(r, ncols)).Value = [vals]


def _apply_matrix_cf(ws, key_col: int, c0: int, c1: int, last_row: int) -> None:
    """Color a ✓/blank matrix block by conditional formatting — green for ✓, red
    for blank — but only on rows that actually have an order number (the key
    column guard), so the empty area below the data isn't painted. The range is
    bounded to the data (+ a buffer for future appends) rather than the whole
    million-row column, which Excel rejects at ~12K rows."""
    if c1 < c0:
        return
    key = get_column_letter(key_col)
    tl = get_column_letter(c0)            # top-left of the CF range, for the relative formula
    bottom = max(last_row + 3000, 3)      # buffer for appends; re-applied each process start
    rng = ws.Range(ws.Cells(2, c0), ws.Cells(bottom, c1))
    last_e = None
    for sep in _cf_separators(ws):
        green_f = f'=AND(${key}2<>""{sep}{tl}2="✓")'
        red_f = f'=AND(${key}2<>""{sep}{tl}2="")'
        try:
            rng.FormatConditions.Delete()
            # Pass Type, Operator, Formula1 BY POSITION with a REAL Operator value —
            # omitting Operator makes late-bound Excel drop Formula1 and reject the
            # rule (the same bug that hid the Live Queue search highlight).
            green = rng.FormatConditions.Add(_XL_EXPRESSION, _XL_EQUAL, green_f)
            green.Interior.Color = _FILL["dwg_yes"]
            red = rng.FormatConditions.Add(_XL_EXPRESSION, _XL_EQUAL, red_f)
            red.Interior.Color = _FILL["dwg_no"]
            return
        except Exception as e:  # noqa: BLE001 - try the next separator
            last_e = e
    log.warning("Matrix conditional formatting failed (%s)", last_e)


def _draw_separator(ws, sep_col: int) -> None:
    try:
        col = ws.Columns(sep_col)
        col.Interior.Color = _FILL["sep"]
        col.ColumnWidth = 2
    except Exception:  # noqa: BLE001
        pass


def apply_order_history(app, wb, name: str, spec: Dict[str, Any], ops: List,
                        key_col: int, freeze: str | None = None,
                        rebuild: bool = False) -> int:
    """Render the Order History log. It is built ONCE — the first time it's empty,
    or rarely when `rebuild` is set (a schema/migration change) — by writing the
    header, bulk-writing every row (chunked), coloring the matrices via
    conditional formatting, and drawing the divider. Every other run it only
    appends new orders / updates the few changed rows; the tab is never wiped."""
    ws = _get_or_make_sheet(wb, name)
    headers = spec["headers"]
    ncols = len(headers)
    keymap, last_row = _read_keymap(ws, key_col)
    populated = last_row >= 2 and bool(keymap)

    if rebuild or not populated:
        ws.Cells.Clear()
        _write_header(ws, headers)
        records = spec["records"]
        nrows = len(records)
        last = 1 + nrows
        if records:
            grid = [[_cell_to_value(c) for c in cells] + [""] * (ncols - len(cells))
                    for _, cells in records]
            # Chunk the bulk write — a ~12K x ~100 array in one COM call overruns
            # Excel (OLE error 0x800AC472). Batches of 1000 rows go through fine.
            for s in range(0, len(grid), 1000):
                block = grid[s:s + 1000]
                r0 = 2 + s
                try:
                    ws.Range(ws.Cells(r0, 1), ws.Cells(r0 + len(block) - 1, ncols)).Value = block
                except Exception as e:  # noqa: BLE001
                    log.warning("Order History write chunk at row %d failed (%s)", r0, e)
        _apply_matrix_cf(ws, key_col, *spec["dwg_range"], last)
        _apply_matrix_cf(ws, key_col, *spec["feat_range"], last)
        _draw_separator(ws, spec["sep_col"])
        if records:                         # AutoFilter so it's sortable/filterable
            try:
                ws.Range(ws.Cells(1, 1), ws.Cells(last, ncols)).AutoFilter()
            except Exception:  # noqa: BLE001
                pass
        try:                                # size columns from a small sample, not all 12K rows
            sample = min(last, 60)
            ws.Range(ws.Cells(1, 1), ws.Cells(sample, ncols)).Columns.AutoFit()
            for col in range(1, ncols + 1):
                if col != spec["sep_col"] and ws.Columns(col).ColumnWidth > 40:
                    ws.Columns(col).ColumnWidth = 40
        except Exception:  # noqa: BLE001
            pass
        if freeze and name not in _FROZEN:
            try:
                ws.Activate()
                app.ActiveWindow.FreezePanes = False
                ws.Range(freeze).Select()
                app.ActiveWindow.FreezePanes = True
                _FROZEN.add(name)
            except Exception:  # noqa: BLE001
                pass
        return nrows

    # Already built -> incremental only: append new orders / update the few whose
    # flags changed. The tab is never wiped here.
    updates = [(k, c) for kind, k, c in ops if kind == "update"]
    appends = [(k, c) for kind, k, c in ops if kind == "append"]
    for k, cells in updates:
        r = keymap.get(k)
        if r:
            _write_oh_row(ws, r, cells, ncols)
        else:
            appends.append((k, cells))
    for k, cells in appends:
        r = keymap.get(k)
        if r:                                  # already present (re-planned after a
            _write_oh_row(ws, r, cells, ncols)  # failed write) -> update, never duplicate
            continue
        last_row += 1
        keymap[k] = last_row
        _write_oh_row(ws, last_row, cells, ncols)
    if appends:   # re-extend AutoFilter to cover the newly appended rows
        try:
            if ws.AutoFilterMode:
                ws.AutoFilterMode = False
            ws.Range(ws.Cells(1, 1), ws.Cells(last_row, ncols)).AutoFilter()
        except Exception:  # noqa: BLE001
            pass
    return len(updates) + len(appends)


# --------------------------------------------------------------------------- #
# Watchdog: never let a busy/modal Excel hang the watcher                       #
# --------------------------------------------------------------------------- #
_EXCEL_WRITE_GUARD = threading.Lock()   # protects _excel_write_active
_excel_write_active = [0.0]             # monotonic start of the in-flight write (0 = none)


def _excel_write_timeout() -> float:
    """Seconds to wait for an Excel write before abandoning it. Generous, so a
    normal (even slow) render never trips it — only a genuine hang does."""
    try:
        return float(os.environ.get("LIVE_WRITE_TIMEOUT", "180"))
    except ValueError:
        return 180.0


def _run_excel_guarded(label: str, func, default, *args, **kwargs):
    """Run a COM/Excel call on a worker thread, bounded by a timeout, so a busy
    or modal Excel (a dialog up, a co-authoring sync in flight, someone typing in
    a cell) can NEVER hang the watcher's main loop.

    All COM work happens inside the worker (its own CoInitialize), so nothing is
    marshaled across threads. A new write is skipped only while a previous one is
    *still within its timeout*; once a write exceeds the timeout it is treated as
    abandoned and the next poll is allowed to try again — so a permanently stuck
    write (Excel left showing a dialog) can't block every future write forever.
    On timeout/skip/error it returns `default` — the same "nothing rendered,
    retry next poll" outcome the callers already expect from a failed write."""
    timeout = _excel_write_timeout()
    now = time.monotonic()
    with _EXCEL_WRITE_GUARD:
        active = _excel_write_active[0]
        if active and (now - active) < timeout:
            log.warning("%s skipped: a previous Excel write is still in progress (Excel is "
                        "likely busy / showing a dialog). Will retry next poll.", label)
            return default
        my_start = now
        _excel_write_active[0] = my_start   # claim the slot (also reclaims a stale one)

    box: dict = {}

    def _runner():
        try:
            try:
                import pythoncom  # type: ignore
                pythoncom.CoInitialize()
            except Exception:  # noqa: BLE001
                pass
            try:
                box["value"] = func(*args, **kwargs)
            except Exception as e:  # noqa: BLE001 - mirror existing best-effort handling
                box["error"] = e
            finally:
                try:
                    import pythoncom  # type: ignore
                    pythoncom.CoUninitialize()
                except Exception:  # noqa: BLE001
                    pass
        finally:
            with _EXCEL_WRITE_GUARD:
                if _excel_write_active[0] == my_start:   # only if a newer write hasn't taken over
                    _excel_write_active[0] = 0.0

    worker = threading.Thread(target=_runner, name="excel-write", daemon=True)
    worker.start()
    worker.join(timeout)
    if worker.is_alive():
        log.warning("%s exceeded %.0fs and was abandoned — Excel is busy or showing a "
                    "dialog (co-authoring conflict, 'file in use', a save prompt). The "
                    "watcher keeps running; the write retries next poll once Excel is free.",
                    label, timeout)
        return default
    if "error" in box:
        log.warning("%s failed (%s)", label, box["error"])
        return default
    return box.get("value", default)


def update_master_workbook(workbook_path: str | Path, lq_payload: Dict[str, Any],
                           oh_payload: Dict[str, Any],
                           changes_sheet: Sheet | None = None,
                           extra_sheets: List[Sheet] | None = None) -> set:
    """Watchdog wrapper: render the master workbook on a bounded worker thread so
    a busy/modal Excel can never hang the watcher. On timeout it returns an empty
    set, so those tabs simply re-render next poll (see _update_master_workbook_impl)."""
    return _run_excel_guarded(
        "Live workbook update", _update_master_workbook_impl, set(),
        workbook_path, lq_payload, oh_payload, changes_sheet, extra_sheets,
    )


def _update_master_workbook_impl(workbook_path: str | Path, lq_payload: Dict[str, Any],
                           oh_payload: Dict[str, Any],
                           changes_sheet: Sheet | None = None,
                           extra_sheets: List[Sheet] | None = None) -> set:
    """Render the master workbook: an incremental upsert for Live Queue, the
    matrix log for Order History, and a full repaint for the Changes snapshot
    (only when changed). Best-effort — any COM error is logged and swallowed.

    Returns the set of tab names that rendered WITHOUT error this cycle, so the
    caller can commit only those tabs' row signatures. If a tab's write fails
    (Excel busy / OLE error) its name is omitted, so its rows are re-planned and
    re-drawn next poll rather than being silently treated as already on the sheet."""
    path = Path(workbook_path)
    try:
        app = _get_excel()
    except Exception as e:  # noqa: BLE001
        log.warning("Could not reach Excel via COM (%s); live workbook not updated. "
                    "On Windows, ensure Excel is installed and signed in.", e)
        return set()
    # Reaching the workbook can be momentarily rejected while Excel is busy — a
    # dialog is up, a co-authoring sync is in flight, or someone is mid-keystroke
    # (surfaces as e.g. 'Excel.Application.Workbooks'). Retry a few times with a
    # short backoff, refreshing the app handle in case it disconnected; if it's
    # still busy we skip this cycle and try again on the next poll.
    wb = last_err = None
    for attempt in range(1, 4):
        try:
            wb = _find_workbook(app, path)
            break
        except Exception as e:  # noqa: BLE001
            last_err = e
            if attempt < 3:
                time.sleep(0.6 * attempt)
                try:
                    app = _get_excel()
                except Exception:  # noqa: BLE001
                    pass
    if wb is None:
        log.warning("Could not open the live workbook after 3 tries: %s: %r. Excel "
                    "may have a dialog open (co-authoring conflict, 'file in use', a "
                    "save prompt) or be mid-sync — clearing that usually fixes it. "
                    "Skipping this cycle; will retry next poll.",
                    type(last_err).__name__, last_err)
        return set()

    # Render each tab independently — a failure on one (e.g. the big Order
    # History) must not stop the others from updating. `ok` collects the tabs that
    # rendered cleanly so the caller commits only their signatures. The whole pass
    # runs with screen updates + automatic recalc suspended (see _tuned): cell-by-
    # cell COM writes with those on recompute/redraw the entire workbook every
    # write, which is the bulk of the CPU and a steady source of Excel's memory growth.
    touched = []
    ok: set = set()
    with _tuned(app):
        try:
            n = apply_upserts(app, wb, lq_payload["name"], lq_payload["headers"], lq_payload["ops"],
                              lq_payload["key_col"], lq_payload["allow_delete"], lq_payload.get("freeze"),
                              sort_col=lq_payload.get("sort_col"), text_cols=lq_payload.get("text_cols"),
                              below=lq_payload.get("below"), header_row=lq_payload.get("header_row", 1),
                              search=lq_payload.get("search", False))
            ok.add(lq_payload["name"])
            if n:
                touched.append(f"{lq_payload['name']}(+{n})")
        except Exception as e:  # noqa: BLE001
            log.warning("Live Queue update failed (%s)", e)
        try:
            n = apply_order_history(app, wb, oh_payload["name"], oh_payload["spec"],
                                    oh_payload["ops"], oh_payload["key_col"], oh_payload.get("freeze"),
                                    rebuild=oh_payload.get("rebuild", False))
            ok.add(oh_payload["name"])
            if n:
                touched.append(f"{oh_payload['name']}(+{n})")
        except Exception as e:  # noqa: BLE001
            log.warning("Order History update failed (%s)", e)
        # Full-repaint tabs (Changes + any extras, e.g. Similar Data/Orders):
        # skipped entirely when unchanged — a repaint would reset a viewer's
        # filter/scroll (and wipe the Similar Orders picker mid-use).
        repaints = ([changes_sheet] if changes_sheet is not None else [])
        repaints += list(extra_sheets or [])
        for model in repaints:
            try:
                fp = _fingerprint(model)
                if _RENDER_CACHE.get(model.name) != fp:
                    render_sheet(app, wb, model)
                    _RENDER_CACHE[model.name] = fp
                    touched.append(model.name)
                else:
                    # Content unchanged — just refresh any 'Last updated' stamp in
                    # place so the tab still reads as live. AutoSave carries it.
                    _refresh_volatile(wb, model)
            except Exception as e:  # noqa: BLE001
                log.warning("%s update failed (%s)", model.name, e)

    if touched:
        try:
            wb.Save()
        except Exception as e:  # noqa: BLE001
            log.debug("wb.Save() raised (likely AutoSave-managed): %s", e)
        log.info("Live workbook updated: %s [%s]", path.name, ", ".join(touched))
    else:
        log.info("Live workbook unchanged this cycle — nothing written.")
    return ok


def update_workbook(sheets: List[Sheet], workbook_path: str | Path) -> bool:
    """Render all sheet models into the live workbook via Excel COM. Returns True
    on success. Best-effort: any COM error is logged and swallowed so the poll
    cycle continues."""
    path = Path(workbook_path)
    try:
        app = _get_excel()
    except Exception as e:  # noqa: BLE001
        log.warning("Could not reach Excel via COM (%s); live workbook not updated. "
                    "On Windows, ensure Excel is installed and signed in.", e)
        return False
    try:
        wb = _find_workbook(app, path)
        order = {n: i for i, n in enumerate(SHEET_ORDER)}
        rendered = []
        with _tuned(app):   # suspend redraw + automatic recalc for the render pass
            for sheet in sorted(sheets, key=lambda s: order.get(s.name, 99)):
                fp = _fingerprint(sheet)
                if _RENDER_CACHE.get(sheet.name) == fp:
                    continue  # unchanged — leave it alone so filters/scroll persist
                render_sheet(app, wb, sheet)
                _RENDER_CACHE[sheet.name] = fp
                rendered.append(sheet.name)
        if rendered:
            try:
                wb.Save()
            except Exception as e:  # noqa: BLE001
                log.debug("wb.Save() raised (likely AutoSave-managed): %s", e)
            log.info("Live workbook updated: %s [%s]", path.name, ", ".join(rendered))
        else:
            log.info("Live workbook unchanged this cycle — nothing repainted.")
        return True
    except Exception as e:  # noqa: BLE001
        log.warning("Live workbook update failed (%s); state + alerts still recorded.", e)
        return False


def save_morning_copy(workbook_path: str | Path, dest: str | Path) -> bool:
    """Freeze a dated copy of the workbook as the morning snapshot, using Excel's
    SaveCopyAs (doesn't disturb the live file or its co-authors)."""
    src, dest = Path(workbook_path), Path(dest)
    try:
        app = _get_excel()
        wb = _find_workbook(app, src)
        dest.parent.mkdir(parents=True, exist_ok=True)
        wb.SaveCopyAs(str(dest))
        log.info("Saved morning snapshot copy: %s", dest)
        return True
    except Exception as e:  # noqa: BLE001
        log.warning("Could not save morning snapshot copy (%s)", e)
        return False


def recycle_workbook(workbook_path: str | Path) -> bool:
    """Close the live workbook so Excel reclaims the memory it accumulated over a
    long watch session, then let the next poll reopen it fresh.

    Excel never gives all of it back on its own: a day (often several days, since
    the watcher attaches to whatever Excel is already running and never quits it)
    of thousands of writes, sorts, AutoFilters and conditional-format passes leaves
    fragmented CF rules, a bloated calc chain and undo/redraw caches behind, and
    the process climbs into the multi-GB range. Closing the workbook frees those
    workbook-bound structures; reopening reads the same file back from
    OneDrive/SharePoint.

    Safe on the co-authored file: AutoSave + co-authoring have already synced every
    edit (we Save once more first to be sure), so nothing is lost, and only the
    BOT'S OWN Excel instance is touched — coworkers have their own open elsewhere
    and are unaffected. The process-level render caches (_HEADER_DONE/_FROZEN/…)
    are deliberately left intact: the reopened file still has the headers, frozen
    panes, rows and formats on disk, so the next poll resumes incremental upserts
    rather than doing a full rebuild. Best-effort — a failure just means we keep
    the current Excel and try again at the next recycle point."""
    path = Path(workbook_path)
    try:
        import win32com.client  # type: ignore
        # Only recycle an ALREADY-running Excel; never launch one just to close it.
        app = win32com.client.GetActiveObject("Excel.Application")
    except Exception as e:  # noqa: BLE001
        log.debug("No running Excel to recycle (%s); skipping.", e)
        return False
    wb = None
    target = os.path.normcase(str(path))
    for w in app.Workbooks:
        try:
            if w.Name == path.name or os.path.normcase(w.FullName) == target:
                wb = w
                break
        except Exception:  # noqa: BLE001
            continue
    if wb is None:
        log.debug("Live workbook not open in Excel; nothing to recycle.")
        return False
    try:
        try:
            wb.Save()                       # flush any unsynced edits before closing
        except Exception as e:  # noqa: BLE001
            log.debug("pre-recycle Save() raised (likely AutoSave-managed): %s", e)
        wb.Close(SaveChanges=True)
        log.info("Recycled the live workbook (closed to reclaim Excel memory; "
                 "the next poll reopens it).")
        return True
    except Exception as e:  # noqa: BLE001
        log.warning("Could not recycle the live workbook (%s); keeping the current "
                    "Excel and will retry at the next recycle point.", e)
        return False
