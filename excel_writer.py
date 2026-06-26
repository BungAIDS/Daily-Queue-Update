"""Build the daily Excel report with two tabs: Changes (first) and Full Queue."""
from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import engineers
from config import OUTPUT_DIR

log = logging.getLogger(__name__)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # due today -> red tomorrow
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
# One step darker than each base fill, used when a row is also new today.
RED_FILL_NEW    = PatternFill(start_color="F4A5A8", end_color="F4A5A8", fill_type="solid")
ORANGE_FILL_NEW = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
YELLOW_FILL_NEW = PatternFill(start_color="F5D750", end_color="F5D750", fill_type="solid")
NEW_FILL        = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=12)
RED_FONT = Font(color="C00000", bold=True)  # rows whose change order landed this run
LINK_FONT = Font(color="0563C1", underline="single")  # hyperlinks (Job # -> SO pdf, Folder)
DRIVE_RUN_FONT = Font(color="C55A11", bold=True)  # highly-custom (has a drive run)
DRIVE_RUN_LINK_FONT = Font(color="C55A11", bold=True, underline="single")  # ^ + links to the PDF
DWG_HAS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green: has the drawing
DWG_NO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # red: doesn't
CENTER_ALIGN = Alignment(horizontal="center")

# Full Queue search bar. Type an order # into the search cell (row 1) and a
# conditional-format rule lights up whichever Job # row matches — pure formula +
# formatting, so no macros and nothing to "enable"; it survives the daily
# openpyxl rewrite. SEARCH_HIT_* style the matched row (bright yellow so it pops
# over every urgency color, bold so it's legible whatever the row was already);
# SEARCH_BOX_* make the input cell look like a box you can type into.
SEARCH_HIT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
SEARCH_HIT_FONT = Font(bold=True, color="000000")
_SEARCH_HIT_SIDE = Side(style="medium", color="C00000")  # red box around the matched row
SEARCH_HIT_BORDER = Border(left=_SEARCH_HIT_SIDE, right=_SEARCH_HIT_SIDE,
                           top=_SEARCH_HIT_SIDE, bottom=_SEARCH_HIT_SIDE)
SEARCH_BOX_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_SEARCH_BOX_SIDE = Side(style="thin", color="BF8F00")
SEARCH_BOX_BORDER = Border(left=_SEARCH_BOX_SIDE, right=_SEARCH_BOX_SIDE,
                           top=_SEARCH_BOX_SIDE, bottom=_SEARCH_BOX_SIDE)
# Full Queue row layout: row 1 is the search bar, the column headers sit on
# row 2, and job rows start on row 3. SEARCH_CELL is what you type into.
SEARCH_ROW, HEADER_ROW, FIRST_DATA_ROW = 1, 2, 3
SEARCH_CELL = "B1"
SEARCH_CELL_ABS = "$B$1"  # absolute form for the conditional-format formula

# Single source of truth for column order. Each entry is (header, key); the key
# names the job field to print, or a special renderer handled in _write_job_row:
#   "job"         -> job number, hyperlinked to its Sales Order pdf (Z: drive)
#   "folder"      -> AutoCAD job folder (or SO archive folder) hyperlink
#   "co"          -> CO# label
#   "drive_run"   -> YES, hyperlinked to the quote-run file (highly-custom fans)
#   "total_price" -> money-formatted cell
#   "flags"       -> flag summary string
# To reorder the report, reorder this list — everything else follows.
COLUMNS = [
    ("Job #", "job"),
    ("Folder", "folder"),
    ("Quote Run", "drive_run"),
    ("CO#", "co"),
    ("Oper", "oper"),
    ("Design", "design"),
    ("Customer", "customer"),
    ("Size", "so_size"),
    ("Arrangement", "so_arrangement"),
    ("Assigned To", "assigned_to"),
    ("Checker", "checker"),
    ("Note", "status_note"),
    ("Engineer", "engineers"),
    ("End Date", "end_date"),
    # --- everything else (original relative order) ---
    ("Description", "so_design_desc"),
    ("Motor Pos", "so_motor_pos"),
    ("Class", "so_class"),
    ("Rotation", "so_rotation"),
    ("Discharge", "so_discharge"),
    ("% Width", "so_pct_width"),
    ("Wheel Type", "so_wheel_type"),
    ("Design Temp", "so_design_temp"),
    ("Max Temp", "so_max_temp"),
    ("Special Temp", "so_special_temp"),
    ("Features", "line_item_tags"),
    ("Primary Rep", "primary_rep"),
    ("Start Date", "start_date"),
    ("FanNet Date", "fannet_date"),
    ("Item", "item"),
    ("Plan Hrs", "plan_hrs"),
    ("Total Price", "total_price"),
    ("Ship With", "ship_with"),
    ("Flags", "flags"),
    ("Status", "status"),
]
QUEUE_HEADERS = [h for h, _ in COLUMNS]
_COL_IDX = {key: i for i, (_, key) in enumerate(COLUMNS, start=1)}
TOTAL_PRICE_COL = _COL_IDX["total_price"]  # 1-based; used by the footer total


def _flags_str(j: Dict[str, Any]) -> str:
    flags = []
    if j.get("unapproved"):
        flags.append("UNAPPROVED")
    if j.get("credit_hold"):
        flags.append("CREDIT HOLD")
    if j.get("has_notes"):
        flags.append("NOTES")
    return ", ".join(flags)


def _parse_date(s: str) -> date | None:
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime((s or "").strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _parse_money(s: str) -> float:
    try:
        return float((s or "").replace("$", "").replace(",", "").strip() or 0)
    except ValueError:
        return 0.0


MONEY_FMT = '"$"#,##0.00'


def _write_money_cell(ws, row: int, col: int, raw: str):
    """Write a price as a real number (so Excel can sum/sort it), formatted as
    currency. Blank stays blank rather than showing $0.00."""
    raw = (raw or "").strip()
    if not raw:
        return ws.cell(row=row, column=col, value="")
    cell = ws.cell(row=row, column=col, value=_parse_money(raw))
    cell.number_format = MONEY_FMT
    return cell


def _autosize(ws, num_cols: int, skip_cells: set | None = None) -> None:
    """Size each column to its widest cell. Cells listed in skip_cells (a set of
    (row, col) tuples) are ignored — used for values we deliberately let overflow
    into an empty neighbor instead of widening their whole column."""
    skip_cells = skip_cells or set()
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if (cell.row, cell.column) in skip_cells:
                continue
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _write_changes_tab(ws, briefing: Dict[str, Any], diff: Dict[str, Any],
                       co_changed_ids: set | None = None) -> None:
    co_changed_ids = co_changed_ids or set()
    overflow: set = set()  # (row, col) cells excluded from autosize so they overrun
    row = 1

    # AI Briefing block
    ws.cell(row=row, column=1, value="AI Briefing").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=briefing.get("briefing", "(no briefing)"))
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 80
    row += 2

    # Anomalies
    anomalies = briefing.get("anomalies", []) or []
    if anomalies:
        ws.cell(row=row, column=1, value="Anomalies").font = SECTION_FONT
        row += 1
        for a in anomalies:
            ws.cell(row=row, column=1, value=f"- {a}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1
        row += 1

    # Action items
    items = briefing.get("action_items", []) or []
    if items:
        ws.cell(row=row, column=1, value="Top Action Items").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Rank").font = HEADER_FONT
        ws.cell(row=row, column=2, value="Job #").font = HEADER_FONT
        ws.cell(row=row, column=3, value="Reason").font = HEADER_FONT
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = HEADER_FILL
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=item.get("rank", ""))
            ws.cell(row=row, column=2, value=str(item.get("job", "")))
            ws.cell(row=row, column=3, value=item.get("reason", ""))
            row += 1
        row += 1

    # New orders
    ws.cell(row=row, column=1, value=f"New orders ({len(diff['new'])})").font = SECTION_FONT
    row += 1
    if diff["new"]:
        for c, h in enumerate(QUEUE_HEADERS, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for j in diff["new"]:
            _write_job_row(ws, row, j, co_changed=j.get("job") in co_changed_ids)
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1
    row += 1

    # Change orders that landed this run (CO# rose since yesterday, including
    # jobs that came back at a higher CO#). Placed right under New orders.
    co_changed = list(diff.get("co_changed", []))
    for j in diff.get("returning", []):
        cr = j.get("_co_returned")
        if cr:
            co_changed.append({"job": j.get("job"), "customer": j.get("customer", ""),
                               "old_co": cr["old_co"], "new_co": cr["new_co"],
                               "co_history": j.get("co_history", []), "_returned": True})
    ws.cell(row=row, column=1, value=f"Change orders this run ({len(co_changed)})").font = SECTION_FONT
    row += 1
    if co_changed:
        for c, h in enumerate(["Job #", "Customer", "Change", "Latest change-order note"], start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for c in co_changed:
            arrow = f"CO#{c['old_co']} -> CO#{c['new_co']}"
            if c.get("_returned"):
                arrow += " (returned)"
            note = c["co_history"][0] if c.get("co_history") else ""
            ws.cell(row=row, column=1, value=c["job"]).font = RED_FONT
            ws.cell(row=row, column=2, value=c["customer"]).font = RED_FONT
            ws.cell(row=row, column=3, value=arrow).font = RED_FONT
            ws.cell(row=row, column=4, value=note).font = RED_FONT
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1
    row += 1

    # Returning orders (came back from history)
    returning = diff.get("returning", [])
    ws.cell(row=row, column=1, value=f"Returning orders — back from history ({len(returning)})").font = SECTION_FONT
    row += 1
    if returning:
        headers = QUEUE_HEADERS + ["Last Seen"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for j in returning:
            _write_job_row(ws, row, j, co_changed=j.get("job") in co_changed_ids)
            ws.cell(row=row, column=len(QUEUE_HEADERS) + 1, value=j.get("_last_seen", ""))
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1
    row += 1

    # Completed/Removed
    ws.cell(row=row, column=1, value=f"Completed / Removed ({len(diff['removed'])})").font = SECTION_FONT
    row += 1
    if diff["removed"]:
        for c, h in enumerate(QUEUE_HEADERS, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for j in diff["removed"]:
            _write_job_row(ws, row, j)
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1
    row += 1

    # Orders that have changed. Customer spans two columns: it's written in col 2
    # with col 3 left blank so a long name overflows into it, and the customer
    # cells are excluded from autosize so they don't force col 2 (shared with the
    # narrow Folder column in the sections above) wide. Field/Old/New shift right.
    ws.cell(row=row, column=1, value=f"Orders that have changed ({len(diff['changed'])})").font = SECTION_FONT
    row += 1
    if diff["changed"]:
        for c, h in enumerate(["Job #", "Customer", "", "Field", "Old value", "New value"], start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for ch in diff["changed"]:
            for (field, old, new) in ch["changes"]:
                ws.cell(row=row, column=1, value=ch["job"])
                ws.cell(row=row, column=2, value=ch["customer"])
                overflow.add((row, 2))  # let the customer name overrun the blank col 3
                ws.cell(row=row, column=4, value=field)
                ws.cell(row=row, column=5, value=old)
                ws.cell(row=row, column=6, value=new)
                row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1
    row += 1

    # Persistent (3+ days)
    ws.cell(row=row, column=1, value=f"Persistent orders — 3+ days in queue ({len(diff['persistent'])})").font = SECTION_FONT
    row += 1
    if diff["persistent"]:
        for c, h in enumerate(["Job #", "Customer", "Days in queue", "End Date", "Assigned To", "Total Price"], start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for p in diff["persistent"]:
            snap = p["snapshot"]
            ws.cell(row=row, column=1, value=p["job"])
            ws.cell(row=row, column=2, value=p["customer"])
            ws.cell(row=row, column=3, value=p["days"])
            ws.cell(row=row, column=4, value=snap.get("end_date", ""))
            ws.cell(row=row, column=5, value=snap.get("assigned_to", ""))
            _write_money_cell(ws, row, 6, snap.get("total_price", ""))
            row += 1
    else:
        ws.cell(row=row, column=1, value="(none)")
        row += 1

    _autosize(ws, num_cols=len(QUEUE_HEADERS), skip_cells=overflow)


def _co_label(j: Dict[str, Any]) -> str:
    co = j.get("co_number") or 0
    return f"CO#{co}" if co else ""


_ARRANGEMENT_RE = re.compile(r"^(A/[A-Za-z0-9]+)(?:\s+(.*))?$")
# Verbose spellings the Sales Order sometimes uses instead of the short code:
# 'Arrangement 4', 'Arr. 9', 'arr 10 belt drive'. Normalize the leading
# 'Arrangement N' to 'A/N'; any trailing text becomes the note.
_ARRANGEMENT_VERBOSE_RE = re.compile(r"^arr(?:angement)?\.?\s*(\d+[A-Za-z]*)(?:\s+(.*))?$", re.I)


def split_arrangement(value: str) -> "tuple[str, str]":
    """Split a raw arrangement into the short 'A/X' code (X = digits/letters) and
    any trailing descriptive text, so the column stays tidy and the detail moves
    to a hover note. 'A/4V C-Face Flange mount (no motor base)' -> ('A/4V',
    'C-Face Flange mount (no motor base)'). The verbose spelling is normalized too:
    'Arrangement 4' -> ('A/4', ''), 'Arr. 9 belt drive' -> ('A/9', 'belt drive').
    Anything that isn't an arrangement (e.g. 'N/A', '') passes through unchanged."""
    s = (value or "").strip()
    m = _ARRANGEMENT_RE.match(s)
    if m:
        return m.group(1), (m.group(2) or "").strip()
    m = _ARRANGEMENT_VERBOSE_RE.match(s)
    if m:
        return "A/" + m.group(1).upper(), (m.group(2) or "").strip()
    return s, ""


# The main size is the leading NUMBER (optionally a fraction, so "13 1/2" stays
# whole). A "-A6"/"-B12"-style code suffix on the number, and any trailing text
# (e.g. "Blade-1800", "(3600 RPM or less)"), are descriptive and move to the note.
# Sizes that aren't number-led (e.g. "H3", "N/A") pass through unchanged.
_SIZE_RE = re.compile(r"^(\d+(?:\s+\d+/\d+)?)(-\S*)?(?:\s+(.*))?$")


def split_size(value: str) -> "tuple[str, str]":
    """Split a raw size into the leading number and the descriptive rest, so the
    column stays tidy and the detail moves to a hover note. '3000-A6 Blade-1800' ->
    ('3000', '-A6 Blade-1800'); '3300-B12' -> ('3300', '-B12'); '2412 (3600 RPM or
    less)' -> ('2412', '(3600 RPM or less)'); a fraction like '13 1/2' stays whole;
    non-number sizes ('H3', '182', '') pass through with no note."""
    s = (value or "").strip()
    m = _SIZE_RE.match(s)
    if not m:
        return s, ""
    note = " ".join(p for p in (m.group(2), m.group(3)) if p).strip()
    return m.group(1), note


def folder_of(path: str) -> str:
    """The containing folder of a Windows/posix file path. Used to link the Job #
    at the Sales Order *folder* rather than a specific PDF: a change order renames
    the SO file (``… (original).pdf`` → ``… CO#1.pdf``), so a link captured before
    the CO would dead-link, but the per-job folder name never changes."""
    p = (path or "").strip()
    if not p:
        return ""
    return p.rsplit("\\", 1)[0] if "\\" in p else os.path.dirname(p)


def _drive_run_label(j: Dict[str, Any]) -> str:
    """"YES" when the job has a quote run; "YES (X)" when X > 1 files matched,
    so someone knows to review which one is the real run."""
    if not j.get("has_drive_run"):
        return ""
    n = j.get("drive_run_count") or 0
    return f"YES ({n})" if n > 1 else "YES"


def _write_job_row(ws, row: int, j: Dict[str, Any], co_changed: bool = False) -> None:
    linked_cols = set()  # hyperlink cells keep their link style, not red
    for col, (_header, key) in enumerate(COLUMNS, start=1):
        if key == "job":
            cell = ws.cell(row=row, column=col, value=j.get("job", ""))
            # Job # links to its latest Sales Order pdf on the Z: drive (when we
            # have one). The daily run downloads the current revision each
            # morning, so this path is always the newest CO's SO.
            so_pdf = (j.get("so_pdf") or "").strip()
            if so_pdf and j.get("job"):
                cell.hyperlink = so_pdf
                cell.font = LINK_FONT
                linked_cols.add(col)
        elif key == "folder":
            # AutoCAD job folder, or the SO archive folder as fallback.
            folder = (j.get("job_folder") or "").strip()
            cell = ws.cell(row=row, column=col, value=(j.get("job_type") or "Open") if folder else "")
            if folder:
                cell.hyperlink = folder
                cell.font = LINK_FONT
                linked_cols.add(col)
        elif key == "co":
            ws.cell(row=row, column=col, value=_co_label(j))
        elif key == "drive_run":
            # YES for a highly-custom fan, linked to the quote-run file. More
            # than one match -> "YES (X)" so someone reviews which is the run.
            dr_pdf = (j.get("drive_run_pdf") or "").strip()
            cell = ws.cell(row=row, column=col, value=_drive_run_label(j))
            if dr_pdf:
                cell.hyperlink = dr_pdf
                cell.font = DRIVE_RUN_LINK_FONT
                linked_cols.add(col)
            elif j.get("has_drive_run"):
                cell.font = DRIVE_RUN_FONT
        elif key == "total_price":
            _write_money_cell(ws, row, col, j.get("total_price", ""))
        elif key == "so_arrangement":
            # Keep the column to the short 'A/X' code; any descriptive suffix
            # moves to a hover note.
            code, note = split_arrangement(j.get("so_arrangement", ""))
            cell = ws.cell(row=row, column=col, value=code)
            if note:
                cell.comment = Comment(note, "Queue")
        elif key == "so_size":
            # Keep the column to the main size; any trailing detail moves to a note.
            main, note = split_size(j.get("so_size", ""))
            cell = ws.cell(row=row, column=col, value=main)
            if note:
                cell.comment = Comment(note, "Queue")
        elif key == "flags":
            ws.cell(row=row, column=col, value=_flags_str(j))
        elif key == "engineers":
            ws.cell(row=row, column=col, value=engineers.cell_text(j))
        else:
            ws.cell(row=row, column=col, value=j.get(key, ""))

    # A change order that landed this run -> the whole row's text goes red,
    # except the hyperlink cells, which keep their link style.
    if co_changed:
        for col in range(1, len(COLUMNS) + 1):
            if col not in linked_cols:
                ws.cell(row=row, column=col).font = RED_FONT


def _dwg_suffixes(jobs: List[Dict[str, Any]]) -> List[str]:
    """Sorted union of every custom-DWG suffix found across the jobs (the -NN
    extras beyond the standard -01/-02), numeric-ordered."""
    seen: set = set()
    for j in jobs:
        seen.update((j.get("dwg_extras") or {}).keys())
    return sorted(seen, key=lambda s: (int(s), s) if s.isdigit() else (10**9, s))


def _append_dwg_matrix(ws, rows: List, start_col: int, header_row: int = 1) -> List[str]:
    """Append the custom-DWG matrix starting at column `start_col`: one header
    per distinct suffix (on `header_row`), then per-row green-✓ (has the drawing)
    / red (doesn't). `rows` is a list of (row_index, job_dict). Returns the
    suffixes so the caller can size the AutoFilter/columns."""
    suffixes = _dwg_suffixes([j for _, j in rows])
    for k, s in enumerate(suffixes, start=start_col):
        cell = ws.cell(row=header_row, column=k, value=f"-{s}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row_i, j in rows:
        extras = j.get("dwg_extras") or {}
        for k, s in enumerate(suffixes, start=start_col):
            cell = ws.cell(row=row_i, column=k)
            if s in extras:
                cell.value, cell.fill, cell.alignment = "✓", DWG_HAS_FILL, CENTER_ALIGN
            else:
                cell.fill = DWG_NO_FILL
    return suffixes


def _write_search_bar(ws) -> None:
    """Row 1: a labelled cell you type an order # into. The conditional-format
    rule that actually highlights the match is added by the caller once the data
    extent is known (see _write_full_queue_tab)."""
    label = ws.cell(row=SEARCH_ROW, column=1, value="Search:")
    label.font = Font(bold=True)
    label.alignment = Alignment(horizontal="right")

    box = ws.cell(row=SEARCH_ROW, column=2)  # B1 — SEARCH_CELL, the input cell
    box.fill = SEARCH_BOX_FILL
    box.border = SEARCH_BOX_BORDER
    box.alignment = CENTER_ALIGN
    box.font = Font(bold=True)
    box.comment = Comment(
        "Type an order # here and press Enter. The matching row below lights up "
        "bright yellow. Clear this cell to remove the highlight.", "Queue")

    hint = ws.cell(row=SEARCH_ROW, column=3,
                   value="← type an order # to highlight its row (clear to remove)")
    hint.font = Font(italic=True, color="808080")


def _write_full_queue_tab(
    ws,
    jobs: List[Dict[str, Any]],
    today: date,
    new_job_ids: set | None = None,
    co_changed_ids: set | None = None,
) -> None:
    new_job_ids = new_job_ids or set()
    co_changed_ids = co_changed_ids or set()
    _write_search_bar(ws)
    for c, h in enumerate(QUEUE_HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    total_price_sum = 0.0
    soon_threshold = today + timedelta(days=3)

    for i, j in enumerate(jobs, start=FIRST_DATA_ROW):
        _write_job_row(ws, i, j, co_changed=j.get("job") in co_changed_ids)
        # Pick a row fill based on End Date urgency; if the order is also new
        # today, step the chosen color one shade darker (or to light gray if
        # there's no urgency fill yet).
        is_new = j.get("job") in new_job_ids
        end = _parse_date(j.get("end_date", ""))
        if end is not None and end < today:
            fill = RED_FILL_NEW if is_new else RED_FILL
        elif end is not None and end == today:           # due today -> red tomorrow
            fill = ORANGE_FILL_NEW if is_new else ORANGE_FILL
        elif end is not None and end <= soon_threshold:
            fill = YELLOW_FILL_NEW if is_new else YELLOW_FILL
        elif is_new:
            fill = NEW_FILL
        else:
            fill = None
        if fill is not None:
            for c in range(1, len(QUEUE_HEADERS) + 1):  # urgency fill: standard cols only
                ws.cell(row=i, column=c).fill = fill
        total_price_sum += _parse_money(j.get("total_price", ""))

    # Summary footer (one blank row below the last job row)
    last_data = len(jobs) + FIRST_DATA_ROW - 1
    footer = last_data + 2
    ws.cell(row=footer, column=1, value=f"Total jobs: {len(jobs)}").font = SECTION_FONT
    ws.cell(row=footer, column=TOTAL_PRICE_COL - 1, value="Total $ in process:").font = SECTION_FONT
    total_cell = ws.cell(row=footer, column=TOTAL_PRICE_COL, value=total_price_sum)
    total_cell.number_format = MONEY_FMT
    total_cell.font = SECTION_FONT

    # Custom-DWG green-✓/red matrix, appended after the standard columns.
    suffixes = _append_dwg_matrix(ws, list(enumerate(jobs, start=FIRST_DATA_ROW)),
                                  len(QUEUE_HEADERS) + 1, header_row=HEADER_ROW)
    total_cols = len(QUEUE_HEADERS) + len(suffixes)

    if jobs:
        last_col = get_column_letter(total_cols)
        # AutoFilter across the data rows (including the DWG columns)
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_data}"
        # Search highlight: light up the whole row whose Job # (col A) matches
        # what's typed into the search cell. TEXT(...,"@") on both sides so a
        # job stored as text still matches a number you type (and vice versa).
        ws.conditional_formatting.add(
            f"A{FIRST_DATA_ROW}:{last_col}{last_data}",
            FormulaRule(
                formula=[f'AND({SEARCH_CELL_ABS}<>"",'
                         f'TEXT($A{FIRST_DATA_ROW},"@")=TEXT({SEARCH_CELL_ABS},"@"))'],
                fill=SEARCH_HIT_FILL, font=SEARCH_HIT_FONT,
                border=SEARCH_HIT_BORDER, stopIfTrue=True))

    # Freeze below the search bar + header so both stay visible while scrolling.
    ws.freeze_panes = f"B{FIRST_DATA_ROW}"
    # Skip the search-bar cells when sizing: the hint deliberately overflows into
    # the empty cells to its right, so it shouldn't widen its whole column.
    _autosize(ws, num_cols=total_cols,
              skip_cells={(SEARCH_ROW, 1), (SEARCH_ROW, 2), (SEARCH_ROW, 3)})


def _write_history_tab(ws, history: Dict[str, Any]) -> None:
    """Archived jobs (left the queue, not yet returned), newest departure first."""
    headers = QUEUE_HEADERS + ["Last Seen"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    if not history:
        ws.cell(row=2, column=1,
                value="(no archived orders yet — a job appears here after it drops off the queue)")
        ws.freeze_panes = "B2"  # keep the header row AND the Job # column visible
        _autosize(ws, num_cols=len(headers))
        return

    entries = sorted(history.values(), key=lambda e: e.get("last_seen", ""), reverse=True)
    for i, entry in enumerate(entries, start=2):
        _write_job_row(ws, i, entry.get("snapshot", {}))
        ws.cell(row=i, column=len(QUEUE_HEADERS) + 1, value=entry.get("last_seen", ""))

    # Custom-DWG matrix, appended after "Last Seen" (same green-✓/red as the Full
    # Queue) so History is a complete per-order log. Uses each archived order's
    # snapshot, so it carries DWG data for orders archived once the scan was live.
    rows = [(i, e.get("snapshot", {})) for i, e in enumerate(entries, start=2)]
    suffixes = _append_dwg_matrix(ws, rows, len(headers) + 1)
    total_cols = len(headers) + len(suffixes)

    last_col = get_column_letter(total_cols)
    ws.auto_filter.ref = f"A1:{last_col}{len(entries) + 1}"
    ws.freeze_panes = "B2"  # keep the header row AND the Job # column visible
    _autosize(ws, num_cols=total_cols)


def build_workbook(
    jobs: List[Dict[str, Any]],
    diff: Dict[str, Any],
    briefing: Dict[str, Any],
    today: date,
    history: Dict[str, Any] | None = None,
) -> Path:
    # Jobs whose change order landed this run (CO# rose, or returned higher).
    co_changed_ids = {c.get("job") for c in diff.get("co_changed", [])}
    co_changed_ids |= {j.get("job") for j in diff.get("returning", []) if j.get("_co_returned")}

    wb = Workbook()
    changes_ws = wb.active
    changes_ws.title = "Changes"
    _write_changes_tab(changes_ws, briefing, diff, co_changed_ids=co_changed_ids)

    full_ws = wb.create_sheet("Full Queue")
    # Mark new + returning orders so they pop on the Full Queue tab too.
    new_ids = {j.get("job") for j in diff.get("new", []) if j.get("job")} | \
              {j.get("job") for j in diff.get("returning", []) if j.get("job")}
    _write_full_queue_tab(full_ws, jobs, today, new_job_ids=new_ids, co_changed_ids=co_changed_ids)

    history_ws = wb.create_sheet("History")
    _write_history_tab(history_ws, history or {})

    path = OUTPUT_DIR / f"queue_{today.isoformat()}.xlsx"
    try:
        wb.save(path)
    except PermissionError:
        # The file is almost always locked because it's open in Excel. Fall
        # back to a timestamped name so the run still produces a report.
        alt = OUTPUT_DIR / f"queue_{today.isoformat()}_{datetime.now():%H%M%S}.xlsx"
        log.warning(
            "Could not write %s (is it open in Excel?). Saving to %s instead.",
            path.name, alt.name,
        )
        wb.save(alt)
        path = alt
    log.info("Wrote Excel report: %s", path)
    return path
