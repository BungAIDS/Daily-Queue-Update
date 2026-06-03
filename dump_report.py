"""Print the rows of the latest (or a given) queue report, so we can review it.

    python dump_report.py                       # newest queue_*.xlsx in OUTPUT_DIR
    python dump_report.py "C:\\path\\queue_2026-06-03.xlsx"

Shows the Full Queue's key columns (Job/Design/Size/Arrangement/CO#/Folder),
the 'Change orders this run' section, and a coverage summary. Paste it back so
we can confirm the integration looks right.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from config import OUTPUT_DIR

COLS = ["Job #", "Design", "Size", "Arrangement", "CO#", "Folder"]


def _latest() -> Path:
    files = sorted(OUTPUT_DIR.glob("queue_*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not files:
        raise SystemExit(f"No queue_*.xlsx found in {OUTPUT_DIR}")
    return files[-1]


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else _latest()
    print(f"Report: {path}\n")
    wb = load_workbook(path)

    fq = wb["Full Queue"]
    headers = [fq.cell(1, c).value for c in range(1, fq.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    n_jobs = n_size = n_co = n_folder = 0
    print("FULL QUEUE")
    print("  " + " | ".join(COLS))
    print("  " + "-" * 70)
    for r in range(2, fq.max_row + 1):
        job = fq.cell(r, idx["Job #"]).value
        if not job:
            continue
        n_jobs += 1
        vals = []
        for c in COLS:
            cell = fq.cell(r, idx[c])
            if c == "Folder":
                v = cell.hyperlink.target if cell.hyperlink else (cell.value or "")
            else:
                v = cell.value
            vals.append("" if v is None else str(v))
        if vals[COLS.index("Size")]:
            n_size += 1
        if vals[COLS.index("CO#")]:
            n_co += 1
        if vals[COLS.index("Folder")]:
            n_folder += 1
        print("  " + " | ".join(vals))

    print("\nCHANGE ORDERS THIS RUN (from Changes tab)")
    ch = wb["Changes"]
    printing = False
    for r in range(1, ch.max_row + 1):
        a = ch.cell(r, 1).value
        if isinstance(a, str) and a.startswith("Persistent") and printing:
            break
        if isinstance(a, str) and a.startswith("Change orders this run"):
            printing = True
        if printing:
            rowvals = [str(ch.cell(r, c).value) for c in range(1, 6) if ch.cell(r, c).value is not None]
            if rowvals:
                print("  " + " | ".join(rowvals))

    print("\nCOVERAGE")
    print(f"  {n_jobs} jobs | {n_size} with Size (from SO pdf) | {n_co} at a change order | {n_folder} with a folder link")
    print(f"  {n_jobs - n_size} missing SO-pdf fields (no SO, e.g. HDX, or a failed download)")


if __name__ == "__main__":
    main()
