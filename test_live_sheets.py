"""Tests for the live master workbook's pure sheet model (live_sheets.py).

No pytest — run directly:

    python test_live_sheets.py

Checks the content/formatting intent of each tab (Live Queue, Changes, History,
Line Items) without any Excel/COM dependency.
"""
from __future__ import annotations

import sys
from datetime import date

import live_sheets as ls
from live_sheets import (FILL_DWG_NO, FILL_DWG_YES, FILL_NEW, FILL_OVERDUE,
                         FILL_DUETODAY, FILL_SOON, F_LINK, F_SECTION)

TODAY = date(2026, 6, 16)


def _job(num, **kw):
    j = {"job": num, "item": "47-0-0000", "design": "47", "customer": "ACME CORP",
         "end_date": "06/20/2026", "total_price": "$1,000.00", "so_pdf": "",
         "dwg_extras": {}, "job_folder": ""}
    j.update(kw)
    return j


def _find(sheet, text):
    for r, row in enumerate(sheet.grid):
        for c, cell in enumerate(row):
            if str(cell.value).startswith(text):
                return r, c
    return None


def test_full_queue_headers_and_added_column():
    sh = ls.full_queue_sheet([_job("421000")], TODAY)
    assert sh.grid[0][0].value == "Added"
    assert sh.grid[0][1].value == "Job #"          # first standard column
    assert sh.name == "Live Queue"
    assert sh.autofilter_a1 and sh.freeze == "C2"


def test_full_queue_overdue_fill_and_job_link():
    so = "Z:\\SO\\421000\\421000 - Sales Order CO#1.pdf"
    j = _job("421000", end_date="06/10/2026", so_pdf=so)
    sh = ls.full_queue_sheet([j], TODAY)
    job_cell = sh.grid[1][1]                        # Added is col0, Job# is col1
    assert job_cell.value == "421000"
    # Links straight to the latest SO PDF (the watcher keeps so_pdf current).
    assert job_cell.link == so and job_cell.font == F_LINK
    # End Date in the past -> overdue fill on the standard cells.
    assert sh.grid[1][1].fill == FILL_OVERDUE


def test_full_queue_new_fill_and_added_label():
    from datetime import datetime
    fs = datetime.now().replace(hour=9, minute=14, second=0, microsecond=0).isoformat()
    j = _job("421001", end_date="12/31/2026", _carried_over=False, _first_seen=fs)
    sh = ls.full_queue_sheet([j], TODAY, new_ids={"421001"})
    assert sh.grid[1][0].value.endswith("AM") or sh.grid[1][0].value.endswith("PM")  # added today -> time
    assert sh.grid[1][1].fill == FILL_NEW          # new, no urgency


def test_full_queue_dwg_matrix():
    a = _job("421000", dwg_extras={"51": "x"})
    b = _job("421001", dwg_extras={})
    sh = ls.full_queue_sheet([a, b], TODAY)
    # Header has a "-51" column at the end; rows show ✓/green and blank/red.
    pos = _find(sh, "-51")
    assert pos is not None and pos[0] == 0
    col = pos[1]
    assert sh.grid[1][col].value == "✓" and sh.grid[1][col].fill == FILL_DWG_YES
    assert sh.grid[2][col].value == "" and sh.grid[2][col].fill == FILL_DWG_NO


def test_full_queue_footer_total():
    sh = ls.full_queue_sheet([_job("421000", total_price="$1,000.00"),
                              _job("421001", total_price="$2,500.00")], TODAY)
    pos = _find(sh, "Total jobs: 2")
    assert pos is not None and sh.grid[pos[0]][pos[1]].font == F_SECTION
    # The money total lives on the footer row at the Total Price column.
    total_cells = [c for c in sh.grid[pos[0]] if isinstance(c.value, (int, float)) and c.value]
    assert any(abs(c.value - 3500.0) < 0.001 for c in total_cells)


def test_removed_block_mirrors_live_queue_row():
    so = "Z:\\SO\\421757\\421757 CO#2.pdf"
    overdue = _job("421757", end_date="06/10/2026", so_pdf=so, co_number=2)   # overdue + CO#
    fresh = _job("421802", end_date="12/31/2026")
    blk = ls.removed_block([(overdue, "2026-06-16T06:26:00"),
                            (fresh, "2026-06-16T07:42:00")],
                           TODAY, new_ids={"421802"}, co_changed_ids={"421757"})
    # Leads with the removal time ('Removed'), then the same data columns as the
    # board (so Job #/End Date and far-right DWG Reuse line up under it).
    assert blk["header_cells"][0].value == "Removed"
    assert [c.value for c in blk["header_cells"]] == ls.LIVE_QUEUE_REMOVED_HEADERS
    assert len(blk["header_cells"]) == len(ls.LIVE_QUEUE_HEADERS)
    assert blk["header_cells"][-1].value == "DWG Reuse"
    r0, r1 = blk["rows"]
    assert len(r0) == len(ls.LIVE_QUEUE_REMOVED_HEADERS)
    assert r0[0].value == "6:26 AM" and r0[0].number_format == "@"   # removal time leads, as text
    job0 = r0[ls.LIVE_QUEUE_KEY_COL - 1]                             # Job # aligns with the board
    assert job0.value == "421757" and job0.link == so               # Job # still links to its SO
    assert r0[ls.LIVE_QUEUE_END_DATE_COL - 1].fill == FILL_OVERDUE   # kept its overdue fill
    assert any(c.font == "red" for c in r0)                          # CO#-changed -> red text
    # the fresh (new-today) order keeps its 'new' shading and isn't red
    assert any(c.fill == FILL_NEW for c in r1) and not any(c.font == "red" for c in r1)


def test_live_queue_last_out_column():
    # A returning order shows its most recent prior departure in 'Last Out';
    # an order that has never left shows blank.
    assert ls.LIVE_QUEUE_HEADERS[ls.LIVE_QUEUE_LAST_OUT_COL - 1] == "Last Out"
    assert ls.LIVE_QUEUE_HEADERS[ls.LIVE_QUEUE_CBC_COL - 1] == "#"
    assert ls.LIVE_QUEUE_HEADERS[-1] == "DWG Reuse"
    returned = _job("421000", _added_iso="2026-06-16T08:00:00", _last_out="2026-06-12T16:30:00")
    cells = ls.live_queue_records([returned], TODAY)[0][1]
    assert cells[ls.LIVE_QUEUE_LAST_OUT_COL - 1].value == "Jun 12, 2026"   # prior departure (earlier day)
    never_left = _job("421001", _added_iso="2026-06-16T08:00:00")          # no _last_out
    blank = ls.live_queue_records([never_left], TODAY)[0][1]
    assert blank[ls.LIVE_QUEUE_LAST_OUT_COL - 1].value == ""


def test_changes_today_log_sections():
    new_today = [_job("421001", _carried_over=False, _first_seen="2026-06-16T09:14:00")]
    events = [
        {"time": "2026-06-16T09:30:00", "job": "420800", "customer": "X",
         "field": "End Date", "old": "06/01/2026", "new": "06/05/2026"},
        {"time": "2026-06-16T10:00:00", "job": "420800", "customer": "X",
         "field": "End Date", "old": "06/05/2026", "new": "06/09/2026"},   # same field, 2nd line
        {"time": "2026-06-16T11:00:00", "job": "420700", "customer": "Y",
         "field": "CO#", "old": "0", "new": "1"},
    ]
    removed_today = [_job("420900")]
    order_lookup = {"420700": {"design": "47", "so_arrangement": "A/9H Belt drive",
                               "co_history": ["C/O #1 06/16/26 KLO: ADDED VFD CONTROLS"]}}
    sh = ls.changes_sheet(new_today, events, removed_today, "2026-06-16",
                          updated_at="Jun 16, 2026 11:05 AM", order_lookup=order_lookup)
    assert _find(sh, "Changes — 2026-06-16") is not None
    assert _find(sh, "Last updated Jun 16, 2026 11:05 AM") is not None   # live stamp near the top
    assert _find(sh, "New orders today (1)") is not None
    assert _find(sh, "Change orders today (1)") is not None          # the CO# event
    # One order changed; End Date moved in two separate polls -> two before/after
    # pairs (instances), so each step's old and new value appears.
    assert _find(sh, "Orders that changed today (1)") is not None
    assert _find(sh, "06/01/2026") is not None    # instance 1 'before'
    assert _find(sh, "06/05/2026") is not None    # instance 1 'after' / instance 2 'before'
    assert _find(sh, "06/09/2026") is not None    # instance 2 'after'
    from live_sheets import FILL_CHANGE1, FILL_CHANGE2
    fills = {c.fill for row in sh.grid for c in row}
    # Two instances -> the second 'after' row is a darker grey than the first.
    assert FILL_CHANGE1 in fills and FILL_CHANGE2 in fills
    assert _find(sh, "Removed / completed today (1)") is not None
    assert _find(sh, "CO#1") is not None             # CO# column shows the current CO#
    # Change-order table reads like the rest: Folder, Quote Run, CO#, Oper, Design,
    # Customer, then the free-text 'What changed'.
    assert _find(sh, "What changed") is not None                     # the new header
    assert _find(sh, "ADDED VFD CONTROLS") is not None               # change description
    # The 'What changed' description overruns instead of widening its column.
    r, c = _find(sh, "ADDED VFD CONTROLS")
    assert sh.grid[r][c].overflow is True


def test_co_change_desc_uses_latest_note_when_exact_revision_is_absent():
    order = {"co_history": [
        "CO#1 070826 AMF - CORRECTED BHP, MOTOR DESCRIPTION AND ENCLOSURE",
    ]}
    assert ls._co_change_desc(order, 2) == (
        "CORRECTED BHP, MOTOR DESCRIPTION AND ENCLOSURE"
    )

    order["co_history"].insert(0, "C/O #2 7/13/26 DG: ADDED OUTLET DAMPER")
    assert ls._co_change_desc(order, 2) == "ADDED OUTLET DAMPER"


def test_verbose_arrangement_normalizes_to_short_code():
    """The Sales Order sometimes spells it out ('Arrangement 4') instead of the
    short 'A/4' code; that should normalize so the column stays narrow."""
    from excel_writer import split_arrangement, QUEUE_HEADERS
    assert split_arrangement("Arrangement 4") == ("A/4", "")
    assert split_arrangement("Arr. 9 belt drive") == ("A/9", "belt drive")
    assert split_arrangement("arrangement 10") == ("A/10", "")
    assert split_arrangement("A/4V C-Face mount") == ("A/4V", "C-Face mount")  # unchanged
    assert split_arrangement("N/A") == ("N/A", "")                             # passthrough

    # And it shows through on the Live Queue cell.
    qi = {h: i for i, h in enumerate(QUEUE_HEADERS)}
    recs = ls.live_queue_records([_job("421884", so_arrangement="Arrangement 4")], TODAY)
    _key, cells = recs[0]
    # live_queue_records prepends the 'Added' column, so the queue cols shift by one.
    assert cells[qi["Arrangement"] + 1].value == "A/4"


def test_change_orders_table_columns_and_abbrev_header():
    """'Change orders today' reads like the other tables: Time, Job #, Folder,
    Quote Run, CO#, Oper, Design, Customer, What changed. And Arrangement headers
    are abbreviated to 'Arr.' to keep the column narrow."""
    events = [{"time": "2026-06-16T11:00:00", "job": "420700", "customer": "ACME",
               "field": "CO#", "old": "0", "new": "1"}]
    lookup = {"420700": {"job": "420700", "design": "47", "oper": "200",
                         "so_arrangement": "A/9H Belt drive",
                         "co_history": ["C/O #1 06/16/26 KLO: ADDED VFD"]}}
    sh = ls.changes_sheet([], events, [], "2026-06-16", updated_at="x",
                          order_lookup=lookup)
    r, _ = _find(sh, "Change orders today")
    hdr = [str(c.value) for c in sh.grid[r + 1]]
    assert hdr == ["Time", "Job #", "Folder", "Quote Run", "CO#", "Oper",
                   "Design", "Customer", "What changed"]
    row = sh.grid[r + 2]
    assert row[1].value == "420700"                 # Job #
    assert str(row[4].value) == "CO#1"              # CO# column shows current CO#
    assert row[6].value == "47"                     # Design
    assert row[7].value == "ACME"                   # Customer

    # The Arrangement header reads 'Arr.' wherever it appears (e.g. the changed
    # table), while the internal column key stays 'Arrangement'.
    sh2 = ls.changes_sheet([_job("421001")], [], [], "2026-06-16", updated_at="x")
    assert _find(sh2, "Arr.") is not None
    assert _find(sh2, "Arrangement") is None
    assert "Arr." in ls.LIVE_QUEUE_HEADERS and "Arrangement" not in ls.LIVE_QUEUE_HEADERS


def test_changes_today_columns_align_across_sections():
    """Every section leads with a Time column in the same format and position as
    'Orders that changed today': Time in column A, Job # in column B, Folder in
    C — New orders show their arrival time, Removed orders their departure time,
    so Folder / Quote Run / CO# line up in the same columns across all three."""
    new_today = [_job("421001", _added_iso="2026-06-16T07:33:00")]
    events = [{"time": "2026-06-16T09:30:00", "job": "420800", "customer": "X",
               "field": "Oper", "old": "10", "new": "20"}]
    removed_today = [_job("420900", _left_iso="2026-06-16T10:15:00")]
    sh = ls.changes_sheet(new_today, events, removed_today, "2026-06-16",
                          updated_at="x", order_lookup={"420800": {"design": "47"}})

    def header_after(title):
        r, _ = _find(sh, title)
        return [str(c.value) for c in sh.grid[r + 1]]

    new_hdr = header_after("New orders today")
    chg_hdr = header_after("Orders that changed today")
    rem_hdr = header_after("Removed / completed today")

    # All three: Time in A, Job # in B, Folder in C; Quote Run / CO# follow.
    for hdr in (new_hdr, chg_hdr, rem_hdr):
        assert hdr[0] == "Time" and hdr[1] == "Job #" and hdr[2] == "Folder"
        assert hdr[2:5] == ["Folder", "Quote Run", "CO#"]

    # Data rows: the arrival / departure time leads in the changed-table format.
    nr, _ = _find(sh, "New orders today")
    new_data = sh.grid[nr + 2]
    assert new_data[0].value == ls.fmt_time("2026-06-16T07:33:00")   # e.g. '7:33 AM'
    assert new_data[1].value == "421001"
    rr, _ = _find(sh, "Removed / completed today")
    rem_data = sh.grid[rr + 2]
    assert rem_data[0].value == ls.fmt_time("2026-06-16T10:15:00")
    assert rem_data[1].value == "420900"

    # No merged spacer anywhere: Time and Job # stay two separate cells.
    for row in (sh.grid[nr + 1], new_data, sh.grid[rr + 1], rem_data,
                sh.grid[_find(sh, "Orders that changed today")[0] + 1]):
        assert row[0].colspan == 1 and row[1].colspan == 1


def test_changes_arrangement_size_suffix_moves_to_comment():
    """Like the Live Queue, the Changes tab trims Arrangement to its 'A/X' code and
    Size to its main value, moving the descriptive suffix to a hover comment so the
    columns stay narrow — in the New/Removed tables and the changed-order rows."""
    from excel_writer import QUEUE_HEADERS
    qi = {h: i for i, h in enumerate(QUEUE_HEADERS)}
    arr = "A/4V C-Face Flange mount (no motor base)"
    new_today = [_job("421884", so_arrangement=arr, so_size="6000-C6 Blade-1800")]
    events = [{"time": "2026-06-16T09:52:00", "job": "421572", "customer": "F",
               "field": "Arrangement", "old": "A/8", "new": arr}]
    lookup = {"421572": {"job": "421572", "so_arrangement": arr}}
    sh = ls.changes_sheet(new_today, events, [], "2026-06-16",
                          updated_at="x", order_lookup=lookup)

    # New orders table: Job # + spacer shift the queue columns right by one cell.
    nr, _ = _find(sh, "New orders today")
    data = sh.grid[nr + 2]
    a_cell, s_cell = data[qi["Arrangement"] + 1], data[qi["Size"] + 1]
    assert a_cell.value == "A/4V" and a_cell.comment == "C-Face Flange mount (no motor base)"
    assert s_cell.value == "6000" and s_cell.comment == "-C6 Blade-1800"

    # The changed-order instance row trims the new Arrangement value the same way.
    cr, _ = _find(sh, "Orders that changed today")
    inst = sh.grid[cr + 3]                       # title, header, 'was', instance
    ic = inst[qi["Arrangement"] + 1]             # leading Time shifts queue cols by one
    assert ic.value == "A/4V" and ic.comment == "C-Face Flange mount (no motor base)"


def test_orders_changed_one_instance_multiple_fields():
    # An order that changes ONCE (one poll) with several fields moving must be a
    # single before/after pair (two rows), not one row per field.
    t = "2026-06-16T09:30:00"
    events = [
        {"time": t, "job": "420800", "customer": "X",
         "field": "Assigned To", "old": "", "new": "DG"},   # '' old still shows as a move
        {"time": t, "job": "420800", "customer": "X",
         "field": "End Date", "old": "06/01/2026", "new": "06/05/2026"},
        {"time": t, "job": "420800", "customer": "X",
         "field": "Plan Hrs", "old": "10", "new": "12"},
        {"time": t, "job": "420800", "customer": "X",
         "field": "Checker", "old": "A", "new": "B"},
    ]
    sh = ls.changes_sheet([], events, [], "2026-06-16", updated_at="x")
    # Find the section and count its data rows (header + before + after only).
    grid = sh.grid
    start = next(i for i, r in enumerate(grid)
                 if r and str(r[0].value).startswith("Orders that changed today"))
    # rows after the section title: header row, then exactly the before/after pair,
    # then a blank separator row.
    body = []
    for r in grid[start + 1:]:
        if not r:           # blank row ends the section
            break
        body.append(r)
    assert len(body) == 3                       # header + before + after (one instance)
    header, before, after = body
    # Same column order as the rest of the workbook: a leading Time, then the queue
    # columns (so Job # sits one in, behind Time).
    assert header[0].value == "Time" and header[1].value == "Job #"
    assert before[0].value == ""                # no time on the start-of-day 'was' row
    assert before[1].value == "420800"          # Job # on the before row (after Time)
    # the 'before' row carries every changed field's OLD value; the 'after' its NEW.
    before_vals = {str(c.value) for c in before}
    after_vals = {str(c.value) for c in after}
    assert "06/01/2026" in before_vals and "10" in before_vals   # OLD values
    assert "06/05/2026" in after_vals and "12" in after_vals     # NEW values
    assert after[0].value == ls.fmt_time(t)     # the change instance is time-stamped
    assert after[0].fill == ls.FILL_CHANGE1     # after row shaded grey


def test_orders_changed_does_not_add_a_cleared_label():
    from excel_writer import QUEUE_HEADERS

    events = [{"time": "2026-06-16T09:30:00", "job": "420800", "customer": "X",
               "field": "Note", "old": "NEEDS CHECKING", "new": ""}]
    sh = ls.changes_sheet([], events, [], "2026-06-16", updated_at="x")
    start = next(i for i, row in enumerate(sh.grid)
                 if row and str(row[0].value).startswith("Orders that changed today"))
    instance = sh.grid[start + 3]               # title, header, was, changed row
    note_col = QUEUE_HEADERS.index("Note") + 1  # leading Time column
    assert instance[note_col].value == ""
    assert instance[note_col].font == ls.F_RED


def test_changes_fingerprint_ignores_timestamp_but_not_content():
    # The 'Last updated' stamp is volatile: it must NOT change the render
    # fingerprint (else the tab fully repaints every poll), but a real content
    # change must.
    from live_excel import _fingerprint
    events = [{"time": "2026-06-16T11:00:00", "job": "420700", "customer": "Y",
               "field": "CO#", "old": "0", "new": "1"}]
    a = ls.changes_sheet([], events, [], "2026-06-16", updated_at="Jun 16, 2026 11:05 AM")
    b = ls.changes_sheet([], events, [], "2026-06-16", updated_at="Jun 16, 2026 11:07 AM")
    assert _fingerprint(a) == _fingerprint(b)            # only the stamp differs -> no repaint
    c = ls.changes_sheet([_job("421001", _carried_over=False,
                                _first_seen="2026-06-16T09:14:00")],
                         events, [], "2026-06-16", updated_at="Jun 16, 2026 11:05 AM")
    assert _fingerprint(a) != _fingerprint(c)            # real content change -> repaint


def test_history_sheet():
    hist = {"420000": {"last_seen": "2026-06-10",
                       "snapshot": _job("420000", dwg_extras={"35": "x"})}}
    sh = ls.history_sheet(hist)
    assert sh.grid[0][0].value == "Job #"
    assert "Last Seen" in [c.value for c in sh.grid[0]]
    assert sh.grid[1][0].value == "420000"


def test_live_queue_records_no_dwg_and_new_today_fill():
    j = _job("421000", end_date="12/31/2026", dwg_extras={"51": "x", "35": "x"},
             _carried_over=False, _first_seen="2026-06-16T09:14:00")
    recs = ls.live_queue_records([j], TODAY, new_ids={"421000"})
    assert len(recs) == 1
    key, cells = recs[0]
    assert key == "421000"
    # Job # sits at the key column (1-based) — confirm the index lines up.
    assert cells[ls.LIVE_QUEUE_KEY_COL - 1].value == "421000"
    # Custom DWGs is no longer on Live Queue (Order History only).
    assert "Custom DWGs" not in ls.LIVE_QUEUE_HEADERS
    assert len(cells) == len(ls.LIVE_QUEUE_HEADERS)
    # new_ids drives the new-today highlight (no urgency on a far-future date).
    assert any(c.fill == FILL_NEW for c in cells)


def test_live_queue_end_date_is_sortable_serial():
    from datetime import date, timedelta
    j = _job("421000", end_date="06/10/2026")
    cells = ls.live_queue_records([j], TODAY)[0][1]
    ed = cells[ls.LIVE_QUEUE_END_DATE_COL - 1]
    # An Excel date serial (int), not a Python date (COM can't marshal a date),
    # formatted as a date and round-tripping to 2026-06-10.
    assert isinstance(ed.value, int)
    assert ed.number_format == "mm/dd/yyyy"
    assert date(1899, 12, 30) + timedelta(days=ed.value) == date(2026, 6, 10)
    # A blank End Date stays empty (sorts to the bottom), not a serial.
    blank = ls.live_queue_records([_job("421001", end_date="")], TODAY)[0][1]
    assert blank[ls.LIVE_QUEUE_END_DATE_COL - 1].value == ""


def test_live_queue_board_position_column():
    # "Added" leads; "#" remains the sort key while DWG Reuse is far right.
    assert ls.LIVE_QUEUE_HEADERS[0] == "Added"
    assert ls.LIVE_QUEUE_HEADERS[ls.LIVE_QUEUE_CBC_COL - 1] == "#"
    assert ls.LIVE_QUEUE_HEADERS[-1] == "DWG Reuse"
    j = _job("421000", dwg_reuse_label="421999 (-51)")
    j["_cbc_pos"] = 5
    cells = ls.live_queue_records([j], TODAY)[0][1]
    cbc = cells[ls.LIVE_QUEUE_CBC_COL - 1]
    assert cbc.value == 5 and cbc.center
    assert cells[-1].value == "421999 (-51)"
    # No board position (off-board / unknown) leaves it blank so it sorts last.
    cells2 = ls.live_queue_records([_job("421001")], TODAY)[0][1]
    assert cells2[ls.LIVE_QUEUE_CBC_COL - 1].value == ""


def test_dwg_reuse_is_far_right_in_changes_tables():
    from excel_writer import QUEUE_HEADERS

    assert QUEUE_HEADERS[-1] == "DWG Reuse"
    new_job = _job("421001", dwg_reuse_label="420999 (-51)",
                   _added_iso="2026-06-16T09:14:00")
    # Drawings is a dynamic changed-field column. DWG Reuse must remain after it.
    events = [{"time": "2026-06-16T09:30:00", "job": "421002", "customer": "X",
               "field": "Drawings", "old": "-51", "new": "-51, -95"}]
    lookup = {"421002": _job("421002", dwg_reuse_label="420998 (-95)")}
    sh = ls.changes_sheet([new_job], events, [], "2026-06-16",
                          updated_at="x", order_lookup=lookup)

    new_title, _ = _find(sh, "New orders today")
    new_header = sh.grid[new_title + 1]
    new_row = sh.grid[new_title + 2]
    assert new_header[-1].value == "DWG Reuse"
    assert new_row[-1].value == "420999 (-51)"

    changed_title, _ = _find(sh, "Orders that changed today")
    changed_header = sh.grid[changed_title + 1]
    changed_was = sh.grid[changed_title + 2]
    assert changed_header[-2].value == "Drawings"
    assert changed_header[-1].value == "DWG Reuse"
    assert changed_was[-1].value == "420998 (-95)"


def test_due_today_is_orange_between_red_and_gold():
    from datetime import timedelta
    def fill_for(end):
        cells = ls.live_queue_records([_job("421000", end_date=end)], TODAY)[0][1]
        return cells[ls.LIVE_QUEUE_KEY_COL - 1].fill        # Job # cell carries the row fill
    assert fill_for(TODAY.strftime("%m/%d/%Y")) == FILL_DUETODAY                       # due today -> orange
    assert fill_for((TODAY - timedelta(days=1)).strftime("%m/%d/%Y")) == FILL_OVERDUE  # past -> red
    assert fill_for((TODAY + timedelta(days=2)).strftime("%m/%d/%Y")) == FILL_SOON     # within 3 days -> gold


def test_prev_business_day_and_added_date():
    from datetime import date as _d
    assert ls.prev_business_day(_d(2026, 6, 22)) == _d(2026, 6, 19)   # Mon -> Fri
    assert ls.prev_business_day(_d(2026, 6, 23)) == _d(2026, 6, 22)   # Tue -> Mon
    assert ls.prev_business_day(_d(2026, 6, 19)) == _d(2026, 6, 18)   # Fri -> Thu
    assert ls.prev_business_day(_d(2026, 6, 21)) == _d(2026, 6, 19)   # Sun -> Fri
    assert ls.added_date({"_added_iso": "2026-06-22T13:28:00"}) == _d(2026, 6, 22)
    assert ls.added_date({"_first_seen": "2026-06-19T09:00:00"}) == _d(2026, 6, 19)
    assert ls.added_date({}) is None


def test_live_queue_co_change_turns_text_red():
    from live_sheets import F_RED
    j = _job("421000", co_number=2)
    cells = ls.live_queue_records([j], TODAY, co_changed_ids={"421000"})[0][1]
    assert cells[0].font == F_RED                       # Added
    assert cells[-1].font == F_RED                       # "#"
    # A non-link standard cell (Customer) goes red too.
    cust = (ls.LIVE_QUEUE_KEY_COL - 1) + ls.QUEUE_HEADERS.index("Customer")
    assert cells[cust].font == F_RED
    # An order without a change order today keeps its normal (non-red) text.
    plain = ls.live_queue_records([_job("421001")], TODAY)[0][1]
    assert plain[cust].font != F_RED


def test_live_queue_arrangement_code_and_comment():
    arr = (ls.LIVE_QUEUE_KEY_COL - 1) + ls.QUEUE_HEADERS.index("Arrangement")
    # A descriptive arrangement is trimmed to 'A/X'; the rest moves to a hover note.
    j = _job("421000", so_arrangement="A/4V C-Face Flange mount (no motor base)")
    c = ls.live_queue_records([j], TODAY)[0][1][arr]
    assert c.value == "A/4V"
    assert c.comment and "C-Face Flange mount (no motor base)" in c.comment
    # A clean code keeps no comment; N/A passes through untouched.
    c2 = ls.live_queue_records([_job("421001", so_arrangement="A/4")], TODAY)[0][1][arr]
    assert c2.value == "A/4" and c2.comment is None
    na = ls.live_queue_records([_job("421002", so_arrangement="N/A")], TODAY)[0][1][arr]
    assert na.value == "N/A" and na.comment is None
    # Order History keeps the full text (no trim, no note) to avoid re-writing the log.
    oh = ls._job_value_cells(j, columns=ls.OH_DATA_COLUMNS)
    oh_arr = oh[[k for _, k in ls.OH_DATA_COLUMNS].index("so_arrangement")]
    assert oh_arr.value == "A/4V C-Face Flange mount (no motor base)" and oh_arr.comment is None


def test_live_queue_size_main_and_comment():
    sz = (ls.LIVE_QUEUE_KEY_COL - 1) + ls.QUEUE_HEADERS.index("Size")
    # The leading number stays; the -code suffix and any description move to a note.
    j = ls.live_queue_records([_job("421000", so_size="3300-B12 Blade-1800")], TODAY)[0][1][sz]
    assert j.value == "3300" and j.comment == "-B12 Blade-1800"
    code = ls.live_queue_records([_job("421004", so_size="3000-A6")], TODAY)[0][1][sz]
    assert code.value == "3000" and code.comment == "-A6"
    paren = ls.live_queue_records([_job("421001", so_size="2412 (3600 RPM or less)")], TODAY)[0][1][sz]
    assert paren.value == "2412" and "(3600 RPM or less)" in paren.comment
    # A fraction stays whole; a plain number / code keeps no comment.
    frac = ls.live_queue_records([_job("421002", so_size="13 1/2")], TODAY)[0][1][sz]
    assert frac.value == "13 1/2" and frac.comment is None
    plain = ls.live_queue_records([_job("421003", so_size="H3")], TODAY)[0][1][sz]
    assert plain.value == "H3" and plain.comment is None
    # Order History keeps the full text (built-once log isn't re-trimmed).
    oh = ls._job_value_cells(_job("421000", so_size="3300-B12 Blade-1800"), columns=ls.OH_DATA_COLUMNS)
    oh_sz = oh[[k for _, k in ls.OH_DATA_COLUMNS].index("so_size")]
    assert oh_sz.value == "3300-B12 Blade-1800" and oh_sz.comment is None


def test_quote_run_link_folder_when_multiple_runs():
    dr_idx = ls.QUEUE_HEADERS.index("Quote Run")
    lead = ls.LIVE_QUEUE_KEY_COL - 1               # 0-based start of standard cells
    # One run -> link straight to the downloaded PDF.
    one = _job("421000", drive_run_pdf="Z:\\QUOTES\\421\\421000\\run1.pdf",
               drive_run_count=1, has_drive_run=True)
    cell = ls.live_queue_records([one], TODAY)[0][1][lead + dr_idx]
    assert cell.link == "Z:\\QUOTES\\421\\421000\\run1.pdf"
    # Multiple runs -> link to the folder that holds them, not just one file.
    many = _job("421001", drive_run_pdf="Z:\\QUOTES\\421\\421001\\run2.pdf",
                drive_run_count=3, has_drive_run=True)
    cell = ls.live_queue_records([many], TODAY)[0][1][lead + dr_idx]
    assert cell.link == "Z:\\QUOTES\\421\\421001"


def test_added_label_today_older_and_no_data():
    from datetime import datetime
    ref = datetime(2026, 6, 17, 12, 0, 0)
    # NO DATA only when there's literally no timestamp.
    assert ls.added_label({}, ref=ref) == "NO DATA"
    assert ls.added_label({"_added_iso": ""}, ref=ref) == "NO DATA"
    today = ls.added_label({"_added_iso": "2026-06-17T09:14:00"}, ref=ref)
    assert today.endswith("AM") and "," not in today          # today -> time only
    older = ls.added_label({"_added_iso": "2026-06-16T15:53:00"}, ref=ref)
    assert older == "Jun 16, 2026" or older == "Jun 16, 2026"  # earlier -> date, no time
    assert ":" not in older and "Jun 16" in older


def test_co_comment_most_recent_first_and_live_queue_cell():
    hist = ["CO#1 06/01 A - first", "CO#3 06/15 C - latest", "CO#2 06/08 B - mid"]
    cm = ls._co_comment({"co_history": hist})
    body = cm.splitlines()[1:]
    assert body[0].startswith("CO#3") and body[1].startswith("CO#2") and body[2].startswith("CO#1")
    assert ls._co_comment({"co_history": []}) is None
    # The CO# cell on a Live Queue row carries that hover note. The standard
    # columns begin after the "#"/"Added" lead pair (Job # sits at the key col).
    j = _job("421000", co_number=3, co_history=hist)
    cells = ls.live_queue_records([j], TODAY)[0][1]
    lead = ls.LIVE_QUEUE_KEY_COL - 1   # 0-based start of the standard column block
    assert cells[lead + ls._CO_IDX].comment is not None and "CO#3" in cells[lead + ls._CO_IDX].comment


def test_order_history_build_matrices_flags_and_separator():
    orders = [
        ("421000", {"on_queue": True, "added": "2026-06-16T09:00:00", "left": None,
                    "job": _job("421000", dwg_extras={"51": "x"},
                               line_items=[{"tags": ["SHAFT SEAL"]}, {"tags": ["COATING"]}])}),
        ("420900", {"on_queue": False, "added": "2026-06-15T08:00:00",
                    "left": "2026-06-16T07:30:00",
                    "job": _job("420900", dwg_extras={}, line_items=[{"tags": ["SHAFT SEAL"]}])}),
    ]
    spec = ls.order_history_build(orders, TODAY)
    h = spec["headers"]
    assert h[ls.ORDER_HISTORY_KEY_COL - 1] == "Job #" and ls.ORDER_HISTORY_KEY_COL == 1
    assert "On Queue" in h and "-51" in h and "SHAFT SEAL" in h and "COATING" in h and ls.OH_SEP_HEADER in h
    # On Queue / Added / Left sit right before the DWG matrix (Left is the column
    # immediately before the first '-suffix' matrix column).
    assert h.index("Left") == spec["dwg_range"][0] - 2
    r0 = dict(zip(h, spec["records"][0][1]))
    assert r0["On Queue"].value == "YES" and r0["-51"].value == "✓" and r0["SHAFT SEAL"].value == "✓"
    r1 = dict(zip(h, spec["records"][1][1]))
    assert r1["On Queue"].value == "NO" and r1["Left"].value == "2026-06-16 07:30"
    assert r1["-51"].value == ""                          # 420900 lacks that drawing
    ds, de = spec["dwg_range"]
    fs, fe = spec["feat_range"]
    assert spec["sep_col"] == de + 1 and fs == spec["sep_col"] + 1 and fe >= fs


def test_order_history_columns_stable_append_only():
    a = ("100", {"on_queue": True, "added": "t", "left": None,
                 "job": _job("100", dwg_extras={"51": "x"}, line_items=[{"tags": ["SHAFT SEAL"]}])})
    s1 = ls.order_history_build([a], TODAY)
    assert s1["columns"] == {"suffixes": ["51"], "tags": ["SHAFT SEAL"]}
    # A new order brings a new suffix (-35) and a new tag (COATING).
    b = ("200", {"on_queue": True, "added": "t", "left": None,
                 "job": _job("200", dwg_extras={"35": "x"}, line_items=[{"tags": ["COATING"]}])})
    s2 = ls.order_history_build([a, b], TODAY, prev_columns=s1["columns"])
    # Prior order kept; new items APPENDED at the end (never reordered) so the
    # existing rows/columns don't shift -> no rebuild needed for unchanged data.
    assert s2["columns"]["suffixes"] == ["51", "35"]
    assert s2["columns"]["tags"] == ["SHAFT SEAL", "COATING"]
    # Same data + same prev columns -> identical columns (rebuild not triggered).
    s3 = ls.order_history_build([a, b], TODAY, prev_columns=s2["columns"])
    assert s3["columns"] == s2["columns"]


def test_order_history_row_sig_stable_across_churny_fields():
    base = {"on_queue": True, "added": "2026-06-16T09:00:00", "left": None}
    e1 = ("421000", {**base, "job": _job("421000", end_date="06/20/2026", total_price="$1,000.00")})
    e2 = ("421000", {**base, "job": _job("421000", end_date="07/01/2026", total_price="$9,999.00")})
    s1 = ls.order_history_build([e1], TODAY)
    s2 = ls.order_history_build([e2], TODAY)
    # Churny board fields aren't on Order History, so the row signature is stable
    # -> the 12K log isn't rewritten when a date/price ticks.
    assert ls.row_sig(s1["records"][0][1]) == ls.row_sig(s2["records"][0][1])


def test_plan_upsert_append_update_delete():
    a = ("100", "sigA", ["cellsA"])
    b = ("200", "sigB", ["cellsB"])
    # 100 unchanged, 200 changed, 300 new, 400 only in existing (-> delete).
    desired = [("100", "sigA", "x"), ("200", "sigB2", "y"), ("300", "sigC", "z")]
    existing = {"100": "sigA", "200": "sigB", "400": "sigD"}
    ops = ls.plan_upsert(desired, existing, allow_delete=True)
    kinds = {(o[0], o[1]) for o in ops}
    assert ("update", "200") in kinds
    assert ("append", "300") in kinds
    assert ("delete", "400") in kinds
    assert not any(o[1] == "100" for o in ops)           # unchanged -> no op
    # Without allow_delete, 400 is left alone.
    ops2 = ls.plan_upsert(desired, existing, allow_delete=False)
    assert not any(o[0] == "delete" for o in ops2)


def test_similar_orders_sheets_layout():
    rows = [{"job": "421900", "similar": "420150", "customer": "UBP", "score": 1.95,
             "dwg": "-95 (PDF)", "shared": "THREADED PLUG", "folder": "Z:\\J\\420150"},
            {"job": "421900", "similar": "419704", "customer": "", "score": 0.99,
             "dwg": "—", "shared": "PO BOX", "folder": ""}]
    data = ls.similar_data_sheet(rows, ["421900", "421901", "421902"])
    assert data.name == ls.SIMILAR_DATA_TAB and not data.hidden and data.freeze == "A2"
    hdr = [c.value for c in data.grid[0]]
    assert hdr[0] == "Queue Order" and hdr[1] == "Similar Order" and hdr[8] == "Queue Orders"
    assert data.grid[1][0].value == "421900" and data.grid[1][1].value == "420150"
    # Group start styled (grey band + bold order #); later group rows plain, but
    # the Queue Order VALUE still repeats — the picker tab's FILTER matches on it.
    assert data.grid[1][0].fill == ls.FILL_NEW and data.grid[1][0].font == ls.F_SECTION
    assert data.grid[2][0].fill is None and data.grid[2][0].value == "421900"
    assert data.grid[1][6].link == "Z:\\J\\420150"    # folder cell click-through
    # The picker list carries EVERY queue order, past the end of the data rows.
    assert [r[8].value for r in data.grid[1:]] == ["421900", "421901", "421902"]
    assert data.grid[3][0].value == ""     # padding row for the longer picker list
    # The Live Queue's 'Similar' cell deep-links via a STABLE defined name (so
    # other groups shifting can never rewrite its row); the data sheet model
    # carries the name -> group-start-cell mapping the renderer (re)points.
    assert ls.similar_anchor(rows, "421900") == "#SIM_421900"
    assert ls.similar_anchor(rows, "421901") == ""
    assert data.names == {"SIM_421900": "'Similar Data'!$A$2"}

    tab = ls.similar_orders_sheet(len(rows), 3)
    assert tab.name == ls.SIMILAR_ORDERS_TAB and not tab.hidden
    assert tab.picker["cell"] == ls.SIMILAR_PICKER_CELL
    assert tab.picker["source"] == "='Similar Data'!$I$2:$I$4"
    assert [c.value for c in tab.grid[2]] == ls.SIMILAR_HEADERS
    f = tab.grid[3][0].value
    assert f.startswith("=IFERROR(FILTER('Similar Data'!$B$2:$G$3")
    assert "$A$2:$A$3" in f and "$B$1" in f


def test_row_sig_ignores_volatile_value_and_live_queue_pos_is_volatile():
    a = [ls.Cell("x"), ls.Cell(3, center=True, volatile=True)]
    b = [ls.Cell("x"), ls.Cell(9, center=True, volatile=True)]
    c = [ls.Cell("y"), ls.Cell(3, center=True, volatile=True)]
    assert ls.row_sig(a) == ls.row_sig(b)     # volatile value ignored...
    assert ls.row_sig(a) != ls.row_sig(c)     # ...but real content still counts
    # The Live Queue '#' cell is volatile, so board-position jitter alone can
    # never force a full row rewrite.
    j = {"job": "421000", "_cbc_pos": 5}
    _, cells = ls.live_queue_records([j], TODAY)[0]
    assert cells[ls.LIVE_QUEUE_CBC_COL - 1].volatile
    j2 = {"job": "421000", "_cbc_pos": 17}
    _, cells2 = ls.live_queue_records([j2], TODAY)[0]
    assert ls.row_sig(cells) == ls.row_sig(cells2)


def test_similar_orders_sheet_empty_has_note_not_formula():
    tab = ls.similar_orders_sheet(0, 0)
    assert tab.picker is None
    assert "No similar-order data yet" in tab.grid[3][0].value


def _so_item(**kw):
    it = {"raw": "1 FAN WHEEL SS SLEEVE 1,000.00 L",
          "norm": "FAN WHEEL STAINLESS STEEL SLEEVE", "qty": "1",
          "price": "1,000.00", "ptype": "L", "section": "",
          "details": ["Product: Wheel"], "tags": ["STAINLESS STEEL"]}
    it.update(kw)
    return it


def test_sales_order_data_sheet_layout():
    a = _job("421001", line_items=[_so_item()], co_number=1,
             co_history=["CO#1 - PRICE CHANGE"], so_pdf="Z:\\SO\\421001.pdf")
    b = _job("421000", line_items=[])
    sh = ls.sales_order_data_sheet([a, b])       # job-number order, not input order
    assert sh.name == ls.SO_DATA_TAB and not sh.hidden and sh.freeze == "A2"
    hdr = [c.value for c in sh.grid[0]]
    assert hdr[0] == "Queue Order"
    assert hdr[1:1 + len(ls.SO_ITEM_HEADERS)] == ls.SO_ITEM_HEADERS
    key = ls._SO_KEY_COL - 1                     # summary block start (0-based)
    assert hdr[key] == "Queue Order" and hdr[key + 1] == "Customer"
    # Item block: 421001's one item aligned from column A, group-banded.
    assert sh.grid[1][0].value == "421001" and sh.grid[1][0].fill == ls.FILL_NEW
    assert sh.grid[1][1].value == 1                            # item #
    assert sh.grid[1][4].value.startswith("1 FAN WHEEL")       # raw Description
    assert sh.grid[1][7].value == "Product: Wheel"             # Details joined
    assert sh.grid[1][8].value == "STAINLESS STEEL"            # Tags joined
    assert sh.grid[1][9].value == "FAN WHEEL STAINLESS STEEL SLEEVE"  # Normalized
    # Summary block: one row per order in job-number order, with the CO history
    # and SO PDF path at the shared column positions the picker tab looks up.
    assert sh.grid[1][key].value == "421000"
    assert sh.grid[2][key].value == "421001"
    assert sh.grid[2][key + 1].value == "ACME CORP"
    assert sh.grid[2][ls._SO_CO_HIST_COL - 1].value == "CO#1 - PRICE CHANGE"
    assert sh.grid[2][ls._SO_PDF_COL - 1].value == "Z:\\SO\\421001.pdf"


def test_sales_order_sheet_picker_and_formulas():
    from openpyxl.utils import get_column_letter
    tab = ls.sales_order_sheet(5, 2)
    key = get_column_letter(ls._SO_KEY_COL)
    assert tab.picker["cell"] == ls.SO_PICKER_CELL
    assert tab.picker["source"] == f"='SO Data'!${key}$2:${key}$3"
    # Summary spill row: FILTER over the summary block keyed on the picker,
    # then the Open-PDF HYPERLINK and the CO-history INDEX alongside it.
    summary = tab.grid[4]
    sum1 = get_column_letter(ls._SO_SUM_FIRST)
    assert f"FILTER('SO Data'!${sum1}$2:" in summary[0].value
    assert "$B$1" in summary[0].value
    pdf = summary[len(ls.SO_SUMMARY_COLUMNS)]
    assert "HYPERLINK" in pdf.value and pdf.font == ls.F_LINK
    assert "INDEX" in summary[-1].value and summary[-1].overflow
    # Line-items spill: FILTER over the item block (cols B..) keyed on col A; a
    # blank picker collapses to blank rather than spilling every row.
    items = tab.grid[8][0].value
    assert items.startswith('=IF($B$1&""="",""')
    assert "FILTER('SO Data'!$B$2:" in items and "$A$2:$A$6" in items


def test_sales_order_sheet_empty_has_note_not_formula():
    tab = ls.sales_order_sheet(0, 0)
    assert tab.picker is None
    assert "No sales-order data yet" in tab.grid[2][0].value


def main() -> int:
    passed = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            fn()
            print(f"  ok  {name}")
            passed += 1
    print(f"\n{passed} tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
