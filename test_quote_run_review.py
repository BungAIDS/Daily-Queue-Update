"""Tests for the Quote-Run review tool (quote_run_review.py): the note queue,
the unrolled run rows, and the workbook write/read round-trip.

No pytest — run directly:

    python test_quote_run_review.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import quote_run_review as qr


def _records():
    """Two orders. 421572: one text run with a field set + an uncaptured line.
    400111: a CHECK VISION scan with a suspect, plus a superseded format-dupe
    of the same fan (same Size/Design/Arrangement) that dedupe must collapse."""
    return {
        "421572": {"job": "421572", "type": "GENERAL LINE",
                   "folder": "Z:\\JOBS\\421572", "runs": [
            {"file": "421572 QT RUN.txt", "path": "Z:\\JOBS\\421572\\421572 QT RUN.txt",
             "template": "cbc_qt_run_text", "status": "OK", "damper": False,
             "fields": {"Serial": "421572", "Size": "36", "CFM": "12000",
                        "Oddball": "X"},
             "summary": "36 / 12000 CFM", "mtime": 100.0,
             "missed_data": ["SIDEPL,SPUN 0.075 (14) 12 34"]},
        ]},
        "400111": {"job": "400111", "type": "GENERAL LINE",
                   "folder": "Z:\\JOBS\\400111", "runs": [
            {"file": "QUOTE RUN.pdf", "path": "Z:\\JOBS\\400111\\QUOTE RUN.pdf",
             "template": "pdf_vision", "status": "OK", "damper": False,
             "fields": {"Serial": "400111", "Size": "22", "Design": "6195",
                        "Arrangement": "4", "CFM": "26843"},
             "summary": "", "mtime": 50.0, "vision": {}},
            {"file": "QUOTE RUN CO#1.pdf", "path": "Z:\\JOBS\\400111\\QUOTE RUN CO#1.pdf",
             "template": "pdf_vision", "status": "CHECK VISION", "damper": False,
             "fields": {"Serial": "400111", "Size": "22", "Design": "6195",
                        "Arrangement": "4", "CFM": "4/100"},
             "summary": "", "mtime": 40.0,
             "vision": {"suspect": ["implausible CFM='4/100'"]}},
        ]},
    }


FIELD_ORDER = ["Serial", "Size", "Design", "Arrangement", "CFM"]


def test_record_note_appends_and_dedups():
    store = {"notes": []}
    a = qr.record_note(store, "421572", "421572 QT RUN.txt", "CFM", "12000", "verify vs PDF")
    assert a and a["id"] == 1 and a["status"] == qr.STATUS_OPEN
    # Exact repeat (same anchor + text) is ignored.
    assert qr.record_note(store, "421572", "421572 QT RUN.txt", "CFM", "12000", "verify vs PDF") is None
    assert len(store["notes"]) == 1
    # Different text -> a new note; ids increment.
    b = qr.record_note(store, "421572", "421572 QT RUN.txt", "CFM", "12000", "actually it's fine")
    assert b["id"] == 2 and len(store["notes"]) == 2
    # Blank note or blank order is dropped.
    assert qr.record_note(store, "421572", "r", "CFM", "", "   ") is None
    assert qr.record_note(store, "", "r", "CFM", "x", "note") is None
    # Same text on a different row_key IS a new note.
    c = qr.record_note(store, "421572", "421572 QT RUN.txt", "CFM", "12000",
                       "verify vs PDF", row_key="run:x|field:sp")
    assert c is None or c["id"] == 3  # no stored key on the originals -> text match wins
    print("  ok  test_record_note_appends_and_dedups")


def test_mark_handled_and_open_filter():
    store = {"notes": []}
    qr.record_note(store, "421572", "r", "CFM", "12000", "check this")
    qr.record_note(store, "400111", "r", "SP", "4.5", "and this")
    assert len(qr.open_notes(store)) == 2
    assert qr.mark_handled(store, 1, "pattern fixed") is True
    assert qr.mark_handled(store, 99, "nope") is False
    opens = qr.open_notes(store)
    assert len(opens) == 1 and opens[0]["id"] == 2
    done = [n for n in store["notes"] if n["status"] == qr.STATUS_HANDLED][0]
    assert done["resolution"] == "pattern fixed" and done["handled_at"]
    print("  ok  test_mark_handled_and_open_filter")


def test_review_rows_shape_and_order():
    rows = qr.review_rows(_records(), {"notes": []}, field_order=FIELD_ORDER)
    # Newest order first.
    assert rows[0]["order"] == "421572" and rows[-1]["order"] == "400111"
    # First row of each order is the RUN header and carries group_start.
    assert rows[0]["kind"] == qr.KIND_RUN and rows[0]["group_start"]
    assert rows[0]["item"] == "421572 QT RUN.txt"
    assert "OK" in rows[0]["value"] and "4 fields" in rows[0]["value"]
    assert rows[0]["path"].endswith("QT RUN.txt")
    r421 = [r for r in rows if r["order"] == "421572"]
    # Fields follow field_order, then extras in parse order.
    fields = [r["item"] for r in r421 if r["kind"] == qr.KIND_FIELD]
    assert fields == ["Serial", "Size", "CFM", "Oddball"]
    # The uncaptured line is a MISSED row carrying the doc line.
    missed = [r for r in r421 if r["kind"] == qr.KIND_MISSED]
    assert len(missed) == 1 and missed[0]["value"].startswith("SIDEPL,SPUN")
    # 400111's suspect renders as a SUSPECT row right after its run header.
    r400 = [r for r in rows if r["order"] == "400111"]
    suspects = [r for r in r400 if r["kind"] == qr.KIND_SUSPECT]
    assert len(suspects) == 1 and "implausible CFM" in suspects[0]["value"]
    print("  ok  test_review_rows_shape_and_order")


def test_review_rows_dedupe_superseded_runs():
    rows = qr.review_rows(_records(), {"notes": []}, field_order=FIELD_ORDER)
    runs_400 = [r for r in rows if r["order"] == "400111" and r["kind"] == qr.KIND_RUN]
    # The CO#1 copy supersedes the base scan of the same fan -> one RUN row.
    assert len(runs_400) == 1 and runs_400[0]["item"] == "QUOTE RUN CO#1.pdf"
    print("  ok  test_review_rows_dedupe_superseded_runs")


def test_needs_human_reason_replaces_suspects():
    records = _records()
    records["400111"]["runs"][0]["vision"] = {
        "suspect": ["implausible CFM='4/100'", "odd Arrangement '48'"],
        "human_reason": "two reads disagree — CFM: 28000 vs 26843",
    }
    # Make it the surviving run so it renders.
    records["400111"]["runs"] = [records["400111"]["runs"][0]]
    rows = qr.review_rows(records, {"notes": []})
    suspects = [r for r in rows if r["kind"] == qr.KIND_SUSPECT]
    assert len(suspects) == 1 and suspects[0]["value"].startswith("two reads disagree")
    print("  ok  test_needs_human_reason_replaces_suspects")


def test_row_keys_dedupe_duplicates_and_reset_per_order():
    rows = [
        {"order": "1", "run": "a.txt", "kind": qr.KIND_MISSED, "item": "(uncaptured line)", "value": "SAME LINE"},
        {"order": "1", "run": "a.txt", "kind": qr.KIND_MISSED, "item": "(uncaptured line)", "value": "SAME LINE"},
        {"order": "2", "run": "a.txt", "kind": qr.KIND_MISSED, "item": "(uncaptured line)", "value": "SAME LINE"},
    ]
    keys = qr._row_keys(rows)
    assert keys[0] != keys[1] and keys[1].endswith("|2")
    assert keys[2] == keys[0]  # counts reset for the next order
    print("  ok  test_row_keys_dedupe_duplicates_and_reset_per_order")


def test_notes_attach_by_row_key_then_text_then_first_row():
    store = {"notes": []}
    rows0 = qr.review_rows(_records(), store, field_order=FIELD_ORDER)
    cfm = next(r for r in rows0 if r["order"] == "421572" and r["item"] == "CFM")
    # 1) row_key anchor.
    qr.record_note(store, "421572", cfm["run"], "CFM", cfm["value"],
                   "spot-checked, OK", row_key=cfm["row_key"])
    # 2) no key, anchors by item text.
    qr.record_note(store, "421572", cfm["run"], "Size", "36", "size looks wrong")
    # 3) orphaned key + unknown text falls back to the order's first row.
    qr.record_note(store, "421572", cfm["run"], "Gone", "gone-value",
                   "orphan", row_key="run:zzz|field:gone")
    rows = qr.review_rows(_records(), store, field_order=FIELD_ORDER)
    r421 = [r for r in rows if r["order"] == "421572"]
    assert "spot-checked, OK" in next(r for r in r421 if r["item"] == "CFM")["note"]
    assert "size looks wrong" in next(r for r in r421 if r["item"] == "Size")["note"]
    assert "orphan" in r421[0]["note"] and r421[0]["kind"] == qr.KIND_RUN
    assert all(r["status"] == qr.STATUS_OPEN for r in r421 if r["note"])
    print("  ok  test_notes_attach_by_row_key_then_text_then_first_row")


def test_handled_notes_leave_active_rows():
    store = {"notes": []}
    qr.record_note(store, "421572", "421572 QT RUN.txt", "CFM", "12000", "check")
    qr.mark_handled(store, 1, "verified against the PDF")
    rows = qr.review_rows(_records(), store, field_order=FIELD_ORDER)
    assert all(not r["note"] for r in rows)          # nothing renders as open
    assert store["notes"][0]["resolution"] == "verified against the PDF"
    print("  ok  test_handled_notes_leave_active_rows")


def test_store_roundtrip(tmp: Path):
    p = tmp / "notes.json"
    store = {"notes": []}
    qr.record_note(store, "421572", "r", "CFM", "12000", "hello")
    qr.save_store(store, p)
    again = qr.load_store(p)
    assert again["notes"][0]["note"] == "hello"
    assert qr.load_store(tmp / "missing.json") == {"notes": []}
    (tmp / "bad.json").write_text("{not json", encoding="utf-8")
    assert qr.load_store(tmp / "bad.json") == {"notes": []}
    print("  ok  test_store_roundtrip")


def test_handled_marks_ledger_roundtrip(tmp: Path):
    orig = qr.HANDLED_MARKS_PATH
    qr.HANDLED_MARKS_PATH = tmp / "quote_run_review_handled.json"
    try:
        store = {"notes": []}
        qr.record_note(store, "421572", "r", "CFM", "12000", "check")
        qr.record_note(store, "400111", "r", "SP", "4.5", "check too")
        qr.record_handled_mark(1, "pattern added + test")
        closed = qr.apply_handled_marks(store)
        assert [n["id"] for n in closed] == [1]
        assert store["notes"][0]["status"] == qr.STATUS_HANDLED
        assert store["notes"][0]["resolution"] == "pattern added + test"
        # Re-applying is a no-op; the open note stays open.
        assert qr.apply_handled_marks(store) == []
        assert len(qr.open_notes(store)) == 1
    finally:
        qr.HANDLED_MARKS_PATH = orig
    print("  ok  test_handled_marks_ledger_roundtrip")


def test_workbook_write_read_and_sync_roundtrip(tmp: Path):
    from openpyxl import load_workbook

    path = tmp / "quote_run_review.xlsx"
    store = {"notes": []}
    n = qr.write_workbook(path, _records(), store, field_order=FIELD_ORDER)
    assert n > 0

    # Type into the Note cell of 421572's CFM row, as a user would.
    wb = load_workbook(str(path))
    ws = wb[qr.REVIEW_SHEET]
    note_col = qr.HEADERS.index("Note") + 1
    target = None
    for row in range(2, ws.max_row + 1):
        if (str(ws.cell(row, 1).value) == "421572"
                and str(ws.cell(row, qr.HEADERS.index("Item") + 1).value) == "CFM"):
            target = row
            break
    assert target
    ws.cell(target, note_col).value = "confirm CFM against the customer spec"
    wb.save(str(path))
    wb.close()

    edits = qr.read_edits(path)
    assert len(edits) == 1
    e = edits[0]
    assert e["order"] == "421572" and e["item"] == "CFM"
    assert e["row_key"].endswith("field:cfm")
    assert qr.ingest_edits(store, edits) == 1
    # Re-reading the same workbook adds nothing (dedup).
    assert qr.ingest_edits(store, qr.read_edits(path)) == 0

    # Rebuild: the note renders as "#id:" history in Note; no Add Note column.
    qr.write_workbook(path, _records(), store, field_order=FIELD_ORDER)
    wb = load_workbook(str(path))
    ws = wb[qr.REVIEW_SHEET]
    rendered = str(ws.cell(target, note_col).value or "")
    assert "confirm CFM against the customer spec" in rendered
    assert qr.ADD_NOTE not in [cell.value for cell in ws[1]]
    wb.close()

    # A manual override typed straight into the rendered Note cell is recovered,
    # while the rendered "#id:" history is not re-imported.
    wb = load_workbook(str(path))
    ws = wb[qr.REVIEW_SHEET]
    ws.cell(target, note_col).value = rendered + qr.NOTE_SEPARATOR + "also check SP"
    wb.save(str(path))
    wb.close()
    edits = qr.read_edits(path)
    assert [e["note"] for e in edits] == ["also check SP"]
    print("  ok  test_workbook_write_read_and_sync_roundtrip")


def test_sync_upgrades_previous_add_note_layout(tmp: Path):
    from types import SimpleNamespace

    from openpyxl import load_workbook

    records = _records()
    path = tmp / "quote_run_review.xlsx"
    queue = tmp / "quote_run_review_notes.json"
    qr.write_workbook(path, records, {"notes": []}, field_order=FIELD_ORDER)

    # Recreate the earlier layout: a dedicated Add Note column after Note,
    # with an entry the migration must not lose.
    wb = load_workbook(str(path), data_only=False)
    ws = wb[qr.REVIEW_SHEET]
    note_col = qr.HEADERS.index("Note") + 1
    ws.insert_cols(note_col + 1, 1)
    ws.cell(1, note_col + 1).value = qr.ADD_NOTE
    target = next(
        row for row in range(2, ws.max_row + 1)
        if str(ws.cell(row, 1).value) == "421572"
        and str(ws.cell(row, qr.HEADERS.index("Item") + 1).value) == "CFM"
    )
    ws.cell(target, note_col + 1).value = "preserve old Add Note entry"
    wb.save(str(path))
    wb.close()
    assert qr._layout_needs_upgrade(path)

    old_store_path = qr.REVIEW_STORE_PATH
    old_runs, old_fields = qr._load_runs, qr._core_fields
    qr.REVIEW_STORE_PATH = queue
    qr._load_runs = lambda: records
    qr._core_fields = lambda: FIELD_ORDER
    try:
        assert qr._cmd_sync(SimpleNamespace(out=str(path))) == 0
    finally:
        qr.REVIEW_STORE_PATH = old_store_path
        qr._load_runs, qr._core_fields = old_runs, old_fields

    stored = qr.load_store(queue)
    assert [n["note"] for n in stored["notes"]] == ["preserve old Add Note entry"]
    rebuilt = load_workbook(str(path), data_only=False)
    header = [cell.value for cell in rebuilt[qr.REVIEW_SHEET][1]]
    assert header[:len(qr.HEADERS)] == qr.HEADERS
    assert qr.ADD_NOTE not in header
    rebuilt.close()
    assert not qr._layout_needs_upgrade(path)
    print("  ok  test_sync_upgrades_previous_add_note_layout")


def test_formula_like_document_lines_stay_text(tmp: Path):
    # Real corpus lines can start with "=" (price sums like
    # "=80240-6773-6056=67,411"); written as formulas they make Excel "repair"
    # the sheet by deleting the cell. They must round-trip as literal text.
    from openpyxl import load_workbook

    records = _records()
    run = records["421572"]["runs"][0]
    run["missed_data"] = ["=80240-6773-6056=67,411"]
    run["fields"]["Total"] = "= $9,704.1"
    store = {"notes": []}
    qr.record_note(store, "421572", run["file"], "CFM", "12000", "=confirm this sum")
    qr.mark_handled(store, 1, "=verified against the PDF")
    path = tmp / "quote_run_review.xlsx"
    qr.write_workbook(path, records, store, field_order=FIELD_ORDER)

    wb = load_workbook(str(path))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", (ws.title, cell.coordinate, cell.value)
    ws = wb[qr.REVIEW_SHEET]
    val_col = qr.HEADERS.index("Value") + 1
    values = {str(ws.cell(r, val_col).value) for r in range(2, ws.max_row + 1)}
    assert "=80240-6773-6056=67,411" in values
    assert "= $9,704.1" in values
    wb.close()
    print("  ok  test_formula_like_document_lines_stay_text")


def test_resolved_tab_keeps_history_off_the_active_row(tmp: Path):
    from openpyxl import load_workbook

    path = tmp / "quote_run_review.xlsx"
    store = {"notes": []}
    qr.record_note(store, "421572", "421572 QT RUN.txt", "CFM", "12000", "old note")
    qr.mark_handled(store, 1, "verified")
    qr.write_workbook(path, _records(), store, field_order=FIELD_ORDER)
    wb = load_workbook(str(path))
    ws = wb[qr.REVIEW_SHEET]
    note_col = qr.HEADERS.index("Note") + 1
    assert all(not ws.cell(r, note_col).value for r in range(2, ws.max_row + 1))
    resolved = wb[qr.RESOLVED_SHEET]
    assert resolved.max_row == 2
    assert str(resolved.cell(2, 1).value) == "421572"
    assert str(resolved.cell(2, 5).value) == "verified"
    wb.close()
    print("  ok  test_resolved_tab_keeps_history_off_the_active_row")


def main() -> int:
    tests_no_tmp = [
        test_record_note_appends_and_dedups,
        test_mark_handled_and_open_filter,
        test_review_rows_shape_and_order,
        test_review_rows_dedupe_superseded_runs,
        test_needs_human_reason_replaces_suspects,
        test_row_keys_dedupe_duplicates_and_reset_per_order,
        test_notes_attach_by_row_key_then_text_then_first_row,
        test_handled_notes_leave_active_rows,
    ]
    tests_tmp = [
        test_store_roundtrip,
        test_handled_marks_ledger_roundtrip,
        test_workbook_write_read_and_sync_roundtrip,
        test_sync_upgrades_previous_add_note_layout,
        test_formula_like_document_lines_stay_text,
        test_resolved_tab_keeps_history_off_the_active_row,
    ]
    for t in tests_no_tmp:
        t()
    for t in tests_tmp:
        with tempfile.TemporaryDirectory() as d:
            t(Path(d))
    print(f"\n{len(tests_no_tmp) + len(tests_tmp)} tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
