"""Tests for the Sales-Order review tool (so_review.py): the note queue, the
hierarchy rows, and the workbook write/read round-trip.

No pytest — run directly:

    python test_so_review.py
"""
from __future__ import annotations

import sys
import tempfile
from contextlib import redirect_stderr
from io import StringIO
from pathlib import Path

import so_review as sr


def _line_items_store():
    # One order with the 421966-shaped IVC family (three lines -> one component)
    # plus a standalone line, so tree_rows yields COMPONENT/FACT/SOURCE rows.
    return {"jobs": {"421966": {"items": [
        {"raw": "Base Fan L 16,649.00", "norm": "BASE FAN", "price": "16,649.00",
         "details": [], "tags": []},
        {"raw": "Inlet Volume Control, Low Leak, Automatic L 3,531.00",
         "norm": "INLET VOLUME CONTROL LOW LEAK AUTOMATIC", "price": "3,531.00",
         "details": ["Actuator Manufacturer: By Others"], "tags": ["DAMPER"],
         "attributes": {"used_on": "IVC", "operation": "Automatic"}},
        {"raw": "Inlet, Flanged, Punched (with IVC) L 1,559.00",
         "norm": "INLET FLANGED PUNCHED WITH IVC", "price": "1,559.00",
         "details": [], "tags": ["FLANGE"], "attributes": {"used_on": "IVC"}},
    ]}}}


def _legacy_picker_workbook(path: Path, line_items=None) -> None:
    """Build the pre-row-grid Notes + formula-picker workbook shape."""
    from openpyxl import load_workbook

    sr.write_workbook(path, line_items or _line_items_store(), {"notes": []})
    book = load_workbook(str(path), data_only=False)
    notes = book[sr.BROWSE_SHEET]
    notes.title = sr.NOTES_SHEET
    note_col = sr.HEADERS.index("Note") + 1
    notes.insert_cols(note_col + 1, 1)
    notes.cell(1, note_col + 1).value = sr.NOTES_ADD_NOTE

    windows = []
    order_col = sr.HEADERS.index("Order") + 1
    start = 2
    while start <= notes.max_row:
        order = str(notes.cell(start, order_col).value or "")
        end = start
        while end + 1 <= notes.max_row and str(notes.cell(end + 1, order_col).value or "") == order:
            end += 1
        if order:
            windows.append((order, start, end - start + 1))
        start = end + 1

    orders = book.create_sheet(sr.ORDERS_SHEET)
    for row in windows:
        orders.append(row)
    orders.sheet_state = "hidden"

    browse = book.create_sheet(sr.BROWSE_SHEET, 0)
    browse["B1"] = windows[0][0] if windows else ""
    for col, header in enumerate(sr.BROWSE_HEADERS, start=1):
        browse.cell(sr.BROWSE_HEADER_ROW, col).value = header
    book.save(str(path))
    book.close()


def test_record_note_appends_and_dedups():
    store = {"notes": []}
    a = sr.record_note(store, "421966", 1, "Base Fan", "keep as-is")
    assert a and a["id"] == 1 and a["status"] == sr.STATUS_OPEN
    # Exact repeat (same order+item+text) is ignored.
    assert sr.record_note(store, "421966", 1, "Base Fan", "keep as-is") is None
    assert len(store["notes"]) == 1
    # Different text -> a new note; ids increment.
    b = sr.record_note(store, "421966", 1, "Base Fan", "actually check the price")
    assert b["id"] == 2 and len(store["notes"]) == 2
    # Blank note or blank order is dropped.
    assert sr.record_note(store, "421966", 1, "Base Fan", "   ") is None
    assert sr.record_note(store, "", 1, "x", "note") is None

    # A pre-row-key derived note dedups by its hierarchy text, without
    # suppressing the same wording on a different derived row.
    legacy = sr.record_note(store, "421966", "", "    motor_hp: 10", "REDUNDANT")
    assert legacy and "row_key" not in legacy
    assert sr.record_note(
        store, "421966", "", "    motor_hp: 10", "REDUNDANT",
        row_key="attribute|motor|motor_hp: 10",
    ) is None
    distinct = sr.record_note(
        store, "421966", "", "    motor_rpm: 1800", "REDUNDANT",
        row_key="attribute|motor|motor_rpm: 1800",
    )
    assert distinct and distinct["id"] == 4


def test_legacy_row_key_counts_reset_for_each_order():
    rows = [
        {"order": "421900", "kind": "COMPONENT", "hierarchy": "[MOTOR]", "item_no": "1"},
        {"order": "421900", "kind": "ATTRIBUTE", "hierarchy": "    motor_hp: 10", "item_no": ""},
        {"order": "421901", "kind": "COMPONENT", "hierarchy": "[MOTOR]", "item_no": "1"},
        {"order": "421901", "kind": "ATTRIBUTE", "hierarchy": "    motor_hp: 10", "item_no": ""},
    ]
    assert sr._hierarchy_row_keys(rows) == [
        "item:1",
        "attribute|[motor]|motor_hp: 10",
        "item:1",
        "attribute|[motor]|motor_hp: 10",
    ]


def test_mark_handled_and_open_filter():
    store = {"notes": []}
    n = sr.record_note(store, "421966", 2, "IVC", "these 3 should be one component")
    assert len(sr.open_notes(store)) == 1
    assert sr.mark_handled(store, n["id"], "confirmed used_on=IVC grouping")
    assert sr.open_notes(store) == []
    done = store["notes"][0]
    assert done["status"] == sr.STATUS_HANDLED and done["handled_at"]
    assert done["resolution"] == "confirmed used_on=IVC grouping"
    assert not sr.mark_handled(store, 999, "no such id")


def test_review_rows_attach_notes_to_line_items_only():
    li = _line_items_store()
    store = {"notes": []}
    # A note on item #2 (the IVC's primary source line).
    sr.record_note(store, "421966", 2, "Inlet Volume Control…", "grouping looks right")
    rows = sr.review_rows(li, store)
    # Derived rows carry no item number but remain annotatable by hierarchy text.
    comp = [r for r in rows if r["kind"] == sr.so_hierarchy.KIND_COMPONENT]
    assert comp and all(c["order"] == "421966" for c in comp)
    attrs = [r for r in rows if r["kind"] == sr.so_hierarchy.KIND_ATTRIBUTE]
    assert attrs and all(a["annotatable"] and a["note"] == "" for a in attrs)
    # The SOURCE row for item #2 carries the recorded note and is annotatable.
    src2 = [r for r in rows if r["kind"] == sr.so_hierarchy.KIND_SOURCE
            and str(r["item_no"]) == "2"]
    assert src2 and src2[0]["annotatable"] and src2[0]["note"] == "#1: grouping looks right"
    # Every row names its order, so the sheet is self-identifying.
    assert all(r["order"] == "421966" for r in rows)
    assert rows[0]["group_start"] and not rows[1]["group_start"]


def test_multiple_notes_on_one_item_remain_visible_and_readable(tmp: Path):
    li = _line_items_store()
    store = {"notes": []}
    sr.record_note(store, "421966", 2, "IVC", "first note for item 2")
    sr.record_note(store, "421966", 2, "IVC", "second note for item 2")

    rows = [
        row for row in sr.review_rows(li, store)
        if str(row["item_no"]) == "2" and row["note"]
    ]
    assert rows
    assert all("first note for item 2" in row["note"] for row in rows)
    assert all("second note for item 2" in row["note"] for row in rows)

    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, li, store)
    # Rendered history is not re-imported as new typing in the current layout.
    assert sr.read_edits(wb) == []


def test_ingest_edits_records_source_and_derived_row_notes():
    store = {"notes": []}
    edits = [
        {"order": "421966", "item_no": "2", "item_text": "IVC", "note": "good"},
        {"order": "421966", "item_no": "", "item_text": "[IVC]", "note": "component note"},
        {"order": "421966", "item_no": "", "item_text": "    used on: IVC", "note": "component note"},
        {"order": "421966", "item_no": "1", "item_text": "Base Fan", "note": ""},   # empty note
    ]
    added = sr.ingest_edits(store, edits)
    assert added == 3
    assert [(n["item_no"], n["note"]) for n in store["notes"]] == [
        ("2", "good"), ("", "component note"), ("", "component note")]
    # Re-ingesting the same edits adds nothing (dedup).
    assert sr.ingest_edits(store, edits) == 0


def test_store_roundtrip(tmp: Path):
    p = tmp / "so_review_notes.json"
    store = {"notes": []}
    sr.record_note(store, "421966", 2, "IVC", "note one")
    sr.save_store(store, p)
    back = sr.load_store(p)
    assert back["notes"][0]["note"] == "note one"
    # A missing/garbage file loads as an empty queue.
    assert sr.load_store(tmp / "nope.json") == {"notes": []}


def test_handled_notes_leave_active_rows_but_remain_in_history():
    li = _line_items_store()
    store = {"notes": []}
    sr.record_note(store, "421966", 1, "Base Fan", "handled one")       # id 1
    sr.record_note(store, "421966", 2, "IVC", "still open")             # id 2
    sr.mark_handled(store, 1, "did the thing")
    rows = sr.review_rows(li, store)
    # A resolved note leaves the active row so another note can be added there.
    r1 = next(r for r in rows if str(r["item_no"]) == "1")
    assert r1["note"] == "" and r1["status"] == "" and r1["resolution"] == ""
    # Its resolution remains in the durable queue history.
    assert store["notes"][0]["note"] == "handled one"
    assert store["notes"][0]["resolution"] == "did the thing"
    # An open note still shows, flagged open (needs attention).
    r2 = next(r for r in rows if str(r["item_no"]) == "2"
              and r["kind"] == sr.so_hierarchy.KIND_SOURCE)
    assert r2["note"] == "#2: still open" and r2["status"] == sr.STATUS_OPEN
    assert r2["resolution"] == ""


def test_orphaned_note_reanchors_to_the_job_row():
    li = _line_items_store()
    store = {"notes": []}
    # A note whose line item no longer exists (item 99 isn't in the parse)
    # re-anchors to the order's first row instead of being dropped.
    sr.record_note(store, "421966", 99, "a line that got reparsed away", "keep me")
    rows = [r for r in sr.review_rows(li, store) if r["order"] == "421966"]
    assert rows[0]["note"] == "#1: keep me"            # landed on the job/first row
    assert not any(r["note"] == "#1: keep me" for r in rows[1:])


def test_handled_marks_ledger_roundtrip(tmp: Path):
    import so_review
    ledger = tmp / "so_review_handled.json"
    old = so_review.HANDLED_MARKS_PATH
    so_review.HANDLED_MARKS_PATH = ledger
    try:
        # Claude records two resolutions in the tracked ledger.
        so_review.record_handled_mark(2, "confirmed IVC grouping")
        so_review.record_handled_mark(5, "fixed the flange rule")
        # Re-recording the same id updates in place (no duplicate).
        so_review.record_handled_mark(2, "confirmed IVC grouping (v2)")
        led = so_review._load_ledger()["handled"]
        assert {m["id"] for m in led} == {2, 5} and len(led) == 2

        # The user's local queue has notes 2 and 5 open; applying the ledger
        # closes exactly those, with Claude's resolutions.
        store = {"notes": [
            {"id": 2, "order": "421966", "item_no": "2", "note": "x", "status": "open"},
            {"id": 5, "order": "421966", "item_no": "15", "note": "y", "status": "open"},
            {"id": 9, "order": "421900", "item_no": "3", "note": "z", "status": "open"},
        ]}
        closed = so_review.apply_handled_marks(store)
        assert {c["id"] for c in closed} == {2, 5}
        assert sr.open_notes(store) == [store["notes"][2]]       # note 9 still open
        assert store["notes"][0]["resolution"] == "confirmed IVC grouping (v2)"
        # Idempotent: applying again closes nothing new.
        assert so_review.apply_handled_marks(store) == []
    finally:
        so_review.HANDLED_MARKS_PATH = old


def test_workbook_write_read_and_sync_roundtrip(tmp: Path):
    li = _line_items_store()
    store = {"notes": []}
    wb = tmp / "review.xlsx"
    n = sr.write_workbook(wb, li, store)
    assert n > 0 and wb.exists()

    # Simulate the human typing a note in Note on item #2's
    # SOURCE row, then reading it back.
    from openpyxl import load_workbook
    book = load_workbook(str(wb))
    ws = book[sr.BROWSE_SHEET]
    note_col = sr.HEADERS.index("Note") + 1
    item_col = sr.HEADERS.index("Item") + 1
    typed = False
    for i in range(2, ws.max_row + 1):
        if str(ws.cell(row=i, column=item_col).value) == "2":
            ws.cell(row=i, column=note_col).value = "parsed IVC grouping is correct"
            typed = True
            break
    assert typed
    book.save(str(wb))

    edits = sr.read_edits(wb)
    assert {"order": "421966", "item_no": "2"} == {k: edits[0][k] for k in ("order", "item_no")}
    added = sr.ingest_edits(store, edits)
    assert added == 1 and sr.open_notes(store)[0]["note"] == "parsed IVC grouping is correct"

    # Rebuild with the note recorded -> it shows on the sheet, marked open.
    sr.write_workbook(wb, li, store)
    book2 = load_workbook(str(wb))
    ws2 = book2[sr.BROWSE_SHEET]
    status_col = sr.HEADERS.index("Status") + 1
    found = [i for i in range(2, ws2.max_row + 1)
             if str(ws2.cell(row=i, column=item_col).value) == "2"
             and ws2.cell(row=i, column=note_col).value]
    assert found and str(ws2.cell(row=found[0], column=status_col).value) == sr.STATUS_OPEN
    assert ws2.cell(found[0], note_col).value == "#1: parsed IVC grouping is correct"


def test_current_note_column_override_on_derived_row_is_recovered(tmp: Path):
    from openpyxl import load_workbook

    line_items = _line_items_store()
    target = next(
        row for row in sr.review_rows(line_items, {"notes": []})
        if row["kind"] == sr.so_hierarchy.KIND_ATTRIBUTE
    )
    assert target["item_no"] == ""

    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, line_items, {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    notes = book[sr.BROWSE_SHEET]
    hierarchy_col = sr.HEADERS.index("Hierarchy") + 1
    note_col = sr.HEADERS.index("Note") + 1
    row_no = next(
        row for row in range(2, notes.max_row + 1)
        if notes.cell(row, hierarchy_col).value == target["hierarchy"]
    )
    notes.cell(row_no, note_col).value = "rename this attribute"
    book.save(str(wb))
    book.close()

    edits = sr.read_edits(wb)
    assert len(edits) == 1
    assert edits[0]["item_no"] == ""
    assert edits[0]["item_text"] == target["hierarchy"].strip()
    assert edits[0]["note"] == "rename this attribute"

    store = {"notes": []}
    assert sr.ingest_edits(store, edits) == 1
    attached = next(row for row in sr.review_rows(line_items, store)
                    if row["hierarchy"] == target["hierarchy"])
    assert attached["note"] == "#1: rename this attribute"
    sr.write_workbook(wb, line_items, store)
    assert sr.read_edits(wb) == []


def test_legacy_duplicate_attribute_labels_keep_separate_parent_targets(tmp: Path):
    from openpyxl import load_workbook

    line_items = {"jobs": {"421959": {"items": [
        {"raw": "Replacement Shaft L 1.00", "norm": "REPLACEMENT SHAFT",
         "price": "1.00", "details": [], "tags": ["SPARE PARTS"],
         "attributes": {"component": "REPLACEMENT SHAFT", "shared": "X"}},
        {"raw": "Replacement Wheel L 1.00", "norm": "REPLACEMENT WHEEL",
         "price": "1.00", "details": [], "tags": ["SPARE PARTS"],
         "attributes": {"component": "REPLACEMENT WHEEL", "shared": "X"}},
    ]}}}


    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, line_items, {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    notes = book[sr.BROWSE_SHEET]
    # Simulate the workbook saved before hidden row keys existed.
    row_key_col = next(cell.column for cell in notes[1] if cell.value == sr.ROW_KEY_HEADER)
    notes.delete_cols(row_key_col)
    hierarchy_col = sr.HEADERS.index("Hierarchy") + 1
    note_col = sr.HEADERS.index("Note") + 1
    targets = [row for row in range(2, notes.max_row + 1)
               if notes.cell(row, hierarchy_col).value == "    shared: X"]
    assert len(targets) == 2
    for row in targets:
        notes.cell(row, note_col).value = "REDUNDANT"
    book.save(str(wb))
    book.close()

    edits = sr.read_edits(wb)
    assert len(edits) == 2
    assert len({edit["row_key"] for edit in edits}) == 2
    store = {"notes": []}
    assert sr.ingest_edits(store, edits) == 2


def test_legacy_notes_column_is_imported_once(tmp: Path):
    from openpyxl import load_workbook

    wb = tmp / "review.xlsx"
    _legacy_picker_workbook(wb)
    book = load_workbook(str(wb), data_only=False)
    notes = book[sr.NOTES_SHEET]
    legacy_header = [cell.value for cell in notes[1]]
    notes.delete_cols(legacy_header.index(sr.NOTES_ADD_NOTE) + 1, 1)
    item_col = sr.HEADERS.index("Item") + 1
    note_col = sr.HEADERS.index("Note") + 1
    row = next(row for row in range(2, notes.max_row + 1)
               if str(notes.cell(row, item_col).value) == "2")
    notes.cell(row, note_col).value = "recover legacy Notes-tab typing"
    book.save(str(wb))
    book.close()

    edits = [edit for edit in sr.read_edits(wb) if edit.get("source") == "notes"]
    assert len(edits) == 1
    assert edits[0]["order"] == "421966" and edits[0]["item_no"] == "2"
    assert edits[0]["note"] == "recover legacy Notes-tab typing"
    assert sr._browse_layout_needs_upgrade(wb)


def test_workbook_has_one_filterable_real_row_grid(tmp: Path):
    from openpyxl import load_workbook

    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, _line_items_store(), {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    assert book.sheetnames == [sr.BROWSE_SHEET, sr.RESOLVED_SHEET]
    assert book.active.title == sr.BROWSE_SHEET
    review = book[sr.BROWSE_SHEET]
    assert [review.cell(1, c).value for c in range(1, len(sr.HEADERS) + 1)] == sr.HEADERS
    assert review.freeze_panes == "A2"
    assert review.auto_filter.ref == f"A1:H{review.max_row}"
    assert review.cell(2, 1).value == "421966"
    assert all(
        not (isinstance(review.cell(row, col).value, str)
             and review.cell(row, col).value.startswith("="))
        for row in range(2, review.max_row + 1)
        for col in range(1, len(sr.HEADERS) + 1)
    )
    assert review.column_dimensions["I"].hidden
    assert review.column_dimensions["J"].hidden
    assert review.conditional_formatting._cf_rules
    assert not sr._browse_layout_needs_upgrade(wb)
    book.close()


def test_resolved_tab_keeps_history_off_the_active_row(tmp: Path):
    from openpyxl import load_workbook

    store = {"notes": []}
    note = sr.record_note(store, "421966", 1, "Base Fan", "fix this grouping")
    sr.mark_handled(store, note["id"], "grouping rule corrected", when="2026-07-15T12:00:00")
    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, _line_items_store(), store)

    book = load_workbook(str(wb), data_only=False)
    notes = book[sr.BROWSE_SHEET]
    item_col = sr.HEADERS.index("Item") + 1
    note_col = sr.HEADERS.index("Note") + 1
    active_row = next(
        row for row in range(2, notes.max_row + 1)
        if str(notes.cell(row, item_col).value) == "1"
    )
    assert notes.cell(active_row, note_col).value is None

    resolved = book[sr.RESOLVED_SHEET]
    assert [resolved.cell(1, c).value for c in range(1, 7)] == sr.RESOLVED_HEADERS
    assert [resolved.cell(2, c).value for c in range(1, 7)] == [
        "421966", "1", "Base Fan", "fix this grouping",
        "grouping rule corrected", "2026-07-15T12:00:00",
    ]
    book.close()


def test_formula_like_note_text_stays_text(tmp: Path):
    # Note/resolution text starting with "=" must be written as literal text —
    # openpyxl would otherwise store it as a broken Excel formula, and Excel
    # "repairs" the sheet by deleting the cell on open.
    from openpyxl import load_workbook

    store = {"notes": []}
    note = sr.record_note(store, "421966", 1, "Base Fan",
                          "=80240-6773-6056=67,411 is a price sum, capture it")
    sr.mark_handled(store, note["id"], "= now captured as text")
    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, _line_items_store(), store)

    book = load_workbook(str(wb))
    for ws in book.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", (ws.title, cell.coordinate, cell.value)
    resolved = book[sr.RESOLVED_SHEET]
    assert str(resolved.cell(2, 4).value).startswith("=80240")
    book.close()


def test_sales_order_note_is_read_back(tmp: Path):
    from openpyxl import load_workbook

    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, _line_items_store(), {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    browse = book[sr.BROWSE_SHEET]
    item_col = sr.HEADERS.index("Item") + 1
    row = next(
        row for row in range(2, browse.max_row + 1)
        if str(browse.cell(row, item_col).value) == "2"
    )
    note_col = sr.HEADERS.index("Note") + 1
    browse.cell(row, note_col).value = "parsed IVC grouping is correct"
    book.save(str(wb))
    book.close()

    edits = sr.read_edits(wb)
    added = [e for e in edits if e.get("source") == "sales_order"]
    assert len(added) == 1
    assert added[0]["order"] == "421966"
    assert added[0]["item_no"] == "2"
    assert added[0]["note"] == "parsed IVC grouping is correct"
    assert "Inlet Volume Control" in added[0]["item_text"]


def test_sales_order_accepts_note_on_derived_row(tmp: Path):
    from openpyxl import load_workbook

    wb = tmp / "review.xlsx"
    sr.write_workbook(wb, _line_items_store(), {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    browse = book[sr.BROWSE_SHEET]
    item_col = sr.HEADERS.index("Item") + 1
    kind_col = sr.HEADERS.index("Kind") + 1
    row = next(
        row for row in range(2, browse.max_row + 1)
        if not browse.cell(row, item_col).value
        and browse.cell(row, kind_col).value == sr.so_hierarchy.KIND_ATTRIBUTE
    )
    note_col = sr.HEADERS.index("Note") + 1
    browse.cell(row, note_col).value = "derived-row note"
    book.save(str(wb))
    book.close()

    edits = [e for e in sr.read_edits(wb) if e.get("source") == "sales_order"]
    assert len(edits) == 1
    assert edits[0]["item_no"] == ""
    assert edits[0]["note"] == "derived-row note"


def test_sync_records_sales_order_input_and_clears_it(tmp: Path):
    from types import SimpleNamespace

    from openpyxl import load_workbook

    line_items = _line_items_store()
    wb = tmp / "review.xlsx"
    queue = tmp / "so_review_notes.json"
    sr.write_workbook(wb, line_items, {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    browse = book[sr.BROWSE_SHEET]
    item_col = sr.HEADERS.index("Item") + 1
    item_row = next(
        row for row in range(2, browse.max_row + 1)
        if str(browse.cell(row, item_col).value) == "2"
    )
    note_col = sr.HEADERS.index("Note") + 1
    browse.cell(item_row, note_col).value = "sync this note"
    book.save(str(wb))
    book.close()

    old_store_path = sr.REVIEW_STORE_PATH
    old_loader = sr._load_line_items
    sr.REVIEW_STORE_PATH = queue
    sr._load_line_items = lambda: line_items
    try:
        assert sr._cmd_sync(SimpleNamespace(out=str(wb))) == 0
    finally:
        sr.REVIEW_STORE_PATH = old_store_path
        sr._load_line_items = old_loader

    stored = sr.load_store(queue)
    assert len(stored["notes"]) == 1
    assert stored["notes"][0]["note"] == "sync this note"
    rebuilt = load_workbook(str(wb), data_only=False)
    rebuilt_browse = rebuilt[sr.BROWSE_SHEET]
    assert sr.NOTES_ADD_NOTE not in [cell.value for cell in rebuilt_browse[1]]
    assert any(
        str(rebuilt_browse.cell(row, item_col).value) == "2"
        and rebuilt_browse.cell(row, note_col).value == "#1: sync this note"
        for row in range(2, rebuilt_browse.max_row + 1)
    )
    rebuilt.close()


def test_sync_upgrades_legacy_notes_tab_input(tmp: Path):
    from types import SimpleNamespace

    from openpyxl import load_workbook

    line_items = _line_items_store()
    wb = tmp / "review.xlsx"
    queue = tmp / "so_review_notes.json"
    _legacy_picker_workbook(wb, line_items)
    book = load_workbook(str(wb), data_only=False)
    notes = book[sr.NOTES_SHEET]
    item_col = sr.HEADERS.index("Item") + 1
    legacy_header = [cell.value for cell in notes[1]]
    add_note_col = legacy_header.index(sr.NOTES_ADD_NOTE) + 1
    row = next(
        row_no for row_no in range(2, notes.max_row + 1)
        if str(notes.cell(row_no, item_col).value) == "2"
    )
    notes.cell(row, add_note_col).value = "clear this Notes-tab input"
    book.save(str(wb))
    book.close()

    old_store_path = sr.REVIEW_STORE_PATH
    old_loader = sr._load_line_items
    sr.REVIEW_STORE_PATH = queue
    sr._load_line_items = lambda: line_items
    try:
        assert sr._cmd_sync(SimpleNamespace(out=str(wb))) == 0
    finally:
        sr.REVIEW_STORE_PATH = old_store_path
        sr._load_line_items = old_loader

    stored = sr.load_store(queue)
    assert [n["note"] for n in stored["notes"]] == ["clear this Notes-tab input"]
    rebuilt = load_workbook(str(wb), data_only=False)
    assert rebuilt.sheetnames == [sr.BROWSE_SHEET, sr.RESOLVED_SHEET]
    rebuilt_notes = rebuilt[sr.BROWSE_SHEET]
    assert sr.NOTES_ADD_NOTE not in [cell.value for cell in rebuilt_notes[1]]
    rebuilt.close()


def test_sync_upgrades_previous_row_grid_add_note_input(tmp: Path):
    from types import SimpleNamespace

    from openpyxl import load_workbook

    line_items = _line_items_store()
    wb = tmp / "review.xlsx"
    queue = tmp / "so_review_notes.json"
    sr.write_workbook(wb, line_items, {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    review = book[sr.BROWSE_SHEET]
    note_col = sr.HEADERS.index("Note") + 1
    review.insert_cols(note_col + 1, 1)
    review.cell(1, note_col + 1).value = sr.NOTES_ADD_NOTE
    item_col = sr.HEADERS.index("Item") + 1
    item_row = next(
        row for row in range(2, review.max_row + 1)
        if str(review.cell(row, item_col).value) == "2"
    )
    review.cell(item_row, note_col + 1).value = "preserve old row-grid note"
    book.save(str(wb))
    book.close()
    assert sr._browse_layout_needs_upgrade(wb)

    old_store_path = sr.REVIEW_STORE_PATH
    old_loader = sr._load_line_items
    sr.REVIEW_STORE_PATH = queue
    sr._load_line_items = lambda: line_items
    try:
        assert sr._cmd_sync(SimpleNamespace(out=str(wb))) == 0
    finally:
        sr.REVIEW_STORE_PATH = old_store_path
        sr._load_line_items = old_loader

    stored = sr.load_store(queue)
    assert [n["note"] for n in stored["notes"]] == ["preserve old row-grid note"]
    rebuilt = load_workbook(str(wb), data_only=False)
    assert [cell.value for cell in rebuilt[sr.BROWSE_SHEET][1]][:len(sr.HEADERS)] == sr.HEADERS
    assert sr.NOTES_ADD_NOTE not in [cell.value for cell in rebuilt[sr.BROWSE_SHEET][1]]
    rebuilt.close()


def test_legacy_sales_order_formula_overwrite_is_recovered(tmp: Path):
    from openpyxl import load_workbook

    wb = tmp / "review.xlsx"
    _legacy_picker_workbook(wb)
    book = load_workbook(str(wb), data_only=False)
    browse = book[sr.BROWSE_SHEET]
    notes = book[sr.NOTES_SHEET]
    orders = book[sr.ORDERS_SHEET]
    browse["B1"] = "421966"
    browse.cell(sr.BROWSE_HEADER_ROW, 5).value = "Note"
    browse.delete_cols(6, 1)

    order, start_row, row_count = next(
        row for row in orders.iter_rows(min_row=1, max_col=3, values_only=True)
        if str(row[0]) == "421966"
    )
    assert str(order) == "421966"
    item_col = sr.HEADERS.index("Item") + 1
    item_offset = next(
        offset for offset in range(int(row_count))
        if str(notes.cell(int(start_row) + offset, item_col).value) == "2"
    )
    browse.cell(sr.BROWSE_FIRST_ROW + item_offset, 5).value = "recover this old note"
    book.save(str(wb))
    book.close()

    edits = sr.read_edits(wb)
    recovered = [e for e in edits if e.get("source") == "sales_order"]
    assert len(recovered) == 1
    assert recovered[0]["order"] == "421966"
    assert recovered[0]["item_no"] == "2"
    assert recovered[0]["note"] == "recover this old note"


def test_sync_upgrades_legacy_browse_without_pending_notes(tmp: Path):
    from types import SimpleNamespace

    from openpyxl import load_workbook

    line_items = _line_items_store()
    wb = tmp / "review.xlsx"
    queue = tmp / "so_review_notes.json"
    _legacy_picker_workbook(wb, line_items)
    book = load_workbook(str(wb), data_only=False)
    browse = book[sr.BROWSE_SHEET]
    browse.cell(sr.BROWSE_HEADER_ROW, 5).value = "Note"
    browse.delete_cols(6, 1)
    book.save(str(wb))
    book.close()
    assert sr._browse_layout_needs_upgrade(wb)

    old_store_path = sr.REVIEW_STORE_PATH
    old_loader = sr._load_line_items
    sr.REVIEW_STORE_PATH = queue
    sr._load_line_items = lambda: line_items
    try:
        assert sr._cmd_sync(SimpleNamespace(out=str(wb))) == 0
    finally:
        sr.REVIEW_STORE_PATH = old_store_path
        sr._load_line_items = old_loader

    rebuilt = load_workbook(str(wb), data_only=False)
    headers = [rebuilt[sr.BROWSE_SHEET].cell(1, col).value
               for col in range(1, len(sr.HEADERS) + 1)]
    rebuilt.close()
    assert headers == sr.HEADERS


def test_refresh_reports_notes_safe_when_workbook_rewrite_fails(tmp: Path):
    from types import SimpleNamespace

    from openpyxl import load_workbook

    line_items = _line_items_store()
    wb = tmp / "review.xlsx"
    queue = tmp / "so_review_notes.json"
    sr.write_workbook(wb, line_items, {"notes": []})
    book = load_workbook(str(wb), data_only=False)
    notes = book[sr.BROWSE_SHEET]
    kind_col = sr.HEADERS.index("Kind") + 1
    note_col = sr.HEADERS.index("Note") + 1
    row = next(
        row_no for row_no in range(2, notes.max_row + 1)
        if str(notes.cell(row_no, kind_col).value) == sr.so_hierarchy.KIND_ATTRIBUTE
    )
    notes.cell(row, note_col).value = "keep derived-row note even if Excel is open"
    book.save(str(wb))
    book.close()

    old_store_path = sr.REVIEW_STORE_PATH
    old_loader = sr._load_line_items
    old_writer = sr.write_workbook
    sr.REVIEW_STORE_PATH = queue
    sr._load_line_items = lambda: line_items

    def blocked_writer(*_args, **_kwargs):
        raise RuntimeError("Could not write review.xlsx because Excel has it open.")

    sr.write_workbook = blocked_writer
    error = StringIO()
    try:
        with redirect_stderr(error):
            assert sr._run(sr._cmd_refresh, SimpleNamespace(out=str(wb))) == 1
    finally:
        sr.REVIEW_STORE_PATH = old_store_path
        sr._load_line_items = old_loader
        sr.write_workbook = old_writer

    message = error.getvalue()
    assert "All 1 workbook note(s) are safely recorded" in message
    assert "workbook refresh could not finish" in message
    stored = sr.load_store(queue)
    assert len(stored["notes"]) == 1
    assert stored["notes"][0]["note"] == "keep derived-row note even if Excel is open"


def main() -> int:
    passed = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            code = fn.__code__
            if "tmp" in code.co_varnames[:code.co_argcount]:
                with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as d:
                    fn(Path(d))
            else:
                fn()
            print(f"  ok  {name}")
            passed += 1
    print(f"\n{passed} tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
