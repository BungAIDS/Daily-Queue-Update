"""Scan the AutoCAD job folders and record each fan's custom drawings.

For every job folder under AUTOCAD_JOBS_DIR (`<type>/<intermediate>/<job>`) this
records which drawings exist:

  - `<job>-01`  = CW   (clockwise)        }  every order has these; we record
  - `<job>-02`  = CCW  (counter-clockwise)}  whether each is PDF, DWG, or both.
  - any other `<job>-NN` suffix (`-51`, `-35`, …) is a *custom* extra drawing.
    Each distinct suffix seen across all jobs becomes its own yes/no column.

A drawing belongs to a job when its filename starts with `<job>-<digits>` and
its extension is .dwg or .pdf (case-insensitive).

This is a pure-filesystem sweep — no login, no network — so it's fast and can
run in parallel with (or independently of) the cbcinsider backfill. Progress is
written to a JSON store after every folder, so a run that's interrupted (or runs
out of time) resumes where it left off instead of starting over.

Usage:
    python autocad_scan.py                      # sweep every job folder, resumable
    python autocad_scan.py 421314 421388        # just these jobs
    python autocad_scan.py --rescan             # ignore saved progress, redo all
    python autocad_scan.py --recursive          # also look in sub-folders
    python autocad_scan.py --out C:\\path\\dwgs.xlsx --limit 500

Outputs (under BACKLOG_DIR):
    autocad_scan_progress.json   resumable per-job store (source of truth)
    autocad_dwgs.xlsx            the matrix: one row per job, one col per suffix
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Set, Tuple

from config import AUTOCAD_JOBS_DIR, BACKLOG_DIR

log = logging.getLogger("autocad-scan")

DRAWING_EXTS = {".dwg", ".pdf"}
CW_SUFFIX = "01"
CCW_SUFFIX = "02"
STD_SUFFIXES = (CW_SUFFIX, CCW_SUFFIX)

PROGRESS_PATH = BACKLOG_DIR / "autocad_scan_progress.json"
WORKBOOK_PATH = BACKLOG_DIR / "autocad_dwgs.xlsx"

# Real job numbers start a bit below 403425; folders with smaller numeric names
# (or non-numeric names) aren't jobs, so the full sweep skips them. Override with
# --min-job once you know your exact lowest. Kept conservative so it never
# excludes a real job by default.
DEFAULT_MIN_JOB = 400000


# --------------------------------------------------------------------------- #
# Pure logic (no I/O — unit-tested)                                           #
# --------------------------------------------------------------------------- #
def job_key(folder_name: str) -> str:
    """The job number used to match drawing files, taken from a folder name.

    Folders are usually just the number ("421314") but may carry a suffix
    ("421314 ACME CORP"); drawings inside are named with the bare number, so we
    key off the leading digit run. Falls back to the whole name if it has no
    leading digits (keeps odd/legacy job ids working)."""
    m = re.match(r"\d+", folder_name.strip())
    return m.group(0) if m else folder_name.strip()


def parse_drawing(filename: str, job: str) -> Optional[Tuple[str, str]]:
    """If `filename` is a drawing for `job`, return (suffix, ext); else None.

    DWG/PDF names take the form `<job>-<suffix><revletter>` (e.g. "421314-01A",
    "421314-51B"). We match `<job>-<digits>` and capture only the digits, so the
    trailing revision letter is dropped — `-01A` and `-01B` both group under
    suffix "01" (one column, different revisions). Suffix is kept as written
    ("01", "51"). (Capturing the revision letter is a later add — see WORKLOG.)"""
    p = Path(filename)
    ext = p.suffix.lower()
    if ext not in DRAWING_EXTS:
        return None
    m = re.match(rf"{re.escape(job)}-(\d+)", p.stem, re.IGNORECASE)
    if not m:
        return None
    return m.group(1), ext.lstrip(".")


def fmt_exts(exts: Set[str]) -> str:
    """Render a set of extensions as a stable label: PDF, DWG, or PDF+DWG."""
    order = [e for e in ("pdf", "dwg") if e in exts]
    order += sorted(e for e in exts if e not in ("pdf", "dwg"))
    return "+".join(e.upper() for e in order)


def scan_files(filenames: List[str], job: str) -> Dict[str, Set[str]]:
    """Map suffix -> set of extensions for every drawing belonging to `job`."""
    found: Dict[str, Set[str]] = {}
    for name in filenames:
        hit = parse_drawing(name, job)
        if hit:
            suffix, ext = hit
            found.setdefault(suffix, set()).add(ext)
    return found


def build_record(job: str, jtype: str, folder: str, found: Dict[str, Set[str]]) -> Dict[str, Any]:
    """Turn a raw {suffix: exts} scan into the stored per-job record."""
    extras = {
        suffix: fmt_exts(exts)
        for suffix, exts in sorted(found.items(), key=_suffix_sort_key)
        if suffix not in STD_SUFFIXES
    }
    return {
        "job": job,
        "type": jtype,
        "folder": folder,
        "cw": fmt_exts(found.get(CW_SUFFIX, set())),
        "ccw": fmt_exts(found.get(CCW_SUFFIX, set())),
        "extras": extras,
        "missing_std": not (found.get(CW_SUFFIX) or found.get(CCW_SUFFIX)),
        "scanned_at": datetime.now().isoformat(timespec="seconds"),
    }


def _suffix_sort_key(item) -> Tuple[int, str]:
    """Sort suffixes numerically when possible ("35" < "51" < "100")."""
    s = item[0] if isinstance(item, tuple) else item
    return (int(s), s) if s.isdigit() else (10**9, s)


def all_extra_suffixes(records: Dict[str, Dict[str, Any]]) -> List[str]:
    """Sorted union of every extra suffix seen across all job records."""
    seen: Set[str] = set()
    for rec in records.values():
        seen.update(rec.get("extras", {}).keys())
    return [s for s, _ in sorted(((s, None) for s in seen), key=_suffix_sort_key)]


# --------------------------------------------------------------------------- #
# Filesystem walk                                                             #
# --------------------------------------------------------------------------- #
def iter_job_folders(root: Path, min_job: int = DEFAULT_MIN_JOB, max_job: int = 0) -> Iterator[Tuple[str, str, Path]]:
    """Yield (job, type, folder) for every `<type>/<intermediate>/<job>` dir whose
    leaf is a real job number. Non-job folders (year/template/archive dirs, etc.)
    have names that aren't digits or fall below min_job, so they're skipped."""
    for path in root.glob("*/*/*"):
        if not path.is_dir():
            continue
        job = job_key(path.name)
        if _is_real_job(job, min_job, max_job):
            yield job, path.relative_to(root).parts[0], path


def _is_real_job(job: str, min_job: int = DEFAULT_MIN_JOB, max_job: int = 0) -> bool:
    """A real job number is all digits, >= min_job, and (if max_job>0) <= max_job."""
    if not job.isdigit():
        return False
    n = int(job)
    return n >= min_job and (max_job <= 0 or n <= max_job)


def scan_one(job: str, jtype: str, folder: Path, recursive: bool) -> Dict[str, Any]:
    """Scan a single job folder and build its record."""
    globber = folder.rglob("*") if recursive else folder.glob("*")
    names = [f.name for f in globber if f.is_file()]
    return build_record(job, jtype, str(folder), scan_files(names, job))


# --------------------------------------------------------------------------- #
# Progress store (resumable)                                                  #
# --------------------------------------------------------------------------- #
def load_progress() -> Dict[str, Dict[str, Any]]:
    if not PROGRESS_PATH.exists():
        return {}
    try:
        return json.loads(PROGRESS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        log.warning("Could not read %s (%s); starting fresh", PROGRESS_PATH, e)
        return {}


def save_progress(records: Dict[str, Dict[str, Any]]) -> None:
    PROGRESS_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = PROGRESS_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(records, indent=2), encoding="utf-8")
    tmp.replace(PROGRESS_PATH)  # atomic: a crash mid-write never corrupts the store


# --------------------------------------------------------------------------- #
# Excel matrix                                                                #
# --------------------------------------------------------------------------- #
def write_workbook(records: Dict[str, Dict[str, Any]], path: Path) -> Path:
    """Write the per-job matrix: fixed columns + one yes/no column per suffix."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")
    present_fill = PatternFill("solid", fgColor="C6EFCE")  # green: job HAS this drawing
    absent_fill = PatternFill("solid", fgColor="FFC7CE")   # red: it doesn't
    missing_font = Font(color="9C0006", bold=True)         # job missing BOTH -01 and -02
    center = Alignment(horizontal="center")

    suffixes = all_extra_suffixes(records)
    # -01/-02 (CW/CCW) are on essentially every job, so they're not shown; only
    # the custom extra suffixes carry signal. A job missing BOTH still gets
    # flagged (red) as the rare exception.
    fixed = ["Job #", "Type", "Extras", "Folder"]
    headers = fixed + [f"-{s}" for s in suffixes]

    wb = Workbook()
    ws = wb.active
    ws.title = "AutoCAD DWGs"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill

    folder_col = len(fixed)  # "Folder" is the last fixed column
    rows = sorted(records.values(), key=lambda r: r.get("job", ""))
    for i, rec in enumerate(rows, start=2):
        ws.cell(i, 1, rec.get("job", ""))
        ws.cell(i, 2, rec.get("type", ""))
        ws.cell(i, 3, len(rec.get("extras", {})))  # how many custom drawings
        folder = (rec.get("folder") or "").strip()
        fcell = ws.cell(i, folder_col, "Open" if folder else "")
        if folder:
            fcell.hyperlink = folder
            fcell.font = link_font
        extras = rec.get("extras", {})
        for k, s in enumerate(suffixes, start=len(fixed) + 1):
            cell = ws.cell(i, k)
            if s in extras:  # green + a tiny check; red + blank when absent
                cell.value, cell.fill, cell.alignment = "✓", present_fill, center
            else:
                cell.fill = absent_fill
        if rec.get("missing_std"):
            ws.cell(i, 1).font = missing_font  # rare job with neither -01 nor -02

    # AutoFilter + freeze + a light auto-size.
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "B2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 6), 40)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #
def _parse_args(argv: List[str]) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Scan AutoCAD job folders for custom drawings.")
    ap.add_argument("jobs", nargs="*", help="Specific job numbers to scan (default: all).")
    ap.add_argument("--root", default=str(AUTOCAD_JOBS_DIR), help="AutoCAD jobs root.")
    ap.add_argument("--out", default=str(WORKBOOK_PATH), help="Excel output path.")
    ap.add_argument("--recursive", action="store_true", help="Also scan sub-folders of each job.")
    ap.add_argument("--rescan", action="store_true", help="Ignore saved progress; redo every job.")
    ap.add_argument("--limit", type=int, default=0,
                    help="Stop after scanning N folders this run (0 = no limit; "
                         "folders already done on an earlier run don't count).")
    ap.add_argument("--min-job", type=int, default=DEFAULT_MIN_JOB,
                    help=f"Skip folders below this job number on a full sweep (default {DEFAULT_MIN_JOB}).")
    ap.add_argument("--max-job", type=int, default=0, help="Skip folders above this job number (0 = no cap).")
    return ap.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    args = _parse_args(sys.argv[1:] if argv is None else argv)
    root = Path(args.root)
    if not root.exists():
        log.error("AutoCAD jobs root not reachable: %s (is Z: mapped?)", root)
        return 1

    records = {} if args.rescan else load_progress()

    if args.jobs:
        # Explicit jobs: locate each folder the same way find_job_folder does.
        targets: List[Tuple[str, str, Path]] = []
        for job in args.jobs:
            matches = list(root.glob(f"*/*/{job}")) or list(root.glob(f"*/*/{job}*"))
            if matches:
                m = matches[0]
                targets.append((job_key(m.name), m.relative_to(root).parts[0], m))
            else:
                log.warning("  %s: folder not found", job)
        folders: Iterator[Tuple[str, str, Path]] = iter(targets)
    else:
        folders = iter_job_folders(root, min_job=args.min_job, max_job=args.max_job)

    t0 = time.monotonic()
    scanned = 0
    for job, jtype, folder in folders:
        if args.limit and scanned >= args.limit:
            break
        if job in records and not args.rescan and not args.jobs:
            continue  # already scanned on an earlier run
        try:
            records[job] = scan_one(job, jtype, folder, args.recursive)
            scanned += 1
        except OSError as e:
            log.warning("  %s: scan failed (%s)", job, e)
            continue
        if scanned % 200 == 0:
            save_progress(records)
            log.info("  scanned %d (%d total) ...", scanned, len(records))

    save_progress(records)
    try:   # fold these custom-DWG findings into the one master store
        import master_sync
        master_sync.run("autocad")
    except Exception as e:  # noqa: BLE001
        log.warning("Could not sync DWGs to the live master (%s)", e)
    out = write_workbook(records, Path(args.out))
    extras = all_extra_suffixes(records)
    missing = sum(1 for r in records.values() if r.get("missing_std"))
    log.info("Done in %.1fs: %d jobs in store (%d scanned this run).", time.monotonic() - t0, len(records), scanned)
    log.info("  %d distinct extra suffixes: %s", len(extras), ", ".join(f"-{s}" for s in extras) or "(none)")
    if missing:
        log.info("  %d job(s) missing both CW(-01) and CCW(-02) — flagged red in the sheet.", missing)
    log.info("  Wrote %s", out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
