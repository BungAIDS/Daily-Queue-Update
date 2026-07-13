"""Tests for the quote-run detection logic (sales_orders) + report labeling.

No pytest needed — run it directly:

    python test_quote_runs.py

The live cbcinsider/Z: parts can't run off the work machine, but the matching
itself — pid types, file-name patterns, folder scan, archive naming, download
guards, YES (X) labels — is plain Python and is fully checked here. The two
document listings come verbatim from real discovery runs (jobs 421473/421492).
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

from sales_orders import (
    _co_number_for_so_doc, _download_error, _is_run_name, _latest_of_type,
    _run_docs, _run_filename, _run_files_in_folder, SO_TYPE,
)
from excel_writer import _drive_run_label
from templates import QuoteRunContext, parse_quote_run


def _doc(t, rev, fn):
    return (f"href::{t}::{fn}", {"fn": fn, "type": t, "rev": rev})


# Real listing: the quote run is a .txt filed under CBC_Inquiry — the file
# name is the only thing that identifies it.
DOCS_421473 = [
    _doc("CS_SalesOrder", 2, "OrderVerificationReportViewer_04ccbc64-2e6e-4c0d"),
    _doc("CBC_OperatorChecksReports", 1, "421473.docx"),
    _doc("CBC_Inquiry", 1, "909-26-1812_special pricing page 2.rtf"),
    _doc("CBC_Inquiry", 1, "909-26-1812_special pricing page 1.rtf"),
    _doc("CBC_Inquiry", 1, "421473_Quote-34592.pdf"),
    _doc("CBC_Inquiry", 1, "421473_909-26-1604 Qt Run.txt"),
    _doc("CBC_Inquiry", 1, "421473_SelectionShortReport_d71ec6a1-1b49-4fc4-8.pdf"),
    _doc("CBC_Inquiry", 1, "909-26-1604_special pricing page 2.rtf"),
    _doc("CBC_Inquiry", 1, "909-26-1604_special pricing page 1.rtf"),
    _doc("CBC_FanCurve", 1, "Selection Short Report.pdf"),
    _doc("CBC_SalesOrder", 1, "421473 - Sales Order.pdf"),
    _doc("CBC_VendorQuote", 1, "Vendor Quote - Outlet flanged.pdf"),
    _doc("CBC_VendorQuote", 1, "Vendor Quote - Model ECP84407TR-5 complete with"),
    _doc("CBC_VendorQuote", 1, "Vendor Quote - Inlet flanged.pdf"),
]

# Real listing: a design-64 fan whose quote run is the wheel-construction xlsx.
DOCS_421492 = [
    _doc("CS_SalesOrder", 2, "OrderVerificationReportViewer_de47e417-9c6f-4271"),
    _doc("CBC_DrawingTransmital", 1, "Drawing Transmittal for Order Number     421492"),
    _doc("CBC_OperatorChecksReports", 1, "421492.docx"),
    _doc("CBC_Inquiry", 1, "421492_314-26-1647 D64 Wheel Construction (Inner).xlsx"),
    _doc("CBC_Inquiry", 1, "421492_SelectionShortReport_c2095388-8c52-4e60-8.pdf"),
    _doc("CBC_Inquiry", 1, "339-26-1647_special pricing page 2.rtf"),
    _doc("CBC_Inquiry", 1, "339-26-1647_special pricing page 1.rtf"),
    _doc("CBC_FanCurve", 1, "Selection Short Report.pdf"),
    _doc("CBC_SalesOrder", 1, "421492 - Sales Order.pdf"),
    _doc("CBC_VendorQuote", 1, "PO 69069.pdf"),
]


def test_qt_run_txt_under_inquiry():
    runs = _run_docs(DOCS_421473)
    assert [d["fn"] for _, d in runs] == ["421473_909-26-1604 Qt Run.txt"], runs


def test_d64_wheel_construction_xlsx():
    runs = _run_docs(DOCS_421492)
    assert [d["fn"] for _, d in runs] == \
        ["421492_314-26-1647 D64 Wheel Construction (Inner).xlsx"], runs


def test_true_cbc_salesorder_outranks_verification_report():
    # CS_SalesOrder is an Order Verification Report with its own revision
    # counter. It must not outrank the actual CBC_SalesOrder document.
    for docs, fn in ((DOCS_421473, "421473 - Sales Order.pdf"),
                     (DOCS_421492, "421492 - Sales Order.pdf")):
        so = _latest_of_type(docs, SO_TYPE)
        assert so[1]["fn"] == fn and so[1]["rev"] == 1


def test_cs_salesorder_is_never_a_sales_order_fallback():
    docs = [_doc("CS_SalesOrder", 3, "OrderVerificationReportViewer_old.pdf")]
    assert _latest_of_type(docs, SO_TYPE) is None


def test_verification_report_revision_is_not_a_change_order():
    assert _co_number_for_so_doc({"type": "CBC_SalesOrder", "rev": 3}) == 2
    assert _co_number_for_so_doc({"type": "CS_SalesOrder", "rev": 3}) == 0
    assert _co_number_for_so_doc(
        {"type": "CS_SalesOrder", "rev": 3}, {"header_co": 1}
    ) == 0


def test_dedicated_pid_type_sorts_first():
    # HDX fans file the run under its own pid type; it outranks a name match.
    both = [_doc("CBC_Inquiry", 1, "x Qt Run.txt"), _doc("CBC_QuoteRun", 2, "whatever.pdf")]
    assert [d["type"] for _, d in _run_docs(both)] == ["CBC_QuoteRun", "CBC_Inquiry"]
    # Any other non-SO type ending in "run" matches as a pid-type fallback.
    assert _run_docs([_doc("CBC_ConstRun", 1, "z.pdf")])[0][1]["type"] == "CBC_ConstRun"
    # Higher revision wins within a type.
    runs = _run_docs([_doc("CBC_QuoteRun", 1, "a.pdf"), _doc("CBC_QuoteRun", 4, "b.pdf")])
    assert runs[0][1]["rev"] == 4


def test_no_false_positives():
    # A board job with no run-like docs yields nothing; the SO never matches.
    assert _run_docs([_doc("CBC_SalesOrder", 2, "x - Sales Order.pdf"),
                      _doc("CBC_FanCurve", 1, "Selection Short Report.pdf")]) == []
    assert _run_docs([]) == []


def test_name_patterns():
    for s in ("Qt Run", "QT  RUN", "qt run", "QtRun", "quote run", "Quote  Run",
              "D64 Wheel Construction"):
        assert _is_run_name(f"123 {s}.txt"), s
    for s in ("Quote-34592", "special pricing page 1", "Vendor Quote - Inlet flanged",
              "SelectionShortReport", "421473.docx", ""):
        assert not _is_run_name(s), s


def test_run_filename_keeps_name_and_extension():
    assert _run_filename("421473", {"fn": "421473_909-26-1604 Qt Run.txt"}) == \
        "421473_909-26-1604 Qt Run.txt"
    assert _run_filename("421473", {"fn": "909-26-1604 Qt Run.txt"}) == \
        "421473 - 909-26-1604 Qt Run.txt"
    # Windows-illegal characters are replaced; empty fn falls back to a label.
    assert "/" not in _run_filename("421473", {"fn": "a/b Qt Run.txt"})
    assert _run_filename("421473", {"fn": "", "rev": 1}) == "421473 - Quote Run.pdf"


def test_download_guard_pdf_vs_other():
    # PDFs are strictly validated; other run types only reject obvious HTML
    # (the doc server returns the login page with HTTP 200 once the session
    # expires — archiving that would poison the dest.exists() cache).
    assert _download_error(200, b"%PDF-1.7 ...", True) is None
    assert _download_error(200, b"plain qt run text", True) is not None
    assert _download_error(200, b"plain qt run text", False) is None
    assert _download_error(200, b"PK\x03\x04 xlsx zip", False) is None
    assert _download_error(200, b"  <!DOCTYPE html><html>login", False) is not None
    assert _download_error(200, b"<HTML><body>err</body>", False) is not None
    assert _download_error(500, b"%PDF-", False) == "HTTP 500"


def test_extensionless_pdf_is_routed_to_pdf_parser(tmp: Path):
    path = tmp / "421462 - QT RUN-CO"
    path.write_bytes(b"%PDF-1.7\r\nplaceholder")

    assert QuoteRunContext(path).ext == ".pdf"
    with patch("drive_run.parse_drive_run_pdf", return_value={
        "text": "Material: 304 SS", "fields": {"Material": "304 SS"},
    }):
        parsed = parse_quote_run(path)

    assert parsed["template"] == "pdf"
    assert parsed["summary"] == "Material=304 SS"


def test_extensionless_binary_payload_is_not_decoded_as_text(tmp: Path):
    path = tmp / "unknown-run"
    path.write_bytes(b"binary\x00payload\x01\x02Label: fake")
    assert QuoteRunContext(path).text() == ""


def test_extensionless_ole_payload_is_marked_unknown(tmp: Path):
    path = tmp / "418120 - RE CBC ORDER"
    path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 32)

    assert QuoteRunContext(path).ext == ".bin"
    parsed = parse_quote_run(path)
    assert parsed["template"] == "unknown"
    assert parsed["fields"] == {}
    assert parsed["summary"] == ""


def test_folder_scan_recursive(tmp: Path):
    # Runs are often tucked in a subfolder (ENG REF), sometimes with quirky
    # spacing; drawings and other files must not match.
    jobdir = tmp / "420410"
    (jobdir / "ENG REF").mkdir(parents=True)
    (jobdir / "ENG REF" / "420410 qt  run.txt").write_text("x")
    (jobdir / "420410-01A.dwg").write_text("x")
    (jobdir / "420410-51B.pdf").write_text("x")
    (jobdir / "notes.txt").write_text("x")
    # A CAD file named like a run is NOT a run document; nor is an Office temp
    # file (~$...). Both must be excluded even though the name matches.
    (jobdir / "qt run-70 HDX LAYOUT.dwg").write_text("x")
    (jobdir / "qt run.SLDASM").write_text("x")
    (jobdir / "~$420410 qt run.docx").write_text("x")
    assert [h.name for h in _run_files_in_folder(jobdir)] == ["420410 qt  run.txt"]
    # A folder with no run files (or a vanished folder) yields [] quietly.
    assert _run_files_in_folder(tmp / "420410" / "ENG REF") != []  # sanity: scan works on subdirs
    assert _run_files_in_folder(tmp / "does-not-exist") == []


def test_drive_run_label():
    assert _drive_run_label({}) == ""
    assert _drive_run_label({"has_drive_run": False, "drive_run_count": 3}) == ""
    assert _drive_run_label({"has_drive_run": True}) == "YES"
    assert _drive_run_label({"has_drive_run": True, "drive_run_count": 1}) == "YES"
    # More than one match -> flag for review.
    assert _drive_run_label({"has_drive_run": True, "drive_run_count": 2}) == "YES (2)"
    assert _drive_run_label({"has_drive_run": True, "drive_run_count": 4}) == "YES (4)"


def test_quote_run_details_column_surfaces_summary():
    """The 'Quote Run Details' report column renders the parsed drive_run_summary
    (the real engineering fields), while the 'Quote Run' flag stays YES."""
    from openpyxl import Workbook
    import excel_writer as ew

    assert "Quote Run Details" in ew.QUEUE_HEADERS
    det_col = ew.QUEUE_HEADERS.index("Quote Run Details") + 1
    flag_col = ew.QUEUE_HEADERS.index("Quote Run") + 1

    job = {
        "job": "421579", "has_drive_run": True, "drive_run_count": 1,
        "drive_run_pdf": r"Z:\jobs\421579 QT RUN.txt",
        "drive_run_summary": "Size=3300; CFM=22500; SP=18.00; BHP=78.3",
    }
    wb = Workbook(); ws = wb.active
    ew._write_job_row(ws, 2, job)
    assert ws.cell(row=2, column=det_col).value == job["drive_run_summary"]
    assert ws.cell(row=2, column=flag_col).value == "YES"   # flag untouched

    # A job with no run leaves the details cell blank, never errors.
    ws2 = wb.create_sheet("no_run")
    ew._write_job_row(ws2, 2, {"job": "400111"})
    assert (ws2.cell(row=2, column=det_col).value or "") == ""


def main() -> int:
    passed = 0
    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)
        for name, fn in sorted(globals().items()):
            if not name.startswith("test_") or not callable(fn):
                continue
            (fn(tmp) if "tmp" in fn.__code__.co_varnames else fn())
            print(f"  ok  {name}")
            passed += 1
    print(f"\n{passed} tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
