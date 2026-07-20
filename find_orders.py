"""Find orders by what's on their Sales Order — search the line-items store.

The store (fed by the daily run, backfill_orders.py, and line_items_scan.py)
holds every captured Sales Order line item in three forms: the verbatim text,
a normalized form (abbreviations expanded, qty/prices stripped), and canonical
feature tags. Search hits all three, so "stainless" finds the order whose SO
says "SS SHAFT SLEEVE".

    python find_orders.py shaft seal           # orders matching BOTH terms
    python find_orders.py --any teflon viton   # ...ANY of the terms
    python find_orders.py --tag "SHAFT SEAL"   # by canonical tag
    python find_orders.py cermic felt --fuzzy  # typo-tolerant (ratio 0.84;
                                               # give --fuzzy AFTER the terms,
                                               # or as --fuzzy=0.8)
    python find_orders.py shaft seal --dwg     # ...only jobs the AutoCAD scan
                                               # found custom drawings for
    python find_orders.py --like 421314        # rank the backlog by how much of
                                               # this job's SO each order shares
    python find_orders.py --like 421314 --dwg  # DWG-reuse candidates: similar
                                               # jobs that have custom drawings
    python find_orders.py --job 421314         # one job's stored items
    python find_orders.py --list-tags          # live tag vocabulary + counts
    python find_orders.py --audit-untagged     # names current rules still miss
    python find_orders.py --audit-review       # rows marked for human/template review
    python find_orders.py shaft seal --xlsx    # ...write matches to Excel too
    python find_orders.py --xlsx               # full inventory workbook:
                                               # "Line Items" (row per item,
                                               # AutoFilter) + "Feature Matrix"
                                               # (jobs x tags, green ✓ = has
                                               # it / red = doesn't — like the
                                               # AutoCAD DWG matrix)

Terms are case-insensitive substrings (multi-word terms must appear as a
phrase in the normalized text; separate words are separate ANDed terms).

Every result view also shows the job's custom AutoCAD drawings (from the DWG
scan store), so a feature search doubles as "who already drew this?"."""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import BACKLOG_DIR, REUSE_MIN_SCORE, REUSE_TOP
import autocad_scan
import line_items as li

log = logging.getLogger("find-orders")

XLSX_DEFAULT = BACKLOG_DIR / "line_items.xlsx"


def _store_stats(store: Dict[str, Any]) -> str:
    n_items = sum(len(r.get("items") or []) for r in store["jobs"].values())
    return f"{len(store['jobs'])} order(s), {n_items} item(s) in {li.store_path()}"


# --- custom-DWG awareness (joins the AutoCAD scan store into every view) ---- #
def _dwg_label(extras: Dict[str, str] | None) -> str:
    """'-07 (DWG), -51 (PDF+DWG)' from a scan record's extras dict."""
    return ", ".join(f"-{s} ({fmt})" if fmt else f"-{s}"
                     for s, fmt in (extras or {}).items())


def attach_dwg(hits: List[Dict[str, Any]], dwg: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Annotate each hit with its AutoCAD scan record (custom drawing suffixes +
    job folder) so every search answers "does someone already have this drawn?"."""
    for h in hits:
        rec = dwg.get(h["job"]) or {}
        h["dwg_extras"] = rec.get("extras") or {}
        h["dwg_folder"] = rec.get("folder", "")
        h["dwg_scanned"] = bool(rec)
    return hits


def _job_num(job: str) -> int:
    return int(job) if str(job).isdigit() else -1


def _item_tags(items: List[Dict[str, Any]] | None) -> set:
    return {t for it in items or [] for t in it.get("tags") or []}


def _item_norms(items: List[Dict[str, Any]] | None) -> set:
    return {it.get("norm", "") for it in items or []} - {""}


_CONTEXT_KEYS = (
    "arrangement", "so_design_desc", "so_size", "so_arrangement",
    "so_motor_pos", "so_class", "so_rotation",
)
_NOISE_WORDS = (
    "admin", "additional", "co number", "customer", "document",
    "inquiry", "job number", "mark", "note", "rep reference", "review",
    "ship", "shipping", "ship to", "source", "warranty",
)

# Tags remain the broad feature signal. Structured engineering facts are more
# specific, while exact normalized lines are deliberately only a tie-breaker.
TAG_WEIGHT = 1.0
CONTEXT_WEIGHT = 1.5
FACT_WEIGHT = 2.0
NORM_WEIGHT = 0.5


def _fact_text(value: Any) -> str:
    if isinstance(value, (list, tuple, set)):
        value = " | ".join(str(x) for x in value)
    return " ".join(str(value or "").upper().split())


def _is_noise_fact(key: str, value: Any) -> bool:
    text = f"{key}={_fact_text(value)}".lower()
    return any(word in text for word in _NOISE_WORDS)


def _item_facts(items: List[Dict[str, Any]] | None) -> set:
    """Return engineering component/attribute facts, excluding parser noise."""
    facts = set()
    for item in items or []:
        section = _fact_text(item.get("section"))
        document_fact = _fact_text(item.get("document_fact"))
        if (_is_noise_fact("section", section)
                or (document_fact and _is_noise_fact("document_fact", document_fact))):
            continue
        if section:
            facts.add(f"section={section}")
        attrs = item.get("attributes")
        if not isinstance(attrs, dict):
            continue
        for key, value in attrs.items():
            key_text = _fact_text(key)
            value_text = _fact_text(value)
            if not key_text or not value_text or _is_noise_fact(key_text, value_text):
                continue
            facts.add(f"{key_text}={value_text}")
    return facts


def _job_context(rec: Dict[str, Any]) -> set:
    context = set()
    for key in _CONTEXT_KEYS:
        value = rec.get(key)
        if value is None or value is False or not _fact_text(value):
            continue
        if not _is_noise_fact(key, value):
            context.add(f"{key}={_fact_text(value)}")
    if rec.get("parts_only") is True:
        context.add("parts_only=TRUE")
    return context


def _item_facts_with_context(rec: Dict[str, Any]) -> set:
    return _item_facts(rec.get("items") or [])


def build_index(store: Dict[str, Any],
                dwg: Dict[str, Dict[str, Any]] | None = None) -> Dict[str, Any]:
    """One pass over the store, including rarity-weighted engineering facts."""
    job_tags: Dict[str, set] = {}
    job_norms: Dict[str, set] = {}
    job_context: Dict[str, set] = {}
    job_facts: Dict[str, set] = {}
    tag_df: Dict[str, int] = {}
    norm_df: Dict[str, int] = {}
    context_df: Dict[str, int] = {}
    fact_df: Dict[str, int] = {}
    jobs = store.get("jobs") or {}
    for j, rec in jobs.items():
        items = rec.get("items") or []
        job_tags[j], job_norms[j] = _item_tags(items), _item_norms(items)
        job_context[j] = _job_context(rec)
        job_facts[j] = _item_facts_with_context(rec)
        for t in job_tags[j]:
            tag_df[t] = tag_df.get(t, 0) + 1
        for n in job_norms[j]:
            norm_df[n] = norm_df.get(n, 0) + 1
        for c in job_context[j]:
            context_df[c] = context_df.get(c, 0) + 1
        for f in job_facts[j]:
            fact_df[f] = fact_df.get(f, 0) + 1
    return {"jobs": jobs, "job_tags": job_tags, "job_norms": job_norms,
            "job_context": job_context, "job_facts": job_facts,
            "tag_df": tag_df, "norm_df": norm_df,
            "context_df": context_df, "fact_df": fact_df,
            "dwg": dwg or {}}


def similar_to_items(index: Dict[str, Any], items: List[Dict[str, Any]] | None,
                     exclude_job: str = "", top: int = 15,
                     require_dwg: bool = False,
                     context: Dict[str, Any] | None = None) -> List[Dict[str, Any]]:
    """Rank the indexed orders by how much of `items` (one order's SO) they share.

    Rarity-weighted overlap uses tags, job/document context, and engineering
    component facts. Exact normalized lines remain a lower-weight tie-breaker.
    With `require_dwg` only jobs whose AutoCAD scan found custom drawings are
    kept: the DWG-reuse shortlist for a new order."""
    t_tags, t_norms = _item_tags(items), _item_norms(items)
    t_facts = _item_facts(items)
    t_context = set()
    if context:
        t_context.update(_job_context(context))
    for item in items or []:
        # Unstored orders can carry context on each parsed item.
        context = item.get("_order_context")
        if isinstance(context, dict):
            t_context.update(_job_context(context))
    tag_df, norm_df = index.get("tag_df", {}), index.get("norm_df", {})
    context_df, fact_df = index.get("context_df", {}), index.get("fact_df", {})
    dwg = index.get("dwg", {})
    out: List[Dict[str, Any]] = []
    for j, rec in index["jobs"].items():
        if j == str(exclude_job):
            continue
        extras = (dwg.get(j) or {}).get("extras") or {}
        if require_dwg and not extras:
            continue
        shared_tags = t_tags & index["job_tags"][j]
        shared_lines = t_norms & index["job_norms"][j]
        shared_context = t_context & index.get("job_context", {}).get(j, set())
        shared_facts = t_facts & index.get("job_facts", {}).get(j, set())
        if not shared_tags and not shared_lines and not shared_context and not shared_facts:
            continue
        score = (TAG_WEIGHT * sum(1.0 / tag_df[t] for t in shared_tags)
                 + CONTEXT_WEIGHT * sum(1.0 / context_df[c] for c in shared_context)
                 + FACT_WEIGHT * sum(1.0 / fact_df[f] for f in shared_facts)
                 + NORM_WEIGHT * sum(1.0 / norm_df[n] for n in shared_lines))
        out.append({
            "job": j, "customer": rec.get("customer", ""),
            "co_number": rec.get("co_number"), "so_pdf": rec.get("so_pdf", ""),
            "score": score,
            "shared_tags": sorted(shared_tags, key=lambda t: (tag_df[t], t)),
            "shared_lines": sorted(shared_lines, key=lambda n: (norm_df[n], n)),
            "shared_context": sorted(shared_context),
            "shared_facts": sorted(shared_facts),
            "dwg_extras": extras, "dwg_folder": (dwg.get(j) or {}).get("folder", ""),
        })
    out.sort(key=lambda r: (-r["score"], -_job_num(r["job"])))
    return out[:top] if top and top > 0 else out


def similar_jobs(store: Dict[str, Any], job: str,
                 dwg: Dict[str, Dict[str, Any]] | None = None,
                 top: int = 15, require_dwg: bool = False) -> Optional[List[Dict[str, Any]]]:
    """`similar_to_items` for an order already in the store (the --like CLI).
    Returns None when `job` has no stored line items."""
    job = str(job)
    target = (store.get("jobs") or {}).get(job)
    if not target or not target.get("items"):
        return None
    return similar_to_items(build_index(store, dwg), target["items"],
                            exclude_job=job, top=top, require_dwg=require_dwg,
                            context=target)


# --- the live suggester: compact reuse shortlist carried on each job dict --- #
def reuse_suggestions(index: Dict[str, Any], items: List[Dict[str, Any]] | None,
                      exclude_job: str = "", min_score: float | None = None,
                      top: int | None = None) -> List[Dict[str, Any]]:
    """The DWG-reuse shortlist for one order, trimmed for storage on its job
    dict (live_master carries every job wholesale, so keep it lean): only
    custom-DWG jobs scoring >= min_score, essentials-only fields."""
    if min_score is None:
        min_score = REUSE_MIN_SCORE
    if top is None:
        top = REUSE_TOP
    res = similar_to_items(index, items, exclude_job=exclude_job,
                           top=top, require_dwg=True)
    return [{
        "job": r["job"], "customer": r["customer"], "score": round(r["score"], 2),
        "suffixes": list(r["dwg_extras"]), "dwg": _dwg_label(r["dwg_extras"]),
        "folder": r["dwg_folder"], "lines": r["shared_lines"][:3],
        "tags": r["shared_tags"][:6],
    } for r in res if r["score"] >= min_score]


def reuse_label(sugg: List[Dict[str, Any]] | None) -> str:
    """The one-cell column form: top candidate + its custom suffixes, e.g.
    '421100 (-07,-51) +2' — details live in the hover note / notification."""
    if not sugg:
        return ""
    r0 = sugg[0]
    sufs = ",".join(f"-{s}" for s in r0.get("suffixes") or [])
    more = f" +{len(sugg) - 1}" if len(sugg) > 1 else ""
    return f"{r0['job']} ({sufs}){more}"


def reuse_note(sugg: List[Dict[str, Any]] | None) -> str:
    """The hover-comment form: every candidate with its drawings, the SO lines
    it shares with this order, and its CAD folder."""
    lines: List[str] = []
    for r in sugg or []:
        cust = f"  {r['customer']}" if r.get("customer") else ""
        lines.append(f"{r['job']}{cust} — score {r['score']:.2f}")
        if r.get("dwg"):
            lines.append(f"  custom DWGs: {r['dwg']}")
        for n in r.get("lines") or []:
            lines.append(f"  = {n}")
        if r.get("folder"):
            lines.append(f"  {r['folder']}")
    return "\n".join(lines)


def _print_similar(job: str, target: Dict[str, Any], results: List[Dict[str, Any]],
                   require_dwg: bool, dwg_loaded: bool) -> None:
    cust = f"  {target['customer']}" if target.get("customer") else ""
    n_items = len(target.get("items") or [])
    what = "job(s) with custom DWGs" if require_dwg else "job(s)"
    print(f"Orders most like {job}{cust}  ({n_items} stored line item(s))")
    print(f"{len(results)} similar {what}; rare shared features score highest, "
          "identical lines count double.")
    if not dwg_loaded:
        print("(No AutoCAD scan store found — run autocad_scan.py to see custom DWGs here.)")
    for i, r in enumerate(results, start=1):
        co = f"  CO#{r['co_number']}" if r.get("co_number") else ""
        cust = f"  {r['customer']}" if r.get("customer") else ""
        print(f"\n{i:3d}. {r['job']}{cust}{co}   score {r['score']:.2f}")
        if r.get("dwg_extras"):
            print(f"       custom DWGs: {_dwg_label(r['dwg_extras'])}")
        if r.get("dwg_folder"):
            print(f"       folder: {r['dwg_folder']}")
        for n in r["shared_lines"][:6]:
            print(f"       = {n}")
        if r["shared_tags"]:
            print(f"       shared tags: {', '.join(r['shared_tags'][:10])}")


def _print_hits(hits: List[Dict[str, Any]], terms: List[str] | None = None,
                all_details: bool = False) -> None:
    """Print matched orders. Detail lines (vendor, motor specs, ...) are shown
    when they're what the search term actually hit — or all of them with
    all_details (the --job view)."""
    terms_u = [t.upper() for t in (terms or [])]
    for h in hits:
        co = f"  CO#{h['co_number']}" if h.get("co_number") else ""
        cust = f"  {h['customer']}" if h.get("customer") else ""
        print(f"\n{h['job']}{cust}{co}   (scanned {h.get('scanned_at', '')[:10]})")
        if h.get("so_pdf"):
            print(f"    SO: {h['so_pdf']}")
        if h.get("dwg_extras"):
            folder = f"   {h['dwg_folder']}" if h.get("dwg_folder") else ""
            print(f"    custom DWGs: {_dwg_label(h['dwg_extras'])}{folder}")
        elif h.get("dwg_scanned"):
            print("    custom DWGs: none (standard drawings only)")
        for it in h["matches"]:
            tags = ", ".join(it.get("tags") or []) or "-"
            print(f"    [{tags}]  {it['raw']}")
            review_flags = _review_flags_label(it)
            if review_flags:
                print(f"        review: {review_flags}")
            attrs = _attrs_label(it)
            if attrs:
                print(f"        attrs: {attrs}")
            for d in it.get("details") or []:
                if all_details or any(t in d.upper() for t in terms_u):
                    print(f"        · {d}")


def _attrs_label(item: Dict[str, Any]) -> str:
    attrs = item.get("attributes") or {}
    if not isinstance(attrs, dict):
        return ""
    keys = [
        "inquiry_num", "note_type", "component", "component_review",
        "warranty_type", "warranty_duration", "warranty_scope", "warranty_start",
        "warranty_source",
        "used_on", "used_on_review", "vendor", "product",
        "split_type",
        "drawing_type", "drawing_scope",
        "motor_enclosure", "motor_mounting", "motor_explosion_class",
        "motor_explosion_groups", "motor_explosion_division", "motor_base",
        "motor_feature", "motor_conduit_box_location", "motor_conduit_box_position",
        "motor_conduit_box_orientation",
        "motor_warranty",
        "motor_nameplate", "motor_nameplate_action",
        "nameplate_type", "nameplate_mount_location", "nameplate_mounting",
        "flange_scope", "flange_type", "flange_location",
        "flex_connector_type",
        "coupling_subcategory", "coupling_type", "fit", "cover_type", "set_screws",
        "leakage_class", "duty_rating",
        "temperature_service", "temperature_direction", "temperature_rating", "grease_type",
        "special_construction_type", "special_construction_scope",
        "special_construction_detail", "effective_diameter_percent", "welding_code",
        "wheel_feature", "wheel_effective_diameter_percent", "wheel_hub_construction",
        "wheel_hub_bore", "wheel_bore",
        "unitary_base_type", "unitary_base_size", "unitary_base_detail", "unitary_base_clearance",
        "vfd_context", "vfd_supplied_by", "motor_vfd_suitability",
        "motor_vfd_operation", "motor_vfd_speed_range",
        "vibration_isolation_type", "isolation_deflection", "isolation_frame",
        "material", "material_grade", "material_scope", "material_treatment",
        "component_material", "component_material_grade", "component_material_scope",
        "drain_type", "drain_closure", "drain_detail",
        "coating_context", "coating_category", "coating_scope", "coating_process",
        "coating_state", "coating_type", "coating_color", "alternate_coating_color", "coats",
        "guard_type", "guard_material", "tach_hole", "tach_hole_location",
        "weather_cover_type", "weather_cover_scope", "weather_cover_used_on",
        "weather_cover_model", "weather_cover_feature",
        "screen_subcategory", "screen_feature", "screen_diameter", "screen_size",
        "shaft_cooler", "shaft_cooler_type", "shaft_cooler_construction",
        "shaft_seal_type", "shaft_sleeve", "shaft_sleeve_type",
        "silencer_subcategory", "silencer_used_on", "silencer_model",
        "silencer_noise_target", "pressure_drop", "silencer_feature",
        "spark_resistant", "spark_resistant_type",
        "spare_part_type", "spare_part_component",
        "testing_type", "testing_status", "testing_duration", "testing_measurements",
        "testing_voltage",
        "shipping_state", "shipping_method", "shipping_instruction", "shipping_scope",
        "balance_type", "balance_grade",
        "bearing_type", "bearing_bore",
        "manufacturer", "model", "size", "operation", "supplied_by", "mounting", "fail_power", "fail_signal",
        "drive_subcategory", "belt_qty", "belt", "selected_drive", "drive_sheave", "drive_bushing",
        "driven_sheave", "driven_bushing", "drive_sheave_bushing", "driven_sheave_bushing",
        "actual_sf", "actual_cd", "service_factor", "center_distance_range",
    ]
    return "; ".join(f"{k}={attrs[k]}" for k in keys if attrs.get(k))


def _review_flags_label(item: Dict[str, Any]) -> str:
    return "; ".join(str(x) for x in item.get("review_flags") or [])


def _print_untagged_audit(store: Dict[str, Any], limit: int) -> None:
    rows = li.audit_untagged(store, limit=limit)
    if not rows:
        print(f"Current rules tag every stored normalized item. ({_store_stats(store)})")
        return
    w = max(len(r["norm"]) for r in rows)
    print(f"Top {len(rows)} normalized item name(s) current rules still do not tag")
    print(f"({'after re-deriving names/tags from stored raw text'}; {_store_stats(store)})")
    print(f"{'COUNT':>5}  {'NORMALIZED'.ljust(min(w, 72))}  JOBS  AI")
    for r in rows:
        norm = r["norm"]
        shown = norm if len(norm) <= 72 else norm[:69] + "..."
        jobs = ",".join(r.get("jobs") or [])
        ai = ", ".join(r.get("ai_tags") or [])
        print(f"{r['count']:5d}  {shown.ljust(min(w, 72))}  {jobs}  {ai}")


def _print_review_audit(store: Dict[str, Any], limit: int, tag: str = "") -> None:
    rows = li.audit_review(store, limit=limit, tag=tag)
    if not rows:
        tag_msg = f" for tag {tag!r}" if tag else ""
        print(f"No line-item review candidates{tag_msg}. ({_store_stats(store)})")
        return
    w = max(len(r["norm"]) for r in rows)
    tag_msg = f" tagged {tag!r}" if tag else ""
    print(f"Top {len(rows)} line-item template/review candidate(s){tag_msg}")
    print(f"({'after re-deriving names/tags from stored raw text'}; {_store_stats(store)})")
    print(f"{'COUNT':>5}  {'NORMALIZED'.ljust(min(w, 64))}  TAGS  REVIEW FLAGS  JOBS")
    for r in rows:
        norm = r["norm"]
        shown = norm if len(norm) <= 64 else norm[:61] + "..."
        tags = ", ".join(r.get("tags") or []) or "-"
        flags = "; ".join(r.get("review_flags") or [])
        jobs = ",".join(r.get("jobs") or [])
        print(f"{r['count']:5d}  {shown.ljust(min(w, 64))}  {tags}  {flags}  {jobs}")


def write_xlsx(hits: List[Dict[str, Any]], path: Path,
               store: Dict[str, Any] | None = None) -> Path:
    """Two sheets: "Line Items" — one row per (job, item) with AutoFilter, the
    filter-in-Excel view — and "Feature Matrix" — one row per job, one column
    per canonical tag, green ✓ when the order has that feature and red when it
    doesn't (same look as the AutoCAD DWG matrix). When `store` is given the
    matrix shows each job's FULL feature profile even if the hits were
    filtered by a search."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")

    headers = ["Job #", "Customer", "CO#", "Tags", "Review Flags", "Attributes",
               "Item (as printed)", "Normalized", "Details", "Qty", "Price", "Section",
               "SO PDF", "Custom DWGs"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Line Items"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for h in hits:
        co = f"CO#{h['co_number']}" if h.get("co_number") else ""
        for it in h["matches"]:
            ws.cell(row, 1, h["job"])
            ws.cell(row, 2, h.get("customer", ""))
            ws.cell(row, 3, co)
            ws.cell(row, 4, ", ".join(it.get("tags") or []))
            ws.cell(row, 5, _review_flags_label(it))
            ws.cell(row, 6, _attrs_label(it))
            ws.cell(row, 7, it.get("raw", ""))
            ws.cell(row, 8, it.get("norm", ""))
            ws.cell(row, 9, " ; ".join(it.get("details") or []))
            ws.cell(row, 10, it.get("qty", ""))
            ws.cell(row, 11, it.get("price", ""))
            ws.cell(row, 12, it.get("section", ""))
            if h.get("so_pdf"):
                cell = ws.cell(row, 13, "Open")
                cell.hyperlink = h["so_pdf"]
                cell.font = link_font
            ws.cell(row, 14, _dwg_label(h.get("dwg_extras")))
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = "B2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 6), 70)

    # ---- Feature Matrix: jobs x tags, green ✓ / red (like the DWG matrix) --
    has_fill = PatternFill("solid", fgColor="C6EFCE")  # green: order HAS the feature
    no_fill = PatternFill("solid", fgColor="FFC7CE")   # red: it doesn't
    center = Alignment(horizontal="center")

    # Each job's full item set — from the store when given, so a search-
    # filtered workbook still shows the whole feature profile per matched job.
    items_by_job: Dict[str, List[Dict[str, Any]]] = {}
    for h in hits:
        rec = (store or {}).get("jobs", {}).get(h["job"]) or {}
        items_by_job[h["job"]] = rec.get("items") or h["matches"]
    tags_by_job = {job: {t for it in items for t in it.get("tags") or []}
                   for job, items in items_by_job.items()}
    counts: Dict[str, int] = {}
    for tags in tags_by_job.values():
        for t in tags:
            counts[t] = counts.get(t, 0) + 1
    all_tags = sorted(counts, key=lambda t: (-counts[t], t))  # most common left

    mx = wb.create_sheet("Feature Matrix")
    fixed = ["Job #", "Customer", "CO#", "Items", "Custom DWGs"]
    for c, h in enumerate(fixed, start=1):
        cell = mx.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
    for k, t in enumerate(all_tags, start=len(fixed) + 1):
        cell = mx.cell(1, k, t)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
    if all_tags:  # rotated headers keep the tag columns narrow
        mx.row_dimensions[1].height = min(150, 14 + 5.4 * max(len(t) for t in all_tags))

    for i, h in enumerate(hits, start=2):
        cell = mx.cell(i, 1, h["job"])
        if h.get("so_pdf"):
            cell.hyperlink = h["so_pdf"]
            cell.font = link_font
        mx.cell(i, 2, h.get("customer", ""))
        mx.cell(i, 3, f"CO#{h['co_number']}" if h.get("co_number") else "")
        mx.cell(i, 4, len(items_by_job[h["job"]]))
        dwg_label = _dwg_label(h.get("dwg_extras"))
        cell = mx.cell(i, 5, dwg_label)
        if dwg_label and h.get("dwg_folder"):
            cell.hyperlink = h["dwg_folder"]  # click through to the job's CAD folder
            cell.font = link_font
        job_tags = tags_by_job[h["job"]]
        for k, t in enumerate(all_tags, start=len(fixed) + 1):
            cell = mx.cell(i, k)
            if t in job_tags:
                cell.value, cell.fill, cell.alignment = "✓", has_fill, center
            else:
                cell.fill = no_fill

    if hits:
        mx.auto_filter.ref = f"A1:{get_column_letter(len(fixed) + len(all_tags))}{len(hits) + 1}"
    mx.freeze_panes = "B2"
    for c in range(1, len(fixed) + 1):
        letter = get_column_letter(c)
        width = max((len(str(cell.value)) for cell in mx[letter] if cell.value is not None), default=8)
        mx.column_dimensions[letter].width = min(max(width + 2, 6), 40)
    for k in range(len(fixed) + 1, len(fixed) + len(all_tags) + 1):
        mx.column_dimensions[get_column_letter(k)].width = 4.5

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def main(argv: Optional[List[str]] = None) -> int:
    logging.basicConfig(level=logging.WARNING, format="[%(levelname)s] %(message)s")
    ap = argparse.ArgumentParser(description="Find orders by their Sales Order line items.")
    ap.add_argument("terms", nargs="*", help="Search terms (all must match; see --any).")
    ap.add_argument("--any", action="store_true", help="Match ANY term instead of all.")
    ap.add_argument("--tag", default="", help='Filter by canonical tag (e.g. "SHAFT SEAL").')
    ap.add_argument("--fuzzy", nargs="?", const=0.84, default=0.0, type=float, metavar="RATIO",
                    help="Also accept close (typo) matches; optional ratio 0-1, default 0.84.")
    ap.add_argument("--job", default="", help="Show the stored items for one job number.")
    ap.add_argument("--like", default="", metavar="JOB",
                    help="Rank the backlog by how much of JOB's Sales Order each other order "
                         "shares (rare features / identical lines score highest). "
                         "Add --dwg for the DWG-reuse shortlist.")
    ap.add_argument("--dwg", action="store_true",
                    help="Only jobs whose AutoCAD scan found custom drawings.")
    ap.add_argument("--top", type=int, default=15,
                    help="How many similar jobs --like shows (0 = all; default 15).")
    ap.add_argument("--list-tags", action="store_true", help="List the tag vocabulary with counts.")
    ap.add_argument("--list-inquiries", action="store_true",
                    help="List parsed inquiry numbers with order/item counts.")
    ap.add_argument("--audit-untagged", action="store_true",
                    help="List the most common normalized item names current rules still do not tag.")
    ap.add_argument("--audit-review", action="store_true",
                    help="List line-item templates marked for human/rule review. Use --tag to narrow it.")
    ap.add_argument("--audit-limit", type=int, default=50,
                    help="How many audit rows to print (0 = all; default 50).")
    ap.add_argument("--xlsx", nargs="?", const=str(XLSX_DEFAULT), default="", metavar="PATH",
                    help=f"Write the result (or, with no filters, the full inventory) to Excel: "
                         f"a per-item sheet + a green-✓/red jobs-x-features matrix "
                         f"(default {XLSX_DEFAULT}).")
    args = ap.parse_args(sys.argv[1:] if argv is None else argv)

    store = li.load_store()
    if not store["jobs"]:
        print(f"The line-items store is empty ({li.store_path()}).\n"
              "It fills from the daily run as orders are parsed — or build it now from\n"
              "the archived Sales Orders:  python line_items_scan.py")
        return 1

    if args.list_tags:
        counts = li.tag_counts(store)
        if not counts:
            print("No tagged items yet — run `python line_items_scan.py --ai` or extend the rules.")
            return 0
        w = max(len(t) for t, _, _ in counts)
        print(f"{'TAG'.ljust(w)}  ORDERS  ITEMS")
        for t, n_jobs, n_items in counts:
            print(f"{t.ljust(w)}  {n_jobs:6d}  {n_items:5d}")
        print(f"\n({_store_stats(store)})")
        return 0

    if args.list_inquiries:
        counts = li.inquiry_counts(store)
        if not counts:
            print("No inquiry numbers parsed yet.")
            return 0
        w = max(len(n) for n, _, _, _ in counts)
        print(f"{'INQUIRY #'.ljust(w)}  ORDERS  ITEMS  JOBS")
        for num, n_jobs, n_items, jobs in counts:
            print(f"{num.ljust(w)}  {n_jobs:6d}  {n_items:5d}  {', '.join(jobs[:12])}")
        print(f"\n({_store_stats(store)})")
        return 0

    if args.audit_untagged:
        _print_untagged_audit(store, args.audit_limit)
        return 0

    if args.audit_review:
        _print_review_audit(store, args.audit_limit, tag=args.tag)
        return 0

    dwg = autocad_scan.load_progress()  # read-only: custom-DWG suffixes + job folders

    if args.like:
        results = similar_jobs(store, args.like, dwg=dwg, top=args.top, require_dwg=args.dwg)
        if results is None:
            print(f"Job {args.like} has no stored line items ({_store_stats(store)}).\n"
                  f"Backfill it first:  python backfill_orders.py {args.like}")
            return 1
        _print_similar(args.like, store["jobs"][str(args.like)], results,
                       require_dwg=args.dwg, dwg_loaded=bool(dwg))
        return 0

    if args.job:
        rec = store["jobs"].get(args.job)
        if not rec:
            print(f"Job {args.job} is not in the store ({_store_stats(store)}).")
            return 1
        hits = [{"job": args.job, **{k: rec.get(k) for k in
                 ("customer", "co_number", "so_pdf", "scanned_at")},
                 "matches": rec.get("items") or []}]
        _print_hits(attach_dwg(hits, dwg), all_details=True)
        return 0

    hits = attach_dwg(li.search(store, args.terms, any_mode=args.any,
                                tag=args.tag, fuzzy=args.fuzzy), dwg)
    if args.dwg:
        hits = [h for h in hits if h["dwg_extras"]]
    if args.terms or args.tag:
        what = " ".join(args.terms) + (f" [tag={args.tag}]" if args.tag else "")
        dwg_msg = " with custom DWGs" if args.dwg else ""
        print(f"{len(hits)} order(s){dwg_msg} match {what!r}   ({_store_stats(store)})")
        _print_hits(hits, terms=args.terms)
    elif not args.xlsx:
        print(f"Nothing to search for — give terms, --tag, --job, --like, --list-tags or --xlsx.\n"
              f"({_store_stats(store)})")
        return 1

    if args.xlsx:
        out = write_xlsx(hits, Path(args.xlsx), store)
        n = sum(len(h["matches"]) for h in hits)
        print(f"\nWrote {n} item row(s) across {len(hits)} order(s) "
              f"(+ the green-check/red Feature Matrix tab) -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
