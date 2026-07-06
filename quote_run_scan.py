"""Sweep ALL AutoCAD job folders for quote runs and parse each one.

A pure-filesystem pass over every job folder under AUTOCAD_JOBS_DIR (no login,
no browser): it finds every quote/construction-run file (the same name patterns
the daily run uses — `*qt run*`, `*quote run*`, `*d64 wheel construction*`,
searched recursively, so copies tucked in subfolders like `ENG REF\\` are
caught, while superseded `history\\` / `hist\\` copies are skipped), runs
each through the template collection in
`templates.py`, and writes an inventory of what was pulled.

Because it only reads the filesystem it's fast and fully resumable — progress is
saved after every batch, so an interrupted run picks up where it left off.

What it canNOT see: a run that exists only in an order's online cbcinsider
documents and was never saved to the job folder. Use `backfill_orders.py` (which
logs in and reads the documents too) for that. In practice most runs live in the
folder, so this catches the large majority.

Usage:
    python quote_run_scan.py                     # sweep every job folder, resumable
    python quote_run_scan.py 421473 421572       # just these jobs
    python quote_run_scan.py --range 420000 421999
    python quote_run_scan.py --list jobs.txt     # job numbers from a file
    python quote_run_scan.py --rescan            # ignore saved progress, redo all
    python quote_run_scan.py --reparse-attention # redo only jobs flagged "needs attention"
    python quote_run_scan.py --limit 500

Outputs (under BACKLOG_DIR):
    quote_run_scan_progress.json   resumable per-job store (source of truth)
    quote_runs.xlsx                one row per run: matched template + key fields
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Tuple

from config import AUTOCAD_JOBS_DIR, BACKLOG_DIR
from autocad_scan import iter_job_folders, job_key, DEFAULT_MIN_JOB
from sales_orders import _run_files_in_folder
from templates import parse_quote_run

log = logging.getLogger("quote-run-scan")

PROGRESS_PATH = BACKLOG_DIR / "quote_run_scan_progress.json"
WORKBOOK_PATH = BACKLOG_DIR / "quote_runs.xlsx"

# The fields worth their own column, in a readable order. Anything a template
# pulls that isn't here lands in the catch-all "Other" column, so adding fields
# to a template never silently drops data from the inventory.
CORE_FIELDS = [
    "Serial", "Size", "Design", "Class", "Fan Type", "Arrangement", "% Width",
    "Discharge", "Rotation", "Effective Wheel Dia", "CFM", "SP", "BHP", "RPM",
    "Air Temp F", "Max Temp F", "Density", "Max HP", "Max RPM", "Tip Speed FPM",
    "Shaft Dia", "Brg Centers", "Critical Speed RPM",
    "Shaft Length", "OH", "BX", "STB", "TG&P", "STH",
    "Bearing Size", "Bearing Series", "Bearing L10 Hr",
    "Drive Brg Mount", "Drive Brg Static Lb", "Other Brg Mount",
    "Other Brg Static Lb", "Brg Thrust Lb",
    "Rotor WR2", "Rotor Max RPM", "Rotor Material",
    "Stress Ratio at Hub", "Stress Ratio at Bearing",
    "Housing Width (N)", "Base to CL (F)",
    "Blade Material", "Blade Gauge", "Sideplate Material", "Sideplate Gauge",
    "Backplate Material", "Backplate Gauge", "Liner Material", "Liner Gauge",
    "Wheel Material", "Blades", "Max RPM Wheel Only", "Wheel Resonance CPM",
    "Wheel Weight Lb", "Wheel Thrust Lb", "Wheel WR2",
    "Housing to Wheel CG", "Housing to Hub Inlet Face",
    "Sheave PD", "Min Sheave PD",
    "Motor Frame", "Motor Position", "Motor Enclosure", "Motor Weight Lb",
    "Housing Construction", "Stiffeners", "Fan Outlet Area FT2",
    "Inlet Box Size", "Shaft Seal", "Shaft Seal Height", "Flanged Inlet",
    "Total Weight Lb", "Total Price",
    "Hub", "Coupling", "Drive", "Engineering Approval", "FEA Analysis",
    "Non-Std Wheel Materials", "Shrink Fit", "Factory Run Test",
]


# --------------------------------------------------------------------------- #
# Pure logic (no I/O — unit-tested)                                           #
# --------------------------------------------------------------------------- #
_DAMPER = re.compile(r"(?i)damper")


def is_damper(filename: str) -> bool:
    """Flag an in-house damper quote run — a dedicated damper run document
    (e.g. '420848 damper quote run.docx'), a separate product line we design.
    Keyed on 'damper' in the *file name* only: the file is already a quote/run
    doc, so a damper-named one is a damper quote. We deliberately ignore the
    body text, since plenty of fan runs merely list a damper as an accessory."""
    return bool(_DAMPER.search(filename))


def classify_status(template: str, fields: Dict[str, Any], raw_lines: List[str], ext: str) -> str:
    """A one-word health flag per run, so the unreadable ones are easy to spot."""
    if fields:
        return "OK"
    if template == "unknown":
        return "UNRECOGNIZED FORMAT"   # an extension no template reads yet
    if ext == ".pdf" and not raw_lines:
        return "PDF (no text layer)"   # a drawing/scanned run, not a parsing gap
    return "NO FIELDS"                 # readable, but no fields pulled — needs tuning


def reparse_stored(records: Dict[str, Dict[str, Any]]) -> int:
    """Re-run the (current, improved) Chicago Blower parser over the text each
    run already carries — raw_lines from the sweep, or the vision transcript —
    so new extraction patterns apply instantly: no Z: access, no API cost.
    Vision-extracted fields are kept and pattern hits merged over them. Returns
    how many runs changed."""
    from templates import _parse_chicago_blower, _cb_summary, SELECTION_PROGRAM_MARKERS
    changed = 0
    for rec in records.values():
        job = rec.get("job", "")
        for run in rec.get("runs", []):
            vision = run.get("vision") or {}
            text = "\n".join(run.get("raw_lines") or []) or vision.get("transcript", "")
            up = text.upper()
            if not text or not any(m in up for m in SELECTION_PROGRAM_MARKERS):
                continue                      # not a selection-program run
            parsed = _parse_chicago_blower(text)
            # Vision runs: the model read the actual IMAGE; the transcript the
            # patterns run on is derivative (OCR noise like "4/100 CFM" happens).
            # So patterns only FILL GAPS on vision runs — the model wins overlaps.
            fields = {**parsed, **(run.get("fields") or {})} if vision else parsed
            if fields and job:
                fields.setdefault("Serial", job)
            if fields == run.get("fields"):
                continue
            run["fields"] = fields
            run["summary"] = _cb_summary(fields)
            if fields:
                run["status"] = "OK"
                if not vision:
                    run["template"] = "cbc_qt_run_text"
            changed += 1
    return changed


def carry_vision_forward(old_rec: Optional[Dict[str, Any]],
                         new_rec: Dict[str, Any]) -> Dict[str, Any]:
    """A --rescan re-parses every file from disk, but pdfplumber still gets
    nothing out of a scanned PDF — so a fresh parse would wipe the (paid-for)
    Claude-vision result. Carry each vision result forward onto the matching
    re-scanned run (same path) whenever the fresh parse pulled no fields."""
    if not old_rec:
        return new_rec
    old_by_path = {r.get("path"): r for r in old_rec.get("runs", []) if r.get("vision")}
    for run in new_rec.get("runs", []):
        old = old_by_path.get(run.get("path"))
        if old and not run.get("fields"):
            for k in ("template", "fields", "summary", "status", "vision"):
                run[k] = old.get(k)
    return new_rec


def run_rows(records: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Flatten the per-job store into one row per run (a job can have several:
    e.g. the quote and the production run; `history\\` copies are not swept)."""
    from run_rank import rank_runs
    rows: List[Dict[str, Any]] = []
    for rec in sorted(records.values(), key=lambda r: r.get("job", "")):
        for run in rank_runs(rec.get("runs", [])):   # most-current run first
            f = run.get("fields", {}) or {}
            other = "; ".join(f"{k}={v}" for k, v in f.items() if k not in CORE_FIELDS)
            rows.append({
                "job": rec.get("job", ""),
                "type": rec.get("type", ""),
                "status": run.get("status", ""),
                "damper": run.get("damper", False),
                "template": run.get("template", ""),
                "file": run.get("file", ""),
                "path": run.get("path", ""),
                "folder": rec.get("folder", ""),
                "summary": run.get("summary", ""),
                "core": {k: f.get(k, "") for k in CORE_FIELDS},
                "other": other,
            })
    return rows


# --------------------------------------------------------------------------- #
# Filesystem scan                                                             #
# --------------------------------------------------------------------------- #
def _mtime(f: Path) -> float:
    try:
        return f.stat().st_mtime
    except OSError:
        return 0.0


def scan_one(job: str, jtype: str, folder: Path) -> Dict[str, Any]:
    """Find and parse every quote run in one job folder."""
    runs: List[Dict[str, Any]] = []
    for f in _run_files_in_folder(folder):
        r = parse_quote_run(f)
        if r["fields"] and "Serial" not in r["fields"]:
            r["fields"]["Serial"] = job   # runs often omit the SN# header
        runs.append({
            "file": f.name,
            "path": str(f),
            "template": r["template"],
            "fields": r["fields"],
            "summary": r["summary"],
            "status": classify_status(r["template"], r["fields"], r["raw_lines"], f.suffix.lower()),
            "damper": is_damper(f.name),
            # The document's own lines — persisted so the store doubles as a
            # re-parsable corpus (new patterns re-extract without re-reading Z:).
            "raw_lines": r["raw_lines"],
            "mtime": _mtime(f),   # recency signal for ranking multiple runs
        })
    return {
        "job": job,
        "type": jtype,
        "folder": str(folder),
        "runs": runs,
        "scanned_at": datetime.now().isoformat(timespec="seconds"),
    }


# --------------------------------------------------------------------------- #
# Progress store (resumable)                                                  #
# --------------------------------------------------------------------------- #
def load_progress() -> Dict[str, Dict[str, Any]]:
    if not PROGRESS_PATH.exists():
        return {}
    try:
        return json.loads(PROGRESS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        log.warning("Could not read %s (%s); starting fresh", PROGRESS_PATH, e)
        return {}


def save_progress(records: Dict[str, Dict[str, Any]]) -> None:
    PROGRESS_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = PROGRESS_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(records, indent=2), encoding="utf-8")
    tmp.replace(PROGRESS_PATH)  # atomic: a crash mid-write never corrupts the store


# --------------------------------------------------------------------------- #
# Excel inventory                                                             #
# --------------------------------------------------------------------------- #
def write_workbook(records: Dict[str, Dict[str, Any]], path: Path) -> Path:
    """One row per run: job, status, matched template, the core fields, then the
    full summary. Unreadable rows (no fields) are tinted so they stand out."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")     # green: fields pulled
    warn_fill = PatternFill("solid", fgColor="FFEB9C")   # amber: readable, no fields
    bad_fill = PatternFill("solid", fgColor="FFC7CE")    # red: unrecognized / no text

    rows = run_rows(records)
    fixed = ["Job #", "Type", "Status", "Template", "Run File", "Damper"]
    headers = fixed + CORE_FIELDS + ["Other", "Summary", "Folder"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Runs"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill

    drawing_fill = PatternFill("solid", fgColor="D9D9D9")  # grey: a drawing, nothing to parse
    status_fill = {"OK": ok_fill, "NO FIELDS": warn_fill,
                   "UNRECOGNIZED FORMAT": bad_fill, "PDF (no text layer)": bad_fill,
                   "DRAWING": drawing_fill}
    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1, row["job"])
        ws.cell(i, 2, row["type"])
        scell = ws.cell(i, 3, row["status"])
        scell.fill = status_fill.get(row["status"], warn_fill)
        ws.cell(i, 4, row["template"])
        fcell = ws.cell(i, 5, row["file"])
        if row["path"]:
            fcell.hyperlink = row["path"]
            fcell.font = link_font
        ws.cell(i, 6, "DAMPER" if row["damper"] else "")
        for k, name in enumerate(CORE_FIELDS, start=len(fixed) + 1):
            ws.cell(i, k, row["core"].get(name, ""))
        base = len(fixed) + len(CORE_FIELDS)
        ws.cell(i, base + 1, row["other"])
        ws.cell(i, base + 2, row["summary"])
        folder = (row["folder"] or "").strip()
        ocell = ws.cell(i, base + 3, "Open" if folder else "")
        if folder:
            ocell.hyperlink = folder
            ocell.font = link_font

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "C2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 6), 40)

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        # The usual cause: the workbook is still open in Excel (Windows locks
        # it). Don't throw away the whole scan — write a timestamped copy.
        alt = path.with_name(f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}")
        wb.save(alt)
        log.warning("%s is locked (close it in Excel) — wrote %s instead.", path, alt)
        return alt


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #
def _resolve_jobs(root: Path, jobs: List[str]) -> Iterator[Tuple[str, str, Path]]:
    """Locate each explicitly-requested job folder under the AutoCAD root."""
    for job in jobs:
        matches = list(root.glob(f"*/*/{job}")) or list(root.glob(f"*/*/{job}*"))
        if matches:
            m = matches[0]
            yield job_key(m.name), m.relative_to(root).parts[0], m
        else:
            log.warning("  %s: folder not found", job)


def _parse_args(argv: List[str]) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Sweep AutoCAD job folders for quote runs and parse them.")
    ap.add_argument("jobs", nargs="*", help="Specific job numbers to scan (default: all).")
    ap.add_argument("--list", metavar="FILE", help="Read job numbers (one per line) from a file.")
    ap.add_argument("--range", nargs=2, type=int, metavar=("FIRST", "LAST"), help="Scan jobs in [FIRST, LAST].")
    ap.add_argument("--root", default=str(AUTOCAD_JOBS_DIR), help="AutoCAD jobs root.")
    ap.add_argument("--out", default=str(WORKBOOK_PATH), help="Excel output path.")
    ap.add_argument("--rescan", action="store_true", help="Ignore saved progress; redo every job.")
    ap.add_argument("--reparse-attention", action="store_true",
                    help="Re-parse only the jobs whose saved results currently need attention "
                         "(any run not flagged OK). Fast way to re-check failures after a parser fix.")
    ap.add_argument("--reparse-stored", action="store_true",
                    help="Re-run the parser over the text already saved in the store (raw lines "
                         "+ vision transcripts) — applies new patterns in seconds, no Z:, no API.")
    ap.add_argument("--limit", type=int, default=0,
                    help="Stop after scanning N folders this run (0 = no limit).")
    ap.add_argument("--min-job", type=int, default=DEFAULT_MIN_JOB,
                    help=f"Skip folders below this job number on a full sweep (default {DEFAULT_MIN_JOB}).")
    ap.add_argument("--max-job", type=int, default=0, help="Skip folders above this job number (0 = no cap).")
    return ap.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    args = _parse_args(sys.argv[1:] if argv is None else argv)
    root = Path(args.root)
    if not root.exists():
        log.error("AutoCAD jobs root not reachable: %s (is Z: mapped?)", root)
        return 1

    prior = load_progress()
    # --rescan redoes every parse, but vision results are carried forward from
    # the prior store (see carry_vision_forward) — they cost API money.
    records = {} if args.rescan else prior

    if args.reparse_stored:
        changed = reparse_stored(records)
        save_progress(records)
        try:
            import master_sync
            master_sync.run("quote_runs")
        except Exception as e:  # noqa: BLE001
            log.warning("Could not sync quote runs to the live master (%s)", e)
        out = write_workbook(records, Path(args.out))
        log.info("Re-parsed stored text: %d run(s) updated. Wrote %s", changed, out)
        return 0

    jobs = list(args.jobs)
    if args.list:
        try:
            jobs += [ln.strip() for ln in Path(args.list).read_text(encoding="utf-8").splitlines() if ln.strip()]
        except OSError as e:
            log.error("Could not read --list file %s (%s)", args.list, e)
            return 1
    if args.reparse_attention:
        # DRAWING is a settled classification (a drawing has no fields to pull),
        # not something a parser fix will ever change — don't re-flag it forever.
        attention = sorted({
            rec.get("job", "") for rec in records.values()
            if any(run.get("status") not in ("OK", "DRAWING")
                   for run in rec.get("runs", []))
        } - {""})
        if attention:
            log.info("Re-parsing %d job(s) flagged as needing attention.", len(attention))
            jobs += attention
        elif not jobs:
            log.info("Nothing currently needs attention — no jobs to re-parse.")
            return 0
    jobs = list(dict.fromkeys(jobs))  # dedupe, keep order

    min_job, max_job = args.min_job, args.max_job
    if args.range:
        min_job, max_job = args.range

    explicit = bool(jobs)
    if explicit:
        folders: Iterator[Tuple[str, str, Path]] = _resolve_jobs(root, jobs)
    else:
        folders = iter_job_folders(root, min_job=min_job, max_job=max_job)

    t0 = time.monotonic()
    scanned = 0
    for job, jtype, folder in folders:
        if args.limit and scanned >= args.limit:
            break
        if job in records and not args.rescan and not explicit:
            continue  # already scanned on an earlier run
        try:
            records[job] = carry_vision_forward(prior.get(job),
                                                scan_one(job, jtype, folder))
            scanned += 1
        except OSError as e:
            log.warning("  %s: scan failed (%s)", job, e)
            continue
        if scanned % 200 == 0:
            save_progress(records)
            log.info("  scanned %d (%d total) ...", scanned, len(records))

    save_progress(records)
    try:   # fold these parsed quote/construction runs into the one master store
        import master_sync
        master_sync.run("quote_runs")
    except Exception as e:  # noqa: BLE001
        log.warning("Could not sync quote runs to the live master (%s)", e)
    out = write_workbook(records, Path(args.out))

    rows = run_rows(records)
    jobs_with_runs = sum(1 for r in records.values() if r.get("runs"))
    by_template = Counter(r["template"] for r in rows)
    by_status = Counter(r["status"] for r in rows)
    log.info("Done in %.1fs: %d jobs in store (%d scanned this run), %d with a run, %d runs total.",
             time.monotonic() - t0, len(records), scanned, jobs_with_runs, len(rows))
    log.info("  by template: %s", dict(by_template) or "(none)")
    needs = {k: v for k, v in by_status.items() if k not in ("OK", "DRAWING")}
    if needs:
        log.info("  runs that need attention (no fields): %s", needs)
    log.info("  Wrote %s", out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
