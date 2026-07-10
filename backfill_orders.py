"""Backfill — look up OLD orders on cbcinsider one at a time, all day, resumably.

The daily run only enriches what's on the board today. This walks a much larger
backlog of historical jobs, opens each one's detail, downloads + parses its
Sales Order and (if present) construction/drive run, merges in the AutoCAD DWG
scan, and writes a master backlog workbook.

It searches exactly one order at a time in one browser page. Every completed
order is checkpointed immediately, so stop it any time and plain
`python backfill_orders.py` resumes from the next unfinished order. It can run
beside watch.py; the two processes coordinate CBC access and shared data writes.

Job source (pick one; default is the AutoCAD folders):
    python backfill_orders.py                     # supported job folders (401000+)
    python backfill_orders.py 421314 421388       # explicit job numbers
    python backfill_orders.py --list jobs.txt     # one job number per line
    python backfill_orders.py --range 420000 421000
Options:
    --limit N     stop after N jobs this run        --delay S     seconds between jobs
    --out PATH    workbook path
    --force       reprocess selected jobs           --rescan     ignore saved progress
    --newest-first process high job numbers first   --retry-not-found recheck misses
    --from-dwg-progress use saved AutoCAD scan job list instead of walking Z:
    --passes N    retry incomplete jobs this many times in the same run
    --publish-every N publish order-data to Git every N attempts (default 25)

Outputs (under BACKLOG_DIR):
    backfill_progress.json   resumable per-job store (source of truth)
    backfill_line_items.json watcher-safe line-item overlay
    backlog.xlsx             master sheet: SO/drive-run fields + DWG suffix matrix

Old orders are opened through the queue page's "search order" / "find order"
box: each job number is typed in and the surfaced order's detail is opened. The
box is auto-detected; if that misses on your layout, set CBC_SEARCH_SELECTOR
(and optionally CBC_SEARCH_BUTTON) in .env — `python discover_documents.py
--probe <job#>` prints the exact selector. The run preflights the box and stops
with a clear message rather than grinding the whole list if it can't be found.
"""
from __future__ import annotations

import argparse
import asyncio
import contextlib
import json
import logging
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urljoin

from config import (
    CBC_URL, CBC_QUEUE_URL, STORAGE_STATE_PATH,
    SALES_ORDER_DIR, DRIVE_RUN_DIR, BACKLOG_DIR, AUTOCAD_JOBS_DIR,
    CBC_SEARCH_SELECTOR, CBC_SEARCH_BUTTON,
)
from templates import parse_quote_run
from sales_orders import (
    _parse_doc, _latest_of_type, _run_docs, _run_filename, _run_files_in_folder,
    _trigger_js, _so_filename, _download_error, SO_TYPE, parse_sales_order_pdf,
)
from scraper import CONTAINER_SELECTOR
import autocad_scan
import line_items
from process_lock import cbc_fetch_lock
from sales_order_validation import (
    accept_existing,
    failed_acceptance,
    finalize_candidate,
    modal_text_matches_job,
    staging_path,
)

log = logging.getLogger("backfill")

PROGRESS_PATH = BACKLOG_DIR / "backfill_progress.json"
WORKBOOK_PATH = BACKLOG_DIR / "backlog.xlsx"
RETRYABLE_STATUSES = {"error", "not-found", "no-SO"}
BACKFILL_SCAN_VERSION = "serial-verified-v1"
REQUIRED_MISS_ATTEMPTS = 2
# CBC's normal Search Order box handles the 401xxx+ population. The earlier
# 400xxx orders need the separate legacy lookup path, which is not implemented.
DEFAULT_CBC_SEARCH_MIN_JOB = 401000
DEFAULT_PUBLISH_EVERY = 25


def _attempt_count(rec: Dict[str, Any]) -> int:
    try:
        return max(0, int(rec.get("backfill_attempts") or 0))
    except (TypeError, ValueError):
        return 0


# --------------------------------------------------------------------------- #
# Job sources                                                                 #
# --------------------------------------------------------------------------- #
def jobs_from_folders(root: Path, min_job: int = DEFAULT_CBC_SEARCH_MIN_JOB,
                      max_job: int = 0) -> List[str]:
    """Every real job number under AUTOCAD_JOBS_DIR — same list the DWG scan walks.
    Folders below min_job (year/template/archive dirs) are skipped."""
    return [job for job, _type, _path in autocad_scan.iter_job_folders(root, min_job, max_job)]


def jobs_from_list(path: Path) -> List[str]:
    return [ln.strip() for ln in path.read_text(encoding="utf-8").splitlines() if ln.strip()]


def jobs_from_range(first: int, last: int) -> List[str]:
    lo, hi = sorted((first, last))
    return [str(n) for n in range(lo, hi + 1)]


# --------------------------------------------------------------------------- #
# Browser helpers (sync — serial by design)                                   #
# --------------------------------------------------------------------------- #
def _jobnum(args_js: str) -> str:
    return args_js.split(",", 1)[0].strip().strip("'\"").split("-", 1)[0]


def _board_args(page) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for c in page.locator(CONTAINER_SELECTOR).all():
        m = re.search(r"loadDetail\((.*?)\)", c.get_attribute("onclick") or "")
        if m:
            out[_jobnum(m.group(1).strip())] = m.group(1).strip()
    return out


def _open_via_loaddetail(page, job: str, args_js: str, doc_timeout_ms: int = 120000) -> bool:
    page.evaluate(_trigger_js(args_js))
    return _wait_for_matching_modal(page, job, doc_timeout_ms)


def _open_matching_card(page, job: str, doc_timeout_ms: int = 120000) -> bool:
    """Open the surfaced order row by clicking it instead of replaying loadDetail().

    Some legacy order rows carry punctuation in their inline loadDetail arguments.
    Browser-clicking the original row is more faithful than rebuilding that
    JavaScript string and avoids syntax errors on odd argument text.
    """
    for c in page.locator(CONTAINER_SELECTOR).all():
        m = re.search(r"loadDetail\((.*?)\)", c.get_attribute("onclick") or "")
        if not m or _jobnum(m.group(1).strip()) != job:
            continue
        try:
            c.evaluate("(el) => el.click()")
            return _wait_for_matching_modal(page, job, doc_timeout_ms)
        except Exception:  # noqa: BLE001
            return False
    return False


def find_search_box(page):
    """Locate the queue page's "search order" / "find order" box, or None.

    Honors CBC_SEARCH_SELECTOR if set; otherwise matches a text input whose
    placeholder / aria-label / title / id / name mentions "search" or "find"
    (preferring one that also mentions "order")."""
    if CBC_SEARCH_SELECTOR:
        loc = page.locator(CBC_SEARCH_SELECTOR)
        return loc.first if loc.count() else None
    best = None
    for el in page.locator("input[type=text], input[type=search], input:not([type])").all():
        blob = " ".join(filter(None, [
            el.get_attribute("placeholder"), el.get_attribute("aria-label"),
            el.get_attribute("title"), el.get_attribute("id"), el.get_attribute("name"),
        ])).lower()
        if "search" in blob or "find" in blob:
            if "order" in blob:
                return el          # best: a "search order" / "find order" box
            best = best or el      # fallback: any search/find input
    return best


def _modal_has_docs(page, job: str) -> bool:
    try:
        if page.locator("#modalDetail a[href*='downloaddoc.aspx']").count() <= 0:
            return False
        return modal_text_matches_job(page.locator("#modalDetail").inner_text(), job)
    except Exception:  # noqa: BLE001
        return False


def _wait_for_matching_modal(page, job: str, timeout_ms: int) -> bool:
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        if _modal_has_docs(page, job):
            return True
        page.wait_for_timeout(250)
    return False


def open_order_detail(page, job: str, search_timeout_s: float = 75.0,
                      doc_timeout_ms: int = 120000) -> bool:
    """Surface `job`'s documents in #modalDetail via the queue's search box.

    Types the job number into the search bar, submits (Enter, plus an optional
    CBC_SEARCH_BUTTON click), then waits for either the board to re-render with
    the order (and opens it via loadDetail) or the detail modal to populate
    directly. Returns True once the documents are present, else False."""
    _close_modal(page)   # never let a previous order's stale links be harvested
    box = find_search_box(page)
    if box is None:
        return False
    try:
        box.click()
        box.fill("")
        box.fill(job)
        box.press("Enter")
    except Exception:  # noqa: BLE001
        return False
    if CBC_SEARCH_BUTTON:
        try:
            btn = page.locator(CBC_SEARCH_BUTTON)
            if btn.count():
                btn.first.click()
        except Exception:  # noqa: BLE001
            pass
    # Poll: the search may re-render the board (postback/filter) or open the
    # detail directly. Either way, wait for the order to surface.
    deadline = time.monotonic() + search_timeout_s
    while time.monotonic() < deadline:
        page.wait_for_timeout(500)
        if _open_matching_card(page, job, doc_timeout_ms):
            return True
        if _modal_has_docs(page, job):
            return True
    return False


def _collect_docs(page) -> List[Dict[str, Any]]:
    docs = []
    for a in page.locator("#modalDetail a").all():
        href = a.get_attribute("href") or ""
        if "downloaddoc.aspx" in href.lower():
            d = _parse_doc(href)
            d["href"] = href
            docs.append((href, d))
    return docs


def _download(context, page_url: str, href: str, dest: Path) -> Optional[str]:
    if dest.exists():
        return str(dest)
    try:
        resp = context.request.get(urljoin(page_url, href), timeout=60000)
        body = resp.body()
        # Never archive an error page / expired-session login page as the PDF —
        # the dest.exists() cache would then skip re-downloading it forever.
        err = _download_error(resp.status, body, dest.suffix.lower() == ".pdf")
        if err:
            log.warning("  download failed for %s: %s", dest.name, err)
            return None
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(body)
        return str(dest)
    except Exception as e:  # noqa: BLE001
        log.warning("  download failed for %s: %s", dest.name, e)
        return None


def _download_sales_order(context, page_url: str, href: str, destination: Path, job: str):
    existing = accept_existing(destination, job)
    if existing and existing.path:
        return existing
    if existing:
        log.warning(
            "  %s: rejected existing Sales Order internal=%s status=%s -> %s",
            job,
            existing.validation.internal_order or "?",
            existing.validation.status,
            existing.quarantine_path,
        )
    staged = staging_path(destination, job)
    downloaded = _download(context, page_url, href, staged)
    if not downloaded:
        return failed_acceptance(job, f"download failed for {destination.name}")
    accepted = finalize_candidate(downloaded, destination, job)
    if not accepted.path:
        log.warning(
            "  %s: rejected downloaded Sales Order internal=%s status=%s -> %s",
            job,
            accepted.validation.internal_order or "?",
            accepted.validation.status,
            accepted.quarantine_path,
        )
    return accepted


def _close_modal(page) -> None:
    """Hide the detail modal AND strip its download links (see
    _close_modal_async: stale links + a reused modal made every job harvest
    the PREVIOUS order's documents)."""
    try:
        page.evaluate(
            "() => {"
            " if (window.jQuery) jQuery('#modalDetail').modal('hide');"
            " document.querySelectorAll(\"#modalDetail a[href*='downloaddoc.aspx' i]\")"
            "  .forEach(a => a.remove());"
            "}")
        page.wait_for_timeout(250)
    except Exception:  # noqa: BLE001
        pass


def _best_run_path(paths: List[Path]) -> str:
    if not paths:
        return ""
    try:
        from run_rank import rank_paths
        ranked = rank_paths(paths)
        return str(ranked[0]) if ranked else str(paths[0])
    except Exception:  # noqa: BLE001 - ranking is nice-to-have, never fatal
        return str(paths[0])


def _quote_run_design(rec: Dict[str, Any]) -> str:
    return str(rec.get("design") or rec.get("so_design") or rec.get("so_design_desc") or "")


def _attach_quote_run_parse(rec: Dict[str, Any], dr_path: str | None) -> None:
    if not dr_path:
        return
    qr = parse_quote_run(dr_path, design=_quote_run_design(rec))
    rec["drive_run"] = qr.get("fields", {})
    rec["drive_run_summary"] = qr.get("summary", "")
    rec["drive_run_template"] = qr.get("template", "")
    if qr.get("design") is not None:
        rec["drive_run_design"] = qr.get("design")


def process_one(page, context, job: str, folder: str = "",
                li_store: Dict[str, Any] | None = None) -> Dict[str, Any]:
    """Open one order via the search box, download + parse its SO and quote run.
    `folder` (the job's AutoCAD folder, when known from the DWG scan) is checked
    for a quote-run file if the order's documents don't carry one. `li_store`
    (the shared line-items store) gets the order's parsed line items — the
    caller saves it alongside the progress file."""
    rec: Dict[str, Any] = {"job": job, "status": "", "scanned_at": datetime.now().isoformat(timespec="seconds")}
    try:
        if not open_order_detail(page, job):
            rec["status"] = "not-found"  # search returned no order/documents for this job#
            return rec
        docs = _collect_docs(page)

        so_pdf = dr_pdf = None
        so = _latest_of_type(docs, SO_TYPE)
        if so:
            href, doc = so
            accepted = _download_sales_order(
                context,
                page.url,
                href,
                SALES_ORDER_DIR / job / _so_filename(job, doc["rev"]),
                job,
            )
            so_pdf = accepted.path
            rec["so_validation"] = accepted.validation.status
            rec["so_internal_order"] = accepted.validation.internal_order
            rec["so_validation_method"] = accepted.validation.method
            rec["so_quarantine"] = accepted.quarantine_path
            if so_pdf:
                rec["co_number"] = (doc["rev"] - 1) if doc["rev"] and doc["rev"] > 1 else 0
                p = parse_sales_order_pdf(so_pdf)
                rec.update({
                    "so_design_desc": p.get("design_desc", ""), "so_size": p.get("size", ""),
                    "so_arrangement": p.get("arrangement", ""), "so_motor_pos": p.get("motor_pos", ""),
                    "so_class": p.get("fan_class", ""), "so_rotation": p.get("rotation", ""),
                    "so_discharge": p.get("discharge", ""), "so_pct_width": p.get("pct_width", ""),
                    "so_wheel_type": p.get("wheel_type", ""),
                    "so_design_temp": p.get("design_temp", ""), "so_max_temp": p.get("max_temp", ""),
                    "so_special_temp": p.get("special_temp", ""),
                    "so_pdf": so_pdf,
                })
                items = p.get("line_items") or []
                rec["line_item_count"] = len(items)
                if li_store is not None:
                    line_items.apply_ai_cache(items, li_store)
                    line_items.record_job(li_store, job, items,
                                          co_number=rec.get("co_number"), so_pdf=so_pdf)

        runs = _run_docs(docs) if not so or so_pdf else []
        dr_count = 0
        if runs:
            rec["has_drive_run"] = True
            dr_count = len(runs)
            for href, doc in runs:
                got = _download(context, page.url, href, DRIVE_RUN_DIR / job / _run_filename(job, doc))
                if got and not dr_pdf:
                    dr_pdf = got
            rec["drive_run_pdf"] = dr_pdf or ""
        if folder:
            # Always scan the AutoCAD folder — cheap rglob, catches runs that
            # only live there and any folder copies alongside document ones.
            hits = _run_files_in_folder(Path(folder))
            if hits:
                if not rec.get("has_drive_run"):
                    rec["has_drive_run"] = True
                    dr_pdf = _best_run_path(hits)
                    rec["drive_run_pdf"] = dr_pdf
                dr_count += len(hits)
        if dr_count:
            rec["drive_run_count"] = dr_count
        if dr_pdf:
            _attach_quote_run_parse(rec, dr_pdf)

        # "ok" only when every document we found actually downloaded — a found-
        # but-failed download stays "error" so the resume retries it, instead of
        # permanently recording the job with empty fields.
        if not so:
            rec["status"] = "no-SO"
        elif so_pdf and (not runs or dr_pdf):
            rec["status"] = "ok"
        else:
            rec["status"] = "error"
    except Exception as e:  # noqa: BLE001 - one bad order never stops the backlog
        log.warning("  %s: error (%s)", job, e)
        rec["status"] = "error"
    finally:
        _close_modal(page)
    return rec


# --------------------------------------------------------------------------- #
# Async browser helpers (serial backfill)                                      #
# --------------------------------------------------------------------------- #
async def _open_backfill_page(context):
    page = await context.new_page()
    last = None
    for attempt in (1, 2, 3):
        try:
            await page.goto(CBC_QUEUE_URL or CBC_URL, wait_until="domcontentloaded", timeout=60000)
            if "login" in page.url.lower() or await page.locator('input[type="password"]').count() > 0:
                raise RuntimeError("Landed on login - session expired. Re-run `python login.py`.")
            # Old-order backfill is driven by the search box. The dispatch list
            # may be empty or filtered away, so requiring normal queue rows here
            # blocks a perfectly usable search page.
            await page.wait_for_selector("body", timeout=45000)
            return page
        except Exception as e:  # noqa: BLE001
            last = e
            if attempt < 3:
                await page.wait_for_timeout(3000 * attempt)
    with contextlib.suppress(Exception):
        await page.close()
    raise RuntimeError(f"Could not open CBC queue page: {last}")


async def _board_args_async(page) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for c in await page.locator(CONTAINER_SELECTOR).all():
        m = re.search(r"loadDetail\((.*?)\)", await c.get_attribute("onclick") or "")
        if m:
            out[_jobnum(m.group(1).strip())] = m.group(1).strip()
    return out


async def _open_via_loaddetail_async(page, job: str, args_js: str,
                                     doc_timeout_ms: int = 120000) -> bool:
    await page.evaluate(_trigger_js(args_js))
    return await _wait_for_matching_modal_async(page, job, doc_timeout_ms)


async def _open_matching_card_async(page, job: str, doc_timeout_ms: int = 120000) -> bool:
    for c in await page.locator(CONTAINER_SELECTOR).all():
        m = re.search(r"loadDetail\((.*?)\)", await c.get_attribute("onclick") or "")
        if not m or _jobnum(m.group(1).strip()) != job:
            continue
        try:
            await c.evaluate("(el) => el.click()")
            return await _wait_for_matching_modal_async(page, job, doc_timeout_ms)
        except Exception:  # noqa: BLE001
            return False
    return False


async def find_search_box_async(page):
    if CBC_SEARCH_SELECTOR:
        loc = page.locator(CBC_SEARCH_SELECTOR)
        return loc.first if await loc.count() else None
    best = None
    for el in await page.locator("input[type=text], input[type=search], input:not([type])").all():
        blob = " ".join(filter(None, [
            await el.get_attribute("placeholder"), await el.get_attribute("aria-label"),
            await el.get_attribute("title"), await el.get_attribute("id"),
            await el.get_attribute("name"),
        ])).lower()
        if "search" in blob or "find" in blob:
            if "order" in blob:
                return el
            best = best or el
    return best


async def _modal_has_docs_async(page, job: str) -> bool:
    try:
        if await page.locator("#modalDetail a[href*='downloaddoc.aspx']").count() <= 0:
            return False
        return modal_text_matches_job(await page.locator("#modalDetail").inner_text(), job)
    except Exception:  # noqa: BLE001
        return False


async def _wait_for_matching_modal_async(page, job: str, timeout_ms: int) -> bool:
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        if await _modal_has_docs_async(page, job):
            return True
        await page.wait_for_timeout(250)
    return False


async def open_order_detail_async(page, job: str, search_timeout_s: float = 75.0,
                                  doc_timeout_ms: int = 120000) -> bool:
    # Clear any lingering modal content FIRST (a prior job that errored before
    # its finally, or a reused page) so this job can never see stale links.
    await _close_modal_async(page)
    box = await find_search_box_async(page)
    if box is None:
        return False
    try:
        await box.click()
        await box.fill("")
        await box.fill(job)
        await box.press("Enter")
    except Exception:  # noqa: BLE001
        return False
    if CBC_SEARCH_BUTTON:
        try:
            btn = page.locator(CBC_SEARCH_BUTTON)
            if await btn.count():
                await btn.first.click()
        except Exception:  # noqa: BLE001
            pass
    deadline = time.monotonic() + search_timeout_s
    while time.monotonic() < deadline:
        await page.wait_for_timeout(500)
        if await _open_matching_card_async(page, job, doc_timeout_ms):
            return True
        if await _modal_has_docs_async(page, job):
            return True
    return False


async def _collect_docs_async(page) -> List[Dict[str, Any]]:
    docs = []
    for a in await page.locator("#modalDetail a").all():
        href = await a.get_attribute("href") or ""
        if "downloaddoc.aspx" in href.lower():
            d = _parse_doc(href)
            d["href"] = href
            docs.append((href, d))
    return docs


async def _download_async(context, page_url: str, href: str, dest: Path) -> Optional[str]:
    if dest.exists():
        return str(dest)
    url = urljoin(page_url, href)
    for attempt in (1, 2, 3):
        try:
            resp = await context.request.get(url, timeout=60000)
            body = await resp.body()
            err = _download_error(resp.status, body, dest.suffix.lower() == ".pdf")
            if err:
                raise RuntimeError(err)
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(body)
            return str(dest)
        except Exception as e:  # noqa: BLE001
            if attempt == 3:
                log.warning("  download failed for %s after %d tries: %s", dest.name, attempt, e)
            else:
                await asyncio.sleep(2 * attempt)
    return None


async def _download_sales_order_async(
    context, page_url: str, href: str, destination: Path, job: str
):
    existing = accept_existing(destination, job)
    if existing and existing.path:
        return existing
    if existing:
        log.warning(
            "  %s: rejected existing Sales Order internal=%s status=%s -> %s",
            job,
            existing.validation.internal_order or "?",
            existing.validation.status,
            existing.quarantine_path,
        )
    staged = staging_path(destination, job)
    downloaded = await _download_async(context, page_url, href, staged)
    if not downloaded:
        return failed_acceptance(job, f"download failed for {destination.name}")
    accepted = finalize_candidate(downloaded, destination, job)
    if not accepted.path:
        log.warning(
            "  %s: rejected downloaded Sales Order internal=%s status=%s -> %s",
            job,
            accepted.validation.internal_order or "?",
            accepted.validation.status,
            accepted.quarantine_path,
        )
    return accepted


async def _close_modal_async(page) -> None:
    """Hide the detail modal AND strip its download links. The site reuses
    #modalDetail: the next order's header lands before its documents section
    refreshes, so leaving the old links in the DOM let the next job pass the
    text-match while harvesting THIS order's documents — every job downloaded
    the previous job's Sales Order (caught by validation as a MISMATCH chain)."""
    try:
        await page.evaluate(
            "() => {"
            " if (window.jQuery) jQuery('#modalDetail').modal('hide');"
            " document.querySelectorAll(\"#modalDetail a[href*='downloaddoc.aspx' i]\")"
            "  .forEach(a => a.remove());"
            "}")
        await page.wait_for_timeout(250)
    except Exception:  # noqa: BLE001
        pass


async def process_one_async(page, context, job: str, folder: str = "",
                            search_timeout_s: float = 75.0,
                            doc_timeout_ms: int = 120000) -> Dict[str, Any]:
    rec: Dict[str, Any] = {
        "job": job,
        "status": "",
        "scanned_at": datetime.now().isoformat(timespec="seconds"),
        "backfill_scan_version": BACKFILL_SCAN_VERSION,
    }
    try:
        if not await open_order_detail_async(page, job, search_timeout_s, doc_timeout_ms):
            rec["status"] = "not-found"
            return rec
        # Pre-arm the server before the first fetch: downloads resolve against
        # the site's session-side "current order", which on EVERY observed job
        # after a session's first still pointed at the previously opened order —
        # each first fetch was a guaranteed wrong-SO quarantine. Closing and
        # re-opening this order's detail (cheap card re-click; full search as
        # fallback) re-points the server for the price of one modal load.
        await _close_modal_async(page)
        rearmed = False
        with contextlib.suppress(Exception):
            rearmed = await _open_matching_card_async(page, job, doc_timeout_ms)
        if not rearmed:
            with contextlib.suppress(Exception):
                await open_order_detail_async(page, job, search_timeout_s, doc_timeout_ms)
        # Brief settle so the documents section finishes rendering — links can
        # appear progressively, and harvesting mid-render could miss the
        # latest CO revision. Best-effort: never worth failing the job over.
        with contextlib.suppress(Exception):
            await page.wait_for_timeout(400)
        docs = await _collect_docs_async(page)

        so_pdf = dr_pdf = None
        so = _latest_of_type(docs, SO_TYPE)
        if so:
            # The site serves documents from its session-side "current order",
            # which can lag the modal: the first fetch after switching orders
            # can return the PREVIOUS order's SO (stripping stale modal links
            # didn't cure it — the wrong doc comes from the SERVER). Validation
            # catches it before anything is saved, so on a MISMATCH re-open the
            # detail (a fresh loadDetail re-points the server) and retry.
            accepted = None
            doc: Dict[str, Any] = {}
            for so_attempt in (1, 2, 3):
                href, doc = so
                accepted = await _download_sales_order_async(
                    context,
                    page.url,
                    href,
                    SALES_ORDER_DIR / job / _so_filename(job, doc["rev"]),
                    job,
                )
                if accepted.path or accepted.validation.status != "MISMATCH":
                    break
                log.warning("  %s: attempt %d fetched another order's SO (href=%s) — "
                            "re-opening the detail and retrying.", job, so_attempt, href)
                await _close_modal_async(page)
                await asyncio.sleep(2.0)
                reopened = False
                with contextlib.suppress(Exception):
                    reopened = await open_order_detail_async(
                        page, job, search_timeout_s, doc_timeout_ms)
                if not reopened:
                    break
                with contextlib.suppress(Exception):
                    await page.wait_for_timeout(1200)   # let the server re-point too
                docs = await _collect_docs_async(page)
                so = _latest_of_type(docs, SO_TYPE)
                if not so:
                    break
            so_pdf = accepted.path
            rec["so_validation"] = accepted.validation.status
            rec["so_internal_order"] = accepted.validation.internal_order
            rec["so_validation_method"] = accepted.validation.method
            rec["so_quarantine"] = accepted.quarantine_path
            if so_pdf:
                rec["co_number"] = (doc["rev"] - 1) if doc["rev"] and doc["rev"] > 1 else 0
                p = parse_sales_order_pdf(so_pdf)
                rec.update({
                    "so_design_desc": p.get("design_desc", ""), "so_size": p.get("size", ""),
                    "so_arrangement": p.get("arrangement", ""), "so_motor_pos": p.get("motor_pos", ""),
                    "so_class": p.get("fan_class", ""), "so_rotation": p.get("rotation", ""),
                    "so_discharge": p.get("discharge", ""), "so_pct_width": p.get("pct_width", ""),
                    "so_wheel_type": p.get("wheel_type", ""),
                    "so_design_temp": p.get("design_temp", ""), "so_max_temp": p.get("max_temp", ""),
                    "so_special_temp": p.get("special_temp", ""),
                    "so_pdf": so_pdf,
                })
                items = p.get("line_items") or []
                rec["line_item_count"] = len(items)
                # Kept private until the serial runner commits this job to the
                # latest shared line-items store under its cross-process lock.
                rec["_line_items"] = items

        runs = _run_docs(docs) if not so or so_pdf else []
        dr_count = 0
        if runs:
            rec["has_drive_run"] = True
            dr_count = len(runs)
            for href, doc in runs:
                got = await _download_async(context, page.url, href, DRIVE_RUN_DIR / job / _run_filename(job, doc))
                if got and not dr_pdf:
                    dr_pdf = got
            rec["drive_run_pdf"] = dr_pdf or ""
        if folder:
            hits = _run_files_in_folder(Path(folder))
            if hits:
                if not rec.get("has_drive_run"):
                    rec["has_drive_run"] = True
                    dr_pdf = _best_run_path(hits)
                    rec["drive_run_pdf"] = dr_pdf
                dr_count += len(hits)
        if dr_count:
            rec["drive_run_count"] = dr_count
        if dr_pdf:
            _attach_quote_run_parse(rec, dr_pdf)

        if not so:
            rec["status"] = "no-SO"
        elif so_pdf and (not runs or dr_pdf):
            rec["status"] = "ok"
        else:
            rec["status"] = "error"
    except Exception as e:  # noqa: BLE001
        log.warning("  %s: error (%s)", job, e)
        rec["status"] = "error"
    finally:
        await _close_modal_async(page)
    return rec


def _commit_line_items(job: str, rec: Dict[str, Any]) -> None:
    items = rec.pop("_line_items", None)
    if items is None:
        return
    line_items.record_jobs_atomic([{
        "job": job,
        "items": items,
        "co_number": rec.get("co_number"),
        "so_pdf": rec.get("so_pdf", ""),
    }], line_items.backfill_store_path())


def _publish_checkpoint() -> bool:
    """Sync current stores and publish them to the configured order-data branch."""
    try:
        import master_sync

        # Publish once below even when DATA_PUSH_ON_CHANGE is enabled, so a
        # checkpoint containing only new miss/attempt metadata is not skipped.
        master_sync.run("backfill", "line_items", publish=False)
    except Exception as e:  # noqa: BLE001 - publishing never stops the scan
        log.warning("Could not sync the backfill checkpoint to the live master (%s)", e)
    try:
        import data_push

        if data_push.push_data():
            log.info("Published the current backfill checkpoint to the order-data branch.")
            return True
    except Exception as e:  # noqa: BLE001 - publishing never stops the scan
        log.warning("Could not publish the backfill checkpoint (%s)", e)
    return False


# Consecutive not-founds that trigger the dead-session probe: when the saved
# CBC session dies mid-run, EVERY search comes back empty — the overnight run
# recorded ~7K bogus misses that way. On a streak this long, re-search a job
# that is KNOWN to resolve. LOG-ONLY by request: the run keeps going either
# way (misses are retried on the next run); the log just gains a clear marker
# of when the session died so nobody has to guess on Monday.
DEAD_SESSION_STREAK = 15


def _probe_job(records: Dict[str, Dict[str, Any]]) -> str:
    """A job number whose search is known to resolve — a serial-verified 'ok'
    first, any previously-ok record otherwise ('' when none exist yet)."""
    fallback = ""
    for j, r in records.items():
        if r.get("status") != "ok":
            continue
        if r.get("backfill_scan_version"):
            return j
        fallback = fallback or j
    return fallback


async def _run_serial_pass(context, jobs: List[str],
                           records: Dict[str, Dict[str, Any]],
                           dwg: Dict[str, Dict[str, Any]],
                           delay: float, pass_no: int,
                           search_timeout_s: float,
                           doc_timeout_ms: int,
                           run_state: Dict[str, int]) -> int:
    """Process one pass in strict sequence and checkpoint every completed job."""
    page = await _open_backfill_page(context)
    completed = 0
    miss_streak = 0
    try:
        for index, job in enumerate(jobs, start=1):
            rec: Dict[str, Any] | None = None
            previous = records.get(job) or {}
            previous_attempts = (
                _attempt_count(previous)
                if previous.get("backfill_scan_version") == BACKFILL_SCAN_VERSION else 0
            )
            try:
                # watch.py may continue polling. It gets the same lock for its
                # short Sales Order batch, so CBC searches never overlap.
                with cbc_fetch_lock():
                    rec = await process_one_async(
                        page,
                        context,
                        job,
                        dwg.get(job, {}).get("folder", ""),
                        search_timeout_s=search_timeout_s,
                        doc_timeout_ms=doc_timeout_ms,
                    )
                _commit_line_items(job, rec)
            except Exception as e:  # noqa: BLE001
                log.warning("Backfill error for %s: %s", job, e)
                rec = rec or {
                    "job": job,
                    "scanned_at": datetime.now().isoformat(timespec="seconds"),
                    "backfill_scan_version": BACKFILL_SCAN_VERSION,
                }
                rec.pop("_line_items", None)
                rec["status"] = "error"

            rec["backfill_attempts"] = previous_attempts + 1
            records[job] = rec
            # This is intentionally every job: Ctrl+C or a reboot loses at most
            # the order currently in flight, never a batch of finished orders.
            save_progress(records)
            completed += 1
            run_state["processed"] = run_state.get("processed", 0) + 1
            log.info("  pass %d backfill %d/%d  (%s -> %s)",
                     pass_no, index, len(jobs), job, rec.get("status"))
            publish_every = max(0, int(run_state.get("publish_every", 0)))
            if publish_every and run_state["processed"] % publish_every == 0:
                log.info("Publishing checkpoint after %d completed attempt(s)...",
                         run_state["processed"])
                if _publish_checkpoint():
                    run_state["last_published"] = run_state["processed"]

            # Dead-session detector (log-only): a long miss streak is either a
            # genuinely thin stretch of the backlog or a dead login — a probe
            # of a known-resolvable order tells them apart. Either way the run
            # CONTINUES; misses recorded meanwhile are retried on the next run.
            miss_streak = miss_streak + 1 if rec.get("status") == "not-found" else 0
            if miss_streak >= DEAD_SESSION_STREAK:
                miss_streak = 0
                probe = _probe_job(records)
                found = False
                if probe:
                    with contextlib.suppress(Exception):
                        with cbc_fetch_lock():
                            found = await open_order_detail_async(
                                page, probe, search_timeout_s, doc_timeout_ms)
                        await _close_modal_async(page)
                if found:
                    log.info("  %d consecutive misses, but known order %s still "
                             "resolves — thin stretch, carrying on.",
                             DEAD_SESSION_STREAK, probe)
                elif not run_state.get("dead_session_warned"):
                    run_state["dead_session_warned"] = True
                    log.error(
                        "%d consecutive not-founds and known-good order %s ALSO "
                        "failed to resolve — the CBC session may be dead from "
                        "here on. Continuing anyway (by request); these misses "
                        "are retried on the next run. `python login.py` + a "
                        "restart gets it fetching again sooner.",
                        DEAD_SESSION_STREAK, probe or "<none on record>")

            if delay:
                await asyncio.sleep(delay)
    finally:
        with contextlib.suppress(Exception):
            await page.close()
    return completed


async def _run_backfill(jobs: List[str], records: Dict[str, Dict[str, Any]],
                        dwg: Dict[str, Dict[str, Any]],
                        delay: float, passes: int,
                        search_timeout_s: float,
                        doc_timeout_ms: int,
                        run_state: Dict[str, int] | None = None) -> tuple[int, int]:
    if not jobs:
        return 0, 0
    from playwright.async_api import async_playwright

    state = run_state if run_state is not None else {"processed": 0}
    rc = 0
    processed_total = 0
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(storage_state=str(STORAGE_STATE_PATH), accept_downloads=True)
        try:
            # Preflight once so a bad selector/login fails before touching the
            # backlog. Coordinate this page access with the live watcher too.
            with cbc_fetch_lock():
                page = await _open_backfill_page(context)
                try:
                    if await find_search_box_async(page) is None:
                        log.error("Could not find the 'search order' box on %s. Set CBC_SEARCH_SELECTOR in .env "
                                  "to its CSS selector (run `python discover_documents.py --probe %s` to print it), "
                                  "then re-run.", CBC_QUEUE_URL or CBC_URL, jobs[0] if jobs else "<job#>")
                        return 0, 1
                finally:
                    await page.close()

            todo = list(jobs)
            for pass_no in range(1, max(1, passes) + 1):
                log.info("Backfill fetch pass %d/%d: %d job(s), one order at a time...",
                         pass_no, max(1, passes), len(todo))
                try:
                    processed = await _run_serial_pass(
                        context, todo, records, dwg, delay, pass_no,
                        search_timeout_s, doc_timeout_ms, state)
                except Exception as e:  # noqa: BLE001
                    rc = 1
                    log.warning("Serial backfill pass stopped early: %s", e)
                    break
                processed_total += processed
                if pass_no >= max(1, passes):
                    break
                retry = [job for job in todo
                         if (records.get(job) or {}).get("status") in RETRYABLE_STATUSES]
                if not retry:
                    break
                log.info("  %d job(s) incomplete after pass %d; retrying those.",
                         len(retry), pass_no)
                todo = retry
        finally:
            with contextlib.suppress(Exception):
                await context.close()
            with contextlib.suppress(Exception):
                await browser.close()
    return processed_total, rc


# --------------------------------------------------------------------------- #
# Progress store                                                              #
# --------------------------------------------------------------------------- #
def load_progress() -> Dict[str, Dict[str, Any]]:
    if not PROGRESS_PATH.exists():
        return {}
    try:
        return json.loads(PROGRESS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        log.warning("Could not read %s (%s); starting fresh", PROGRESS_PATH, e)
        return {}


def save_progress(records: Dict[str, Dict[str, Any]]) -> None:
    PROGRESS_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = PROGRESS_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(records, indent=2), encoding="utf-8")
    tmp.replace(PROGRESS_PATH)  # atomic — a crash mid-write never corrupts the store


def _is_done(rec: Dict[str, Any], retry_not_found: bool = False) -> bool:
    """Skip trusted answers and misses made by this serial scanner version.

    The old parallel run's ``not-found``/``no-SO`` rows are deliberately not
    trusted. New serial misses need two completed attempts before a restart
    skips them, so interrupting between retry passes still resumes correctly.
    """
    if not rec:
        return False
    if rec.get("status") == "ok":
        return True
    if rec.get("status") in {"not-found", "no-SO"}:
        return (not retry_not_found
                and rec.get("backfill_scan_version") == BACKFILL_SCAN_VERSION
                and _attempt_count(rec) >= REQUIRED_MISS_ATTEMPTS)
    return False


# --------------------------------------------------------------------------- #
# Master workbook (SO/drive-run fields + DWG suffix matrix)                   #
# --------------------------------------------------------------------------- #
def write_workbook(records: Dict[str, Dict[str, Any]], dwg: Dict[str, Dict[str, Any]], path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from excel_writer import _drive_run_label

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")
    dr_font = Font(color="C55A11", bold=True)
    dr_link_font = Font(color="C55A11", bold=True, underline="single")  # Drive Run -> PDF link
    present_fill = PatternFill("solid", fgColor="C6EFCE")  # green: job HAS this drawing
    absent_fill = PatternFill("solid", fgColor="FFC7CE")   # red: it doesn't
    center = Alignment(horizontal="center")

    suffixes = autocad_scan.all_extra_suffixes(dwg)
    # -01/-02 (CW/CCW) aren't shown — nearly every job has them; only the custom
    # extra suffixes carry signal.
    fixed = ["Job #", "Type", "Description", "Size", "Arrangement", "Motor Pos", "Class",
             "Rotation", "Discharge", "% Width", "Special Temp", "CO#", "Quote Run",
             "Quote Run Summary", "Folder", "Order Status"]
    headers = fixed + [f"-{s}" for s in suffixes]
    folder_col = fixed.index("Folder") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill

    jobs = sorted(set(records) | set(dwg))
    for i, job in enumerate(jobs, start=2):
        r = records.get(job, {})
        d = dwg.get(job, {})
        vals = [
            job, d.get("type", ""), r.get("so_design_desc", ""), r.get("so_size", ""),
            r.get("so_arrangement", ""), r.get("so_motor_pos", ""), r.get("so_class", ""),
            r.get("so_rotation", ""), r.get("so_discharge", ""), r.get("so_pct_width", ""),
            r.get("so_special_temp", ""), (f"CO#{r['co_number']}" if r.get("co_number") else ""),
            _drive_run_label(r), r.get("drive_run_summary", ""),
            "", r.get("status", ""),  # "" = Folder placeholder (hyperlinked below)
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(i, c, v)
        # Drive Run cell links to the archived drive-run PDF when we have it.
        if r.get("has_drive_run"):
            dr_cell = ws.cell(i, fixed.index("Quote Run") + 1)
            if r.get("drive_run_pdf"):
                dr_cell.hyperlink = r["drive_run_pdf"]
                dr_cell.font = dr_link_font
            else:
                dr_cell.font = dr_font
        # Link the AutoCAD folder if we scanned one, else the SO archive folder.
        folder = d.get("folder") or (str(Path(r["so_pdf"]).parent) if r.get("so_pdf") else "")
        if folder:
            fcell = ws.cell(i, folder_col, "Open")
            fcell.hyperlink = folder
            fcell.font = link_font
        extras = d.get("extras", {})
        for k, s in enumerate(suffixes, start=len(fixed) + 1):
            cell = ws.cell(i, k)
            if s in extras:  # green + a tiny check; red + blank when absent
                cell.value, cell.fill, cell.alignment = "✓", present_fill, center
            else:
                cell.fill = absent_fill

    if jobs:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs) + 1}"
    ws.freeze_panes = "B2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 6), 48)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #
def _resolve_jobs(args: argparse.Namespace) -> List[str]:
    if args.jobs:
        return args.jobs
    if args.list:
        return jobs_from_list(Path(args.list))
    if args.range:
        return jobs_from_range(args.range[0], args.range[1])
    if args.from_dwg_progress:
        return list(autocad_scan.load_progress().keys())
    return jobs_from_folders(Path(args.root), args.min_job, args.max_job)


def _job_num_key(job: str) -> tuple[int, str]:
    return (int(job), job) if str(job).isdigit() else (-1, str(job))


def _inside_job_caps(job: str, min_job: int = 0, max_job: int = 0) -> bool:
    if not str(job).isdigit():
        return False
    n = int(job)
    return n >= min_job and (not max_job or n <= max_job)


def main(argv: Optional[List[str]] = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    ap = argparse.ArgumentParser(description="Backfill old orders from cbcinsider search.")
    ap.add_argument("jobs", nargs="*", help="Explicit job numbers (default: supported 401000+ AutoCAD folders).")
    ap.add_argument("--list", help="File of job numbers, one per line.")
    ap.add_argument("--range", nargs=2, type=int, metavar=("FIRST", "LAST"), help="Numeric job range.")
    ap.add_argument("--root", default=str(AUTOCAD_JOBS_DIR), help="AutoCAD jobs root for folder enumeration.")
    ap.add_argument("--out", default=str(WORKBOOK_PATH), help="Master workbook path.")
    ap.add_argument("--delay", type=float, default=1.0, help="Seconds to pause between orders.")
    ap.add_argument("--limit", type=int, default=0, help="Stop after N orders this run (0 = no limit).")
    ap.add_argument("--publish-every", type=int, default=DEFAULT_PUBLISH_EVERY,
                    help=f"Publish a Git order-data checkpoint every N completed attempts "
                         f"(default {DEFAULT_PUBLISH_EVERY}; 0 disables publishing).")
    ap.add_argument("--passes", type=int, default=2,
                    help="Retry not-found/no-SO/error jobs within this run (default 2).")
    ap.add_argument("--search-timeout", type=float, default=40.0,
                    help="Seconds to wait for search results/detail to surface per job "
                         "(default 40 — a real hit surfaces in 10-20s, and a miss burns "
                         "the FULL wait twice, so this dominates a miss-heavy sweep's cost).")
    ap.add_argument("--doc-timeout", type=float, default=120.0,
                    help="Seconds to wait for document links after opening detail (default 120).")
    ap.add_argument("--min-job", type=int, default=DEFAULT_CBC_SEARCH_MIN_JOB,
                    help=f"On a folder sweep, skip job numbers below this (default {DEFAULT_CBC_SEARCH_MIN_JOB}; "
                         "older orders need the legacy CBC lookup).")
    ap.add_argument("--max-job", type=int, default=0, help="Skip job numbers above this (0 = no cap).")
    ap.add_argument("--force", action="store_true",
                    help="Reprocess selected jobs even if saved progress says they are done.")
    ap.add_argument("--newest-first", action="store_true",
                    help="Process higher job numbers first instead of the folder enumeration order.")
    ap.add_argument("--retry-not-found", action="store_true",
                    help="Force another check of misses already made by the current serial scanner.")
    ap.add_argument("--from-dwg-progress", action="store_true",
                    help="Use the saved AutoCAD scan progress job list instead of walking the root folder.")
    ap.add_argument("--rescan", action="store_true", help="Ignore saved progress.")
    args = ap.parse_args(sys.argv[1:] if argv is None else argv)

    if not STORAGE_STATE_PATH.exists():
        log.error("No saved session at %s. Run `python login.py` first.", STORAGE_STATE_PATH)
        return 1

    targets = _resolve_jobs(args)
    if args.from_dwg_progress:
        targets = [j for j in targets if _inside_job_caps(j, args.min_job, args.max_job)]
    if args.newest_first:
        targets = sorted(targets, key=_job_num_key, reverse=True)
    log.info("Backfill target set: %d job(s).", len(targets))
    records = {} if args.rescan else load_progress()
    dwg = autocad_scan.load_progress()  # read-only merge of the DWG scan, if it's been run
    if dwg:
        log.info("Merging AutoCAD DWG scan for %d job(s).", len(dwg))

    rc = 0
    pending: List[str] = []
    for job in targets:
        if args.limit and len(pending) >= args.limit:
            break
        if not args.rescan and not args.force and _is_done(records.get(job, {}),
                                                           retry_not_found=args.retry_not_found):
            continue
        pending.append(job)

    trusted_done = sum(
        1 for job in targets
        if not args.rescan and not args.force
        and _is_done(records.get(job, {}), retry_not_found=args.retry_not_found)
    )
    log.info("Resume checkpoint: %d trusted complete; %d pending this run%s.",
             trusted_done, len(pending), f" (limit {args.limit})" if args.limit else "")

    processed = 0
    interrupted = False
    run_state = {"processed": 0, "publish_every": max(0, args.publish_every)}
    if pending:
        try:
            processed, rc = asyncio.run(_run_backfill(
                pending, records, dwg, args.delay, args.passes,
                args.search_timeout, int(args.doc_timeout * 1000), run_state))
        except KeyboardInterrupt:
            interrupted = True
            processed = run_state["processed"]
            rc = 130
            log.info("Backfill stopped. Every completed order is saved; run the same command to resume.")
    else:
        log.info("No pending jobs to backfill.")

    save_progress(records)
    out: Path | None = None
    if not interrupted:
        try:
            out = write_workbook(records, dwg, Path(args.out))
        except PermissionError as e:
            rc = 1
            log.warning("Could not write %s (%s). Close the workbook and rerun to refresh it; "
                        "progress JSON and line-item stores were already saved.", args.out, e)
    if (args.publish_every > 0
            and run_state.get("last_published") != run_state["processed"]):
        _publish_checkpoint()
    by_status: Dict[str, int] = {}
    for r in records.values():
        by_status[r.get("status", "?")] = by_status.get(r.get("status", "?"), 0) + 1
    log.info("Done: processed %d this run; store has %d job(s).", processed, len(records))
    log.info("  status breakdown: %s", ", ".join(f"{k}={v}" for k, v in sorted(by_status.items())))
    if by_status.get("not-found"):
        log.info("  %d job(s) not found via search - if that seems high, the search box may need "
                 "CBC_SEARCH_SELECTOR / CBC_SEARCH_BUTTON set (see `discover_documents.py --probe`).",
                 by_status["not-found"])
    if out:
        log.info("  Wrote %s", out)
    return rc


if __name__ == "__main__":
    sys.exit(main())
