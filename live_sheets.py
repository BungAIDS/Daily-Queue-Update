"""Pure 'sheet model' for the live master workbook — what each tab contains and
how each cell should look, with NO Excel/COM dependency.

This is the single source of truth for the live workbook's content, kept
separate from how it's pushed into Excel (live_excel.py drives the desktop app
via COM). Keeping it pure means it's unit-tested directly (test_live_sheets.py)
and the daily openpyxl report and the live COM report can't drift — both lean on
excel_writer's column list and label helpers.

A tab is a `Sheet`: a grid of `Cell`s plus a freeze pane and AutoFilter extent.
Each Cell carries a value and *named* style intents (fills/fonts as strings);
the renderer maps those names to concrete Excel colors. The names mirror
excel_writer so the look matches the daily report.
"""
from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional

import engineers
from excel_writer import (COLUMNS, QUEUE_HEADERS, MONEY_FMT, _co_label,
                          _drive_run_label, _flags_str, _parse_date,
                          _parse_money, _dwg_suffixes, folder_of, split_arrangement,
                          split_size)

# --- named styles (resolved to real colors by live_excel) ------------------- #
F_HEADER = "header"           # blue header bg + white bold
F_SECTION = "section"         # bold section title
F_LINK = "link"               # blue underline hyperlink
F_DRIVE_RUN = "drive_run"     # orange bold (highly-custom)
F_DRIVE_RUN_LINK = "drive_run_link"
F_RED = "red"                 # red bold (a change order landed)
F_NOTE = "note"               # muted gray (e.g. the Changes 'last updated' stamp)

FILL_HEADER = "header"
FILL_OVERDUE = "overdue"      # End Date in the past (overdue / red)
FILL_DUETODAY = "duetoday"    # End Date today -> overdue (red) tomorrow (orange)
FILL_SOON = "soon"            # End Date within 3 days (gold)
FILL_NEW = "new"              # new/arrived today (no urgency)
FILL_OVERDUE_NEW = "overdue_new"
FILL_DUETODAY_NEW = "duetoday_new"
FILL_SOON_NEW = "soon_new"
FILL_DWG_YES = "dwg_yes"      # green ✓
FILL_DWG_NO = "dwg_no"        # red (missing)
FILL_SEP = "sep"              # the vertical divider column between the two matrices
# Grey shading on the 'after' row of a change instance, getting darker for each
# later instance the same order picks up through the day.
FILL_CHANGE1 = "chg1"
FILL_CHANGE2 = "chg2"
FILL_CHANGE3 = "chg3"
FILL_CHANGE4 = "chg4"
_CHANGE_FILLS = [FILL_CHANGE1, FILL_CHANGE2, FILL_CHANGE3, FILL_CHANGE4]


@dataclass
class Cell:
    value: Any = ""
    fill: Optional[str] = None
    font: Optional[str] = None
    link: Optional[str] = None
    number_format: Optional[str] = None
    center: bool = False
    comment: Optional[str] = None   # hover note (e.g. the CO# change-order history)
    volatile: bool = False          # value changes every cycle (e.g. the 'Last updated'
                                    # stamp): excluded from the render fingerprint so it
                                    # doesn't force a full repaint, refreshed in place
    overflow: bool = False          # free-text cell meant to overrun its empty right-hand
                                    # neighbors (e.g. the Changes tab's 'What changed'):
                                    # excluded from its column's autofit so the column
                                    # stays sized to its other content and this spills
    colspan: int = 1                # merge this cell across N columns (the cells it covers
                                    # stay in the grid as positional spacers): e.g. the
                                    # Changes 'Job #' header spanning its blank spacer col


@dataclass
class Sheet:
    name: str
    grid: List[List[Cell]] = field(default_factory=list)
    freeze: Optional[str] = "B2"
    autofilter_a1: Optional[str] = None   # e.g. "A1:AB57"; None = no filter
    hidden: bool = False                  # data-only sheet, hidden from the tab bar
    picker: Optional[Dict[str, str]] = None  # input cell w/ dropdown: {cell, source,
                                             # comment} — styled + validated by the
                                             # renderer; its typed value survives repaints
    names: Optional[Dict[str, str]] = None   # workbook defined names to (re)point at this
                                             # sheet, {name: "'Tab'!$A$5"} — stable link
                                             # targets that survive rows shifting

    # -- builders the sheet functions below use to assemble the grid --
    def row(self, cells: List[Cell]) -> None:
        self.grid.append(cells)

    def blank(self, n: int = 1) -> None:
        for _ in range(n):
            self.grid.append([])

    @property
    def ncols(self) -> int:
        return max((len(r) for r in self.grid), default=0)

    @property
    def nrows(self) -> int:
        return len(self.grid)


# --------------------------------------------------------------------------- #
# Shared helpers                                                               #
# --------------------------------------------------------------------------- #
def added_label(job: Dict[str, Any], ref: Optional[datetime] = None) -> str:
    """The 'Added' label: the AM/PM time if it was added today, just the date if
    it was added earlier, or 'NO DATA' only when there's no add timestamp at all."""
    iso = job.get("_added_iso") or job.get("_first_seen") or ""
    try:
        dt = datetime.fromisoformat(iso)
    except (ValueError, TypeError):
        return "NO DATA"
    ref = ref or datetime.now()
    if dt.date() == ref.date():
        return dt.strftime("%#I:%M %p") if _is_windows() else dt.strftime("%-I:%M %p")
    return dt.strftime("%b %#d, %Y") if _is_windows() else dt.strftime("%b %-d, %Y")


def last_out_label(job: Dict[str, Any], ref: Optional[datetime] = None) -> str:
    """The 'Last Out' label: when the order was MOST RECENTLY removed before its
    current stint (the entry's last_out, injected as `_last_out`) — a time if that
    was today, else a short date. Blank if the order has never left the board."""
    iso = job.get("_last_out") or ""
    try:
        dt = datetime.fromisoformat(iso)
    except (ValueError, TypeError):
        return ""
    ref = ref or datetime.now()
    if dt.date() == ref.date():
        return dt.strftime("%#I:%M %p") if _is_windows() else dt.strftime("%-I:%M %p")
    return dt.strftime("%b %#d, %Y") if _is_windows() else dt.strftime("%b %-d, %Y")


def added_date(job: Dict[str, Any]) -> Optional[date]:
    """The date an order most recently came onto the board (from _added_iso /
    _first_seen) — i.e. the date the 'Added' column reflects — or None if unknown."""
    iso = job.get("_added_iso") or job.get("_first_seen") or ""
    try:
        return datetime.fromisoformat(iso).date()
    except (ValueError, TypeError):
        return None


def prev_business_day(d: date) -> date:
    """The most recent business day strictly before `d` (weekends skipped): Mon ->
    Fri, Sun -> Fri, otherwise the day before. Holidays are not accounted for."""
    wd = d.weekday()                  # Mon=0 .. Sun=6
    if wd == 0:
        return d - timedelta(days=3)
    if wd == 6:
        return d - timedelta(days=2)
    return d - timedelta(days=1)


def _is_windows() -> bool:
    import os
    return os.name == "nt"


# Display-only header abbreviations (the full label stays the internal key, so
# change-log field matching and column lookups are unaffected). Keeps wide columns
# from being held open by a long header when the data itself is short.
_HEADER_ABBR = {"Arrangement": "Arr.", "Quote Run Details": "Run Details"}


def _abbrev_header(h: str) -> str:
    return _HEADER_ABBR.get(h, h)


def _header_cells(headers: List[str]) -> List[Cell]:
    return [Cell(_abbrev_header(h), fill=FILL_HEADER, font=F_HEADER) for h in headers]


def _money_cell(raw: str) -> Cell:
    raw = (raw or "").strip()
    if not raw:
        return Cell("")
    return Cell(_parse_money(raw), number_format=MONEY_FMT)


def _job_value_cells(j: Dict[str, Any], columns: Optional[List] = None,
                     co_changed: bool = False, arrange_comment: bool = False) -> List[Cell]:
    """The cells for one job across `columns` (defaults to the full COLUMNS) —
    mirrors excel_writer._write_job_row (hyperlinks, CO#, drive-run label, money,
    flags), as style-tagged Cells. `arrange_comment` trims the Arrangement cell to
    its short 'A/X' code and moves any descriptive suffix to a hover note (used on
    the display tabs; Order History keeps the full text to avoid re-writing the
    whole log)."""
    columns = COLUMNS if columns is None else columns
    cells: List[Cell] = []
    linked_idx = set()
    for idx, (_h, key) in enumerate(columns):
        if key == "job":
            c = Cell(j.get("job", ""))
            so = (j.get("so_pdf") or "").strip()
            if so and j.get("job"):
                # Links straight to the latest Sales Order PDF. The watcher keeps
                # so_pdf pointed at the newest revision (sales_orders
                # .refresh_sales_orders / a re-fetch on a change order), so a CO
                # renaming the file (… (original).pdf -> … CO1.pdf) follows here.
                c.link, c.font = so, F_LINK
                linked_idx.add(idx)
        elif key == "folder":
            folder = (j.get("job_folder") or "").strip()
            c = Cell((j.get("job_type") or "Open") if folder else "")
            if folder:
                c.link, c.font = folder, F_LINK
                linked_idx.add(idx)
        elif key == "co":
            c = Cell(_co_label(j))
        elif key == "drive_run":
            c = Cell(_drive_run_label(j))
            dr = (j.get("drive_run_pdf") or "").strip()
            if dr:
                # With more than one run, link to the folder that holds the
                # downloaded runs, not just one of the files.
                c.link = folder_of(dr) if (j.get("drive_run_count") or 0) > 1 else dr
                c.font = F_DRIVE_RUN_LINK
                linked_idx.add(idx)
            elif j.get("has_drive_run"):
                c.font = F_DRIVE_RUN
        elif key == "total_price":
            c = _money_cell(j.get("total_price", ""))
        elif key == "so_arrangement" and arrange_comment:
            # Keep the column to 'A/X'; hover the cell for the descriptive suffix.
            code, note = split_arrangement(j.get("so_arrangement", ""))
            c = Cell(code)
            if note:
                c.comment = note
        elif key == "so_size" and arrange_comment:
            # Keep the column to the main size; hover for any trailing detail.
            main, note = split_size(j.get("so_size", ""))
            c = Cell(main)
            if note:
                c.comment = note
        elif key == "flags":
            c = Cell(_flags_str(j))
        elif key == "engineers":
            c = Cell(engineers.cell_text(j))
        elif key == "dwg_reuse_label":
            # Backlog order(s) with custom DWGs for this order's rare features —
            # linked to the top candidate's CAD folder, full shortlist on hover.
            c = Cell(j.get("dwg_reuse_label", ""))
            sugg = j.get("dwg_reuse") or []
            if c.value and sugg:
                if j.get("dwg_reuse_note"):
                    c.comment = j["dwg_reuse_note"]
                folder = (sugg[0].get("folder") or "").strip()
                if folder:
                    c.link, c.font = folder, F_LINK
                    linked_idx.add(idx)
        else:
            c = Cell(j.get(key, ""))
        cells.append(c)
    # A change order that landed -> non-link cells go red.
    if co_changed:
        for idx, c in enumerate(cells):
            if idx not in linked_idx:
                c.font = F_RED
    return cells


def _row_fill(j: Dict[str, Any], today: date, is_new: bool) -> Optional[str]:
    end = _parse_date(j.get("end_date", ""))
    soon = today + timedelta(days=3)
    if end is not None and end < today:
        return FILL_OVERDUE_NEW if is_new else FILL_OVERDUE        # past due -> red
    if end is not None and end == today:                          # due today -> red tomorrow
        return FILL_DUETODAY_NEW if is_new else FILL_DUETODAY      #             -> orange
    if end is not None and end <= soon:
        return FILL_SOON_NEW if is_new else FILL_SOON             # within 3 days -> gold
    return FILL_NEW if is_new else None


def _dwg_header_cells(suffixes: List[str]) -> List[Cell]:
    return [Cell(f"-{s}", fill=FILL_HEADER, font=F_HEADER) for s in suffixes]


def _dwg_row_cells(j: Dict[str, Any], suffixes: List[str]) -> List[Cell]:
    extras = j.get("dwg_extras") or {}
    out = []
    for s in suffixes:
        if s in extras:
            out.append(Cell("✓", fill=FILL_DWG_YES, center=True))
        else:
            out.append(Cell("", fill=FILL_DWG_NO))
    return out


def _a1(rows: int, cols: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"A1:{get_column_letter(max(cols, 1))}{max(rows, 1)}"


# --------------------------------------------------------------------------- #
# Live Queue (the master board)                                               #
# --------------------------------------------------------------------------- #
def full_queue_sheet(
    jobs: List[Dict[str, Any]],
    today: date,
    new_ids: Optional[set] = None,
    co_changed_ids: Optional[set] = None,
    ref: Optional[datetime] = None,
    name: str = "Live Queue",
) -> Sheet:
    """The full board: an 'Added' column, every Full Queue column, then the
    green-✓/red DWG matrix. Urgency + new-order fills, hyperlinks, a totals
    footer, freeze panes, and AutoFilter — same look as the daily Full Queue."""
    new_ids = new_ids or set()
    co_changed_ids = co_changed_ids or set()
    suffixes = _dwg_suffixes(jobs)
    headers = ["Added"] + list(QUEUE_HEADERS) + [f"-{s}" for s in suffixes]

    sh = Sheet(name)
    sh.row([Cell("Added", fill=FILL_HEADER, font=F_HEADER)]
           + _header_cells(list(QUEUE_HEADERS)) + _dwg_header_cells(suffixes))

    total = 0.0
    for j in jobs:
        is_new = j.get("job") in new_ids
        fill = _row_fill(j, today, is_new)
        added = Cell(added_label(j, ref=ref))
        std = _job_value_cells(j, co_changed=j.get("job") in co_changed_ids, arrange_comment=True)
        # Apply the urgency/new row fill across Added + standard columns (not the
        # DWG cells, which keep their own green/red).
        if fill is not None:
            added.fill = fill
            for c in std:
                c.fill = fill
        sh.row([added] + std + _dwg_row_cells(j, suffixes))
        total += _parse_money(j.get("total_price", ""))

    # Totals footer two rows below the data.
    sh.blank(2)
    footer = [Cell("") for _ in headers]
    footer[0] = Cell(f"Total jobs: {len(jobs)}", font=F_SECTION)
    price_col = 1 + QUEUE_HEADERS.index("Total Price")  # +1 for the Added column
    if price_col - 1 >= 0:
        footer[price_col - 1] = Cell("Total $ in process:", font=F_SECTION)
    footer[price_col] = Cell(total, number_format=MONEY_FMT, font=F_SECTION)
    sh.row(footer)

    sh.freeze = "C2"  # keep header row + the Added & Job# columns visible
    sh.autofilter_a1 = _a1(len(jobs) + 1, len(headers))
    return sh


# --------------------------------------------------------------------------- #
# Changes (both baselines, each date-labeled)                                 #
# --------------------------------------------------------------------------- #
def _job_table(sh: Sheet, title: str, jobs: List[Dict[str, Any]],
               extra_headers: Optional[List[str]] = None,
               extra: Optional[Any] = None,
               time: Optional[Any] = None) -> None:
    """A titled mini-table of full job rows (used by the Changes sections).

    With `time` (a job -> ISO-timestamp callback) the table leads with a Time
    column in the same format and position as 'Orders that changed today':
    Time in column A, Job # in column B, Folder in column C. Without it, a
    blank spacer column is inserted right after Job # instead, so Folder still
    lands in column C — either way Folder, Quote Run, CO#, … align across all
    the tab's sections."""
    sh.row([Cell(f"{title} ({len(jobs)})", font=F_SECTION)])
    if not jobs:
        sh.row([Cell("(none)")])
        sh.blank()
        return
    if time is not None:
        sh.row(_header_cells(["Time"] + list(QUEUE_HEADERS) + (extra_headers or [])))
        for j in jobs:
            cells = _job_value_cells(j, co_changed=False, arrange_comment=True)
            row = [Cell(_fmt_time(str(time(j) or "")))] + cells
            if extra is not None:
                row.append(Cell(extra(j)))
            sh.row(row)
        sh.blank()
        return
    # [Job #][spacer][Folder, Quote Run, …]. The spacer keeps Folder in column C
    # (aligned with the changed table); Job # is merged across it (colspan=2) so
    # there's no cell wall between Job # and the spacer. 'Orders that changed today'
    # is left alone — its Time and Job # stay as two separate cells.
    headers = [QUEUE_HEADERS[0], ""] + list(QUEUE_HEADERS[1:]) + (extra_headers or [])
    header_cells = _header_cells(headers)
    header_cells[0].colspan = 2
    sh.row(header_cells)
    for j in jobs:
        cells = _job_value_cells(j, co_changed=False, arrange_comment=True)
        cells[0].colspan = 2                       # Job # spans the blank spacer (col B)
        cells = [cells[0], Cell("")] + cells[1:]
        if extra is not None:
            cells = cells + [Cell(extra(j))]
        sh.row(cells)
    sh.blank()


def fmt_time(iso: str) -> str:
    """An ISO timestamp as an AM/PM time, e.g. '3:53 PM'."""
    try:
        dt = datetime.fromisoformat(iso)
    except (ValueError, TypeError):
        return iso or ""
    return dt.strftime("%#I:%M %p") if _is_windows() else dt.strftime("%-I:%M %p")


_fmt_time = fmt_time   # internal alias


def fmt_datetime(when: "datetime | str") -> str:
    """A date + AM/PM time, e.g. 'Jun 18, 2026 3:53 PM' — for the Changes tab's
    'last updated' stamp. Accepts a datetime or an ISO string."""
    if isinstance(when, str):
        try:
            when = datetime.fromisoformat(when)
        except (ValueError, TypeError):
            return when or ""
    fmt = "%b %#d, %Y %#I:%M %p" if _is_windows() else "%b %-d, %Y %-I:%M %p"
    return when.strftime(fmt)


def _events_table(sh: Sheet, title: str, headers: List[str],
                  rows: List[List[Any]]) -> None:
    sh.row([Cell(f"{title} ({len(rows)})", font=F_SECTION)])
    if not rows:
        sh.row([Cell("(none)")])
        sh.blank()
        return
    sh.row(_header_cells(headers))
    for r in rows:
        sh.row([c if isinstance(c, Cell) else Cell(c) for c in r])
    sh.blank()


def _suffix_comment_cell(label: str, value: Any, **cell_kw) -> Cell:
    """A changed-field cell that mirrors the display tabs: for Arrangement / Size
    show the short code and move the descriptive suffix to a hover note (so the
    column stays narrow); any other field is shown as-is. Extra Cell kwargs
    (fill/font) pass through."""
    if label == "Arrangement":
        code, note = split_arrangement(value)
    elif label == "Size":
        code, note = split_size(value)
    else:
        return Cell(value, **cell_kw)
    c = Cell(code, **cell_kw)
    if note:
        c.comment = note
    return c


def _orders_changed_table(sh: Sheet, field_events: List[Dict[str, Any]],
                          order_lookup: Optional[Dict[str, Dict[str, Any]]] = None) -> None:
    """'Orders that changed today' as a running history, laid out in the same
    column order as the rest of the workbook: a leading change Time, then every
    Live Queue column. For each order — a white 'was' row showing the order's full
    row with each field that changed today rolled back to its start-of-day value
    (in red), then ONE row per change instance (a poll in which the order changed),
    each stamped with that poll's time and showing only the fields that moved in it
    (red); the prior row's value carries forward as the implied 'old', so a value
    is never repeated. Instance rows shade progressively darker; white marks the
    order. Changed fields with no queue column (e.g. Drawings, Unapproved / Credit
    Hold) are appended as extra trailing columns so no change is lost. `order_lookup`
    (job# -> the order's job dict) supplies the full 'was' row; an order missing
    from it falls back to just its Job # and Customer."""
    order_lookup = order_lookup or {}
    by_job: Dict[str, Dict[str, Any]] = {}
    for e in sorted(field_events, key=lambda x: x.get("time", "")):   # oldest first
        jn, label = str(e.get("job", "") or ""), e.get("field", "") or ""
        old, new = e.get("old", ""), e.get("new", "")
        if not jn or not label or label == "Customer" or str(old) == str(new):
            continue
        rec = by_job.setdefault(jn, {"customer": "", "instances": {}})
        rec["customer"] = e.get("customer", "") or rec["customer"]
        # An instance = one poll (same timestamp); collect every field that moved
        # in it as {field: (old, new)}.
        rec["instances"].setdefault(e.get("time", ""), {})[label] = (old, new)

    orders: Dict[str, Dict[str, Any]] = {}
    for jn, rec in by_job.items():
        times = [t for t in sorted(rec["instances"]) if rec["instances"][t]]   # oldest first
        if not times:
            continue
        first_old: Dict[str, Any] = {}
        for t in times:                                # oldest first -> first old per field
            for label, (old, _new) in rec["instances"][t].items():
                first_old.setdefault(label, old)
        orders[jn] = {"customer": rec["customer"], "time": times[-1],
                      "fields": set(first_old),
                      # each instance as (poll time, {field: new value}), oldest first
                      "steps": [(t, {l: n for l, (_o, n) in rec["instances"][t].items()})
                                for t in times],
                      "baseline": first_old}

    sh.row([Cell(f"Orders that changed today ({len(orders)})", font=F_SECTION)])
    if not orders:
        sh.row([Cell("(none)")])
        sh.blank()
        return
    changed: set = set()
    for o in orders.values():
        changed |= o["fields"]
    qh_idx = {h: i for i, h in enumerate(QUEUE_HEADERS)}
    dwg_idx = qh_idx["DWG Reuse"]
    queue_headers = [h for h in QUEUE_HEADERS if h != "DWG Reuse"]
    extra_cols = sorted(changed - set(QUEUE_HEADERS))    # changed fields with no queue column
    sh.row(_header_cells(["Time"] + queue_headers + extra_cols + ["DWG Reuse"]))
    for jn in sorted(orders, key=lambda j: orders[j]["time"], reverse=True):  # most recent first
        o = orders[jn]
        # 'was' row (white): the order's full row (so it reads like every other
        # tab), with each field that changed today rolled back to its start-of-day
        # value in red. No timestamp — it's the state before today's first change.
        base_job = dict(order_lookup.get(jn) or {})
        base_job.setdefault("job", jn)
        base_job.setdefault("customer", o["customer"])
        was = _job_value_cells(base_job, co_changed=False, arrange_comment=True)
        for label in o["fields"]:
            ci = qh_idx.get(label)
            if ci is not None:
                was[ci] = _suffix_comment_cell(label, o["baseline"][label], font=F_RED)
        extra_was = [Cell(o["baseline"][label], font=F_RED) if label in o["fields"] else Cell("")
                     for label in extra_cols]
        dwg_was = was.pop(dwg_idx)
        sh.row([Cell("")] + was + extra_was + [dwg_was])
        # one row per instance, progressively darker, stamped with the poll time and
        # showing only the fields that moved in that poll.
        for step_i, (t, step) in enumerate(o["steps"]):
            fill = _CHANGE_FILLS[min(step_i, len(_CHANGE_FILLS) - 1)]
            cells = [Cell("", fill=fill) for _ in QUEUE_HEADERS]
            for label, newval in step.items():
                ci = qh_idx.get(label)
                if ci is not None:
                    cells[ci] = _suffix_comment_cell(label, newval, fill=fill, font=F_RED)
            extra_cells = [_suffix_comment_cell(label, step[label], fill=fill,
                                                font=F_RED) if label in step
                           else Cell("", fill=fill) for label in extra_cols]
            dwg_cell = cells.pop(dwg_idx)
            sh.row([Cell(_fmt_time(t), fill=fill)] + cells + extra_cells + [dwg_cell])
    sh.blank()


def _co_change_desc(order: Dict[str, Any], co_num: Any) -> str:
    """The change-order description for CO#<co_num> from the order's co_history
    (the part after the prefix). Some CBC revisions are numbered one higher than
    the newest printed CO note; in that case use the closest preceding note rather
    than leaving the Changes table blank."""
    cn = str(co_num or "").strip()
    target = int(cn) if cn.isdigit() else None
    candidates: List[tuple[int, str]] = []

    def description(line: str) -> str:
        if ":" in line:
            return line.split(":", 1)[1].strip()
        dash = re.search(r"\s+-\s+", line)
        return line[dash.end():].strip() if dash else line.strip()

    for line in (order.get("co_history") or []):
        m = re.match(r"\s*C\s*/?\s*O\s*#?\s*(\d+)", str(line), re.I)
        if not m:
            continue
        number = int(m.group(1))
        text = str(line).strip()
        if m.group(1) == cn:
            return description(text)
        candidates.append((number, text))
    if not candidates:
        return ""
    if target is None:
        return description(max(candidates, key=lambda item: item[0])[1])
    preceding = [item for item in candidates if item[0] < target]
    chosen = (max(preceding, key=lambda item: item[0]) if preceding
              else min(candidates, key=lambda item: abs(item[0] - target)))
    return description(chosen[1])


def changes_sheet(
    new_today: List[Dict[str, Any]],
    change_events: List[Dict[str, Any]],
    removed_today: List[Dict[str, Any]],
    date_str: str,
    updated_at: Optional[str] = None,
    order_lookup: Optional[Dict[str, Dict[str, Any]]] = None,
    name: str = "Changes",
) -> Sheet:
    """Today's activity log:
      - New orders today (new as of today, with their Added time).
      - Change orders today (CO# increases — restored change-order tracking).
      - Orders that changed today: each changed order as a full queue row rolled
        back to its start-of-day state, then a time-stamped line per change
        instance showing what moved (newest order first), in the same column
        order as the rest of the workbook.
      - Removed / completed today.
    `change_events` is the day's change log (see change_log.py); `updated_at` is a
    display string stamped at the top so users can see the tab is live;
    `order_lookup` maps job# -> the order's job dict (the full row for the
    change-order and orders-that-changed tables)."""
    order_lookup = order_lookup or {}
    sh = Sheet(name, freeze=None)
    sh.row([Cell(f"Changes — {date_str}", font=F_SECTION)])
    if updated_at:
        sh.row([Cell(f"Last updated {updated_at}", font=F_NOTE, volatile=True)])
    sh.blank()

    # Leads with the arrival time (same format/position as the changed table's
    # Time column), which replaces the old trailing 'Added' column.
    _job_table(sh, "New orders today", new_today,
               time=lambda j: j.get("_added_iso") or j.get("_first_seen") or "")

    newest_first = sorted(change_events, key=lambda e: e.get("time", ""), reverse=True)
    co_events = [e for e in newest_first if e.get("field") == "CO#"]
    qh_idx = {h: i for i, h in enumerate(QUEUE_HEADERS)}

    def _co_row(e: Dict[str, Any]) -> List[Any]:
        o = order_lookup.get(str(e.get("job", "")), {})
        # Reuse the standard cell builder for the linked Folder / Quote Run cells so
        # they read exactly like the other tables; pull just those two out by index.
        full = _job_value_cells(o, arrange_comment=True) if o else None
        col = (lambda name: full[qh_idx[name]]) if full else (lambda name: Cell(""))
        return [_fmt_time(e.get("time", "")), e.get("job", ""),
                col("Folder"), col("Quote Run"),
                # CO# column like the other tables: just the current CO# this change
                # landed on (formatted via _co_label), not a CO#old -> CO#new string.
                Cell(_co_label({"co_number": e.get("new")})),
                o.get("oper", ""), o.get("design", ""), e.get("customer", ""),
                # Free-text description: let it overrun the empty cells to its right
                # rather than widen the column (it's the table's last column).
                Cell(_co_change_desc(o, e.get("new")), overflow=True)]
    _events_table(sh, "Change orders today",
                  ["Time", "Job #", "Folder", "Quote Run", "CO#", "Oper", "Design",
                   "Customer", "What changed"],
                  [_co_row(e) for e in co_events])

    field_events = [e for e in newest_first if e.get("field") != "CO#"]
    _orders_changed_table(sh, field_events, order_lookup=order_lookup)

    _job_table(sh, "Removed / completed today", removed_today,
               time=lambda j: j.get("_left_iso") or "")
    return sh


# --------------------------------------------------------------------------- #
# History                                                                      #
# --------------------------------------------------------------------------- #
def history_sheet(history: Dict[str, Any], name: str = "Order History") -> Sheet:
    """Archived orders (left the queue, not yet returned), newest departure
    first, with the DWG matrix appended like the Full Queue."""
    entries = sorted((history or {}).values(),
                     key=lambda e: e.get("last_seen", ""), reverse=True)
    snaps = [e.get("snapshot", {}) for e in entries]
    suffixes = _dwg_suffixes(snaps)
    headers = list(QUEUE_HEADERS) + ["Last Seen"]

    sh = Sheet(name)
    sh.row(_header_cells(headers) + _dwg_header_cells(suffixes))
    if not entries:
        sh.row([Cell("(no archived orders yet — a job appears here after it drops "
                     "off the queue)")])
        return sh
    for e in entries:
        snap = e.get("snapshot", {})
        sh.row(_job_value_cells(snap, co_changed=False)
               + [Cell(e.get("last_seen", ""))] + _dwg_row_cells(snap, suffixes))
    sh.autofilter_a1 = _a1(len(entries) + 1, len(headers) + len(suffixes))
    return sh


# --------------------------------------------------------------------------- #
# Similar Orders: pick a queue order at the top -> its ranked lookalikes       #
# appear instantly (an Excel FILTER spill over the hidden Similar Data sheet,  #
# so no macros and no waiting for the next poll). The watcher refreshes the    #
# data sheet; the visible tab's model is layout-only, so a repaint (which      #
# would wipe the picked value) only happens when the layout itself changes —   #
# and the renderer preserves the picker cell's value even then.               #
# --------------------------------------------------------------------------- #
SIMILAR_ORDERS_TAB = "Similar Orders"
SIMILAR_DATA_TAB = "Similar Data"
SIMILAR_HEADERS = ["Similar Order", "Customer", "Score", "Custom DWGs",
                   "Shared Sales-Order items", "CAD Folder"]
SIMILAR_PICKER_CELL = "B1"
_SIM_PICKER_COL = 9   # column I on the data sheet: the dropdown's source list


def _sim_name(job: str) -> str:
    """The workbook defined name for a job's group on Similar Data. Stable, so
    the Live Queue 'Similar' links never change when rows shift — one order
    joining/leaving the board used to rewrite ~every row below it just to
    re-point row-number anchors."""
    return "SIM_" + re.sub(r"[^0-9A-Za-z_]", "_", str(job))


def similar_data_sheet(rows: List[Dict[str, Any]], queue_orders: List[str]) -> Sheet:
    """The flat table behind the Similar Orders tab: one row per (queue order,
    similar order) pair — grouped by queue order, best score first, each group's
    first row shaded so the Live Queue's 'Similar' column can deep-link straight
    to it. Column I carries every on-board order as the picker's dropdown list
    (so an order with no matches is still pickable). The Queue Order value
    repeats on every row on purpose: the picker tab's FILTER matches on it."""
    sh = Sheet(SIMILAR_DATA_TAB, freeze="A2")
    sh.names = {}
    sh.row(_header_cells(["Queue Order"] + SIMILAR_HEADERS)
           + [Cell(""), _header_cells(["Queue Orders"])[0]])
    prev = None
    for i in range(max(len(rows), len(queue_orders))):
        r = rows[i] if i < len(rows) else None
        if r:
            first = r["job"] != prev
            prev = r["job"]
            if first:  # stable deep-link target for this order's group
                sh.names[_sim_name(r["job"])] = f"'{SIMILAR_DATA_TAB}'!$A${i + 2}"
            fill = FILL_NEW if first else None   # grey band starts each group
            folder = (r.get("folder") or "").strip()
            vals = [Cell(r["job"], fill=fill, font=F_SECTION if first else None),
                    Cell(r["similar"], fill=fill), Cell(r.get("customer", ""), fill=fill),
                    Cell(r.get("score", ""), fill=fill), Cell(r.get("dwg", ""), fill=fill),
                    Cell(r.get("shared", ""), fill=fill, overflow=True),
                    Cell(folder, fill=fill, link=folder or None,
                         font=F_LINK if folder else None)]
        else:
            vals = [Cell("")] * 7
        picker = Cell(queue_orders[i]) if i < len(queue_orders) else Cell("")
        sh.row(vals + [Cell(""), picker])
    return sh


def similar_anchor(rows: List[Dict[str, Any]], job: str) -> str:
    """The internal link the Live Queue 'Similar' cell opens ('' when the job
    has no rows): a defined name pointing at the job's group on Similar Data
    (see `_sim_name`), so the link — and therefore the row's signature — never
    changes when other groups grow/shrink and shift the row numbers."""
    job = str(job)
    return f"#{_sim_name(job)}" if any(r["job"] == job for r in rows) else ""


def similar_orders_sheet(n_rows: int, n_queue: int) -> Sheet:
    """The visible picker tab: yellow input cell in B1 (dropdown of the queue's
    orders; typing any order # also works) and one FILTER formula that spills
    that order's ranked lookalikes from the Similar Data sheet."""
    sh = Sheet(SIMILAR_ORDERS_TAB, freeze="A4")
    if n_queue:
        sh.picker = {
            "cell": SIMILAR_PICKER_CELL,
            "source": f"='{SIMILAR_DATA_TAB}'!$I$2:$I${n_queue + 1}",
            "comment": ("Pick a queue order from the dropdown (or type any order "
                        "#) and press Enter — its most similar past orders appear "
                        "below, best match first. Clear the cell to clear the list."),
        }
    sh.row([Cell("Order:", font=F_SECTION), Cell(""),
            Cell("← pick a queue order (or type any order #) — its most similar "
                 "past orders appear below, best match first",
                 font=F_NOTE, overflow=True)])
    sh.blank()
    sh.row(_header_cells(SIMILAR_HEADERS))
    if n_rows:
        last = n_rows + 1
        d = SIMILAR_DATA_TAB
        sh.row([Cell(f"=IFERROR(FILTER('{d}'!$B$2:$G${last},"
                     f"('{d}'!$A$2:$A${last}&\"\")=($B$1&\"\"),"
                     f"\"no matches for that order # yet\"),\"\")")])
    else:
        sh.row([Cell("No similar-order data yet — each order gets its list as "
                     "it is enriched.", font=F_NOTE, overflow=True)])
    return sh


# --------------------------------------------------------------------------- #
# Sales Order: pick a queue order at the top -> everything captured from its   #
# Sales Order appears instantly — the parsed spec fields (plus an Open-PDF     #
# link and the CO history) up top, and EVERY captured line item below. Same    #
# no-macro mechanics as Similar Orders: the visible tab holds only the picker  #
# and FILTER/INDEX formulas that spill from the flat SO Data sheet, so picking #
# an order never waits for the next poll and repaints never wipe the pick.     #
# --------------------------------------------------------------------------- #
SALES_ORDER_TAB = "Sales Order"
SO_DATA_TAB = "SO Data"
SO_PICKER_CELL = "B1"
# One row per captured line item (see line_items.extract_items): its position on
# the SO, the section it printed under, qty, the raw line exactly as printed,
# price / L-C-N type letter, the unpriced detail lines, the canonical feature
# tags, the normalized form the matching runs on, and any parser review flags —
# everything needed to judge how a line was captured and sorted.
SO_ITEM_HEADERS = ["#", "Section", "Qty", "Description", "Price", "L/C/N",
                   "Details", "Tags", "Normalized", "Review"]
# The parsed SO spec fields shown in the tab's one-row summary (header, job key);
# "co" renders via _co_label like every other tab.
SO_SUMMARY_COLUMNS = [
    ("Customer", "customer"), ("CO#", "co"), ("Total Price", "total_price"),
    ("Design", "design"), ("Description", "so_design_desc"), ("Size", "so_size"),
    ("Arrangement", "so_arrangement"), ("Motor Pos", "so_motor_pos"),
    ("Class", "so_class"), ("Rotation", "so_rotation"),
    ("Discharge", "so_discharge"), ("% Width", "so_pct_width"),
    ("Wheel Type", "so_wheel_type"), ("Design Temp", "so_design_temp"),
    ("Max Temp", "so_max_temp"), ("Special Temp", "so_special_temp"),
    ("Primary Rep", "primary_rep"),
]
# The component-hierarchy block (see so_hierarchy.tree_rows): the rolled-up
# tree the tab shows beside the flat capture table. Item # cross-references
# the flat table's '#' column; Kind identifies the row's role
# (COMPONENT/ATTRIBUTE/REVIEW/SOURCE).
SO_TREE_HEADERS = ["Hierarchy", "Price", "Kind", "Item #"]
# SO Data column geometry (1-based), shared by both builders so the visible
# tab's formulas always aim at the right data columns: the line-item block
# (Queue Order + SO_ITEM_HEADERS), a blank gap, the per-order summary block
# (Queue Order — the picker's dropdown source — the spec fields, the CO
# history, and the SO PDF path the Open-PDF link looks up), a blank gap, then
# the hierarchy block (Queue Order + SO_TREE_HEADERS).
_SO_ITEM_NCOLS = 1 + len(SO_ITEM_HEADERS)
_SO_KEY_COL = _SO_ITEM_NCOLS + 2
_SO_SUM_FIRST = _SO_KEY_COL + 1
_SO_SUM_LAST = _SO_SUM_FIRST + len(SO_SUMMARY_COLUMNS) - 1
_SO_CO_HIST_COL = _SO_SUM_LAST + 1
_SO_PDF_COL = _SO_CO_HIST_COL + 1
_SO_TREE_KEY_COL = _SO_PDF_COL + 2
_SO_TREE_FIRST = _SO_TREE_KEY_COL + 1
_SO_TREE_LAST = _SO_TREE_KEY_COL + len(SO_TREE_HEADERS)


def _job_num_key(job: str) -> tuple:
    """Sort key: numeric job numbers ascending, then anything else."""
    return (0, int(job), job) if job.isdigit() else (1, 0, job)


def sales_order_item_rows(job: Dict[str, Any]) -> List[List[Any]]:
    """One flat value row per captured line item, in SO print order, matching
    SO_ITEM_HEADERS."""
    rows = []
    for i, it in enumerate(job.get("line_items") or [], start=1):
        rows.append([i, it.get("section", ""), it.get("qty", ""), it.get("raw", ""),
                     it.get("price", ""), it.get("ptype", ""),
                     "; ".join(str(d) for d in it.get("details") or []),
                     ", ".join(it.get("tags") or []),
                     it.get("norm", ""),
                     "; ".join(it.get("review_flags") or [])])
    return rows


def _so_summary_values(j: Dict[str, Any]) -> List[Any]:
    return [_co_label(j) if key == "co" else j.get(key, "")
            for _h, key in SO_SUMMARY_COLUMNS]


def sales_order_data_sheet(jobs: List[Dict[str, Any]]) -> Sheet:
    """The flat table behind the Sales Order tab: the line-item block on the
    left (one row per captured item, grouped by order), one summary row per
    on-board order in the middle — the parsed spec fields plus the CO history
    and the SO PDF path — and the component-hierarchy block on the right (one
    row per so_hierarchy tree row). The summary block's Queue Order column
    doubles as the picker's dropdown source. Ordered by JOB NUMBER, not board
    position, so the board reshuffling every poll doesn't repaint this tab."""
    import so_hierarchy

    jobs = sorted(jobs, key=lambda j: _job_num_key(str(j.get("job") or "")))
    item_rows: List[List[Any]] = []
    tree_vals: List[List[Any]] = []
    tree_first: set = set()                    # each order's first tree row (banded)
    for j in jobs:
        jn = str(j.get("job") or "")
        for r in sales_order_item_rows(j):
            item_rows.append([jn] + r)
        first = True
        for r in so_hierarchy.tree_rows(j.get("line_items") or []):
            if first:
                tree_first.add(len(tree_vals))
                first = False
            tree_vals.append([jn, so_hierarchy.indent_text(r), r["price"],
                              r["kind"], r["item_no"] or ""])

    sh = Sheet(SO_DATA_TAB, freeze="A2")
    sh.row(_header_cells(["Queue Order"] + SO_ITEM_HEADERS) + [Cell("")]
           + _header_cells(["Queue Order"] + [h for h, _ in SO_SUMMARY_COLUMNS]
                           + ["CO History", "SO PDF"]) + [Cell("")]
           + _header_cells(["Queue Order"] + SO_TREE_HEADERS))
    n_summary = 1 + len(SO_SUMMARY_COLUMNS) + 2    # key + fields + CO hist + PDF
    prev = None
    for i in range(max(len(item_rows), len(jobs), len(tree_vals))):
        if i < len(item_rows):
            vals = item_rows[i]
            first = vals[0] != prev            # grey band starts each order's group
            prev = vals[0]
            fill = FILL_NEW if first else None
            cells = [Cell(v, fill=fill) for v in vals]
        else:
            cells = [Cell("")] * _SO_ITEM_NCOLS
        cells.append(Cell(""))
        if i < len(jobs):
            j = jobs[i]
            cells += ([Cell(str(j.get("job") or ""))]
                      + [Cell(v) for v in _so_summary_values(j)]
                      + [Cell(" | ".join(str(x) for x in j.get("co_history") or [])),
                         Cell((j.get("so_pdf") or "").strip())])
        elif i < len(tree_vals):               # pad so the tree block stays aligned
            cells += [Cell("")] * n_summary
        if i < len(tree_vals):
            fill = FILL_NEW if i in tree_first else None
            cells += [Cell("")] + [Cell(v, fill=fill) for v in tree_vals[i]]
        sh.row(cells)
    return sh


def sales_order_sheet(n_data_rows: int, n_orders: int) -> Sheet:
    """The visible Sales Order tab: yellow input cell in B1 (dropdown of the
    queue's orders; typing an order # also works), a one-row spill of the
    order's parsed SO spec — with an Open-PDF link and its CO history — then
    the line items TWICE: the component hierarchy (so_hierarchy's rollup) on
    the left and the flat capture table on the right, cross-referenced by
    item #. `n_data_rows` is the SO Data sheet's data row count (its nrows
    minus the header); `n_orders` the number of on-board orders (= summary
    rows)."""
    from openpyxl.utils import get_column_letter

    sh = Sheet(SALES_ORDER_TAB, freeze="A9")   # pin the pick + summary + item headers
    d = SO_DATA_TAB
    key = get_column_letter(_SO_KEY_COL)
    olast = n_orders + 1
    if n_orders:
        sh.picker = {
            "cell": SO_PICKER_CELL,
            "source": f"='{d}'!${key}$2:${key}${olast}",
            "comment": ("Pick a queue order from the dropdown (or type its order "
                        "#) and press Enter — everything captured from its Sales "
                        "Order (the spec fields and every line item) appears "
                        "below. Clear the cell to clear the view."),
        }
    sh.row([Cell("Order:", font=F_SECTION), Cell(""),
            Cell("← pick a queue order (or type its order #) — everything from "
                 "its Sales Order appears below, line items included",
                 font=F_NOTE, overflow=True)])
    sh.blank()
    if not n_data_rows:
        sh.row([Cell("No sales-order data yet — an order appears here once its "
                     "Sales Order has been fetched and parsed.",
                     font=F_NOTE, overflow=True)])
        return sh

    last = n_data_rows + 1
    sum1, sum2 = get_column_letter(_SO_SUM_FIRST), get_column_letter(_SO_SUM_LAST)
    hist, pdf = get_column_letter(_SO_CO_HIST_COL), get_column_letter(_SO_PDF_COL)
    # Both sides of every comparison/lookup are coerced to text (&"") so a job
    # stored as a number still matches digits typed in the picker, and the whole
    # view collapses to blank while the picker is empty (no all-rows spill).
    lookup = f"MATCH($B$1&\"\",'{d}'!${key}$2:${key}${olast}&\"\",0)"

    sh.row([Cell("Order details", font=F_SECTION)])
    sh.row(_header_cells([h for h, _ in SO_SUMMARY_COLUMNS]
                         + ["Sales Order PDF", "CO History"]))
    summary = [Cell(f"=IF($B$1&\"\"=\"\",\"\",IFERROR("
                    f"FILTER('{d}'!${sum1}$2:${sum2}${olast},"
                    f"('{d}'!${key}$2:${key}${olast}&\"\")=($B$1&\"\"),"
                    f"\"no data for that order # (not on the queue?)\"),\"\"))")]
    summary += [Cell("") for _ in range(len(SO_SUMMARY_COLUMNS) - 1)]  # spill room
    # Blank stays blank when the order has no PDF path on file — a HYPERLINK to
    # "" would render a link that errors on click.
    pdf_rng = f"'{d}'!${pdf}$2:${pdf}${olast}"
    summary.append(Cell(f"=IF($B$1&\"\"=\"\",\"\",IFERROR("
                        f"IF(INDEX({pdf_rng},{lookup})&\"\"=\"\",\"\","
                        f"HYPERLINK(INDEX({pdf_rng},{lookup}),\"Open PDF\")),\"\"))",
                        font=F_LINK))
    # CO history last, so the free text overruns the empty cells to its right.
    summary.append(Cell(f"=IF($B$1&\"\"=\"\",\"\",IFERROR("
                        f"INDEX('{d}'!${hist}$2:${hist}${olast},{lookup}),\"\"))",
                        overflow=True))
    sh.row(summary)
    sh.blank()

    # Line items, two synchronized views: the component HIERARCHY on the left
    # (so_hierarchy's rollup — one group per component, its attributes, review
    # flags and source lines beneath it) and the flat CAPTURE table on the right
    # (one row per stored item, exactly as parsed). The tree's 'Item #' column
    # matches the flat table's '#' column, so a funky tree row can be traced
    # straight to the captured line that produced it.
    flat_c0 = len(SO_TREE_HEADERS) + 2         # flat block starts after a gap col
    title = [Cell("Line items — hierarchy", font=F_SECTION)]
    title += [Cell("")] * (flat_c0 - len(title) - 1)
    title.append(Cell("Captured lines (flat)", font=F_SECTION))
    sh.row(title)
    sh.row(_header_cells(SO_TREE_HEADERS) + [Cell("")] + _header_cells(SO_ITEM_HEADERS))
    tkey = get_column_letter(_SO_TREE_KEY_COL)
    t1, t2 = get_column_letter(_SO_TREE_FIRST), get_column_letter(_SO_TREE_LAST)
    item2 = get_column_letter(_SO_ITEM_NCOLS)
    tree_f = Cell(f"=IF($B$1&\"\"=\"\",\"\",IFERROR("
                  f"FILTER('{d}'!${t1}$2:${t2}${last},"
                  f"('{d}'!${tkey}$2:${tkey}${last}&\"\")=($B$1&\"\"),"
                  f"\"no line items captured for that order # yet\"),\"\"))")
    flat_f = Cell(f"=IF($B$1&\"\"=\"\",\"\",IFERROR("
                  f"FILTER('{d}'!$B$2:${item2}${last},"
                  f"('{d}'!$A$2:$A${last}&\"\")=($B$1&\"\"),"
                  f"\"no line items captured for that order # yet\"),\"\"))")
    sh.row([tree_f] + [Cell("")] * (flat_c0 - 2) + [flat_f])
    return sh


# --------------------------------------------------------------------------- #
# Incremental "master log" tabs (Live Queue + Order History)                   #
#                                                                             #
# These are upserted row-by-row (keyed on the order number) instead of being   #
# repainted, so a coworker's filter/sort/scroll survives. Live Queue is the     #
# on-board board (churny fields, full formatting); Order History is the stable  #
# log of every order with its two ✓/red matrices (built below). A record is     #
# (order#, [Cell, ...]); the renderer writes/append/updates the row whose key   #
# matches.                                                                     #
# --------------------------------------------------------------------------- #
# "Added" (last time it came onto the board), then the Full Queue columns except
# DWG Reuse, then the utility columns, and finally DWG Reuse at the far right.
# The tab is still sorted by "#" so the order matches cbcinsider; moving that sort
# key away from the edge does not change the renderer's behavior.
_DWG_REUSE_HEADER = "DWG Reuse"
_DWG_REUSE_IDX = QUEUE_HEADERS.index(_DWG_REUSE_HEADER)
_LIVE_QUEUE_STANDARD_HEADERS = [
    _abbrev_header(h) for h in QUEUE_HEADERS if h != _DWG_REUSE_HEADER
]
_LIVE_QUEUE_UTILITY_HEADERS = ["Last Out", "Similar", "#"]
LIVE_QUEUE_HEADERS = (["Added"] + _LIVE_QUEUE_STANDARD_HEADERS
                      + _LIVE_QUEUE_UTILITY_HEADERS + [_DWG_REUSE_HEADER])
LIVE_QUEUE_CBC_COL = LIVE_QUEUE_HEADERS.index("#") + 1          # board-position sort key
LIVE_QUEUE_SIMILAR_COL = LIVE_QUEUE_HEADERS.index("Similar") + 1
LIVE_QUEUE_LAST_OUT_COL = LIVE_QUEUE_HEADERS.index("Last Out") + 1
LIVE_QUEUE_DWG_REUSE_COL = LIVE_QUEUE_HEADERS.index(_DWG_REUSE_HEADER) + 1
LIVE_QUEUE_KEY_COL = 2 + QUEUE_HEADERS.index("Job #")          # 1-based col of Job # (Added is col 1)
LIVE_QUEUE_END_DATE_COL = 2 + QUEUE_HEADERS.index("End Date")  # 1-based col of End Date
# The 'Removed since this morning' block lines its data columns up under the board
# above, but LEADS with the removal time (the "Removed" column) instead of "Added"
# — the most relevant fact for an order that just left. Blank utility headers keep
# DWG Reuse aligned with the main grid's far-right column.
LIVE_QUEUE_REMOVED_HEADERS = (["Removed"] + _LIVE_QUEUE_STANDARD_HEADERS
                              + [""] * len(_LIVE_QUEUE_UTILITY_HEADERS)
                              + [_DWG_REUSE_HEADER])
_END_DATE_IDX = QUEUE_HEADERS.index("End Date")               # its index within the standard cells
_CO_IDX = [i for i, (_, k) in enumerate(COLUMNS) if k == "co"][0]   # CO# cell index


def _co_comment(job: Dict[str, Any]) -> Optional[str]:
    """A hover note for the CO# cell: the order's change-order history, most
    recent first, so hovering shows what the latest change order was."""
    hist = job.get("co_history") or []
    if not hist:
        return None

    def _co_num(line: str) -> int:
        m = re.search(r"C\s*/?\s*O\s*#?\s*(\d+)", line, re.I)
        return int(m.group(1)) if m else 0

    ordered = sorted(hist, key=_co_num, reverse=True)
    return "Change orders (most recent first):\n" + "\n".join(ordered)


_EXCEL_EPOCH = date(1899, 12, 30)   # Excel's day 0 (1900 date system)


def _excel_serial(d: date) -> int:
    """An Excel date serial (days since 1899-12-30). Written as a plain int with a
    date number-format so the cell sorts chronologically and shows as a date —
    avoids passing a Python date through COM (pywin32 won't marshal a bare date)."""
    return (d - _EXCEL_EPOCH).days


def _fmt_dt(iso: Optional[str]) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso).strftime("%Y-%m-%d %H:%M")
    except (ValueError, TypeError):
        return iso


def row_sig(cells: List[Cell]) -> str:
    """A stable string signature of a row's content+style, so the renderer can
    tell whether a row actually changed and skip writing it if not. Returned as a
    hex digest (not a tuple) so it survives a JSON round-trip in the master store
    — change detection has to work across watcher restarts.

    Volatile cells contribute their STYLE but not their value: the Live Queue's
    '#' board position jitters on most scrapes (the site reorders ties), and
    hashing it made 5-15 full row rewrites per poll for rows whose data hadn't
    changed — most of the render cost. The renderer refreshes those columns in
    one bulk write instead (apply_upserts `positions`)."""
    payload = [["\x00VOLATILE" if c.volatile else c.value,
                c.fill, c.font, c.link, c.number_format, c.center] for c in cells]
    return hashlib.md5(json.dumps(payload, default=str, sort_keys=True).encode()).hexdigest()


def _board_row_cells(j: Dict[str, Any], today: date, co_changed: bool, is_new: bool,
                     ref: Optional[datetime], leading: Cell, trailing: List[Cell]) -> List[Cell]:
    """`leading` cell + every Full Queue column + `trailing` cells, styled exactly
    like a Live Queue row: real-date End Date, CO# hover note, CO#-red text, and the
    urgency / new-today row fill. `leading` is the Added time on the board or the
    removal time in the Removed block; `trailing` is [Last Out, Similar, '#'] on
    the board and [] in the Removed block. DWG Reuse is always moved after those
    cells so it remains the far-right column."""
    std = _job_value_cells(j, co_changed=co_changed, arrange_comment=True)
    # Write End Date as a real date so it sorts/filters as a date in Excel.
    ed = _parse_date(j.get("end_date", ""))
    if ed is not None:
        std[_END_DATE_IDX].value = _excel_serial(ed)
        std[_END_DATE_IDX].number_format = "mm/dd/yyyy"
    # Hover the CO# cell to see the change-order history (most recent first).
    std[_CO_IDX].comment = _co_comment(j)
    if co_changed:                       # carry the red over to the non-std cells
        leading.font = F_RED
        for t in trailing:
            t.font = F_RED
    dwg_reuse = std.pop(_DWG_REUSE_IDX)
    cells = [leading] + std + list(trailing) + [dwg_reuse]
    fill = _row_fill(j, today, is_new=is_new)
    if fill:
        for c in cells:
            if c.fill is None:
                c.fill = fill
    return cells


def removed_block(removed: List, today: date, new_ids: Optional[set] = None,
                  co_changed_ids: Optional[set] = None, ref: Optional[datetime] = None,
                  title: str = "Removed from the queue since this morning") -> Dict[str, Any]:
    """The Live Queue's 'Removed since this morning' section as a styled block —
    each removed order drawn exactly like its Live Queue row (same columns, the
    urgency/new-today fill and CO#-red text it had on the board). The leading Added
    slot becomes the removal time; utility cells are blank, and DWG Reuse remains
    aligned at the far right. `removed` is a list of (job_dict, left_iso). Returns
    a payload the COM renderer draws below the board."""
    new_ids = new_ids or set()
    co_changed_ids = co_changed_ids or set()
    rows = []
    for j, left in removed:
        jn = str(j.get("job") or "")
        rem = Cell(fmt_time(left), number_format="@")   # when it left -> the LEADING column
        utility_spacers = [Cell("") for _ in _LIVE_QUEUE_UTILITY_HEADERS]
        rows.append(_board_row_cells(j, today, jn in co_changed_ids, jn in new_ids, ref,
                                     leading=rem, trailing=utility_spacers))
    return {"title": title,
            "header_cells": _header_cells(LIVE_QUEUE_REMOVED_HEADERS),
            "rows": rows}


def live_queue_records(jobs: List[Dict[str, Any]], today: date,
                       new_ids: Optional[set] = None,
                       co_changed_ids: Optional[set] = None,
                       ref: Optional[datetime] = None) -> List:
    """(order#, cells) per on-board order: Added + every Full Queue column + the
    "#" board position, with urgency / new-today row fills and hyperlinks.
    `new_ids` is the set of order numbers new today (not in the previous
    snapshot); `co_changed_ids` is the set that had a change order (CO#) land
    today — their text goes red."""
    new_ids = new_ids or set()
    co_changed_ids = co_changed_ids or set()
    out = []
    for j in jobs:
        jn = str(j.get("job") or "")
        added = Cell(added_label(j, ref=ref), number_format="@")
        last_out = Cell(last_out_label(j, ref=ref), number_format="@")   # most recent prior departure
        # "Similar" = how many lookalike past orders this one has; clicking jumps
        # to its group on the Similar Data tab (watch stamps _sim_count/_sim_anchor).
        sim = Cell(j.get("_sim_count") or "", center=True)
        if sim.value and j.get("_sim_anchor"):
            sim.link, sim.font = j["_sim_anchor"], F_LINK
        # "#" = the cbcinsider board position (the sort key for board order).
        # Volatile: position jitter must not force a full row rewrite — the
        # renderer bulk-refreshes this column each cycle instead.
        pos = j.get("_cbc_pos")
        cbc = Cell(pos if isinstance(pos, int) else "", center=True, volatile=True)
        cells = _board_row_cells(j, today, jn in co_changed_ids, jn in new_ids, ref,
                                 leading=added, trailing=[last_out, sim, cbc])
        out.append((jn, cells))
    return out


# Order History is a stable *log*: only the order's identity + SO spec + the two
# matrices + the presence flags — NOT the churny board fields (dates, price,
# assignee, status), which live on Live Queue. Keeping only stable columns means
# a row's signature changes solely when the order is added or its On Queue/Left
# flags flip, so the 12K-row log isn't rewritten on every field tick. Layout:
# the data columns first (Job # leads, so it's the pinned column), then the
# On Queue / Added / Left flags right before the two ✓/red matrices.
OH_DATA_COLUMNS = [
    ("Job #", "job"), ("Folder", "folder"), ("Quote Run", "drive_run"), ("CO#", "co"),
    ("Design", "design"), ("Description", "so_design_desc"), ("Size", "so_size"),
    ("Arrangement", "so_arrangement"), ("Motor Pos", "so_motor_pos"), ("Class", "so_class"),
    ("Rotation", "so_rotation"), ("Discharge", "so_discharge"), ("% Width", "so_pct_width"),
    ("Wheel Type", "so_wheel_type"), ("Design Temp", "so_design_temp"),
    ("Max Temp", "so_max_temp"), ("Special Temp", "so_special_temp"),
    ("Customer", "customer"), ("Engineer", "engineers"),
    ("Primary Rep", "primary_rep"), ("Item", "item"),
]
OH_FLAG_HEADERS = ["On Queue", "Added", "Left"]
OH_SEP_HEADER = "│"
ORDER_HISTORY_KEY_COL = 1   # Job # is the first column (pinned)


def _order_tags(j: Dict[str, Any]) -> set:
    """The canonical feature tags on an order, from its captured line items."""
    out = set()
    for it in j.get("line_items") or []:
        for t in it.get("tags") or []:
            out.add(t)
    return out


def _suffix_key(s: str):
    return (int(s), s) if str(s).isdigit() else (10 ** 9, s)


# Fields the Order History rows pull straight from the job dict (cell value =
# j.get(key)). Derived from OH_DATA_COLUMNS so a new PLAIN column is folded into
# the fingerprint automatically; the computed columns (job/folder/co/drive_run/
# engineers) are handled explicitly in order_history_fingerprint. Keep the two in
# step — test_order_history_fingerprint_* guards them against drift.
_OH_FINGERPRINT_COMPUTED = ("job", "folder", "co", "drive_run", "engineers")
_OH_PLAIN_KEYS = [k for _, k in OH_DATA_COLUMNS if k not in _OH_FINGERPRINT_COMPUTED]


def order_history_fingerprint(orders: List) -> str:
    """A cheap, EXACT digest of everything the Order History tab is built from —
    the order set and each order's stable spec, its On Queue/Added/Left flags,
    and the two ✓/red matrices (custom DWGs, feature tags). Lets the poll loop
    skip the ~9s order_history_build when nothing that shows on the tab changed.

    By construction it excludes the churny board fields (price, dates, assignee,
    status, notes) — they are not in OH_DATA_COLUMNS — so a normal intraday tick
    never forces a rebuild. It DOES move when an order is added/departs/returns,
    a CO# or SO-spec value lands, an engineer is (re)assigned, or a new DWG
    suffix / feature tag appears (which is also what grows the matrix columns).
    `orders` is the (job#, entry) list order_history_build consumes."""
    h = hashlib.md5()
    for jn, e in orders:
        j = e.get("job") or {}
        h.update(str(jn).encode())
        h.update(("\x00%s|%s|%s|%s|%s|%s|%s|%s|%s|%s" % (
            e.get("on_queue"), e.get("added"), e.get("left"),
            j.get("co_number"), j.get("job_folder") or "", j.get("job_type") or "",
            j.get("has_drive_run"), j.get("drive_run_pdf") or "",
            j.get("drive_run_count"), engineers.cell_text(j))).encode())
        for k in _OH_PLAIN_KEYS:
            h.update(b"|")
            h.update(str(j.get(k) or "").encode())
        h.update(("|" + ",".join(sorted(j.get("dwg_extras") or {}))).encode())
        tags = sorted({t for it in (j.get("line_items") or [])
                       for t in (it.get("tags") or [])})
        h.update(("|" + ",".join(tags) + "\n").encode())
    return h.hexdigest()


def order_history_build(orders: List, today: date,
                        prev_columns: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Build the Order History tab: one row per order with the stable data
    columns (Job # first), then the On Queue / Added / Left flags, then the
    AutoCAD **DWG matrix** (green ✓ / red), a vertical divider, and the line-item
    **Feature matrix** (green ✓ / red).

    Column order is STABLE so the tab isn't rebuilt in normal operation: it keeps
    `prev_columns`' order and only APPENDS brand-new suffixes/tags at the end (it
    never drops or reorders). `spec["columns"]` is the order used — the caller
    persists it; the headers change only when a new suffix/tag appears, which is
    the only time the tab is rebuilt.

    `orders` is a list of (job#, entry) with entry = {on_queue, added, left, job}.
    Returns headers, records [(key, cells)], the matrix/separator column ranges,
    and the column order."""
    jobs = [e.get("job", {}) for _, e in orders]
    cur_suffixes = set(_dwg_suffixes(jobs))
    cur_tags: set = set()
    for j in jobs:
        cur_tags |= _order_tags(j)

    prev = prev_columns or {}
    prev_suffixes = list(prev.get("suffixes") or [])
    prev_tags = list(prev.get("tags") or [])
    suffixes = prev_suffixes + sorted(cur_suffixes - set(prev_suffixes), key=_suffix_key)
    tags = prev_tags + sorted(cur_tags - set(prev_tags))
    columns = {"suffixes": suffixes, "tags": tags}

    headers = ([h for h, _ in OH_DATA_COLUMNS] + OH_FLAG_HEADERS
               + [f"-{s}" for s in suffixes] + [OH_SEP_HEADER] + list(tags))
    n_data_flags = len(OH_DATA_COLUMNS) + len(OH_FLAG_HEADERS)
    dwg_range = (n_data_flags + 1, n_data_flags + len(suffixes))
    sep_col = n_data_flags + len(suffixes) + 1
    feat_range = (sep_col + 1, sep_col + len(tags))

    records = []
    for jn, e in orders:
        j = e.get("job", {})
        onq = bool(e.get("on_queue"))
        data = _job_value_cells(j, columns=OH_DATA_COLUMNS)
        flags = [Cell("YES" if onq else "NO"),
                 Cell(_fmt_dt(e.get("added"))), Cell(_fmt_dt(e.get("left")))]
        de = j.get("dwg_extras") or {}
        # Matrix cells carry only the ✓ / blank value; the renderer colors them
        # via conditional formatting (cheap enough for ~12K rows).
        dwg = [Cell("✓" if s in de else "", center=True) for s in suffixes]
        sep = [Cell("", fill=FILL_SEP)]
        otags = _order_tags(j)
        feat = [Cell("✓" if t in otags else "", center=True) for t in tags]
        records.append((str(jn), data + flags + dwg + sep + feat))

    return {"headers": headers, "records": records, "columns": columns,
            "dwg_range": dwg_range, "sep_col": sep_col, "feat_range": feat_range}


def plan_upsert(desired: List, existing_sigs: Dict[str, tuple],
                allow_delete: bool = False) -> List:
    """Diff desired rows against what's already on the sheet (by key) and return
    the minimal op list: ('append', key, cells) for keys not present,
    ('update', key, cells) for keys whose signature changed, and (with
    allow_delete) ('delete', key, None) for keys no longer desired. Unchanged
    rows produce no op — so they're never rewritten and filters/scroll persist.

    `desired` is a list of (key, sig, cells); `existing_sigs` is {key: sig}.
    """
    ops, seen = [], set()
    for key, sig, cells in desired:
        if not key:
            continue
        seen.add(key)
        prev = existing_sigs.get(key)
        if prev is None:
            ops.append(("append", key, cells))
        elif prev != sig:
            ops.append(("update", key, cells))
    if allow_delete:
        for key in existing_sigs:
            if key not in seen:
                ops.append(("delete", key, None))
    return ops


